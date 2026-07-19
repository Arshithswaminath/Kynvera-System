"""
Operations Module — Over Time records + Trading Invoices (with Client master).
URL prefix: /operations

v1 is simple CRUD (no approval workflow):
  - Over Time: manual add / list / edit / delete + Excel import + downloadable template.
  - Clients: customer master CRUD (used by trading invoices).
  - Trading Invoices: header + line items, server-computed totals, branded PDF.
"""
import uuid
import logging
from io import BytesIO
from datetime import datetime, date

from flask import (
    Blueprint, render_template, redirect, request, send_file, current_app,
)
from flask_jwt_extended import jwt_required, get_jwt_identity
from sqlalchemy import func, case

from app.models import (
    db, User, OvertimeRecord, OvertimeSettings, Client, TradingInvoice, TradingInvoiceItem,
    ChequeRequest, ChequeRequestItem, ChequeStatusLog, ChequeNotificationConfig,
    CHEQUE_STATUSES,
)
from common.error_responses import success_response, error_response

logger = logging.getLogger(__name__)

operations_bp = Blueprint('operations_bp', __name__, url_prefix='/operations',
                          template_folder='templates')


# ── Access control ───────────────────────────────────────────────────────────

def _has_operations_access(user):
    if not user:
        return False
    if user.role == 'admin':
        return True
    if bool(getattr(user, 'access_operations', False)):
        return True
    if hasattr(user, 'has_any_operations_submodule'):
        return bool(user.has_any_operations_submodule())
    return False


def _has_ops_sub_access(user, sub):
    """sub: overtime | invoices | clients | cheques"""
    if not user:
        return False
    if hasattr(user, 'has_operations_submodule'):
        return bool(user.has_operations_submodule(sub))
    if user.role == 'admin':
        return True
    return bool(getattr(user, 'access_operations', False))


def _can_manage_invoices(user):
    """Who may create/edit/delete trading invoices (vs. view-only).

    Admins always have full access. Everyone else needs Operations access AND
    the per-user "full access" flag set by an admin; otherwise they are view-only."""
    if not user:
        return False
    if user.role == 'admin':
        return True
    return _has_operations_access(user) and bool(getattr(user, 'access_operations_manage', False))


def _can_edit_overtime_settings(user):
    """Admin or operations-manage users may edit OT weekday/weekend rates."""
    return _can_manage_invoices(user) and _has_ops_sub_access(user, 'overtime')


def _current_user():
    return User.query.get(get_jwt_identity())


# ── Helpers ──────────────────────────────────────────────────────────────────

def _gen_id(prefix):
    return f'{prefix}-' + uuid.uuid4().hex[:8].upper()


def _to_float(value, default=None):
    if value is None or value == '':
        return default
    try:
        return float(str(value).replace(',', '').strip())
    except (ValueError, TypeError):
        return default


def _parse_date(value):
    """Parse a date from ISO, DD/MM/YYYY, or a datetime/date object."""
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # Last resort: ISO with time component
    try:
        return datetime.fromisoformat(s).date()
    except ValueError:
        return None


def _normalize_overtime_settings(raw):
    """Merge and validate OT settings against defaults."""
    defaults = dict(OvertimeSettings.DEFAULT_CONFIG)
    cfg = dict(defaults)
    if isinstance(raw, dict):
        cfg.update(raw)

    weekday_rate = _to_float(cfg.get('weekday_rate'), defaults['weekday_rate'])
    weekend_rate = _to_float(cfg.get('weekend_rate'), defaults['weekend_rate'])
    if weekday_rate is None or weekday_rate < 0:
        weekday_rate = defaults['weekday_rate']
    if weekend_rate is None or weekend_rate < 0:
        weekend_rate = defaults['weekend_rate']

    weekend_days = cfg.get('weekend_days', defaults['weekend_days'])
    if not isinstance(weekend_days, (list, tuple)):
        weekend_days = defaults['weekend_days']
    cleaned_days = []
    for d in weekend_days:
        try:
            di = int(d)
        except (TypeError, ValueError):
            continue
        if 0 <= di <= 6 and di not in cleaned_days:
            cleaned_days.append(di)
    if not cleaned_days:
        cleaned_days = list(defaults['weekend_days'])

    return {
        'weekday_rate': float(weekday_rate),
        'weekend_rate': float(weekend_rate),
        'weekend_days': cleaned_days,
    }


def _get_overtime_settings():
    """Load single-row OT settings, seeding defaults if missing."""
    row = OvertimeSettings.query.first()
    if not row:
        row = OvertimeSettings(config_json=dict(OvertimeSettings.DEFAULT_CONFIG))
        db.session.add(row)
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            row = OvertimeSettings.query.first()
            if not row:
                return _normalize_overtime_settings(None)
    return _normalize_overtime_settings(row.config_json or {})


def _is_weekend(ot_date, settings=None):
    settings = settings or _get_overtime_settings()
    if not ot_date:
        return False
    return ot_date.weekday() in set(settings.get('weekend_days') or [])


def _rate_for_date(ot_date, settings=None):
    settings = settings or _get_overtime_settings()
    if _is_weekend(ot_date, settings):
        return float(settings['weekend_rate'])
    return float(settings['weekday_rate'])


def _day_type_for_date(ot_date, settings=None):
    return 'weekend' if _is_weekend(ot_date, settings) else 'weekday'


def _enrich_overtime_dict(rec, settings=None):
    data = rec.to_dict()
    settings = settings or _get_overtime_settings()
    day_type = _day_type_for_date(rec.date, settings) if rec.date else None
    data['day_type'] = day_type
    data['weekday_name'] = rec.date.strftime('%a') if rec.date else None
    return data


def _resolve_overtime_rate(ot_date, hours, data, settings=None):
    """Return (rate, total) using auto rate unless client sends an override.

    Auto-rate when:
      - use_auto_rate is true, OR
      - rate_per_hour is omitted / blank
    Explicit numeric rate_per_hour is treated as an override.
    """
    settings = settings or _get_overtime_settings()
    use_auto = data.get('use_auto_rate')
    if use_auto is True or str(use_auto).lower() in ('1', 'true', 'yes'):
        rate = _rate_for_date(ot_date, settings)
    elif 'rate_per_hour' not in data or data.get('rate_per_hour') in (None, ''):
        rate = _rate_for_date(ot_date, settings)
    else:
        rate = _to_float(data.get('rate_per_hour'))
    total = round(hours * rate, 2) if rate is not None and hours is not None else None
    return rate, total


# ── Hub ──────────────────────────────────────────────────────────────────────

@operations_bp.route('/')
@jwt_required()
def operations_dashboard():
    user = _current_user()
    if not _has_operations_access(user):
        return redirect('/dashboard')
    return render_template('operations_dashboard.html', user=user)


# ── Over Time pages ──────────────────────────────────────────────────────────

@operations_bp.route('/overtime')
@jwt_required()
def overtime_page():
    user = _current_user()
    if not _has_ops_sub_access(user, 'overtime'):
        return redirect('/dashboard')
    return render_template(
        'overtime.html',
        user=user,
        can_edit_settings=_can_edit_overtime_settings(user),
    )


@operations_bp.route('/api/overtime/settings', methods=['GET'])
@jwt_required()
def api_get_overtime_settings():
    user = _current_user()
    if not _has_ops_sub_access(user, 'overtime'):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
    settings = _get_overtime_settings()
    return success_response({
        'settings': settings,
        'can_edit': _can_edit_overtime_settings(user),
    })


@operations_bp.route('/api/overtime/settings', methods=['PUT'])
@jwt_required()
def api_put_overtime_settings():
    user = _current_user()
    if not _has_ops_sub_access(user, 'overtime'):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
    if not _can_edit_overtime_settings(user):
        return error_response('Only admins or operations managers can edit overtime settings',
                              status_code=403, error_code='ACCESS_DENIED')
    data = request.get_json(silent=True) or {}
    normalized = _normalize_overtime_settings(data.get('settings') or data)
    row = OvertimeSettings.query.first()
    if not row:
        row = OvertimeSettings(config_json=normalized)
        db.session.add(row)
    else:
        row.config_json = normalized
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error('put_overtime_settings: %s', e, exc_info=True)
        return error_response('Database error', status_code=500, error_code='DATABASE_ERROR')
    return success_response({'settings': normalized}, message='Overtime settings saved.')


@operations_bp.route('/api/overtime', methods=['GET'])
@jwt_required()
def api_list_overtime():
    user = _current_user()
    if not _has_ops_sub_access(user, 'overtime'):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
    settings = _get_overtime_settings()
    records = OvertimeRecord.query.order_by(OvertimeRecord.date.desc(),
                                            OvertimeRecord.id.desc()).all()
    return success_response({
        'records': [_enrich_overtime_dict(r, settings) for r in records],
        'settings': settings,
    })


@operations_bp.route('/api/overtime', methods=['POST'])
@jwt_required()
def api_create_overtime():
    user = _current_user()
    if not _has_ops_sub_access(user, 'overtime'):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
    data = request.get_json(silent=True) or {}

    staff_name = (data.get('staff_name') or '').strip()
    ot_date = _parse_date(data.get('date'))
    hours = _to_float(data.get('hours'))
    if not staff_name:
        return error_response('Staff name is required', status_code=400, error_code='VALIDATION_ERROR')
    if not ot_date:
        return error_response('A valid date is required', status_code=400, error_code='VALIDATION_ERROR')
    if hours is None:
        return error_response('Hours is required', status_code=400, error_code='VALIDATION_ERROR')

    settings = _get_overtime_settings()
    rate, total = _resolve_overtime_rate(ot_date, hours, data, settings)
    rec = OvertimeRecord(
        record_id=_gen_id('OT'),
        staff_name=staff_name,
        employee_id=(data.get('employee_id') or '').strip() or None,
        department=(data.get('department') or '').strip() or None,
        date=ot_date,
        hours=hours,
        rate_per_hour=rate,
        total_amount=total,
        reason=(data.get('reason') or '').strip() or None,
        status=(data.get('status') or 'recorded').strip() or 'recorded',
        created_by_id=user.id,
    )
    db.session.add(rec)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error('create_overtime: %s', e, exc_info=True)
        return error_response('Database error', status_code=500, error_code='DATABASE_ERROR')
    return success_response(
        {'record': _enrich_overtime_dict(rec, settings)},
        message='Overtime record created.',
        status_code=201,
    )


@operations_bp.route('/api/overtime/<int:rec_id>', methods=['PUT'])
@jwt_required()
def api_update_overtime(rec_id):
    user = _current_user()
    if not _has_ops_sub_access(user, 'overtime'):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
    rec = OvertimeRecord.query.get(rec_id)
    if not rec:
        return error_response('Record not found', status_code=404, error_code='NOT_FOUND')
    data = request.get_json(silent=True) or {}

    if 'staff_name' in data:
        name = (data.get('staff_name') or '').strip()
        if not name:
            return error_response('Staff name cannot be empty', status_code=400, error_code='VALIDATION_ERROR')
        rec.staff_name = name
    if 'employee_id' in data:
        rec.employee_id = (data.get('employee_id') or '').strip() or None
    if 'department' in data:
        rec.department = (data.get('department') or '').strip() or None
    if 'date' in data:
        d = _parse_date(data.get('date'))
        if not d:
            return error_response('Invalid date', status_code=400, error_code='VALIDATION_ERROR')
        rec.date = d
    if 'hours' in data:
        h = _to_float(data.get('hours'))
        if h is None:
            return error_response('Invalid hours', status_code=400, error_code='VALIDATION_ERROR')
        rec.hours = h
    if 'reason' in data:
        rec.reason = (data.get('reason') or '').strip() or None
    if 'status' in data:
        rec.status = (data.get('status') or 'recorded').strip() or 'recorded'

    settings = _get_overtime_settings()
    # Rebuild rate from settings unless an explicit override is provided
    rate_payload = {
        'use_auto_rate': data.get('use_auto_rate'),
        'rate_per_hour': data.get('rate_per_hour') if 'rate_per_hour' in data else None,
    }
    # If neither override flag nor rate sent, re-auto from current date (settings may have changed)
    if 'use_auto_rate' not in data and 'rate_per_hour' not in data:
        rate_payload['use_auto_rate'] = True
    rate, total = _resolve_overtime_rate(rec.date, rec.hours, rate_payload, settings)
    rec.rate_per_hour = rate
    rec.total_amount = total
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error('update_overtime: %s', e, exc_info=True)
        return error_response('Database error', status_code=500, error_code='DATABASE_ERROR')
    return success_response(
        {'record': _enrich_overtime_dict(rec, settings)},
        message='Overtime record updated.',
    )


@operations_bp.route('/api/overtime/<int:rec_id>', methods=['DELETE'])
@jwt_required()
def api_delete_overtime(rec_id):
    user = _current_user()
    if not _has_ops_sub_access(user, 'overtime'):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
    rec = OvertimeRecord.query.get(rec_id)
    if not rec:
        return error_response('Record not found', status_code=404, error_code='NOT_FOUND')
    db.session.delete(rec)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error('delete_overtime: %s', e, exc_info=True)
        return error_response('Database error', status_code=500, error_code='DATABASE_ERROR')
    return success_response(message='Overtime record deleted.')


@operations_bp.route('/api/overtime/import-excel', methods=['POST'])
@jwt_required()
def api_import_overtime_excel():
    user = _current_user()
    if not _has_ops_sub_access(user, 'overtime'):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
    try:
        import openpyxl
    except ImportError:
        return error_response('openpyxl is required for Excel import (pip install openpyxl)',
                              status_code=500, error_code='DEPENDENCY_ERROR')

    if 'file' not in request.files:
        return error_response('No file uploaded', status_code=400, error_code='VALIDATION_ERROR')
    f = request.files['file']
    if not f.filename:
        return error_response('No file selected', status_code=400, error_code='VALIDATION_ERROR')

    try:
        wb = openpyxl.load_workbook(BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return error_response(f'Could not read Excel file: {e}', status_code=400, error_code='PARSE_ERROR')

    if not rows:
        return error_response('Excel file is empty', status_code=400, error_code='VALIDATION_ERROR')

    header_row = [str(c or '').strip().lower().replace(' ', '_') for c in rows[0]]
    COL = {h: i for i, h in enumerate(header_row)}

    def _get(row, key, alt=None):
        idx = COL.get(key, COL.get(alt))
        if idx is None:
            return None
        v = row[idx] if idx < len(row) else None
        if v is None:
            return None
        s = str(v).strip()
        return s if s else None

    settings = _get_overtime_settings()
    created, skipped, errors = 0, 0, []
    for row_num, row in enumerate(rows[1:], start=2):
        if all(c is None or str(c).strip() == '' for c in row):
            continue
        staff_name = _get(row, 'staff_name', 'name')
        raw_date = _get(row, 'date')
        raw_hours = _get(row, 'hours')
        if not staff_name:
            errors.append(f'Row {row_num}: staff name is required — skipped')
            continue
        ot_date = _parse_date(raw_date)
        if not ot_date:
            errors.append(f'Row {row_num}: invalid/missing date — skipped')
            continue
        hours = _to_float(raw_hours)
        if hours is None:
            errors.append(f'Row {row_num}: invalid/missing hours — skipped')
            continue
        raw_rate = _get(row, 'rate_per_hour', 'rate')
        rate_payload = {'rate_per_hour': raw_rate} if raw_rate not in (None, '') else {'use_auto_rate': True}
        rate, total = _resolve_overtime_rate(ot_date, hours, rate_payload, settings)
        rec = OvertimeRecord(
            record_id=_gen_id('OT'),
            staff_name=staff_name,
            employee_id=_get(row, 'employee_id', 'emp_id'),
            department=_get(row, 'department'),
            date=ot_date,
            hours=hours,
            rate_per_hour=rate,
            total_amount=total,
            reason=_get(row, 'reason', 'notes'),
            status='recorded',
            imported_from_excel=True,
            created_by_id=user.id,
        )
        db.session.add(rec)
        created += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error('import_overtime_excel commit: %s', e, exc_info=True)
        return error_response('Database error during import', status_code=500, error_code='DATABASE_ERROR')

    return success_response(
        {'created': created, 'skipped': skipped, 'errors': errors},
        message=f'{created} overtime record(s) imported.',
    )


@operations_bp.route('/api/overtime/export', methods=['GET'])
@jwt_required()
def api_export_overtime_excel():
    user = _current_user()
    if not _has_ops_sub_access(user, 'overtime'):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return error_response('openpyxl is required (pip install openpyxl)',
                              status_code=500, error_code='DEPENDENCY_ERROR')

    records = OvertimeRecord.query.order_by(OvertimeRecord.date.desc(),
                                            OvertimeRecord.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Overtime'

    headers = ['Staff Name', 'Employee ID', 'Department', 'Date',
               'Hours', 'Rate Per Hour', 'Total Amount', 'Status', 'Reason']
    header_fill = PatternFill('solid', fgColor='a8121e')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    col_widths = [26, 16, 22, 14, 10, 14, 14, 14, 36]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 26

    row_fill = PatternFill('solid', fgColor='FFF5F5')
    data_font = Font(size=10)
    data_align = Alignment(vertical='center')

    for row_num, rec in enumerate(records, start=2):
        row_data = [
            rec.staff_name,
            rec.employee_id or '',
            rec.department or '',
            str(rec.date) if rec.date else '',
            rec.hours,
            rec.rate_per_hour,
            rec.total_amount,
            rec.status or '',
            rec.reason or '',
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border
            if row_num % 2 == 0:
                cell.fill = row_fill

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = date.today().strftime('%Y-%m-%d')
    return send_file(
        buf,
        as_attachment=True,
        download_name=f'overtime_export_{today}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@operations_bp.route('/api/overtime/template', methods=['GET'])
@jwt_required()
def api_overtime_template():
    user = _current_user()
    if not _has_ops_sub_access(user, 'overtime'):
        return redirect('/dashboard')
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return error_response('openpyxl is required (pip install openpyxl)',
                              status_code=500, error_code='DEPENDENCY_ERROR')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Overtime'

    headers = ['Staff Name', 'Employee ID', 'Department', 'Date',
               'Hours', 'Rate Per Hour', 'Reason']
    header_fill = PatternFill('solid', fgColor='a8121e')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    col_widths = [26, 16, 22, 14, 10, 14, 36]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 26

    # Example row leaves Rate blank so import auto-applies weekday/weekend settings
    example = ['John Smith', 'EMP-001', 'MEP', '2026-06-01', '3', '', 'Emergency AC repair']
    note_fill = PatternFill('solid', fgColor='FDECEE')
    note_font = Font(color='5A1118', size=10, italic=True)
    for col_idx, v in enumerate(example, start=1):
        cell = ws.cell(row=2, column=col_idx, value=v)
        cell.fill = note_fill
        cell.font = note_font
        cell.alignment = Alignment(vertical='center')
        cell.border = border

    settings = _get_overtime_settings()
    day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    weekend_labels = ', '.join(day_names[d] for d in settings['weekend_days'] if 0 <= d <= 6)

    ws2 = wb.create_sheet('Instructions')
    ws2['A1'] = 'Overtime Import — Instructions'
    ws2['A1'].font = Font(bold=True, size=13, color='a8121e')
    notes = [
        ('Staff Name', 'Required. Full name of the staff member.'),
        ('Employee ID', 'Optional. Staff/employee identifier (free text).'),
        ('Department', 'Optional. Team/department (e.g. MEP, Civil, Cleaning).'),
        ('Date', 'Required. Format: YYYY-MM-DD or DD/MM/YYYY. Used to pick weekday vs weekend rate.'),
        ('Hours', 'Required. Overtime hours worked — numbers only.'),
        ('Rate Per Hour',
         'Optional override. Leave blank to auto-apply settings: '
         f"weekday AED {settings['weekday_rate']:.2f}/h, weekend AED {settings['weekend_rate']:.2f}/h "
         f"(weekend days: {weekend_labels or 'none'}). Total = Hours × Rate."),
        ('Reason', 'Optional. Reason / notes for the overtime.'),
        ('Note',
         'Changing rates in Overtime Settings does not recalculate past imported rows; '
         'it applies to new imports and newly saved records.'),
    ]
    for row_i, (col, desc) in enumerate(notes, start=3):
        ws2.cell(row=row_i, column=1, value=col).font = Font(bold=True)
        ws2.cell(row=row_i, column=2, value=desc)
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 88

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name='overtime_import_template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ── Catalog materials (for invoice line-item search) ─────────────────────────

@operations_bp.route('/api/catalog-materials', methods=['GET'])
@jwt_required()
def api_catalog_materials():
    """Proxy the procurement catalog for operations users (who may lack procurement access)."""
    user = _current_user()
    if not _has_ops_sub_access(user, 'invoices'):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
    from app.models import Submission
    q = (request.args.get('q') or '').strip().lower()
    subs = Submission.query.filter_by(module_type='catalog_material').all()
    results = []
    for sub in subs:
        fd = sub.form_data or {}
        name = fd.get('material_name', '')
        if q and q not in name.lower() and q not in (fd.get('brand') or '').lower():
            continue
        try:
            unit_price = float(fd.get('unit_price') or 0)
        except Exception:
            unit_price = 0.0
        results.append({
            'name': name,
            'brand': fd.get('brand', ''),
            'uom': fd.get('uom', 'PCS'),
            'unit_price': unit_price,
        })
    results.sort(key=lambda x: x['name'])
    limit = 40 if q else 50
    return success_response({'materials': results[:limit]})


# ── Clients ──────────────────────────────────────────────────────────────────

@operations_bp.route('/clients')
@jwt_required()
def clients_page():
    user = _current_user()
    if not _has_ops_sub_access(user, 'clients'):
        return redirect('/dashboard')
    return render_template('clients.html', user=user)


@operations_bp.route('/clients/<client_ref>')
@jwt_required()
def client_detail_page(client_ref):
    user = _current_user()
    if not _has_ops_sub_access(user, 'clients'):
        return redirect('/dashboard')
    client = Client.query.filter_by(client_id=client_ref).first()
    if not client and client_ref.isdigit():
        client = Client.query.get(int(client_ref))
    if not client:
        return redirect('/operations/clients')
    return render_template('client_detail.html', user=user, client=client,
                           can_manage=_can_manage_invoices(user))


@operations_bp.route('/api/clients', methods=['GET'])
@jwt_required()
def api_list_clients():
    user = _current_user()
    if not _has_ops_sub_access(user, 'clients'):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')

    agg_rows = (db.session.query(
            TradingInvoice.client_id.label('cid'),
            func.count(TradingInvoice.id).label('invoice_count'),
            func.coalesce(func.sum(TradingInvoice.grand_total), 0.0).label('total_revenue'),
            func.coalesce(func.sum(case(
                (TradingInvoice.status == 'issued', TradingInvoice.grand_total),
                else_=0.0)), 0.0).label('outstanding'),
            func.max(TradingInvoice.invoice_date).label('last_invoice_date'),
        ).group_by(TradingInvoice.client_id).all())
    agg = {r.cid: r for r in agg_rows}

    query = Client.query
    q = (request.args.get('q') or '').strip()
    status = (request.args.get('status') or '').strip().lower()
    if q:
        like = f'%{q}%'
        query = query.filter(db.or_(
            Client.client_name.ilike(like), Client.contact_person.ilike(like),
            Client.email.ilike(like), Client.city.ilike(like)))
    if status in ('active', 'inactive'):
        query = query.filter(Client.status == status)
    clients = query.order_by(Client.client_name.asc()).all()

    out = []
    for c in clients:
        d = c.to_dict()
        r = agg.get(c.id)
        d['invoice_count'] = r.invoice_count if r else 0
        d['total_revenue'] = round(float(r.total_revenue), 2) if r else 0.0
        d['outstanding'] = round(float(r.outstanding), 2) if r else 0.0
        d['last_invoice_date'] = r.last_invoice_date.isoformat() if (r and r.last_invoice_date) else None
        out.append(d)
    return success_response({'clients': out})


@operations_bp.route('/api/clients', methods=['POST'])
@jwt_required()
def api_create_client():
    user = _current_user()
    if not _has_ops_sub_access(user, 'clients'):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
    data = request.get_json(silent=True) or {}
    name = (data.get('client_name') or '').strip()
    if not name:
        return error_response('Client name is required', status_code=400, error_code='VALIDATION_ERROR')
    c = Client(
        client_id=_gen_id('CLI'),
        client_name=name,
        contact_person=(data.get('contact_person') or '').strip() or None,
        email=(data.get('email') or '').strip() or None,
        phone=(data.get('phone') or '').strip() or None,
        billing_address=(data.get('billing_address') or '').strip() or None,
        city=(data.get('city') or '').strip() or None,
        country=(data.get('country') or '').strip() or None,
        tax_id=(data.get('tax_id') or '').strip() or None,
        status=(data.get('status') or 'active').strip() or 'active',
        notes=(data.get('notes') or '').strip() or None,
        created_by_id=user.id,
    )
    db.session.add(c)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error('create_client: %s', e, exc_info=True)
        return error_response('Database error', status_code=500, error_code='DATABASE_ERROR')
    return success_response({'client': c.to_dict()}, message='Client created.', status_code=201)


@operations_bp.route('/api/clients/<int:client_id>', methods=['PUT'])
@jwt_required()
def api_update_client(client_id):
    user = _current_user()
    if not _has_ops_sub_access(user, 'clients'):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
    c = Client.query.get(client_id)
    if not c:
        return error_response('Client not found', status_code=404, error_code='NOT_FOUND')
    data = request.get_json(silent=True) or {}
    if 'client_name' in data:
        name = (data.get('client_name') or '').strip()
        if not name:
            return error_response('Client name cannot be empty', status_code=400, error_code='VALIDATION_ERROR')
        c.client_name = name
    for field in ('contact_person', 'email', 'phone', 'billing_address',
                  'city', 'country', 'tax_id', 'notes'):
        if field in data:
            setattr(c, field, (data.get(field) or '').strip() or None)
    if 'status' in data:
        c.status = (data.get('status') or 'active').strip() or 'active'
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error('update_client: %s', e, exc_info=True)
        return error_response('Database error', status_code=500, error_code='DATABASE_ERROR')
    return success_response({'client': c.to_dict()}, message='Client updated.')


@operations_bp.route('/api/clients/<int:client_id>', methods=['DELETE'])
@jwt_required()
def api_delete_client(client_id):
    user = _current_user()
    if not _has_ops_sub_access(user, 'clients'):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
    c = Client.query.get(client_id)
    if not c:
        return error_response('Client not found', status_code=404, error_code='NOT_FOUND')
    if c.trading_invoices.count() > 0:
        return error_response('Cannot delete a client with existing invoices.',
                              status_code=400, error_code='VALIDATION_ERROR')
    db.session.delete(c)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error('delete_client: %s', e, exc_info=True)
        return error_response('Database error', status_code=500, error_code='DATABASE_ERROR')
    return success_response(message='Client deleted.')


@operations_bp.route('/api/clients/<int:client_id>/details', methods=['GET'])
@jwt_required()
def api_client_details(client_id):
    user = _current_user()
    if not _has_ops_sub_access(user, 'clients'):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
    c = Client.query.get(client_id)
    if not c:
        return error_response('Client not found', status_code=404, error_code='NOT_FOUND')

    invoices = (TradingInvoice.query.filter_by(client_id=c.id)
                .order_by(TradingInvoice.invoice_date.desc(), TradingInvoice.id.desc()).all())

    total_revenue = round(sum(inv.grand_total or 0.0 for inv in invoices), 2)
    outstanding = round(sum(inv.grand_total or 0.0 for inv in invoices if inv.status == 'issued'), 2)
    paid_total = round(sum(inv.grand_total or 0.0 for inv in invoices if inv.status == 'paid'), 2)
    by_status = {'draft': 0, 'issued': 0, 'paid': 0}
    for inv in invoices:
        if inv.status in by_status:
            by_status[inv.status] += 1
    inv_dates = [inv.invoice_date for inv in invoices if inv.invoice_date]
    stats = {
        'invoice_count': len(invoices),
        'total_revenue': total_revenue,
        'outstanding': outstanding,
        'paid_total': paid_total,
        'draft_count': by_status['draft'],
        'issued_count': by_status['issued'],
        'paid_count': by_status['paid'],
        'avg_invoice_value': round(total_revenue / len(invoices), 2) if invoices else 0.0,
        'first_invoice_date': min(inv_dates).isoformat() if inv_dates else None,
        'last_invoice_date': max(inv_dates).isoformat() if inv_dates else None,
    }

    # Last 12 calendar months, oldest → newest, zero-filled (Python-side bucketing
    # keeps this portable across SQLite and Postgres).
    month_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                    'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    y, m = date.today().year, date.today().month
    months = []
    for _ in range(12):
        months.append((y, m))
        m -= 1
        if m == 0:
            y, m = y - 1, 12
    months.reverse()
    buckets = {f'{yy:04d}-{mm:02d}': 0.0 for yy, mm in months}
    for inv in invoices:
        if inv.invoice_date:
            key = f'{inv.invoice_date.year:04d}-{inv.invoice_date.month:02d}'
            if key in buckets:
                buckets[key] = round(buckets[key] + (inv.grand_total or 0.0), 2)
    monthly_revenue = [{'month': f'{yy:04d}-{mm:02d}',
                        'label': month_labels[mm - 1],
                        'total': buckets[f'{yy:04d}-{mm:02d}']} for yy, mm in months]

    return success_response({
        'client': c.to_dict(),
        'stats': stats,
        'monthly_revenue': monthly_revenue,
        'invoices': [inv.to_dict(include_items=False) for inv in invoices],
    })


# ── Trading Invoices ─────────────────────────────────────────────────────────

def _recompute_totals(invoice, tax_pct):
    subtotal = round(sum((it.total_price or 0.0) for it in invoice.items), 2)
    tax_pct = _to_float(tax_pct, 0.0) or 0.0
    tax_amount = round(subtotal * tax_pct / 100.0, 2)
    invoice.subtotal = subtotal
    invoice.tax_pct = tax_pct
    invoice.tax_amount = tax_amount
    invoice.grand_total = round(subtotal + tax_amount, 2)


def _build_items(invoice, items_data):
    """Replace invoice.items from a list of dicts; returns error string or None."""
    invoice.items.clear()
    for idx, raw in enumerate(items_data or [], start=1):
        material = (raw.get('material_name') or '').strip()
        if not material:
            continue
        qty = _to_float(raw.get('quantity'), 0.0) or 0.0
        unit_price = _to_float(raw.get('unit_price'), 0.0) or 0.0
        invoice.items.append(TradingInvoiceItem(
            material_name=material,
            description=(raw.get('description') or '').strip() or None,
            quantity=qty,
            unit=(raw.get('unit') or '').strip() or None,
            unit_price=unit_price,
            total_price=round(qty * unit_price, 2),
        ))
    if not invoice.items:
        return 'At least one line item with a material name is required.'
    return None


@operations_bp.route('/invoices')
@jwt_required()
def invoices_page():
    user = _current_user()
    if not _has_ops_sub_access(user, 'invoices'):
        return redirect('/dashboard')
    return render_template('trading_invoices.html', user=user, can_manage=_can_manage_invoices(user))


@operations_bp.route('/invoices/<invoice_no>')
@jwt_required()
def invoice_detail_page(invoice_no):
    user = _current_user()
    if not _has_ops_sub_access(user, 'invoices'):
        return redirect('/dashboard')
    invoice = TradingInvoice.query.filter_by(invoice_no=invoice_no).first()
    if not invoice:
        return redirect('/operations/invoices')
    return render_template('trading_invoice_detail.html', user=user, invoice=invoice,
                           can_manage=_can_manage_invoices(user))


@operations_bp.route('/api/invoices', methods=['GET'])
@jwt_required()
def api_list_invoices():
    user = _current_user()
    if not _has_ops_sub_access(user, 'invoices'):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
    invoices = TradingInvoice.query.order_by(TradingInvoice.created_at.desc()).all()
    return success_response({'invoices': [inv.to_dict(include_items=False) for inv in invoices]})


@operations_bp.route('/api/invoices/<int:invoice_id>', methods=['GET'])
@jwt_required()
def api_get_invoice(invoice_id):
    user = _current_user()
    if not _has_ops_sub_access(user, 'invoices'):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
    inv = TradingInvoice.query.get(invoice_id)
    if not inv:
        return error_response('Invoice not found', status_code=404, error_code='NOT_FOUND')
    return success_response({'invoice': inv.to_dict()})


@operations_bp.route('/api/invoices', methods=['POST'])
@jwt_required()
def api_create_invoice():
    user = _current_user()
    if not _has_ops_sub_access(user, 'invoices') or not _can_manage_invoices(user):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
    data = request.get_json(silent=True) or {}

    client_id = data.get('client_id')
    client = Client.query.get(client_id) if client_id else None
    if not client:
        return error_response('A valid client is required', status_code=400, error_code='VALIDATION_ERROR')

    inv = TradingInvoice(
        invoice_no=_gen_id('TRD-INV'),
        client_id=client.id,
        invoice_date=_parse_date(data.get('invoice_date')) or date.today(),
        due_date=_parse_date(data.get('due_date')),
        status=(data.get('status') or 'draft').strip() or 'draft',
        notes=(data.get('notes') or '').strip() or None,
        created_by_id=user.id,
    )
    err = _build_items(inv, data.get('items'))
    if err:
        return error_response(err, status_code=400, error_code='VALIDATION_ERROR')
    _recompute_totals(inv, data.get('tax_pct'))

    db.session.add(inv)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error('create_invoice: %s', e, exc_info=True)
        return error_response('Database error', status_code=500, error_code='DATABASE_ERROR')
    return success_response({'invoice': inv.to_dict()}, message='Trading invoice created.', status_code=201)


@operations_bp.route('/api/invoices/<int:invoice_id>', methods=['PUT'])
@jwt_required()
def api_update_invoice(invoice_id):
    user = _current_user()
    if not _has_ops_sub_access(user, 'invoices') or not _can_manage_invoices(user):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
    inv = TradingInvoice.query.get(invoice_id)
    if not inv:
        return error_response('Invoice not found', status_code=404, error_code='NOT_FOUND')
    data = request.get_json(silent=True) or {}

    if 'client_id' in data:
        client = Client.query.get(data.get('client_id'))
        if not client:
            return error_response('A valid client is required', status_code=400, error_code='VALIDATION_ERROR')
        inv.client_id = client.id
    if 'invoice_date' in data:
        d = _parse_date(data.get('invoice_date'))
        if d:
            inv.invoice_date = d
    if 'due_date' in data:
        inv.due_date = _parse_date(data.get('due_date'))
    if 'status' in data:
        inv.status = (data.get('status') or 'draft').strip() or 'draft'
    if 'notes' in data:
        inv.notes = (data.get('notes') or '').strip() or None
    if 'items' in data:
        err = _build_items(inv, data.get('items'))
        if err:
            return error_response(err, status_code=400, error_code='VALIDATION_ERROR')
    # tax_pct may change independently of items; fall back to existing value
    _recompute_totals(inv, data.get('tax_pct', inv.tax_pct))

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error('update_invoice: %s', e, exc_info=True)
        return error_response('Database error', status_code=500, error_code='DATABASE_ERROR')
    return success_response({'invoice': inv.to_dict()}, message='Trading invoice updated.')


@operations_bp.route('/api/invoices/<int:invoice_id>', methods=['DELETE'])
@jwt_required()
def api_delete_invoice(invoice_id):
    user = _current_user()
    if not _has_ops_sub_access(user, 'invoices') or not _can_manage_invoices(user):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
    inv = TradingInvoice.query.get(invoice_id)
    if not inv:
        return error_response('Invoice not found', status_code=404, error_code='NOT_FOUND')
    db.session.delete(inv)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error('delete_invoice: %s', e, exc_info=True)
        return error_response('Database error', status_code=500, error_code='DATABASE_ERROR')
    return success_response(message='Trading invoice deleted.')


@operations_bp.route('/invoices/<invoice_no>/pdf', methods=['GET'])
@jwt_required()
def invoice_pdf(invoice_no):
    user = _current_user()
    if not _has_ops_sub_access(user, 'invoices'):
        return redirect('/dashboard')
    inv = TradingInvoice.query.filter_by(invoice_no=invoice_no).first()
    if not inv:
        return error_response('Invoice not found', status_code=404, error_code='NOT_FOUND')

    try:
        from module_operations.trading_invoice_builder import build_trading_invoice_pdf
    except Exception as e:
        logger.error('trading_invoice_builder import: %s', e, exc_info=True)
        return error_response('PDF generator unavailable', status_code=500, error_code='DEPENDENCY_ERROR')

    buf = BytesIO()
    try:
        build_trading_invoice_pdf(inv, inv.client, list(inv.items), buf)
    except Exception as e:
        logger.error('build_trading_invoice_pdf: %s', e, exc_info=True)
        return error_response('Could not generate PDF', status_code=500, error_code='PDF_ERROR')
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f'{inv.invoice_no}.pdf',
        mimetype='application/pdf',
    )


# ── Cheque Preparation ───────────────────────────────────────────────────────
#
# Digitized "Cheque Preparation / Request Form". Lifecycle:
#   requested → verified → approved → prepared → submitted → cleared
# plus rejected / cancelled from any non-terminal status.
# All mutations are admin-only; Operations users are view-only.

CHEQUE_FLOW = ['requested', 'verified', 'approved', 'prepared', 'submitted', 'cleared']
CHEQUE_TERMINAL = {'cleared', 'rejected', 'cancelled'}

STATUS_LABELS = {
    'requested': 'Requested', 'verified': 'Verified', 'approved': 'Approved',
    'prepared': 'Prepared', 'submitted': 'Submitted', 'cleared': 'Cleared',
    'rejected': 'Rejected', 'cancelled': 'Cancelled',
}


def _is_admin(user):
    return bool(user) and user.role == 'admin'


def _normalize_signature(raw):
    """Accept a base64 data-URL signature, or clear with empty/null. Reject garbage."""
    if raw is None:
        return None
    sig = str(raw).strip()
    if not sig:
        return None
    if not sig.startswith('data:image'):
        return None
    # Cap ~1.5MB of base64 text so a bad client can't bloat the DB
    if len(sig) > 1_500_000:
        return None
    return sig


def _allowed_next_statuses(current):
    """Forward move to the next stage, or reject/cancel while not terminal."""
    if current in CHEQUE_TERMINAL:
        return []
    try:
        idx = CHEQUE_FLOW.index(current)
    except ValueError:
        return []
    nxt = [CHEQUE_FLOW[idx + 1]] if idx + 1 < len(CHEQUE_FLOW) else []
    return nxt + ['rejected', 'cancelled']


def _user_display_name(user):
    return getattr(user, 'full_name', None) or getattr(user, 'username', None) or 'Unknown'


def _apply_cheque_items(cheque, items_data):
    """Replace all line items from a list of dicts; returns error string or None."""
    cheque.items.clear()
    total = 0.0
    for i, item in enumerate(items_data, start=1):
        supplier = (item.get('supplier') or '').strip()
        if not supplier:
            return f'Line {i}: supplier is required'
        amount = _to_float(item.get('amount'))
        if amount is None or amount < 0:
            return f'Line {i}: a valid amount is required'
        total += amount
        cheque.items.append(ChequeRequestItem(
            sn=i,
            supplier=supplier,
            amount=round(amount, 2),
            cheque_date=_parse_date(item.get('cheque_date')),
            remarks=(item.get('remarks') or '').strip() or None,
        ))
    cheque.total_amount = round(total, 2)
    return None


def _cheque_status_email_html(cheque, from_status, to_status, changed_by_name, note):
    supplier_rows = ''.join(
        f"<tr>"
        f"<td style='padding:6px 10px;border:1px solid #e5e7eb;'>{it.sn}</td>"
        f"<td style='padding:6px 10px;border:1px solid #e5e7eb;'>{it.supplier}</td>"
        f"<td style='padding:6px 10px;border:1px solid #e5e7eb;text-align:right;'>{it.amount:,.2f}</td>"
        f"<td style='padding:6px 10px;border:1px solid #e5e7eb;'>{it.cheque_date.strftime('%d/%m/%Y') if it.cheque_date else '—'}</td>"
        f"</tr>"
        for it in cheque.items
    )
    note_html = (f"<p style='margin:12px 0 0;color:#374151;'><strong>Note:</strong> {note}</p>"
                 if note else '')
    from_label = STATUS_LABELS.get(from_status, from_status or '—')
    to_label = STATUS_LABELS.get(to_status, to_status)
    return f"""
<div style="font-family:Arial,Helvetica,sans-serif;max-width:640px;margin:0 auto;border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;">
  <div style="background:#d21725;color:#ffffff;padding:18px 24px;">
    <h2 style="margin:0;font-size:18px;">Cheque Request Update — {cheque.reference_no}</h2>
    <p style="margin:4px 0 0;font-size:13px;opacity:.9;">Injaaz Operations · Cheque Preparation</p>
  </div>
  <div style="padding:20px 24px;color:#111827;">
    <p style="margin:0 0 14px;">
      Status changed from <strong>{from_label}</strong> to
      <strong style="color:#d21725;">{to_label}</strong> by <strong>{changed_by_name}</strong>.
    </p>
    <table style="border-collapse:collapse;font-size:13px;margin:0 0 6px;">
      <tr><td style="padding:3px 12px 3px 0;color:#6b7280;">Department</td><td>{cheque.department or '—'}</td></tr>
      <tr><td style="padding:3px 12px 3px 0;color:#6b7280;">Total amount</td><td><strong>AED {cheque.total_amount:,.2f}</strong></td></tr>
      <tr><td style="padding:3px 12px 3px 0;color:#6b7280;">Requested by</td><td>{cheque.requested_by_name or '—'}</td></tr>
    </table>
    <table style="border-collapse:collapse;font-size:13px;width:100%;margin-top:10px;">
      <tr style="background:#f9fafb;">
        <th style="padding:6px 10px;border:1px solid #e5e7eb;text-align:left;">SN</th>
        <th style="padding:6px 10px;border:1px solid #e5e7eb;text-align:left;">Supplier</th>
        <th style="padding:6px 10px;border:1px solid #e5e7eb;text-align:right;">Amount (AED)</th>
        <th style="padding:6px 10px;border:1px solid #e5e7eb;text-align:left;">Date</th>
      </tr>
      {supplier_rows}
    </table>
    {note_html}
  </div>
  <div style="background:#f9fafb;padding:12px 24px;font-size:12px;color:#6b7280;border-top:1px solid #e5e7eb;">
    This is an automated notification from the Injaaz Operations — Cheque Preparation module.
  </div>
</div>
"""


def _send_cheque_status_email(cheque, from_status, to_status, changed_by_name, note=None):
    """Email configured department recipients about a status change.

    Returns True if an email was sent, False otherwise (not configured / failure).
    Never raises — a mail problem must not block the status change."""
    try:
        from common.email_service import send_email, is_email_configured
        cfg = ChequeNotificationConfig.query.filter_by(status=to_status).first()
        to_list = [e.strip() for e in (cfg.to_emails if cfg else '').split(',') if e.strip()]
        cc_list = [e.strip() for e in (cfg.cc_emails if cfg else '').split(',') if e.strip()]
        if not to_list:
            logger.info('cheque %s: no recipients configured for status %s — email skipped',
                        cheque.reference_no, to_status)
            return False
        if not is_email_configured():
            logger.warning('cheque %s: email not configured — skipped', cheque.reference_no)
            return False

        to_label = STATUS_LABELS.get(to_status, to_status)
        subject = f'[Cheque {to_label}] {cheque.reference_no} — AED {cheque.total_amount:,.2f}'
        body = (
            f'Cheque request {cheque.reference_no} status changed from '
            f'{STATUS_LABELS.get(from_status, from_status or "—")} to {to_label} by {changed_by_name}.\n'
            f'Department: {cheque.department or "—"}\n'
            f'Total amount: AED {cheque.total_amount:,.2f}\n'
            + (f'Note: {note}\n' if note else '')
        )
        html_body = _cheque_status_email_html(cheque, from_status, to_status, changed_by_name, note)
        return bool(send_email(recipient=to_list, subject=subject, body=body,
                               html_body=html_body, cc=cc_list or None))
    except Exception as e:
        logger.error('cheque status email failed for %s: %s', cheque.reference_no, e, exc_info=True)
        return False


# ── Cheque pages ─────────────────────────────────────────────────────────────

@operations_bp.route('/cheques')
@jwt_required()
def cheques_page():
    user = _current_user()
    if not _has_ops_sub_access(user, 'cheques'):
        return redirect('/dashboard')
    return render_template('operations_cheques.html', user=user, is_admin=_is_admin(user))


@operations_bp.route('/cheques/<reference_no>/pdf', methods=['GET'])
@jwt_required()
def cheque_pdf(reference_no):
    """Printable Cheque Preparation/Request Form — regenerated live from the
    current record on every request, so preview and download are always
    up to date with the latest edits/status."""
    user = _current_user()
    if not _has_ops_sub_access(user, 'cheques'):
        return redirect('/dashboard')
    cheque = ChequeRequest.query.filter_by(reference_no=reference_no).first()
    if not cheque:
        return error_response('Cheque request not found', status_code=404, error_code='NOT_FOUND')

    try:
        from module_operations.cheque_pdf_builder import build_cheque_pdf
    except Exception as e:
        logger.error('cheque_pdf_builder import: %s', e, exc_info=True)
        return error_response('PDF generator unavailable', status_code=500, error_code='DEPENDENCY_ERROR')

    buf = BytesIO()
    try:
        build_cheque_pdf(cheque, buf)
    except Exception as e:
        logger.error('build_cheque_pdf: %s', e, exc_info=True)
        return error_response('Could not generate PDF', status_code=500, error_code='PDF_ERROR')
    buf.seek(0)
    as_attachment = request.args.get('download') in ('1', 'true', 'yes')
    return send_file(
        buf,
        as_attachment=as_attachment,
        download_name=f'{cheque.reference_no}.pdf',
        mimetype='application/pdf',
    )


@operations_bp.route('/cheques/<reference_no>')
@jwt_required()
def cheque_detail_page(reference_no):
    user = _current_user()
    if not _has_ops_sub_access(user, 'cheques'):
        return redirect('/dashboard')
    cheque = ChequeRequest.query.filter_by(reference_no=reference_no).first()
    if not cheque:
        return redirect('/operations/cheques')
    return render_template('operations_cheque_detail.html', user=user,
                           cheque=cheque, is_admin=_is_admin(user))


# ── Cheque APIs ──────────────────────────────────────────────────────────────

@operations_bp.route('/api/cheques', methods=['GET'])
@jwt_required()
def api_list_cheques():
    user = _current_user()
    if not _has_ops_sub_access(user, 'cheques'):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')

    query = ChequeRequest.query
    q = (request.args.get('q') or '').strip()
    status = (request.args.get('status') or '').strip().lower()
    if q:
        like = f'%{q}%'
        query = query.filter(db.or_(
            ChequeRequest.reference_no.ilike(like),
            ChequeRequest.department.ilike(like),
            ChequeRequest.requested_by_name.ilike(like),
            ChequeRequest.items.any(ChequeRequestItem.supplier.ilike(like)),
        ))
    if status in CHEQUE_STATUSES:
        query = query.filter(ChequeRequest.status == status)
    cheques = query.order_by(ChequeRequest.created_at.desc()).all()

    counts = dict(db.session.query(ChequeRequest.status, func.count(ChequeRequest.id))
                  .group_by(ChequeRequest.status).all())
    return success_response({
        'cheques': [c.to_dict(include_items=True, include_signatures=False) for c in cheques],
        'counts': {s: counts.get(s, 0) for s in CHEQUE_STATUSES},
        'is_admin': _is_admin(user),
    })


@operations_bp.route('/api/cheques/<reference_no>', methods=['GET'])
@jwt_required()
def api_get_cheque(reference_no):
    user = _current_user()
    if not _has_ops_sub_access(user, 'cheques'):
        return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
    cheque = ChequeRequest.query.filter_by(reference_no=reference_no).first()
    if not cheque:
        return error_response('Cheque request not found', status_code=404, error_code='NOT_FOUND')
    data = cheque.to_dict(include_items=True, include_logs=True)
    data['allowed_next_statuses'] = _allowed_next_statuses(cheque.status)
    return success_response({'cheque': data, 'is_admin': _is_admin(user)})


@operations_bp.route('/api/cheques', methods=['POST'])
@jwt_required()
def api_create_cheque():
    user = _current_user()
    if not _is_admin(user):
        return error_response('Admin access required', status_code=403, error_code='ACCESS_DENIED')
    data = request.get_json(silent=True) or {}

    items_data = data.get('items') or []
    if not items_data:
        return error_response('At least one supplier line is required',
                              status_code=400, error_code='VALIDATION_ERROR')

    requested_sig = _normalize_signature(data.get('requested_signature'))
    if not requested_sig:
        return error_response('Requested-by signature is required',
                              status_code=400, error_code='VALIDATION_ERROR')

    cheque = ChequeRequest(
        reference_no=_gen_id('CHQ'),
        office=(data.get('office') or '').strip() or None,
        department=(data.get('department') or 'Finance').strip() or 'Finance',
        status='requested',
        remarks=(data.get('remarks') or '').strip() or None,
        attached_documents=(data.get('attached_documents') or '').strip() or None,
        requested_by_name=(data.get('requested_by_name') or '').strip() or _user_display_name(user),
        requested_date=_parse_date(data.get('requested_date')) or datetime.utcnow().date(),
        requested_signature=requested_sig,
        created_by_id=user.id,
    )
    err = _apply_cheque_items(cheque, items_data)
    if err:
        return error_response(err, status_code=400, error_code='VALIDATION_ERROR')

    log = ChequeStatusLog(from_status=None, to_status='requested',
                          changed_by_id=user.id, note='Cheque request created')
    cheque.status_logs.append(log)
    db.session.add(cheque)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error('create_cheque: %s', e, exc_info=True)
        return error_response('Database error', status_code=500, error_code='DATABASE_ERROR')

    log.email_sent = _send_cheque_status_email(cheque, None, 'requested', _user_display_name(user))
    db.session.commit()
    return success_response({'cheque': cheque.to_dict()},
                            message='Cheque request created.', status_code=201)


@operations_bp.route('/api/cheques/<reference_no>', methods=['PUT'])
@jwt_required()
def api_update_cheque(reference_no):
    user = _current_user()
    if not _is_admin(user):
        return error_response('Admin access required', status_code=403, error_code='ACCESS_DENIED')
    cheque = ChequeRequest.query.filter_by(reference_no=reference_no).first()
    if not cheque:
        return error_response('Cheque request not found', status_code=404, error_code='NOT_FOUND')
    data = request.get_json(silent=True) or {}

    for field in ('office', 'remarks', 'attached_documents',
                  'requested_by_name', 'verified_by_name', 'approved_by_name'):
        if field in data:
            setattr(cheque, field, (data.get(field) or '').strip() or None)
    if 'department' in data:
        cheque.department = (data.get('department') or 'Finance').strip() or 'Finance'
    for field in ('requested_date', 'verified_date', 'approved_date'):
        if field in data:
            setattr(cheque, field, _parse_date(data.get(field)))
    for sig_field in ('requested_signature', 'verified_signature', 'approved_signature'):
        if sig_field in data:
            raw = data.get(sig_field)
            if raw in ('', None):
                setattr(cheque, sig_field, None)
            else:
                sig = _normalize_signature(raw)
                if not sig:
                    return error_response(f'{sig_field} must be a valid image signature',
                                          status_code=400, error_code='VALIDATION_ERROR')
                setattr(cheque, sig_field, sig)
    if 'items' in data:
        items_data = data.get('items') or []
        if not items_data:
            return error_response('At least one supplier line is required',
                                  status_code=400, error_code='VALIDATION_ERROR')
        err = _apply_cheque_items(cheque, items_data)
        if err:
            return error_response(err, status_code=400, error_code='VALIDATION_ERROR')

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error('update_cheque: %s', e, exc_info=True)
        return error_response('Database error', status_code=500, error_code='DATABASE_ERROR')
    return success_response({'cheque': cheque.to_dict()}, message='Cheque request updated.')


@operations_bp.route('/api/cheques/<reference_no>', methods=['DELETE'])
@jwt_required()
def api_delete_cheque(reference_no):
    user = _current_user()
    if not _is_admin(user):
        return error_response('Admin access required', status_code=403, error_code='ACCESS_DENIED')
    cheque = ChequeRequest.query.filter_by(reference_no=reference_no).first()
    if not cheque:
        return error_response('Cheque request not found', status_code=404, error_code='NOT_FOUND')
    db.session.delete(cheque)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error('delete_cheque: %s', e, exc_info=True)
        return error_response('Database error', status_code=500, error_code='DATABASE_ERROR')
    return success_response(message='Cheque request deleted.')


@operations_bp.route('/api/cheques/<reference_no>/status', methods=['POST'])
@jwt_required()
def api_change_cheque_status(reference_no):
    user = _current_user()
    if not _is_admin(user):
        return error_response('Admin access required', status_code=403, error_code='ACCESS_DENIED')
    cheque = ChequeRequest.query.filter_by(reference_no=reference_no).first()
    if not cheque:
        return error_response('Cheque request not found', status_code=404, error_code='NOT_FOUND')
    data = request.get_json(silent=True) or {}

    new_status = (data.get('status') or '').strip().lower()
    note = (data.get('note') or '').strip() or None
    allowed = _allowed_next_statuses(cheque.status)
    if new_status not in allowed:
        return error_response(
            f'Cannot move from "{STATUS_LABELS.get(cheque.status, cheque.status)}" to '
            f'"{STATUS_LABELS.get(new_status, new_status)}". Allowed: '
            + (', '.join(STATUS_LABELS.get(s, s) for s in allowed) or 'none (terminal status)'),
            status_code=400, error_code='INVALID_TRANSITION')

    old_status = cheque.status
    cheque.status = new_status
    actor = _user_display_name(user)
    today = datetime.utcnow().date()
    # Stamp the matching signatory block on the paper form (name + signature from UI).
    if new_status == 'verified':
        sig = _normalize_signature(data.get('signature') or data.get('verified_signature'))
        if not sig:
            return error_response('A signature is required to mark as Verified',
                                  status_code=400, error_code='VALIDATION_ERROR')
        cheque.verified_by_name = (data.get('signatory_name') or '').strip() or actor
        cheque.verified_date = today
        cheque.verified_signature = sig
    elif new_status == 'approved':
        sig = _normalize_signature(data.get('signature') or data.get('approved_signature'))
        if not sig:
            return error_response('A signature is required to mark as Approved',
                                  status_code=400, error_code='VALIDATION_ERROR')
        cheque.approved_by_name = (data.get('signatory_name') or '').strip() or actor
        cheque.approved_date = today
        cheque.approved_signature = sig
    elif new_status == 'requested':
        # Re-open / initial stamp path — optional signature if provided
        sig = _normalize_signature(data.get('signature') or data.get('requested_signature'))
        if sig:
            cheque.requested_signature = sig
            if not cheque.requested_by_name:
                cheque.requested_by_name = (data.get('signatory_name') or '').strip() or actor
            if not cheque.requested_date:
                cheque.requested_date = today

    log = ChequeStatusLog(from_status=old_status, to_status=new_status,
                          changed_by_id=user.id, note=note)
    cheque.status_logs.append(log)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error('change_cheque_status: %s', e, exc_info=True)
        return error_response('Database error', status_code=500, error_code='DATABASE_ERROR')

    log.email_sent = _send_cheque_status_email(cheque, old_status, new_status, actor, note)
    db.session.commit()

    data = cheque.to_dict(include_items=True, include_logs=True)
    data['allowed_next_statuses'] = _allowed_next_statuses(cheque.status)
    return success_response({'cheque': data},
                            message=f'Status changed to {STATUS_LABELS.get(new_status, new_status)}.'
                                    + ('' if log.email_sent else ' (No notification email sent — check recipients/config.)'))


@operations_bp.route('/api/cheques-notification-config', methods=['GET'])
@jwt_required()
def api_get_cheque_notification_config():
    user = _current_user()
    if not _is_admin(user):
        return error_response('Admin access required', status_code=403, error_code='ACCESS_DENIED')
    rows = {c.status: c.to_dict() for c in ChequeNotificationConfig.query.all()}
    config = [rows.get(s, {'status': s, 'to_emails': '', 'cc_emails': ''}) for s in CHEQUE_STATUSES]
    return success_response({'config': config})


@operations_bp.route('/api/cheques-notification-config', methods=['PUT'])
@jwt_required()
def api_update_cheque_notification_config():
    user = _current_user()
    if not _is_admin(user):
        return error_response('Admin access required', status_code=403, error_code='ACCESS_DENIED')
    data = request.get_json(silent=True) or {}
    entries = data.get('config') or []

    def _clean_emails(raw):
        return ', '.join(e.strip() for e in str(raw or '').split(',') if e.strip())

    for entry in entries:
        status = (entry.get('status') or '').strip().lower()
        if status not in CHEQUE_STATUSES:
            continue
        row = ChequeNotificationConfig.query.filter_by(status=status).first()
        if not row:
            row = ChequeNotificationConfig(status=status)
            db.session.add(row)
        row.to_emails = _clean_emails(entry.get('to_emails'))
        row.cc_emails = _clean_emails(entry.get('cc_emails'))
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error('update_cheque_notification_config: %s', e, exc_info=True)
        return error_response('Database error', status_code=500, error_code='DATABASE_ERROR')
    return success_response(message='Notification recipients saved.')
