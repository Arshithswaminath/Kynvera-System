"""Operations attendance — biometric Excel upload + day-status editing."""
from __future__ import annotations

import uuid
from datetime import datetime, date
from io import BytesIO

from flask import render_template, redirect, request, send_file
from flask_jwt_extended import jwt_required
from openpyxl import Workbook, load_workbook

from app.models import (
    db, Technician, AttendanceImportBatch, AttendanceEntry, ATTENDANCE_DAY_STATUSES,
)
from common.error_responses import success_response, error_response


def _gen_batch():
    return 'ATT-' + uuid.uuid4().hex[:8].upper()


def _parse_date(value):
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_hhmm(value):
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.strftime('%H:%M')
    if hasattr(value, 'hour'):
        try:
            return f'{int(value.hour):02d}:{int(value.minute):02d}'
        except Exception:
            pass
    s = str(value).strip().replace('.', ':')
    parts = s.split(':')
    if len(parts) < 2:
        return None
    try:
        h, m = int(float(parts[0])), int(float(parts[1]))
        return f'{h:02d}:{m:02d}'
    except (TypeError, ValueError):
        return None


def _hours_between(start_hhmm, end_hhmm):
    if not start_hhmm or not end_hhmm:
        return None
    sh, sm = map(int, start_hhmm.split(':'))
    eh, em = map(int, end_hhmm.split(':'))
    start_m = sh * 60 + sm
    end_m = eh * 60 + em
    if end_m <= start_m:
        end_m += 24 * 60
    return round((end_m - start_m) / 60.0, 2)


def _normalize_status(raw):
    s = str(raw or 'present').strip().lower().replace(' ', '_')
    aliases = {
        'present': 'present', 'p': 'present',
        'off': 'off_day', 'off_day': 'off_day', 'offday': 'off_day',
        'emergency': 'emergency_leave', 'emergency_leave': 'emergency_leave',
        'sick': 'sick_leave', 'sick_leave': 'sick_leave',
        'half': 'half_day', 'half_day': 'half_day',
    }
    return aliases.get(s, s if s in ATTENDANCE_DAY_STATUSES else 'present')


def _match_technician(employee_code, staff_name):
    tech = None
    code = (employee_code or '').strip()
    if code:
        tech = Technician.query.filter(
            db.func.lower(Technician.employee_id) == code.lower()
        ).first()
    if not tech and staff_name:
        tech = Technician.query.filter(
            db.func.lower(Technician.full_name) == staff_name.strip().lower()
        ).first()
    return tech


def register_attendance_routes(bp, *, current_user_fn, has_ops_sub_access_fn):
    @bp.route('/attendance')
    @jwt_required()
    def attendance_page():
        user = current_user_fn()
        if not has_ops_sub_access_fn(user, 'attendance'):
            return redirect('/dashboard')
        return render_template(
            'attendance.html', user=user, day_statuses=list(ATTENDANCE_DAY_STATUSES),
        )

    @bp.route('/api/attendance/template', methods=['GET'])
    @jwt_required()
    def api_attendance_template():
        user = current_user_fn()
        if not has_ops_sub_access_fn(user, 'attendance'):
            return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
        wb = Workbook()
        ws = wb.active
        ws.title = 'Attendance'
        ws.append([
            'Employee ID', 'Staff Name', 'Project', 'Team Size', 'Date',
            'Status', 'Start Time', 'End Time', 'Notes',
        ])
        ws.append([
            'TECH-001', 'Sample Technician', 'Ajman Tower', 4, '2026-07-01',
            'present', '08:00', '17:00', '',
        ])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(
            bio, as_attachment=True,
            download_name='attendance_import_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    @bp.route('/api/attendance/import', methods=['POST'])
    @jwt_required()
    def api_attendance_import():
        user = current_user_fn()
        if not has_ops_sub_access_fn(user, 'attendance'):
            return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
        f = request.files.get('file')
        if not f:
            return error_response('Excel file is required', status_code=400, error_code='FILE_REQUIRED')
        try:
            wb = load_workbook(f, data_only=True)
            ws = wb.active
        except Exception as e:
            return error_response(f'Could not read Excel: {e}', status_code=400, error_code='BAD_FILE')

        batch = AttendanceImportBatch(
            batch_id=_gen_batch(),
            source_filename=getattr(f, 'filename', None),
            period_label=(request.form.get('period_label') or '').strip() or None,
            imported_by_id=user.id,
        )
        db.session.add(batch)
        db.session.flush()

        matched = created = 0
        for raw in ws.iter_rows(min_row=2, values_only=True):
            if not raw or all(c is None or str(c).strip() == '' for c in raw):
                continue
            emp_code = str(raw[0] or '').strip() if len(raw) > 0 else ''
            staff = str(raw[1] or '').strip() if len(raw) > 1 else ''
            project = str(raw[2] or '').strip() if len(raw) > 2 else ''
            try:
                team_size = int(raw[3]) if len(raw) > 3 and raw[3] not in (None, '') else None
            except (TypeError, ValueError):
                team_size = None
            work_date = _parse_date(raw[4] if len(raw) > 4 else None)
            status = _normalize_status(raw[5] if len(raw) > 5 else 'present')
            start_time = _parse_hhmm(raw[6] if len(raw) > 6 else None)
            end_time = _parse_hhmm(raw[7] if len(raw) > 7 else None)
            notes = str(raw[8] or '').strip() if len(raw) > 8 else ''
            if (not staff and not emp_code) or not work_date:
                continue
            tech = _match_technician(emp_code, staff)
            if tech:
                matched += 1
                staff = staff or tech.full_name
                emp_code = emp_code or tech.employee_id
            hours = _hours_between(start_time, end_time)
            if status == 'half_day' and hours is None:
                hours = 4.0
            if status == 'off_day':
                hours = 0.0
            db.session.add(AttendanceEntry(
                batch_id_fk=batch.id,
                technician_id=tech.id if tech else None,
                employee_code=emp_code or None,
                staff_name=staff or emp_code or 'Unknown',
                project_name=project or None,
                team_size=team_size,
                work_date=work_date,
                day_status=status,
                start_time=start_time,
                end_time=end_time,
                hours=hours,
                notes=notes or None,
            ))
            created += 1

        batch.row_count = created
        batch.matched_count = matched
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return error_response(f'Database error: {e}', status_code=500, error_code='DATABASE_ERROR')
        return success_response({
            'batch': batch.to_dict(), 'created': created, 'matched': matched,
        }, message=f'Imported {created} rows ({matched} matched to technician master).')

    @bp.route('/api/attendance', methods=['GET'])
    @jwt_required()
    def api_list_attendance():
        user = current_user_fn()
        if not has_ops_sub_access_fn(user, 'attendance'):
            return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
        q = AttendanceEntry.query.order_by(
            AttendanceEntry.work_date.desc(), AttendanceEntry.id.desc()
        )
        project = (request.args.get('project') or '').strip()
        if project:
            q = q.filter(AttendanceEntry.project_name.ilike(f'%{project}%'))
        rows = q.limit(1000).all()
        return success_response({
            'entries': [r.to_dict() for r in rows],
            'day_statuses': list(ATTENDANCE_DAY_STATUSES),
        })

    @bp.route('/api/attendance/<int:entry_id>', methods=['PUT'])
    @jwt_required()
    def api_update_attendance(entry_id):
        user = current_user_fn()
        if not has_ops_sub_access_fn(user, 'attendance'):
            return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
        row = db.session.get(AttendanceEntry, entry_id)
        if not row:
            return error_response('Entry not found', status_code=404, error_code='NOT_FOUND')
        data = request.get_json(silent=True) or {}
        if 'day_status' in data:
            row.day_status = _normalize_status(data.get('day_status'))
        if 'start_time' in data:
            row.start_time = _parse_hhmm(data.get('start_time'))
        if 'end_time' in data:
            row.end_time = _parse_hhmm(data.get('end_time'))
        if 'project_name' in data:
            row.project_name = (data.get('project_name') or '').strip() or None
        if 'notes' in data:
            row.notes = (data.get('notes') or '').strip() or None
        if row.day_status == 'off_day':
            row.hours = 0.0
        elif row.day_status == 'half_day':
            row.hours = _hours_between(row.start_time, row.end_time) or 4.0
        elif row.day_status in ('emergency_leave', 'sick_leave'):
            row.hours = _hours_between(row.start_time, row.end_time) or 0.0
        else:
            row.hours = _hours_between(row.start_time, row.end_time)
        db.session.commit()
        return success_response({'entry': row.to_dict()}, message='Attendance updated.')
