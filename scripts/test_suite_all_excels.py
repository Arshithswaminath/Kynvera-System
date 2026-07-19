#!/usr/bin/env python3
"""
Suite 2 — Generate every product Excel workbook with test data.

Covers:
  • Fire Systems inspection Excel report
  • Finance monthly multi-sheet report
  • Ticketing analytics reports (all REPORT_SPECS keys)
  • Operations overtime export (sample rows)
  • Store materials sample export
  • Finance contracts import template

Usage (from project root):
  python scripts/test_suite_all_excels.py

Output:
  test_output/excel_suite_YYYYMMDD_HHMMSS/
"""
from __future__ import annotations

import os
import sys
from datetime import date
from io import BytesIO

_SCRIPTS = os.path.dirname(os.path.abspath(__file__))
if _SCRIPTS not in sys.path:
    sys.path.insert(0, _SCRIPTS)

from test_suite_common import (  # noqa: E402
    SuiteResult,
    ensure_root_on_path,
    inspection_sample_data,
    make_outdir,
    sample_finance_payload,
    sample_ticketing_report_rows,
    write_stream,
)


def _assert_xlsx(path: str) -> None:
    if not os.path.isfile(path):
        raise RuntimeError(f"missing file: {path}")
    size = os.path.getsize(path)
    if size < 100:
        raise RuntimeError(f"xlsx too small ({size} bytes)")
    # ZIP magic for xlsx
    with open(path, "rb") as f:
        magic = f.read(2)
    if magic != b"PK":
        raise RuntimeError(f"not an xlsx/zip (got {magic!r})")
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    try:
        if not wb.sheetnames:
            raise RuntimeError("workbook has no sheets")
    finally:
        wb.close()


def _gen_inspection(out_dir: str, result: SuiteResult) -> None:
    """Fire Systems only — this product does not ship Civil / Cleaning forms."""
    insp_dir = os.path.join(out_dir, "inspection")
    os.makedirs(insp_dir, exist_ok=True)
    samples = inspection_sample_data()
    label = "inspection/fire_systems"
    try:
        from module_hvac_mep.hvac_generators import create_excel_report
        path = create_excel_report(samples["hvac"], insp_dir)
        full = path if os.path.isabs(path) else os.path.join(insp_dir, path)
        _assert_xlsx(full)
        final = os.path.join(insp_dir, "fire_systems.xlsx")
        if os.path.abspath(full) != os.path.abspath(final):
            if os.path.isfile(final):
                os.remove(final)
            os.replace(full, final)
        _assert_xlsx(final)
        result.pass_(label)
    except Exception as exc:
        result.fail(label, str(exc))


def _gen_finance(out_dir: str, result: SuiteResult) -> None:
    fin_dir = os.path.join(out_dir, "finance")
    os.makedirs(fin_dir, exist_ok=True)

    label = "finance/monthly_report"
    try:
        from module_finance.finance_report_builder import build_excel
        from openpyxl import load_workbook
        buf = build_excel(sample_finance_payload())
        path = os.path.join(fin_dir, "finance_monthly_report.xlsx")
        write_stream(path, buf)
        _assert_xlsx(path)
        wb = load_workbook(path, read_only=True)
        try:
            if "Dashboard" not in wb.sheetnames:
                raise RuntimeError(f"expected Dashboard sheet, got {wb.sheetnames}")
        finally:
            wb.close()
        result.pass_(label)
    except Exception as exc:
        result.fail(label, str(exc))

    # Contracts import template (mirrors finance routes sample)
    label = "finance/contracts_template"
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contracts"
        headers = [
            "Client Name", "Contract Ref", "Start Date", "End Date",
            "Value (AED)", "Account Handler", "Status", "Notes",
        ]
        fill = PatternFill("solid", fgColor="A8121E")
        font = Font(bold=True, color="FFFFFF")
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")
        example = [
            "Demo Client LLC", "CTR-TEST-001", "2026-01-01", "2026-12-31",
            120000, "Arshith", "active", "Suite sample row",
        ]
        for i, v in enumerate(example, start=1):
            ws.cell(row=2, column=i, value=v)
        path = os.path.join(fin_dir, "finance_contracts_sample_template.xlsx")
        wb.save(path)
        _assert_xlsx(path)
        result.pass_(label)
    except Exception as exc:
        result.fail(label, str(exc))


def _gen_ticketing(out_dir: str, result: SuiteResult) -> None:
    tkt_dir = os.path.join(out_dir, "ticketing")
    os.makedirs(tkt_dir, exist_ok=True)
    try:
        from module_ticketing.report_builders import REPORT_SPECS, build_report_excel
    except Exception as exc:
        result.fail("ticketing/reports", f"import failed: {exc}")
        return

    for key, spec in REPORT_SPECS.items():
        label = f"ticketing/report_{key}"
        try:
            rows, summary = sample_ticketing_report_rows(key)
            buf = BytesIO()
            build_report_excel(spec, rows, summary, "Test suite filters", buf)
            path = os.path.join(tkt_dir, f"report_{key}.xlsx")
            write_stream(path, buf)
            _assert_xlsx(path)
            result.pass_(label)
        except Exception as exc:
            result.fail(label, str(exc))


def _gen_overtime(out_dir: str, result: SuiteResult) -> None:
    ops_dir = os.path.join(out_dir, "operations")
    os.makedirs(ops_dir, exist_ok=True)
    label = "operations/overtime_export"
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Overtime"
        headers = [
            "Staff Name", "Employee ID", "Department", "Date",
            "Hours", "Rate Per Hour", "Total Amount", "Status", "Reason",
        ]
        fill = PatternFill("solid", fgColor="a8121e")
        font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        rows = [
            ["Arshith Swaminath P", "EMP-001", "MEP", str(date.today()),
             3, 50, 150, "approved", "Emergency AC repair — suite sample"],
            ["Ahmed Hassan", "EMP-042", "Operations", str(date.today()),
             2, 45, 90, "pending", "Weekend site cover"],
        ]
        for r_i, row in enumerate(rows, start=2):
            for c_i, val in enumerate(row, start=1):
                cell = ws.cell(row=r_i, column=c_i, value=val)
                cell.border = border

        path = os.path.join(ops_dir, "overtime_export_sample.xlsx")
        wb.save(path)
        _assert_xlsx(path)
        result.pass_(label)
    except Exception as exc:
        result.fail(label, str(exc))


def _gen_store(out_dir: str, result: SuiteResult) -> None:
    store_dir = os.path.join(out_dir, "store")
    os.makedirs(store_dir, exist_ok=True)
    label = "store/materials_sample"
    try:
        import pandas as pd

        df = pd.DataFrame([
            {
                "Material Name": "HEPA Filter 24x24",
                "Category": "HVAC",
                "Unit": "pcs",
                "Unit Price": 75.0,
                "Quantity": 20,
                "Supplier": "Emirates Supplies LLC",
                "Notes": "Suite sample row",
            },
            {
                "Material Name": "Coil Cleaner",
                "Category": "Chemicals",
                "Unit": "L",
                "Unit Price": 45.0,
                "Quantity": 10,
                "Supplier": "CleanCo",
                "Notes": "Suite sample row",
            },
        ])
        path = os.path.join(store_dir, "procurement_import_sample.xlsx")
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Materials")
        _assert_xlsx(path)
        result.pass_(label)
    except Exception as exc:
        result.fail(label, str(exc))


def main() -> int:
    ensure_root_on_path()
    out_dir = make_outdir("excel_suite")
    result = SuiteResult("Excel suite")
    print(f"\n{'=' * 70}")
    print("  Suite 2 — All Excel sheets with test data")
    print(f"  Output: {out_dir}")
    print(f"{'=' * 70}\n")

    print("── Inspection ──────────────────────────────────────────────────────")
    _gen_inspection(out_dir, result)
    print("\n── Finance ─────────────────────────────────────────────────────────")
    _gen_finance(out_dir, result)
    print("\n── Ticketing ───────────────────────────────────────────────────────")
    _gen_ticketing(out_dir, result)
    print("\n── Operations ──────────────────────────────────────────────────────")
    _gen_overtime(out_dir, result)
    print("\n── Store ───────────────────────────────────────────────────────────")
    _gen_store(out_dir, result)

    print(f"\nOutput folder: {out_dir}")
    return result.summary()


if __name__ == "__main__":
    sys.exit(main())
