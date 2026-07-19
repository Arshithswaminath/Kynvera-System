"""
Finance monthly report builder — period resolution, invoice lifecycle buckets,
closed-jobs lens, and professional Excel / PDF export.
"""
from __future__ import annotations

import io
import logging
import os
from calendar import monthrange
from datetime import date, datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.pdfgen.canvas import Canvas
from sqlalchemy import or_

from app.models import FinanceContract, Ticket, db

logger = logging.getLogger(__name__)

BRAND_DARK = colors.HexColor('#a8121e')
BRAND_ACCENT = colors.HexColor('#d21725')
ROW_ALT = colors.HexColor('#f8fafc')
GRID = colors.HexColor('#e2e8f0')
MUTED = colors.HexColor('#6b7280')
PAGE_W, PAGE_H = A4
MARGIN = 13 * mm

FINANCE_STATUSES = ('pending_finance', 'pending_gm_approval')

ROW_HEADERS = [
    'Work Order', 'Title', 'Project', 'Property', 'Account Handler',
    'Status', 'Chargeable', 'Amount (AED)', 'Invoice Ref', 'Date',
]


def _utcnow():
    return datetime.now(timezone.utc)


def _as_date(value) -> Optional[date]:
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s).date()
    except ValueError:
        return None


def resolve_period(
    month: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    year: Optional[int] = None,
    month_num: Optional[int] = None,
) -> Tuple[date, date, str]:
    """Resolve period_start, period_end, period_label.

    Custom date_from/date_to override month bounds when both are provided.
    """
    df = _as_date(date_from)
    dt = _as_date(date_to)
    if df and dt:
        if df > dt:
            df, dt = dt, df
        if df == dt:
            label = df.strftime('%d %b %Y')
        elif df.year == dt.year and df.month == dt.month:
            label = f"{df.strftime('%d')}–{dt.strftime('%d %b %Y')}"
        else:
            label = f"{df.strftime('%d %b %Y')} – {dt.strftime('%d %b %Y')}"
        return df, dt, label

    if month:
        try:
            y_str, m_str = month.split('-', 1)
            year = int(y_str)
            month_num = int(m_str)
        except (ValueError, AttributeError):
            pass

    today = date.today()
    year = int(year or today.year)
    month_num = int(month_num or today.month)
    start = date(year, month_num, 1)
    _, last = monthrange(year, month_num)
    end = date(year, month_num, last)
    return start, end, start.strftime('%B %Y')


def _ticket_amount(t: Ticket) -> float:
    return float(t.selling_price or t.total_cost or 0.0)


def _in_period(dt_val, start: date, end: date) -> bool:
    d = _as_date(dt_val)
    if not d:
        return False
    return start <= d <= end


def _row_from_ticket(t: Ticket, contracts_by_id: Dict[int, FinanceContract],
                     date_field: str = 'closed_at') -> Dict[str, Any]:
    contract = None
    if t.finance_contract_id:
        contract = contracts_by_id.get(t.finance_contract_id)
    handler = (contract.account_handler if contract else '') or ''
    amount = round(_ticket_amount(t), 2)
    date_val = getattr(t, date_field, None) or t.closed_at or t.finance_confirmed_at or t.gm_rejected_at
    return {
        'ticket_id': t.ticket_id,
        'title': t.title or '',
        'project': t.project or '',
        'property': t.property_name or '',
        'account_handler': handler,
        'status': t.status or '',
        'is_chargeable': bool(t.is_chargeable),
        'amount': amount,
        'selling_price': round(float(t.selling_price or 0), 2),
        'total_cost': round(float(t.total_cost or 0), 2),
        'invoice_ref': t.finance_invoice_ref or '',
        'finance_contract_id': t.finance_contract_id,
        'finance_confirmed_at': t.finance_confirmed_at.isoformat() if t.finance_confirmed_at else None,
        'gm_rejected_at': t.gm_rejected_at.isoformat() if t.gm_rejected_at else None,
        'closed_at': t.closed_at.isoformat() if t.closed_at else None,
        'date': date_val.isoformat() if date_val else None,
        'service_group': t.service_group or '',
    }


def _sum_amount(rows: List[Dict]) -> float:
    return round(sum(float(r.get('amount') or 0) for r in rows), 2)


def query_relevant_tickets(
    period_start: date,
    period_end: date,
    project: Optional[str] = None,
    account_handler: Optional[str] = None,
) -> List[Ticket]:
    """Load tickets that touch the finance pipeline or closed-jobs lens in period."""
    q = Ticket.query.filter(
        or_(
            # Currently pending finance/GM — scoped by updated_at within the period
            db.and_(
                Ticket.status.in_(list(FINANCE_STATUSES)),
                db.func.date(Ticket.updated_at) >= period_start,
                db.func.date(Ticket.updated_at) <= period_end,
            ),
            db.and_(
                Ticket.finance_confirmed_at.isnot(None),
                db.func.date(Ticket.finance_confirmed_at) >= period_start,
                db.func.date(Ticket.finance_confirmed_at) <= period_end,
            ),
            db.and_(
                Ticket.gm_rejected_at.isnot(None),
                db.func.date(Ticket.gm_rejected_at) >= period_start,
                db.func.date(Ticket.gm_rejected_at) <= period_end,
            ),
            db.and_(
                Ticket.status.in_(['closed', 'resolved']),
                Ticket.closed_at.isnot(None),
                db.func.date(Ticket.closed_at) >= period_start,
                db.func.date(Ticket.closed_at) <= period_end,
            ),
        )
    )

    if project:
        q = q.filter(Ticket.project == project)

    tickets = q.all()

    if account_handler:
        handler_norm = account_handler.strip().lower()
        # Prefetch contracts for linked tickets + project-name matches
        contract_ids = {t.finance_contract_id for t in tickets if t.finance_contract_id}
        contracts = {
            c.id: c for c in FinanceContract.query.filter(
                FinanceContract.id.in_(contract_ids)
            ).all()
        } if contract_ids else {}
        # Also find contracts with this handler (for project-name join)
        handler_contracts = FinanceContract.query.filter(
            db.func.lower(FinanceContract.account_handler) == handler_norm
        ).all()
        handler_projects = {
            (c.client_name or '').strip().lower()
            for c in handler_contracts if c.client_name
        }

        filtered = []
        for t in tickets:
            contract = contracts.get(t.finance_contract_id) if t.finance_contract_id else None
            if contract and (contract.account_handler or '').strip().lower() == handler_norm:
                filtered.append(t)
                continue
            proj = (t.project or '').strip().lower()
            if proj and proj in handler_projects:
                filtered.append(t)
        tickets = filtered

    return tickets


def classify_and_aggregate(
    tickets: List[Ticket],
    period_start: date,
    period_end: date,
    period_label: str,
    filters: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Classify tickets into lifecycle buckets + closed-jobs lens."""
    contract_ids = {t.finance_contract_id for t in tickets if t.finance_contract_id}
    contracts_by_id = {
        c.id: c for c in FinanceContract.query.filter(
            FinanceContract.id.in_(contract_ids)
        ).all()
    } if contract_ids else {}

    # Map project → handler from contracts for unlinked tickets
    all_handlers = FinanceContract.query.filter(
        FinanceContract.account_handler.isnot(None),
        FinanceContract.account_handler != '',
    ).all()
    project_handler = {}
    for c in all_handlers:
        key = (c.client_name or '').strip().lower()
        if key and key not in project_handler:
            project_handler[key] = c.account_handler

    created: List[Dict] = []
    registered: List[Dict] = []
    not_registered: List[Dict] = []
    rejected: List[Dict] = []
    closed_jobs: List[Dict] = []
    seen_created = set()
    seen_registered = set()
    seen_not_reg = set()
    seen_rejected = set()
    seen_closed = set()

    for t in tickets:
        is_registered = _in_period(t.finance_confirmed_at, period_start, period_end)
        is_rejected = _in_period(t.gm_rejected_at, period_start, period_end)
        is_pending = (
            t.status in FINANCE_STATUSES
            and not t.finance_confirmed_at
            and _in_period(t.updated_at, period_start, period_end)
        )
        is_created = is_pending or is_registered or is_rejected
        is_closed = (
            t.status in ('closed', 'resolved')
            and _in_period(t.closed_at, period_start, period_end)
        )

        if is_created and t.id not in seen_created:
            seen_created.add(t.id)
            row = _row_from_ticket(
                t, contracts_by_id,
                date_field='finance_confirmed_at' if is_registered
                else ('gm_rejected_at' if is_rejected else 'closed_at'),
            )
            if not row['account_handler']:
                row['account_handler'] = project_handler.get(
                    (t.project or '').strip().lower(), '')
            created.append(row)

        if is_registered and t.id not in seen_registered:
            seen_registered.add(t.id)
            row = _row_from_ticket(t, contracts_by_id, date_field='finance_confirmed_at')
            if not row['account_handler']:
                row['account_handler'] = project_handler.get(
                    (t.project or '').strip().lower(), '')
            registered.append(row)

        if is_pending and t.id not in seen_not_reg:
            seen_not_reg.add(t.id)
            row = _row_from_ticket(t, contracts_by_id)
            if not row['account_handler']:
                row['account_handler'] = project_handler.get(
                    (t.project or '').strip().lower(), '')
            not_registered.append(row)

        if is_rejected and t.id not in seen_rejected:
            seen_rejected.add(t.id)
            row = _row_from_ticket(t, contracts_by_id, date_field='gm_rejected_at')
            if not row['account_handler']:
                row['account_handler'] = project_handler.get(
                    (t.project or '').strip().lower(), '')
            rejected.append(row)

        if is_closed and t.id not in seen_closed:
            seen_closed.add(t.id)
            row = _row_from_ticket(t, contracts_by_id, date_field='closed_at')
            if not row['account_handler']:
                row['account_handler'] = project_handler.get(
                    (t.project or '').strip().lower(), '')
            row['invoiced'] = bool(t.finance_confirmed_at)
            row['gm_rejected'] = bool(t.gm_rejected_at)
            closed_jobs.append(row)

    closed_invoiced = [r for r in closed_jobs if r.get('invoiced')]
    closed_not_invoiced = [r for r in closed_jobs if not r.get('invoiced')]
    closed_rejected = [r for r in closed_jobs if r.get('gm_rejected')]

    filters = filters or {}
    summary = {
        'period_label': period_label,
        'period_start': period_start.isoformat(),
        'period_end': period_end.isoformat(),
        'created_count': len(created),
        'created_value': _sum_amount(created),
        'registered_count': len(registered),
        'registered_value': _sum_amount(registered),
        'not_registered_count': len(not_registered),
        'not_registered_value': _sum_amount(not_registered),
        'rejected_count': len(rejected),
        'rejected_value': _sum_amount(rejected),
        'closed_jobs_count': len(closed_jobs),
        'closed_jobs_value': _sum_amount(closed_jobs),
        'closed_invoiced_count': len(closed_invoiced),
        'closed_invoiced_value': _sum_amount(closed_invoiced),
        'closed_not_invoiced_count': len(closed_not_invoiced),
        'closed_not_invoiced_value': _sum_amount(closed_not_invoiced),
        'closed_rejected_count': len(closed_rejected),
        'filters': {
            'project': filters.get('project') or '',
            'account_handler': filters.get('account_handler') or '',
            'month': filters.get('month') or '',
            'date_from': filters.get('date_from') or '',
            'date_to': filters.get('date_to') or '',
        },
    }

    return {
        'summary': summary,
        'created': created,
        'registered': registered,
        'not_registered': not_registered,
        'rejected': rejected,
        'closed_jobs': closed_jobs,
    }


def build_dashboard_payload(
    month: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    year: Optional[int] = None,
    month_num: Optional[int] = None,
    project: Optional[str] = None,
    account_handler: Optional[str] = None,
) -> Dict[str, Any]:
    period_start, period_end, period_label = resolve_period(
        month=month, date_from=date_from, date_to=date_to,
        year=year, month_num=month_num,
    )
    filters = {
        'project': (project or '').strip() or None,
        'account_handler': (account_handler or '').strip() or None,
        'month': month or '',
        'date_from': date_from or '',
        'date_to': date_to or '',
    }
    tickets = query_relevant_tickets(
        period_start, period_end,
        project=filters['project'],
        account_handler=filters['account_handler'],
    )
    return classify_and_aggregate(
        tickets, period_start, period_end, period_label, filters=filters,
    )


def get_filter_options() -> Dict[str, List[str]]:
    projects = (
        db.session.query(Ticket.project)
        .filter(Ticket.project.isnot(None), Ticket.project != '')
        .distinct()
        .order_by(Ticket.project)
        .all()
    )
    handlers = (
        db.session.query(FinanceContract.account_handler)
        .filter(
            FinanceContract.account_handler.isnot(None),
            FinanceContract.account_handler != '',
        )
        .distinct()
        .order_by(FinanceContract.account_handler)
        .all()
    )
    return {
        'projects': [p[0] for p in projects if p[0]],
        'account_handlers': [h[0] for h in handlers if h[0]],
    }


def filters_label(summary: Dict) -> str:
    parts = [summary.get('period_label') or '']
    f = summary.get('filters') or {}
    if f.get('project'):
        parts.append(f"Project: {f['project']}")
    if f.get('account_handler'):
        parts.append(f"Handler: {f['account_handler']}")
    if f.get('date_from') and f.get('date_to'):
        parts.append(f"{f['date_from']} → {f['date_to']}")
    return ' · '.join(p for p in parts if p)


# ─── Excel ───────────────────────────────────────────────────────────

XL_BRAND = 'A8121E'
XL_HEAD_FILL = PatternFill('solid', fgColor=XL_BRAND)
XL_ALT_FILL = PatternFill('solid', fgColor='F8FAFC')
XL_META_FONT = Font(name='Calibri', size=9, color='6B7280')
XL_HEAD_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
XL_TITLE_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=14)
XL_KPI_FONT = Font(name='Calibri', bold=True, size=20, color=XL_BRAND)
XL_LABEL_FONT = Font(name='Calibri', size=9, color='6B7280')
XL_THIN = Side(style='thin', color='E2E8F0')
XL_BORDER = Border(left=XL_THIN, right=XL_THIN, top=XL_THIN, bottom=XL_THIN)


def _rows_to_excel_data(rows: List[Dict]) -> List[List]:
    out = []
    for r in rows:
        out.append([
            r.get('ticket_id') or '',
            r.get('title') or '',
            r.get('project') or '',
            r.get('property') or '',
            r.get('account_handler') or '',
            (r.get('status') or '').replace('_', ' ').title(),
            'Yes' if r.get('is_chargeable') else 'No',
            float(r.get('amount') or 0),
            r.get('invoice_ref') or '',
            (r.get('date') or '')[:10],
        ])
    return out


def _write_data_sheet(wb: Workbook, name: str, title: str, rows: List[Dict], meta: str):
    safe = name[:31]
    ws = wb.create_sheet(title=safe)
    ncols = len(ROW_HEADERS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=f'AMAAN — {title}')
    c.font = XL_TITLE_FONT
    c.fill = XL_HEAD_FILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28
    for i in range(2, ncols + 1):
        ws.cell(row=1, column=i).fill = XL_HEAD_FILL

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    m = ws.cell(row=2, column=1, value=f'{meta}  ·  {len(rows)} rows')
    m.font = XL_META_FONT

    for i, h in enumerate(ROW_HEADERS, start=1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = XL_HEAD_FONT
        cell.fill = XL_HEAD_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = XL_BORDER

    for ri, row in enumerate(_rows_to_excel_data(rows)):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=5 + ri, column=ci, value=val)
            cell.border = XL_BORDER
            cell.font = Font(name='Calibri', size=9)
            if ri % 2:
                cell.fill = XL_ALT_FILL
            if ci == 8 and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

    widths = [14, 28, 18, 16, 16, 14, 10, 12, 14, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if not rows:
        ws.cell(row=5, column=1, value='No data for the selected filters.').font = XL_META_FONT


def _embed_excel_logo(ws, row=1, col=1):
    """Place AMAAN logo in the worksheet header band."""
    logo_path = _find_logo()
    if not logo_path:
        return False
    try:
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(logo_path)
        # Keep header-friendly size
        img.width = 72
        img.height = 44
        # Anchor at A1 (col/row are 1-based for cell ref)
        cell = f'{get_column_letter(col)}{row}'
        ws.add_image(img, cell)
        return True
    except Exception as exc:
        logger.warning('Could not embed Excel logo: %s', exc)
        return False


def build_excel(payload: Dict[str, Any]) -> io.BytesIO:
    """Multi-sheet workbook: Dashboard + bucket sheets."""
    summary = payload['summary']
    meta = filters_label(summary)
    wb = Workbook()

    # Dashboard sheet — white logo plate (A) + red title band (B–F)
    ws = wb.active
    ws.title = 'Dashboard'
    ws.row_dimensions[1].height = 48
    logo_cell = ws.cell(row=1, column=1, value='')
    logo_cell.fill = PatternFill('solid', fgColor='FFFFFF')
    logo_cell.border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB'),
    )
    ws.merge_cells('B1:F1')
    c = ws.cell(row=1, column=2, value='AMAAN — Finance Monthly Report Dashboard')
    c.font = XL_TITLE_FONT
    c.fill = XL_HEAD_FILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    for i in range(3, 7):
        ws.cell(row=1, column=i).fill = XL_HEAD_FILL
    ws.column_dimensions['A'].width = 14
    _embed_excel_logo(ws, row=1, col=1)

    ws.merge_cells('A2:F2')
    gen_on = _utcnow().strftime('%d %b %Y, %H:%M UTC')
    ws.cell(row=2, column=1, value=f'Generated {gen_on}  ·  {meta}').font = XL_META_FONT

    ws.cell(row=4, column=1, value='Invoice Lifecycle').font = Font(
        name='Calibri', bold=True, size=13, color=XL_BRAND)

    kpis = [
        ('Created', summary['created_count'], summary['created_value']),
        ('Registered', summary['registered_count'], summary['registered_value']),
        ('Not Registered', summary['not_registered_count'], summary['not_registered_value']),
        ('Rejected', summary['rejected_count'], summary['rejected_value']),
    ]
    for i, (label, count, value) in enumerate(kpis):
        col = 1 + i
        ws.cell(row=5, column=col, value=label).font = XL_LABEL_FONT
        cell = ws.cell(row=6, column=col, value=count)
        cell.font = XL_KPI_FONT
        cell.alignment = Alignment(horizontal='center')
        ws.cell(row=7, column=col, value=f'AED {value:,.2f}').font = Font(
            name='Calibri', size=9, color='374151')
        for r in (5, 6, 7):
            ws.cell(row=r, column=col).fill = XL_ALT_FILL
            ws.cell(row=r, column=col).border = XL_BORDER
            ws.cell(row=r, column=col).alignment = Alignment(horizontal='center')

    ws.cell(row=9, column=1, value='Closed Jobs Lens').font = Font(
        name='Calibri', bold=True, size=13, color=XL_BRAND)

    closed_kpis = [
        ('Closed Jobs', summary['closed_jobs_count'], summary['closed_jobs_value']),
        ('Invoiced', summary['closed_invoiced_count'], summary['closed_invoiced_value']),
        ('Not Invoiced', summary['closed_not_invoiced_count'], summary['closed_not_invoiced_value']),
        ('GM Rejected', summary['closed_rejected_count'], 0),
    ]
    for i, (label, count, value) in enumerate(closed_kpis):
        col = 1 + i
        ws.cell(row=10, column=col, value=label).font = XL_LABEL_FONT
        cell = ws.cell(row=11, column=col, value=count)
        cell.font = XL_KPI_FONT
        cell.alignment = Alignment(horizontal='center')
        if value or label != 'GM Rejected':
            ws.cell(row=12, column=col, value=f'AED {value:,.2f}').font = Font(
                name='Calibri', size=9, color='374151')
        for r in (10, 11, 12):
            ws.cell(row=r, column=col).fill = XL_ALT_FILL
            ws.cell(row=r, column=col).border = XL_BORDER
            ws.cell(row=r, column=col).alignment = Alignment(horizontal='center')

    # Distribution table for chart data
    ws.cell(row=14, column=1, value='Distribution').font = Font(
        name='Calibri', bold=True, size=12, color=XL_BRAND)
    dist = [
        ('Registered', summary['registered_count']),
        ('Not Registered', summary['not_registered_count']),
        ('Rejected', summary['rejected_count']),
    ]
    ws.cell(row=15, column=1, value='Bucket').font = XL_HEAD_FONT
    ws.cell(row=15, column=1).fill = XL_HEAD_FILL
    ws.cell(row=15, column=2, value='Count').font = XL_HEAD_FONT
    ws.cell(row=15, column=2).fill = XL_HEAD_FILL
    for i, (label, count) in enumerate(dist):
        ws.cell(row=16 + i, column=1, value=label).border = XL_BORDER
        ws.cell(row=16 + i, column=2, value=count).border = XL_BORDER

    try:
        from openpyxl.chart import PieChart, Reference
        from openpyxl.chart.label import DataLabelList
        pie = PieChart()
        pie.title = 'Invoice Status'
        labels = Reference(ws, min_col=1, min_row=16, max_row=18)
        data = Reference(ws, min_col=2, min_row=15, max_row=18)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.width = 12
        pie.height = 8
        ws.add_chart(pie, 'D14')
    except Exception as exc:
        logger.warning('Finance dashboard pie chart skipped: %s', exc)

    for i, w in enumerate([18, 14, 14, 14, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _write_data_sheet(wb, 'Created', 'Created Invoices', payload.get('created') or [], meta)
    _write_data_sheet(wb, 'Registered', 'Registered Invoices', payload.get('registered') or [], meta)
    _write_data_sheet(wb, 'Not Registered', 'Not Registered Invoices',
                      payload.get('not_registered') or [], meta)
    _write_data_sheet(wb, 'Rejected', 'Rejected Invoices', payload.get('rejected') or [], meta)
    _write_data_sheet(wb, 'Closed Jobs', 'Closed Jobs', payload.get('closed_jobs') or [], meta)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── PDF ─────────────────────────────────────────────────────────────

def _find_logo() -> Optional[str]:
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    for rel in (
        'static/icons/Amaan-logo-report.png',
        'static/icons/Amaan-logo-tight.png',
        'static/icons/Amaan-mark.png',
        'static/logo.png',
        'static/icons/Amaan.png',
        'static/img/logo.png',
    ):
        path = os.path.join(root, rel)
        if os.path.exists(path):
            return path
    return None


class _FinanceReportCanvas(Canvas):
    def __init__(self, *args, **kwargs):
        self._logo_path = kwargs.pop('logo_path', None)
        self._report_title = kwargs.pop('report_title', 'Finance Report')
        Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self._draw_chrome(num_pages)
            Canvas.showPage(self)
        Canvas.save(self)

    def _draw_chrome(self, num_pages=1):
        self.saveState()
        bar_h = 26 * mm
        bar_bottom = PAGE_H - bar_h
        bar_mid = bar_bottom + bar_h / 2

        # Split header: white logo plate (left) + red title band (right)
        white_w = 48 * mm
        self.setFillColor(colors.white)
        self.rect(0, bar_bottom, white_w, bar_h, fill=1, stroke=0)
        self.setFillColor(BRAND_DARK)
        self.rect(white_w, bar_bottom, PAGE_W - white_w, bar_h, fill=1, stroke=0)

        # Subtle seam between white / red
        self.setStrokeColor(colors.HexColor('#e5e7eb'))
        self.setLineWidth(0.6)
        self.line(white_w, bar_bottom, white_w, PAGE_H)

        # Logo centered in the white plate
        logo_w, logo_h = 30 * mm, 20 * mm
        logo_drawn = False
        if self._logo_path and os.path.exists(self._logo_path):
            try:
                self.drawImage(
                    self._logo_path,
                    (white_w - logo_w) / 2,
                    bar_mid - logo_h / 2,
                    width=logo_w, height=logo_h,
                    preserveAspectRatio=True, mask='auto',
                )
                logo_drawn = True
            except Exception:
                pass

        if not logo_drawn:
            self.setFillColor(BRAND_DARK)
            self.setFont('Helvetica-Bold', 12)
            self.drawCentredString(white_w / 2, bar_mid - 2 * mm, 'AMAAN')

        # Title on the red side
        text_x = white_w + 8 * mm
        self.setFillColor(colors.white)
        self.setFont('Helvetica-Bold', 12)
        self.drawString(text_x, bar_mid + 1.2 * mm, (self._report_title or 'Finance Report').upper())
        self.setFont('Helvetica', 8)
        self.setFillColor(colors.Color(1, 1, 1, alpha=0.82))
        self.drawString(text_x, bar_mid - 4.2 * mm, 'AMAAN Finance')

        # Footer
        self.setStrokeColor(GRID)
        self.setLineWidth(0.5)
        self.line(MARGIN, 12 * mm, PAGE_W - MARGIN, 12 * mm)
        self.setFillColor(MUTED)
        self.setFont('Helvetica', 8)
        self.drawString(MARGIN, 7 * mm, 'Amaan Application · Finance Report')
        self.drawRightString(PAGE_W - MARGIN, 7 * mm, f'Page {self._pageNumber} of {num_pages}')
        self.restoreState()


def _pdf_escape(text) -> str:
    return (
        str(text or '')
        .replace('&', '&amp;')
        .replace('<', '&lt;')
        .replace('>', '&gt;')
    )


def _section_table(rows: List[Dict], cell_style, head_style, num_style) -> Table:
    cols = [
        ('ticket_id', 'Work Order', 1.8),
        ('title', 'Title', 2.2),
        ('project', 'Project', 1.6),
        ('account_handler', 'Handler', 1.4),
        ('status', 'Status', 1.3),
        ('amount', 'Amount', 1.2),
        ('invoice_ref', 'Invoice Ref', 1.3),
    ]
    avail = PAGE_W - 2 * MARGIN
    total_w = sum(w for _, _, w in cols)
    col_widths = [avail * w / total_w for _, _, w in cols]

    data = [[Paragraph(h, head_style) for _, h, _ in cols]]
    for r in rows:
        data.append([
            Paragraph(_pdf_escape(r.get('ticket_id')), cell_style),
            Paragraph(_pdf_escape((r.get('title') or '')[:60]), cell_style),
            Paragraph(_pdf_escape(r.get('project')), cell_style),
            Paragraph(_pdf_escape(r.get('account_handler')), cell_style),
            Paragraph(_pdf_escape((r.get('status') or '').replace('_', ' ').title()), cell_style),
            Paragraph(f"{float(r.get('amount') or 0):,.2f}", num_style),
            Paragraph(_pdf_escape(r.get('invoice_ref')), cell_style),
        ])
    if len(data) == 1:
        data.append(
            [Paragraph('No data for the selected filters.', cell_style)]
            + [Paragraph('', cell_style)] * (len(cols) - 1)
        )

    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), BRAND_DARK),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, ROW_ALT]),
        ('LINEBELOW', (0, 0), (-1, 0), 0.8, BRAND_DARK),
        ('LINEBELOW', (0, 1), (-1, -1), 0.4, GRID),
        ('BOX', (0, 0), (-1, -1), 0.6, GRID),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 3.5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3.5),
    ]))
    return tbl


def build_pdf(payload: Dict[str, Any]) -> io.BytesIO:
    summary = payload['summary']
    meta = filters_label(summary)
    buf = io.BytesIO()

    base = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        'FCell', parent=base['Normal'], fontSize=7.5, leading=9.5,
        textColor=colors.HexColor('#1f2937'))
    head_style = ParagraphStyle(
        'FHead', parent=base['Normal'], fontSize=7.5, leading=9,
        fontName='Helvetica-Bold', textColor=colors.white)
    num_style = ParagraphStyle('FNum', parent=cell_style, alignment=2)
    meta_style = ParagraphStyle(
        'FMeta', parent=base['Normal'], fontSize=8, leading=11, textColor=MUTED)
    h_style = ParagraphStyle(
        'FH', parent=base['Heading2'], fontSize=12, leading=15,
        textColor=BRAND_DARK, spaceBefore=8, spaceAfter=6)
    label_style = ParagraphStyle(
        'ChipLabel', parent=base['Normal'], fontSize=7, leading=8.5,
        alignment=1, fontName='Helvetica-Bold', textColor=MUTED)

    def value_style(text):
        length = len(str(text))
        fs = 16 if length <= 4 else (13 if length <= 8 else 10)
        return ParagraphStyle(
            'ChipValue', parent=base['Normal'], fontSize=fs, leading=fs + 1,
            alignment=1, fontName='Helvetica-Bold', textColor=colors.HexColor('#111827'))

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=34 * mm, bottomMargin=16 * mm,
        title=f"Finance Report — {summary.get('period_label')}",
    )

    story = []
    gen_on = _utcnow().strftime('%d %b %Y, %H:%M UTC')
    story.append(Paragraph(
        f"Generated {gen_on} &nbsp;·&nbsp; {_pdf_escape(meta)}", meta_style))
    story.append(Spacer(1, 5 * mm))

    # Primary KPI chips
    primary = [
        ('Created', summary['created_count']),
        ('Registered', summary['registered_count']),
        ('Not Registered', summary['not_registered_count']),
        ('Rejected', summary['rejected_count']),
    ]
    n = len(primary)
    gap = 3 * mm
    avail = PAGE_W - 2 * MARGIN
    chip_w = (avail - gap * (n - 1)) / n
    cells, col_widths = [], []
    for i, (label, value) in enumerate(primary):
        cells.append([
            Paragraph(label.upper(), label_style),
            Spacer(1, 3),
            Paragraph(str(value), value_style(value)),
        ])
        col_widths.append(chip_w)
        if i < n - 1:
            cells.append('')
            col_widths.append(gap)

    chip_tbl = Table([cells], colWidths=col_widths)
    style = [
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 9),
    ]
    for ci in range(0, len(col_widths), 2):
        style.append(('BACKGROUND', (ci, 0), (ci, 0), colors.HexColor('#fbfcfe')))
        style.append(('BOX', (ci, 0), (ci, 0), 0.7, GRID))
        style.append(('LINEABOVE', (ci, 0), (ci, 0), 2, BRAND_ACCENT))
    chip_tbl.setStyle(TableStyle(style))
    story.append(chip_tbl)
    story.append(Spacer(1, 4 * mm))

    # Value line
    story.append(Paragraph(
        f"Registered value: <b>AED {summary['registered_value']:,.2f}</b>"
        f" &nbsp;·&nbsp; Closed jobs: <b>{summary['closed_jobs_count']}</b>"
        f" (AED {summary['closed_jobs_value']:,.2f})"
        f" &nbsp;·&nbsp; Invoiced: <b>{summary['closed_invoiced_count']}</b>"
        f" &nbsp;·&nbsp; Not invoiced: <b>{summary['closed_not_invoiced_count']}</b>",
        meta_style,
    ))
    story.append(Spacer(1, 6 * mm))

    sections = [
        ('Created', payload.get('created') or []),
        ('Registered', payload.get('registered') or []),
        ('Not Registered', payload.get('not_registered') or []),
        ('Rejected', payload.get('rejected') or []),
        ('Closed Jobs', payload.get('closed_jobs') or []),
    ]

    for idx, (title, rows) in enumerate(sections):
        if idx > 0:
            story.append(PageBreak())
        story.append(Paragraph(
            f"{title} ({len(rows)}) — AED {_sum_amount(rows):,.2f}", h_style))
        story.append(_section_table(rows, cell_style, head_style, num_style))

    logo = _find_logo()
    doc.build(
        story,
        canvasmaker=lambda *a, **kw: _FinanceReportCanvas(
            *a, logo_path=logo,
            report_title=f"Finance — {summary.get('period_label', '')}",
            **kw),
    )
    buf.seek(0)
    return buf


def jobs_compat_list(payload: Dict[str, Any]) -> List[Dict]:
    """Backward-compatible job list (closed jobs) for email / legacy UI."""
    jobs = []
    for r in payload.get('closed_jobs') or []:
        jobs.append({
            'ticket_id': r.get('ticket_id'),
            'title': r.get('title'),
            'project': r.get('project'),
            'property': r.get('property'),
            'service_group': r.get('service_group'),
            'is_chargeable': r.get('is_chargeable'),
            'total_cost': r.get('total_cost'),
            'selling_price': r.get('selling_price') or r.get('amount'),
            'margin_pct': None,
            'closed_at': r.get('closed_at'),
            'finance_contract_id': r.get('finance_contract_id'),
            'account_handler': r.get('account_handler'),
            'invoice_ref': r.get('invoice_ref'),
            'invoiced': r.get('invoiced'),
            'gm_rejected': r.get('gm_rejected'),
        })
    return jobs
