"""Excel register export for visible work orders."""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from module_ticketing.tz_utils import to_gst

HEADER_FILL = PatternFill("solid", fgColor="FF8E68")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN = Border(
    left=Side(style="thin", color="E9EAEE"),
    right=Side(style="thin", color="E9EAEE"),
    top=Side(style="thin", color="E9EAEE"),
    bottom=Side(style="thin", color="E9EAEE"),
)

HEADERS = (
    "Ticket ID",
    "Title",
    "Project",
    "Location",
    "Service Group",
    "Fault",
    "Priority",
    "Status",
    "Reporter",
    "Supervisor",
    "Assigned To",
    "Created (GST)",
    "Closed (GST)",
    "Chargeable",
    "Actual Price",
    "Selling Price",
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _cell(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _dt(v) -> str:
    gst = to_gst(v)
    if gst is None:
        return ""
    try:
        return gst.strftime("%d %b %Y %H:%M")
    except Exception:
        return _cell(v)


def _money(v) -> str:
    if v is None:
        return ""
    try:
        return f"{float(v):.2f}"
    except (TypeError, ValueError):
        return ""


def _location_path(ticket) -> str:
    return " / ".join(
        p for p in [
            getattr(ticket, "property_name", None),
            getattr(ticket, "zone", None),
            getattr(ticket, "sub_zone", None),
            getattr(ticket, "base_unit", None),
        ] if p
    )


def _name(user) -> str:
    if user is None:
        return ""
    return (user.full_name or user.username or "").strip()


def build_ticket_register(tickets: list) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"
    for i, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, i, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    ws.row_dimensions[1].height = 22

    for r, t in enumerate(tickets or [], 2):
        row = [
            _cell(t.ticket_id),
            _cell(t.title),
            _cell(t.project),
            _location_path(t),
            _cell(t.service_group),
            _cell(t.fault_type),
            _cell(t.priority),
            _cell(t.status).replace("_", " "),
            _name(getattr(t, "reporter", None)),
            _name(getattr(t, "supervisor", None)),
            _name(getattr(t, "assigned_to", None)),
            _dt(t.created_at),
            _dt(t.closed_at),
            "Yes" if t.is_chargeable else "No",
            _money(t.actual_price if t.actual_price is not None else t.total_cost),
            _money(t.selling_price),
        ]
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.border = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=c in (2, 4))

    widths = [14, 32, 18, 28, 16, 18, 10, 16, 16, 16, 16, 18, 18, 12, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
