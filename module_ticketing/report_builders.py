"""
Generic report engine for the ticketing module.

Each report is described by a spec in REPORT_SPECS (title, columns, tier).
Two spec-driven builders render any report to PDF (ReportLab, landscape A4)
or Excel (openpyxl) with consistent AMAAN branding — adding a new report
only requires a new spec plus a row-builder in routes.py.
"""
import os
from datetime import datetime, timezone

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfgen.canvas import Canvas

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Brand colours (match ticket_pdf_builder.py)
BRAND_DARK = colors.HexColor('#a8121e')
BRAND_ACCENT = colors.HexColor('#d21725')
ROW_ALT = colors.HexColor('#f8fafc')
GRID = colors.HexColor('#e2e8f0')
MUTED = colors.HexColor('#6b7280')

PAGE_W, PAGE_H = A4
MARGIN = 13 * mm


# ---------------------------------------------------------------------------
# Report catalog
# ---------------------------------------------------------------------------
# Column spec: (key, header, weight, format)
#   format: 'text' | 'money' | 'int' | 'pct'
REPORT_SPECS = {
    'work_orders': {
        'title': 'Work Orders Summary',
        'description': 'Every work order in scope with status, priority, team and key dates.',
        'tier': 'basic',
        'filters': ['date', 'project', 'status'],
        'columns': [
            ('ticket_id', 'Ticket ID', 2.25, 'text'),
            ('title', 'Title', 1.95, 'text'),
            ('project', 'Project', 1.9, 'text'),
            ('service_group', 'Service Group', 1.7, 'text'),
            ('priority', 'Priority', 1.65, 'text'),
            ('status', 'Status', 1.8, 'text'),
            ('supervisor', 'Supervisor', 1.9, 'text'),
            ('technician', 'Technician', 1.8, 'text'),
            ('created', 'Created', 1.7, 'text'),
            ('closed', 'Closed', 1.7, 'text'),
        ],
    },
    'financial': {
        'title': 'Financial / Revenue Report',
        'description': 'Chargeable work orders with costs, markup, selling price, margin and ERP invoice references.',
        'tier': 'finance',
        'filters': ['date', 'project'],
        'columns': [
            ('ticket_id', 'Ticket ID', 1.6, 'text'),
            ('project', 'Project', 1.9, 'text'),
            ('title', 'Title', 2.1, 'text'),
            ('total_cost', 'Cost (AED)', 1.2, 'money'),
            ('actual_price', 'Actual (AED)', 1.2, 'money'),
            ('markup_pct', 'Markup %', 1.0, 'pct'),
            ('selling_price', 'Selling (AED)', 1.2, 'money'),
            ('margin', 'Margin (AED)', 1.2, 'money'),
            ('invoice_ref', 'ERP Invoice', 1.3, 'text'),
            ('closed', 'Closed', 1.2, 'text'),
        ],
    },
    'projects': {
        'title': 'Project Performance',
        'description': 'Per-project totals: work order counts, revenue, cost and average days to close.',
        'tier': 'finance',
        'filters': ['date'],
        'columns': [
            ('project', 'Project', 2.2, 'text'),
            ('total', 'Work Orders', 1.0, 'int'),
            ('open', 'Active', 0.9, 'int'),
            ('closed', 'Closed', 0.9, 'int'),
            ('revenue', 'Revenue (AED)', 1.2, 'money'),
            ('cost', 'Cost (AED)', 1.2, 'money'),
            ('margin', 'Margin (AED)', 1.2, 'money'),
            ('avg_close_days', 'Avg Days to Close', 1.2, 'text'),
        ],
    },
    'team': {
        'title': 'Team Performance',
        'description': 'Per-person workload: assigned and completed orders, hours logged and labour cost.',
        'tier': 'basic',
        'filters': ['date', 'project'],
        'columns': [
            ('name', 'Team Member', 1.8, 'text'),
            ('role', 'Role', 1.2, 'text'),
            ('assigned', 'Assigned', 0.9, 'int'),
            ('completed', 'Completed', 0.9, 'int'),
            ('hours', 'Hours Logged', 1.0, 'text'),
            ('labour_cost', 'Labour Cost (AED)', 1.3, 'money'),
        ],
    },
    'materials': {
        'title': 'Materials Usage',
        'description': 'Aggregated materials consumption: quantities, average unit price and total spend.',
        'tier': 'basic',
        'filters': ['date', 'project'],
        'columns': [
            ('material', 'Material', 2.4, 'text'),
            ('unit', 'Unit', 0.8, 'text'),
            ('quantity', 'Total Qty', 1.0, 'text'),
            ('avg_price', 'Avg Unit Price (AED)', 1.3, 'money'),
            ('spend', 'Total Spend (AED)', 1.3, 'money'),
            ('tickets', 'Used On (orders)', 1.1, 'int'),
        ],
    },
}


def _find_logo():
    for c in ('static/icons/AMAAN Logo - Edited.png', 'static/icons/icon-144x144.png'):
        if os.path.exists(c):
            return c
    return None


def _fmt_cell(value, fmt):
    if value is None or value == '':
        return '—'
    if fmt == 'money':
        try:
            return f'{float(value):,.2f}'
        except (TypeError, ValueError):
            return str(value)
    if fmt == 'pct':
        try:
            return f'{float(value):.0f}%'
        except (TypeError, ValueError):
            return str(value)
    return str(value)


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

class _ReportCanvas(Canvas):
    """Landscape canvas with brand header bar, logo and page numbering."""

    def __init__(self, *args, **kwargs):
        self._logo_path = kwargs.pop('logo_path', None)
        self._report_title = kwargs.pop('report_title', '')
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self._draw_page(num_pages)
            super().showPage()
        super().save()

    def _draw_page(self, num_pages):
        self.saveState()

        bar_h = 26 * mm
        bar_top = PAGE_H
        bar_bottom = PAGE_H - bar_h
        bar_mid = bar_bottom + bar_h / 2

        # Header background — solid red band
        self.setFillColor(BRAND_DARK)
        self.rect(0, bar_bottom, PAGE_W, bar_h, fill=1, stroke=0)

        # Logo — left, vertically centered within the band, sized to fill it
        logo_w, logo_h = 34 * mm, 21 * mm
        logo_drawn = False
        if self._logo_path and os.path.exists(self._logo_path):
            try:
                self.drawImage(self._logo_path, MARGIN, bar_mid - logo_h / 2,
                               width=logo_w, height=logo_h,
                               preserveAspectRatio=True, mask='auto')
                logo_drawn = True
            except Exception:
                pass

        text_x = MARGIN + (logo_w + 10 * mm if logo_drawn else 0)

        # Thin divider between logo and title
        if logo_drawn:
            self.setStrokeColor(colors.Color(1, 1, 1, alpha=0.35))
            self.setLineWidth(0.6)
            self.line(text_x - 5.5 * mm, bar_mid - 7.5 * mm,
                      text_x - 5.5 * mm, bar_mid + 7.5 * mm)

        # Title + tagline — vertically centered as a stack
        self.setFillColor(colors.white)
        self.setFont('Helvetica-Bold', 15)
        self.drawString(text_x, bar_mid + 0.8 * mm, self._report_title.upper())
        self.setFont('Helvetica', 8)
        self.setFillColor(colors.Color(1, 1, 1, alpha=0.82))
        self.drawString(text_x, bar_mid - 4.8 * mm, 'AMAAN Service Tickets')

        # Footer
        self.setStrokeColor(GRID)
        self.setLineWidth(0.5)
        self.line(MARGIN, 11 * mm, PAGE_W - MARGIN, 11 * mm)
        self.setFillColor(MUTED)
        self.setFont('Helvetica', 7)
        self.drawString(MARGIN, 7.5 * mm, 'Amaan Facilities Management — Confidential')
        self.drawRightString(PAGE_W - MARGIN, 7.5 * mm, f'Page {self._pageNumber} of {num_pages}')

        self.restoreState()


def build_report_pdf(spec, rows, summary, filters_label, output_stream):
    """Render a report spec + rows to a portrait A4 PDF on output_stream."""
    cols = spec['columns']
    # Adaptive sizing — denser font for wide tables so cells don't wrap awkwardly.
    if len(cols) >= 9:
        cell_fs, head_fs = 7, 7.2
    elif len(cols) >= 7:
        cell_fs, head_fs = 8, 8.2
    else:
        cell_fs, head_fs = 8.8, 9

    base = getSampleStyleSheet()
    cell_style = ParagraphStyle('RCell', parent=base['Normal'], fontSize=cell_fs,
                                leading=cell_fs + 2, textColor=colors.HexColor('#1f2937'))
    head_style = ParagraphStyle('RHead', parent=base['Normal'], fontSize=head_fs,
                                leading=head_fs + 1.5, fontName='Helvetica-Bold',
                                textColor=colors.white)
    meta_style = ParagraphStyle('RMeta', parent=base['Normal'], fontSize=8, leading=11,
                                textColor=MUTED)

    doc = SimpleDocTemplate(
        output_stream, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=33 * mm, bottomMargin=16 * mm,
        title=spec['title'],
    )

    story = []
    gen_on = datetime.now(timezone.utc).strftime('%d %b %Y, %H:%M UTC')
    meta = f'Generated {gen_on}'
    if filters_label:
        meta += f' &nbsp;·&nbsp; Filters: {filters_label}'
    meta += f' &nbsp;·&nbsp; {len(rows)} row{"s" if len(rows) != 1 else ""}'
    story.append(Paragraph(meta, meta_style))
    story.append(Spacer(1, 6 * mm))

    # Summary stat cards — centered label + value, separated, with a red top accent
    if summary:
        n = len(summary)
        gap = 3 * mm
        avail = PAGE_W - 2 * MARGIN
        chip_w = (avail - gap * (n - 1)) / n

        label_style = ParagraphStyle('ChipLabel', parent=base['Normal'], fontSize=7,
                                     leading=8.5, alignment=1, fontName='Helvetica-Bold',
                                     textColor=MUTED)

        def _value_style(text):
            length = len(str(text))
            fs = 17 if length <= 4 else (14 if length <= 8 else 11)
            return ParagraphStyle('ChipValue', parent=base['Normal'], fontSize=fs,
                                  leading=fs + 1, alignment=1, fontName='Helvetica-Bold',
                                  textColor=BRAND_DARK)

        cells, col_widths = [], []
        for i, (label, value) in enumerate(summary):
            cells.append([
                Paragraph(str(label).upper(), label_style),
                Spacer(1, 3),
                Paragraph(str(value), _value_style(value)),
            ])
            col_widths.append(chip_w)
            if i < n - 1:
                cells.append('')
                col_widths.append(gap)

        chip_tbl = Table([cells], colWidths=col_widths)
        style = [
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]
        for ci in range(0, len(col_widths), 2):  # only the chip columns
            style.append(('BACKGROUND', (ci, 0), (ci, 0), colors.HexColor('#fbfcfe')))
            style.append(('BOX', (ci, 0), (ci, 0), 0.7, GRID))
            style.append(('LINEABOVE', (ci, 0), (ci, 0), 2, BRAND_ACCENT))
        chip_tbl.setStyle(TableStyle(style))
        story.append(chip_tbl)
        story.append(Spacer(1, 7 * mm))

    # Data table — numeric columns right-aligned for a clean ledger look.
    num_style = ParagraphStyle('RNum', parent=cell_style, alignment=2)  # TA_RIGHT
    total_weight = sum(w for _, _, w, _ in cols)
    avail = PAGE_W - 2 * MARGIN
    col_widths = [avail * w / total_weight for _, _, w, _ in cols]

    def _cell_style_for(fmt):
        return num_style if fmt in ('money', 'int', 'pct') else cell_style

    table_data = [[Paragraph(header, head_style) for _, header, _, _ in cols]]
    for row in rows:
        table_data.append([
            Paragraph(_fmt_cell(row.get(key), fmt), _cell_style_for(fmt))
            for key, _, _, fmt in cols
        ])

    if len(table_data) == 1:
        table_data.append([Paragraph('No data for the selected filters.', cell_style)]
                          + [Paragraph('', cell_style)] * (len(cols) - 1))

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), BRAND_DARK),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, ROW_ALT]),
        # Outer frame + soft horizontal rules only (no busy vertical grid)
        ('LINEBELOW', (0, 0), (-1, 0), 0.8, BRAND_DARK),
        ('LINEBELOW', (0, 1), (-1, -1), 0.4, GRID),
        ('BOX', (0, 0), (-1, -1), 0.6, GRID),
        ('TOPPADDING', (0, 0), (-1, 0), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 4.5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4.5),
    ]
    tbl.setStyle(TableStyle(style))
    story.append(tbl)

    logo = _find_logo()
    doc.build(
        story,
        canvasmaker=lambda *a, **kw: _ReportCanvas(
            *a, logo_path=logo, report_title=spec['title'], **kw),
    )


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

XL_BRAND = 'A8121E'
XL_HEAD_FILL = PatternFill('solid', fgColor=XL_BRAND)
XL_ALT_FILL = PatternFill('solid', fgColor='F8FAFC')
XL_META_FONT = Font(name='Calibri', size=9, color='6B7280')
XL_HEAD_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
XL_TITLE_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=14)
XL_THIN = Side(style='thin', color='E2E8F0')
XL_BORDER = Border(left=XL_THIN, right=XL_THIN, top=XL_THIN, bottom=XL_THIN)


def build_report_excel(spec, rows, summary, filters_label, output_stream):
    """Render a report spec + rows to a styled .xlsx on output_stream."""
    wb = Workbook()
    ws = wb.active
    safe_title = spec['title']
    for ch in '[]:*?/\\':
        safe_title = safe_title.replace(ch, '-')
    ws.title = safe_title[:31]
    cols = spec['columns']
    ncols = len(cols)

    # Title band
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=f"AMAAN — {spec['title']}")
    c.font = XL_TITLE_FONT
    c.fill = XL_HEAD_FILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28
    for i in range(2, ncols + 1):
        ws.cell(row=1, column=i).fill = XL_HEAD_FILL

    # Meta line
    gen_on = datetime.now(timezone.utc).strftime('%d %b %Y, %H:%M UTC')
    meta = f'Generated {gen_on}'
    if filters_label:
        meta += f'  ·  Filters: {filters_label}'
    meta += f'  ·  {len(rows)} rows'
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    m = ws.cell(row=2, column=1, value=meta)
    m.font = XL_META_FONT
    m.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    r = 3
    # Summary row
    if summary:
        r += 1
        for i, (label, value) in enumerate(summary, start=1):
            if i > ncols:
                break
            cell = ws.cell(row=r, column=i, value=f'{label}: {value}')
            cell.font = Font(name='Calibri', bold=True, size=9, color=XL_BRAND)
            cell.fill = XL_ALT_FILL
            cell.border = XL_BORDER
        r += 1

    # Header row
    r += 1
    header_row = r
    for i, (_, header, _, _) in enumerate(cols, start=1):
        cell = ws.cell(row=r, column=i, value=header)
        cell.font = XL_HEAD_FONT
        cell.fill = XL_HEAD_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = XL_BORDER
    ws.row_dimensions[r].height = 22

    # Data rows
    for ridx, row in enumerate(rows):
        r += 1
        for i, (key, _, _, fmt) in enumerate(cols, start=1):
            value = row.get(key)
            cell = ws.cell(row=r, column=i)
            if fmt in ('money', 'pct', 'int') and value is not None and value != '':
                try:
                    cell.value = float(value) if fmt != 'int' else int(value)
                    cell.number_format = '#,##0.00' if fmt == 'money' else ('0"%"' if fmt == 'pct' else '#,##0')
                except (TypeError, ValueError):
                    cell.value = str(value)
            else:
                cell.value = '' if value is None else str(value)
            cell.font = Font(name='Calibri', size=9)
            cell.border = XL_BORDER
            if ridx % 2 == 1:
                cell.fill = XL_ALT_FILL
            if fmt in ('money', 'pct', 'int'):
                cell.alignment = Alignment(horizontal='right')

    # Freeze panes below header, set sensible column widths
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    total_weight = sum(w for _, _, w, _ in cols)
    for i, (_, _, w, _) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(10, round(110 * w / total_weight))

    wb.save(output_stream)
