"""Excel template / export / import for the location hierarchy.

Workbook sheets (headers match CRM SpreadsheetML imports):
  Property   — Property Code, Property Name, Area, City, Country, Client Name, Status, Latitude, Longitude
  Zone       — Zone Code, Zone Name, Property
  Sub Zone   — Sub Zone Code, Sub Zone Name, Property, Zone
  Base Unit  — Base Unit Code, Base Unit Name, Property, Zone, Sub Zone, Latitude, Longitude
"""
from __future__ import annotations

import re
from io import BytesIO
from typing import Any, BinaryIO

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from common.kynvera_excel_brand import (
    InstructionSpec,
    apply_column_widths,
    style_data_cell,
    write_header_row,
    write_instructions_sheet,
)
from module_ticketing.location_catalog import detect_kind, import_location_rows

SHEETS: tuple[tuple[str, str, tuple[str, ...]], ...] = (
    ("Property", "property", (
        "Property Code", "Property Name", "Area", "City", "Country", "Client Name", "Status",
        "Latitude", "Longitude",
    )),
    ("Zone", "zone", ("Zone Code", "Zone Name", "Property")),
    ("Sub Zone", "sub_zone", ("Sub Zone Code", "Sub Zone Name", "Property", "Zone")),
    ("Base Unit", "base_unit", (
        "Base Unit Code", "Base Unit Name", "Property", "Zone", "Sub Zone",
        "Latitude", "Longitude",
    )),
)

SHEET_KIND = {
    "property": "property",
    "zone": "zone",
    "sub zone": "sub_zone",
    "subzone": "sub_zone",
    "base unit": "base_unit",
    "baseunit": "base_unit",
}

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _cell(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _coord_cell(v: Any) -> str:
    if v is None or v == "":
        return ""
    try:
        return f"{float(v):.6f}".rstrip("0").rstrip(".")
    except (TypeError, ValueError):
        return _cell(v)


def tree_to_rows(properties: list[dict] | None) -> dict[str, list[dict[str, str]]]:
    """Flatten a location-tree payload into sheet row dicts."""
    out: dict[str, list[dict[str, str]]] = {
        "property": [],
        "zone": [],
        "sub_zone": [],
        "base_unit": [],
    }
    for prop in properties or []:
        out["property"].append({
            "Property Code": _cell(prop.get("code")),
            "Property Name": _cell(prop.get("name")),
            "Area": _cell(prop.get("area")),
            "City": _cell(prop.get("city")),
            "Country": _cell(prop.get("country")),
            "Client Name": _cell(prop.get("client_name")),
            "Status": _cell(prop.get("status")) or ("Active" if prop.get("is_active", True) else "Inactive"),
            "Latitude": _coord_cell(prop.get("latitude")),
            "Longitude": _coord_cell(prop.get("longitude")),
        })
        pname = _cell(prop.get("name"))
        for zone in prop.get("zones") or []:
            zname = _cell(zone.get("name"))
            out["zone"].append({
                "Zone Code": _cell(zone.get("code")),
                "Zone Name": zname,
                "Property": pname,
            })
            for sz in zone.get("sub_zones") or []:
                szname = _cell(sz.get("name"))
                out["sub_zone"].append({
                    "Sub Zone Code": _cell(sz.get("code")),
                    "Sub Zone Name": szname,
                    "Property": pname,
                    "Zone": zname,
                })
                for unit in sz.get("base_units") or []:
                    out["base_unit"].append({
                        "Base Unit Code": _cell(unit.get("code")),
                        "Base Unit Name": _cell(unit.get("name")),
                        "Property": pname,
                        "Zone": zname,
                        "Sub Zone": szname,
                        "Latitude": _coord_cell(unit.get("latitude")),
                        "Longitude": _coord_cell(unit.get("longitude")),
                    })
    return out


LOCATION_EXAMPLE = {
    "property": [{
        "Property Code": "HQ",
        "Property Name": "HQ Building",
        "Area": "Business Bay",
        "City": "Dubai",
        "Country": "UAE",
        "Client Name": "Example Client",
        "Status": "Active",
        "Latitude": "25.2048",
        "Longitude": "55.2708",
    }],
    "zone": [{
        "Zone Code": "HQ-L1",
        "Zone Name": "First Floor",
        "Property": "HQ Building",
    }],
    "sub_zone": [{
        "Sub Zone Code": "HQ-L1-MR",
        "Sub Zone Name": "Meeting Rooms",
        "Property": "HQ Building",
        "Zone": "First Floor",
    }],
    "base_unit": [{
        "Base Unit Code": "HQ-BRD",
        "Base Unit Name": "Boardroom",
        "Property": "HQ Building",
        "Zone": "First Floor",
        "Sub Zone": "Meeting Rooms",
        "Latitude": "25.2049",
        "Longitude": "55.2709",
    }],
}


def _write_sheet(
    ws: Worksheet,
    headers: tuple[str, ...],
    rows: list[dict[str, str]],
    *,
    example: bool = False,
) -> None:
    write_header_row(ws, headers, row=1)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for r, rec in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(r, c, rec.get(h, "") or "")
            style_data_cell(cell, example=example)
    widths = [max(12, min(36, len(h) + 4)) for h in headers]
    for rec in rows[:80]:
        for i, h in enumerate(headers):
            widths[i] = max(widths[i], min(40, len(_cell(rec.get(h))) + 2))
    apply_column_widths(ws, widths)


def _location_instruction_spec() -> InstructionSpec:
    return InstructionSpec(
        title='Location hierarchy template',
        module_label='Service Tickets',
        about=(
            'Import Property → Zone → Sub Zone → Base Unit for ticket maps and work orders.',
            'Every row needs a unique code and a name. Child rows must use the parent names exactly as they appear on the parent sheet.',
            'Leave unused sheets empty (headers only) if you are not changing that level.',
        ),
        how_to=(
            'Fill Property first, then Zone, Sub Zone, and Base Unit. Keep each coral header row.',
            'Child rows: Zone.Property must match Property Name; Sub Zone must match Property + Zone Name; Base Unit must match all three parent names.',
            'Optional Latitude / Longitude on Property and Base Unit pin the site on New Work Order and ticket maps.',
            'Save as .xlsx, then use Import on the Locations page. Existing codes are updated; new codes are created.',
        ),
        columns=(
            ('Property Code / Name', 'Required on Property. Unique code; name is the parent key for child sheets.'),
            ('Area / City / Country / Client Name', 'Optional property metadata.'),
            ('Status', 'Optional. Active or Inactive (defaults to Active).'),
            ('Latitude / Longitude', 'Optional on Property and Base Unit. Decimal degrees.'),
            ('Zone Code / Name / Property', 'Required on Zone. Property must match a Property Name exactly.'),
            ('Sub Zone Code / Name / Property / Zone', 'Required on Sub Zone. Parents must match names exactly.'),
            ('Base Unit Code / Name / Property / Zone / Sub Zone', 'Required on Base Unit. Parents must match names exactly.'),
        ),
        example_headers=('Sheet', 'Code', 'Name', 'Parent names'),
        example_rows=(
            ('Property', 'HQ', 'HQ Building', '—'),
            ('Zone', 'HQ-L1', 'First Floor', 'Property = HQ Building'),
            ('Sub Zone', 'HQ-L1-MR', 'Meeting Rooms', 'Property = HQ Building, Zone = First Floor'),
            ('Base Unit', 'HQ-BRD', 'Boardroom', 'Property / Zone / Sub Zone as above'),
        ),
        import_rules=(
            'Upsert by code: existing codes are updated, new codes are created.',
            'Parent names must match exactly (including spaces and punctuation).',
            'Export downloads the locations currently on the page in this same layout.',
            'The Instructions sheet is ignored on import.',
        ),
    )


def build_location_workbook(properties: list[dict] | None = None) -> BytesIO:
    """Build a template (empty sheets) or an export of the given tree."""
    rows = tree_to_rows(properties)
    template_only = not any(rows.values())
    wb = Workbook()
    # Placeholder active sheet is replaced by Instructions below.
    wb.active.title = "Property"
    for title, kind, headers in SHEETS:
        ws = wb[title] if title in wb.sheetnames else wb.create_sheet(title)
        data = rows.get(kind) or []
        example = False
        if template_only and not data:
            data = LOCATION_EXAMPLE.get(kind) or []
            example = True
        _write_sheet(ws, headers, data, example=example)
    write_instructions_sheet(wb, _location_instruction_spec())
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _sheet_kind(title: str, headers: list[str]) -> str | None:
    detected = detect_kind(headers)
    if detected:
        return detected
    key = re.sub(r"\s+", " ", (title or "").strip().lower())
    return SHEET_KIND.get(key)


def parse_location_xlsx(source: bytes | BinaryIO) -> dict[str, list[dict[str, str]]]:
    """Read Property / Zone / Sub Zone / Base Unit sheets into import row dicts."""
    data = source.read() if hasattr(source, "read") else source
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    parsed: dict[str, list[dict[str, str]]] = {
        "property": [],
        "zone": [],
        "sub_zone": [],
        "base_unit": [],
    }
    try:
        for ws in wb.worksheets:
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                continue
            headers = [_cell(c) for c in header_row]
            if not any(headers):
                continue
            kind = _sheet_kind(ws.title, headers)
            if kind not in parsed:
                continue
            for raw in rows_iter:
                rec: dict[str, str] = {}
                empty = True
                for h, v in zip(headers, raw or ()):
                    if not h:
                        continue
                    val = _cell(v)
                    rec[h] = val
                    if val:
                        empty = False
                if not empty:
                    parsed[kind].append(rec)
    finally:
        wb.close()
    if not any(parsed.values()):
        raise ValueError("No location rows found. Use the Excel template sheets: Property, Zone, Sub Zone, Base Unit.")
    return parsed


def import_location_xlsx(
    source: bytes | BinaryIO,
    *,
    project_id: int | None = None,
    standalone: bool = False,
) -> dict:
    parsed = parse_location_xlsx(source)
    return import_location_rows(
        property_rows=parsed["property"],
        zone_rows=parsed["zone"],
        sub_zone_rows=parsed["sub_zone"],
        base_unit_rows=parsed["base_unit"],
        project_id=project_id,
        standalone=standalone,
    )


def download_filename(label: str, *, kind: str = "locations") -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", (label or "locations").strip().lower()).strip("_") or "locations"
    return f"{slug}_{kind}.xlsx"
