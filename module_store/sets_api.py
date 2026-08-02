"""Material Sets master + catalog for tickets (Meeting 2)."""
from __future__ import annotations

import uuid
from io import BytesIO

from flask import request, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from openpyxl import Workbook, load_workbook

from app.models import db, User, MaterialSet, MaterialSetItem
from common.error_responses import success_response, error_response


def _gen_set_id():
    return 'SET-' + uuid.uuid4().hex[:8].upper()


def _current_user():
    return db.session.get(User, get_jwt_identity())


def _can_manage_store(user):
    if not user:
        return False
    if user.role == 'admin':
        return True
    return bool(getattr(user, 'access_procurement_module', False))


def _can_see_prices(user):
    if not user:
        return False
    if user.role == 'admin':
        return True
    if bool(getattr(user, 'access_finance', False)):
        return True
    if bool(getattr(user, 'access_procurement_module', False)):
        return True
    desig = str(getattr(user, 'designation', '') or '').strip().lower()
    # Technicians / supervisors must not see unit prices
    if desig in ('technician', 'supervisor', 'field_technician'):
        return False
    return desig in ('general_manager', 'operations_manager', 'finance', 'finance_manager')


def register_sets_routes(bp):
    @bp.route('/api/sets', methods=['GET'])
    @jwt_required()
    def api_list_sets():
        user = _current_user()
        if not user:
            return error_response('User not found', status_code=404, error_code='NOT_FOUND')
        include_prices = _can_see_prices(user)
        rows = MaterialSet.query.filter_by(is_active=True).order_by(MaterialSet.name.asc()).all()
        return success_response({
            'sets': [r.to_dict(include_prices=include_prices, include_items=True) for r in rows],
            'can_see_prices': include_prices,
        })

    @bp.route('/api/sets', methods=['POST'])
    @jwt_required()
    def api_create_set():
        user = _current_user()
        if not _can_manage_store(user):
            return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
        data = request.get_json(silent=True) or {}
        name = (data.get('name') or '').strip()
        if not name:
            return error_response('Set name is required', status_code=400, error_code='VALIDATION')
        row = MaterialSet(
            set_id=_gen_set_id(),
            name=name,
            material_type=(data.get('material_type') or '').strip() or None,
            description=(data.get('description') or '').strip() or None,
            is_active=True,
            created_by_id=user.id,
        )
        for it in (data.get('items') or []):
            mname = (it.get('material_name') or '').strip()
            if not mname:
                continue
            row.items.append(MaterialSetItem(
                material_name=mname,
                unit=(it.get('unit') or '').strip() or None,
                quantity=float(it.get('quantity') or 1),
                unit_price=float(it.get('unit_price') or 0),
                procurement_ref=(it.get('procurement_ref') or '').strip() or None,
            ))
        db.session.add(row)
        db.session.commit()
        return success_response({'set': row.to_dict()}, message='Material set created.', status_code=201)

    @bp.route('/api/sets/<set_id>', methods=['PUT'])
    @jwt_required()
    def api_update_set(set_id):
        user = _current_user()
        if not _can_manage_store(user):
            return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
        row = MaterialSet.query.filter_by(set_id=set_id).first()
        if not row:
            return error_response('Set not found', status_code=404, error_code='NOT_FOUND')
        data = request.get_json(silent=True) or {}
        if 'name' in data:
            row.name = (data.get('name') or '').strip() or row.name
        if 'material_type' in data:
            row.material_type = (data.get('material_type') or '').strip() or None
        if 'description' in data:
            row.description = (data.get('description') or '').strip() or None
        if 'is_active' in data:
            row.is_active = bool(data.get('is_active'))
        if 'items' in data:
            row.items.clear()
            for it in (data.get('items') or []):
                mname = (it.get('material_name') or '').strip()
                if not mname:
                    continue
                row.items.append(MaterialSetItem(
                    material_name=mname,
                    unit=(it.get('unit') or '').strip() or None,
                    quantity=float(it.get('quantity') or 1),
                    unit_price=float(it.get('unit_price') or 0),
                    procurement_ref=(it.get('procurement_ref') or '').strip() or None,
                ))
        db.session.commit()
        return success_response({'set': row.to_dict()}, message='Material set updated.')

    @bp.route('/api/sets/template', methods=['GET'])
    @jwt_required()
    def api_sets_template():
        user = _current_user()
        if not _can_manage_store(user):
            return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sets'
        ws.append(['Set Name', 'Material Type', 'Item Name', 'Unit', 'Quantity', 'Unit Price'])
        ws.append(['Sprinkler Set A', 'Fire', 'Sprinkler head', 'pcs', 5, 12.5])
        ws.append(['Sprinkler Set A', 'Fire', 'Pipe clamp', 'pcs', 5, 3.0])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(
            bio, as_attachment=True, download_name='material_sets_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    @bp.route('/api/sets/import', methods=['POST'])
    @jwt_required()
    def api_import_sets():
        user = _current_user()
        if not _can_manage_store(user):
            return error_response('Access denied', status_code=403, error_code='ACCESS_DENIED')
        f = request.files.get('file')
        if not f:
            return error_response('Excel file required', status_code=400, error_code='FILE_REQUIRED')
        try:
            wb = load_workbook(f, data_only=True)
            ws = wb.active
        except Exception as e:
            return error_response(f'Bad Excel: {e}', status_code=400, error_code='BAD_FILE')

        buckets = {}
        for raw in ws.iter_rows(min_row=2, values_only=True):
            if not raw:
                continue
            set_name = str(raw[0] or '').strip()
            if not set_name:
                continue
            buckets.setdefault(set_name, {
                'material_type': str(raw[1] or '').strip() or None,
                'items': [],
            })
            item_name = str(raw[2] or '').strip() if len(raw) > 2 else ''
            if not item_name:
                continue
            try:
                qty = float(raw[4]) if len(raw) > 4 and raw[4] not in (None, '') else 1.0
            except (TypeError, ValueError):
                qty = 1.0
            try:
                price = float(raw[5]) if len(raw) > 5 and raw[5] not in (None, '') else 0.0
            except (TypeError, ValueError):
                price = 0.0
            buckets[set_name]['items'].append({
                'material_name': item_name,
                'unit': str(raw[3] or '').strip() if len(raw) > 3 else None,
                'quantity': qty,
                'unit_price': price,
            })

        created = 0
        for set_name, payload in buckets.items():
            row = MaterialSet.query.filter(
                db.func.lower(MaterialSet.name) == set_name.lower()
            ).first()
            if not row:
                row = MaterialSet(
                    set_id=_gen_set_id(), name=set_name,
                    material_type=payload['material_type'],
                    created_by_id=user.id, is_active=True,
                )
                db.session.add(row)
                created += 1
            else:
                row.items.clear()
                row.material_type = payload['material_type'] or row.material_type
            for it in payload['items']:
                row.items.append(MaterialSetItem(
                    material_name=it['material_name'],
                    unit=it['unit'] or None,
                    quantity=it['quantity'],
                    unit_price=it['unit_price'],
                ))
        db.session.commit()
        return success_response({'created_or_updated': len(buckets), 'new_sets': created},
                                message=f'Imported {len(buckets)} sets.')

    @bp.route('/api/catalog/sets', methods=['GET'])
    @jwt_required()
    def api_catalog_sets_for_tickets():
        """Ticket picker endpoint — hide prices for technicians/supervisors."""
        user = _current_user()
        if not user:
            return error_response('User not found', status_code=404, error_code='NOT_FOUND')
        include_prices = _can_see_prices(user)
        rows = MaterialSet.query.filter_by(is_active=True).order_by(MaterialSet.name.asc()).all()
        return success_response({
            'sets': [r.to_dict(include_prices=include_prices, include_items=True) for r in rows],
            'can_see_prices': include_prices,
        })
