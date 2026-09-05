"""PDF and Excel exports that follow the procurement page layouts."""
from __future__ import annotations

import io
import logging
import math
from datetime import datetime

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    Flowable, KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)
from reportlab.pdfgen.canvas import Canvas

from common import kynvera_excel_brand as xls
from common import kynvera_pdf_brand as brand
from common.kynvera_brand import COMPANY_NAME
from module_procurement import pr_pdf
from module_procurement import service as svc
from module_procurement.models import ProcPurchaseRequest, ProcStock, _utcnow

logger = logging.getLogger(__name__)

PAGE_W, PAGE_H = A4
L_MARGIN = 16 * mm
R_MARGIN = 16 * mm
T_MARGIN = 16 * mm
B_MARGIN = 18 * mm
AVAIL_W = PAGE_W - L_MARGIN - R_MARGIN
INK = brand.TEXT_DARK
MUTED = brand.TEXT_MID
RULE = brand.HAIRLINE
OK_BG = colors.HexColor('#dcfce7')
OK_FG = colors.HexColor('#15803d')
WARN_BG = colors.HexColor('#ffedd5')
WARN_FG = colors.HexColor('#c2410c')
NEUTRAL_BG = colors.HexColor('#f4f4f5')
NEUTRAL_FG = colors.HexColor('#52525b')
DOT = colors.HexColor('#d4d4d8')
SPARK = colors.HexColor('#ff8e68')
TRADE_ORDER = ('HVAC', 'Cleaning', 'Electrical', 'Plumbing', 'Other')
TRADE_COLORS = {
    'HVAC': colors.HexColor('#18181b'),
    'Cleaning': colors.HexColor('#ff8e68'),
    'Electrical': colors.HexColor('#c4c4cc'),
    'Plumbing': colors.HexColor('#71717a'),
    'Other': colors.HexColor('#e4e4e7'),
}
GRAIN_LABEL = {'week': 'Weekly', 'month': 'Monthly', 'year': 'Yearly'}

STOCK_STATUS = {
    'ok': 'In stock',
    'low': 'Low stock',
    'refill': 'Refill needed',
    'returned': 'Returned',
}
PR_STATUS = dict(pr_pdf.STATUS_LABELS)
PR_STATUS.update({
    'submitted': 'Submitted',
    'cancelled': 'Cancelled',
    'awaiting_quotation': 'Awaiting quotation',
})
REPORTS = {
    'dashboard': 'Procurement overview',
    'usage-log': 'Usage log',
    'materials': 'Property stock',
    'properties': 'Property management',
    'refill': 'Refill queue',
    'purchase-requests': 'Purchase requests',
}


def _p(text, size=8.5, bold=False, color=INK, align=TA_LEFT, leading=None):
    fn = 'Helvetica-Bold' if bold else 'Helvetica'
    return Paragraph(
        _esc(text) if text not in (None, '') else '—',
        ParagraphStyle(
            '_', fontName=fn, fontSize=size, textColor=color,
            alignment=align, leading=leading or (size * 1.35),
        ),
    )


def _esc(text):
    return (
        str(text if text is not None else '')
        .replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    )


def _fmt_aed(value, digits=0):
    n = float(value or 0)
    return f'AED {n:,.{digits}f}'


def _fmt_qty(value, uom=''):
    n = float(value or 0)
    body = f'{n:,.0f}' if abs(n - round(n)) < 0.05 else f'{n:,.1f}'
    return f'{body} {uom}'.strip()


def _fmt_when(iso):
    raw = str(iso or '')[:19].replace('T', ' ')
    if len(raw) >= 16:
        try:
            dt = datetime.strptime(raw[:16], '%Y-%m-%d %H:%M')
            return f"{dt.day} {dt.strftime('%b %Y, %H:%M')}"
        except ValueError:
            return raw[:16]
    return raw or '—'


def _pct_vs(delta):
    d = float(delta or 0)
    if abs(d) < 0.0005:
        return 'No change vs last period'
    sign = '+' if d > 0 else '−'
    return f'{sign}{abs(d) * 100:.1f}% vs last period'


def _stamp():
    return _utcnow().strftime('%d %b %Y %H:%M')


def _nice_ymax(value):
    v = max(float(value or 0), 1.0)
    mag = 10 ** math.floor(math.log10(v))
    for m in (1, 2, 2.5, 5, 10):
        if mag * m >= v * 1.08:
            return mag * m
    return mag * 10


def _series_sum(series):
    return sum(float(v or 0) for vals in (series or {}).values() for v in (vals or []))


def _fmt_range_chip(start_iso, end_iso, fallback=''):
    def one(iso):
        raw = str(iso or '')[:10]
        try:
            d = datetime.strptime(raw, '%Y-%m-%d')
        except ValueError:
            return ''
        return f"{d.day} {d.strftime('%b')}"
    a, b = one(start_iso), one(end_iso)
    if a and b:
        return f'{a} – {b}'
    return fallback


class _CanvasChart(Flowable):
    def __init__(self, width, height, paint):
        super().__init__()
        self.width = width
        self.height = height
        self._paint = paint

    def wrap(self, *_args):
        return self.width, self.height

    def draw(self):
        self._paint(self.canv, self.width, self.height)


def _spark_chart(values, width, height):
    vals = [max(0.0, float(v or 0)) for v in (values or [])]

    def paint(c, w, h):
        if not vals:
            return
        mx = max(vals) or 1.0
        n = len(vals)
        slot = w / n
        bar_w = max(1.4, slot * 0.55)
        c.setFillColor(SPARK)
        for i, v in enumerate(vals):
            bh = max(1.4, (v / mx) * (h - 0.5))
            c.roundRect(i * slot + (slot - bar_w) / 2, 0, bar_w, bh, 0.6, fill=1, stroke=0)

    return _CanvasChart(width, height, paint)


def _x_label_indexes(n, *, ends_only=False):
    if n <= 0:
        return []
    if ends_only or n > 14:
        return [0] if n == 1 else [0, n - 1]
    return list(range(n))


def _draw_swatch_legend(c, items, x, y):
    if not items:
        return
    c.setFont('Helvetica', 6)
    cx = x
    for color, label in items:
        c.setFillColor(color)
        c.roundRect(cx, y + 0.4, 4.2, 4.2, 0.6, fill=1, stroke=0)
        c.setFillColor(MUTED)
        c.drawString(cx + 6.2, y, label)
        cx += 8 + c.stringWidth(label, 'Helvetica', 6) + 8


def _paint_axes(c, pad_l, pad_b, plot_w, plot_h, ymax, labels, *, dots=False, ends_only=False):
    plot_top = pad_b + plot_h
    plot_right = pad_l + plot_w
    if dots:
        c.setFillColor(DOT)
        step = 7
        x = pad_l
        while x <= plot_right + 0.5:
            y = pad_b
            while y <= plot_top + 0.5:
                c.circle(x, y, 0.4, fill=1, stroke=0)
                y += step
            x += step
    c.saveState()
    c.setStrokeColor(colors.HexColor('#ececee'))
    c.setLineWidth(0.4)
    c.setDash(1.2, 2.2)
    for t in (0.0, ymax / 2, ymax):
        y = pad_b + (t / ymax) * plot_h
        if t > 0:
            c.line(pad_l, y, plot_right, y)
    c.restoreState()
    c.setFillColor(MUTED)
    c.setFont('Helvetica', 6)
    for t in (0.0, ymax / 2, ymax):
        y = pad_b + (t / ymax) * plot_h
        c.drawRightString(pad_l - 3, y - 2, f'{t:,.0f}')
    n = max(len(labels), 1)
    slot = plot_w / n
    c.setFont('Helvetica-Bold', 6)
    idxs = _x_label_indexes(min(n, len(labels)) if labels else 0, ends_only=ends_only)
    for i in idxs:
        text = str(labels[i])
        x = pad_l + slot * (i + 0.5)
        tw = c.stringWidth(text, 'Helvetica-Bold', 6)
        if i == 0 or x - tw / 2 < pad_l:
            c.drawString(pad_l, 2.5, text)
        elif i == n - 1 or x + tw / 2 > plot_right:
            c.drawRightString(plot_right, 2.5, text)
        else:
            c.drawCentredString(x, 2.5, text)


def _paint_util(c, w, h, labels, series, legend=None):
    pad_l, pad_b, pad_t, pad_r = 22, 14, 12, 6
    if legend:
        _draw_swatch_legend(c, legend, pad_l, h - 9)
        pad_t = 16
    plot_w = max(w - pad_l - pad_r, 10)
    plot_h = max(h - pad_b - pad_t, 10)
    plot_top = pad_b + plot_h
    n = max(len(labels), 1)
    totals = [0.0] * n
    stacks = []
    for trade in TRADE_ORDER:
        vals = list(series.get(trade) or [])
        while len(vals) < n:
            vals.append(0.0)
        stacks.append((trade, vals))
        for i in range(n):
            totals[i] += float(vals[i] or 0)
    ymax = _nice_ymax(max(totals) if totals else 0)
    _paint_axes(c, pad_l, pad_b, plot_w, plot_h, ymax, labels, dots=True)
    slot = plot_w / n
    bar_w = slot * 0.52
    cell, gap = 2.2, 0.9
    pitch = cell + gap
    for i in range(n):
        cx = pad_l + slot * (i + 0.5)
        cols = max(1, int(bar_w / pitch))
        grid_w = cols * pitch - gap
        left = cx - grid_w / 2
        y_bot = pad_b
        for trade, vals in stacks:
            v = float(vals[i] or 0)
            if v <= 0 or y_bot >= plot_top - 0.4:
                continue
            seg_h = (v / ymax) * plot_h
            y_top = min(plot_top, y_bot + seg_h)
            c.setFillColor(TRADE_COLORS[trade])
            row = 0
            while True:
                py = y_bot + gap * 0.3 + row * pitch
                if py + cell > y_top + 0.15 or py + cell > plot_top:
                    break
                for col in range(cols):
                    c.roundRect(left + col * pitch, py, cell, cell, 0.4, fill=1, stroke=0)
                row += 1
                if row > 60:
                    break
            y_bot = y_top


def _paint_breakdown(c, w, h, labels, this_period, last_period, legend=None):
    pad_l, pad_b, pad_t, pad_r = 22, 14, 12, 8
    if legend:
        _draw_swatch_legend(c, legend, pad_l, h - 9)
        pad_t = 16
    plot_w = max(w - pad_l - pad_r, 10)
    plot_h = max(h - pad_b - pad_t, 10)
    plot_top = pad_b + plot_h
    n = max(len(labels), 1)
    this_vals = [float(v or 0) for v in (this_period or [])]
    last_vals = [float(v or 0) for v in (last_period or [])]
    while len(this_vals) < n:
        this_vals.append(0.0)
    while len(last_vals) < n:
        last_vals.append(0.0)
    this_vals, last_vals = this_vals[:n], last_vals[:n]
    ymax = _nice_ymax(max((a + b) for a, b in zip(this_vals, last_vals)) if n else 0)
    _paint_axes(c, pad_l, pad_b, plot_w, plot_h, ymax, labels, dots=False, ends_only=True)
    slot = plot_w / n
    bar_w = min(5.5, max(1.0, slot * 0.42))
    ink = colors.HexColor('#18181b')
    gray = colors.HexColor('#e4e4e7')
    for i in range(n):
        x = pad_l + slot * (i + 0.5) - bar_w / 2
        h_this = min((this_vals[i] / ymax) * plot_h, plot_h)
        h_last = min((last_vals[i] / ymax) * plot_h, plot_h - h_this)
        if h_this > 0.4:
            c.setFillColor(ink)
            c.rect(x, pad_b, bar_w, h_this, fill=1, stroke=0)
        if h_last > 0.4:
            c.setFillColor(gray)
            c.rect(x, pad_b + h_this, bar_w, min(h_last, plot_top - pad_b - h_this), fill=1, stroke=0)


def _chart_card(title, meta, chart, width):
    inner = Table(
        [
            [_p(title, 8, bold=True, leading=10), _p(meta, 7, color=MUTED, align=TA_RIGHT, leading=10)],
            [chart, ''],
        ],
        colWidths=[width * 0.58, width * 0.42],
    )
    inner.setStyle(TableStyle([
        ('SPAN', (0, 1), (-1, 1)),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (0, 0), 6),
        ('TOPPADDING', (0, 1), (-1, 1), 4),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 6),
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ('BOX', (0, 0), (-1, -1), 0.4, RULE),
    ]))
    return inner


class _ReportCanvas(Canvas):
    def __init__(self, *args, **kwargs):
        self._report_title = kwargs.pop('report_title', '')
        super().__init__(*args, **kwargs)
        self._saved = []

    def showPage(self):
        self._saved.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total = len(self._saved)
        for state in self._saved:
            self.__dict__.update(state)
            brand.draw_simple_footer(
                self, self._pageNumber, total,
                left_margin=L_MARGIN, right_margin=R_MARGIN,
                footer_left=f'{COMPANY_NAME}  ·  Procurement  ·  Confidential',
            )
            super().showPage()
        super().save()


def _chip(text, *, ok=False, warn=False):
    fg, bg = (OK_FG, OK_BG) if ok else ((WARN_FG, WARN_BG) if warn else (NEUTRAL_FG, NEUTRAL_BG))
    chip = Table(
        [[_p(text, 7, bold=True, color=fg, leading=9)]],
        colWidths=[22 * mm],
    )
    chip.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), bg),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    return chip


def _status_text(status):
    key = (status or '').lower()
    label = STOCK_STATUS.get(key, (status or '—').replace('_', ' ').title())
    color = OK_FG if key == 'ok' else (WARN_FG if key in ('low', 'refill') else NEUTRAL_FG)
    return _p(label, 7, bold=True, color=color)


def _stock_chip(status):
    return _status_text(status)


def _pr_chip(status):
    key = (status or '').strip().lower()
    label = PR_STATUS.get(key, (status or '—').replace('_', ' ').title())
    ok = key in ('approved', 'ordered', 'received', 'closed')
    warn = key in ('procurement_review', 'gm_review', 'submitted', 'awaiting_quotation', 'rejected')
    return _chip(label, ok=ok, warn=warn)


def _hero(title, subtitle):
    wordmark = brand.wordmark_flowable(max_width=30 * mm, max_height=7.5 * mm)
    left = Table(
        [
            [wordmark or _p(COMPANY_NAME, 12, bold=True, color=brand.PRIMARY, leading=14)],
            [_p('Procurement', 8, bold=True, color=MUTED, leading=10)],
        ],
        colWidths=[AVAIL_W * 0.42],
    )
    left.setStyle(TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 2),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    right = Table(
        [
            [_p(title, 12, bold=True, color=INK, align=TA_RIGHT, leading=15)],
            [_p(subtitle, 8, color=MUTED, align=TA_RIGHT, leading=11)],
            [_p(f'Generated {_stamp()}', 7.5, color=MUTED, align=TA_RIGHT, leading=10)],
        ],
        colWidths=[AVAIL_W * 0.58],
    )
    right.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    hero = Table([[left, right]], colWidths=[AVAIL_W * 0.42, AVAIL_W * 0.58])
    hero.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LINEBELOW', (0, 0), (-1, -1), 0.6, RULE),
    ]))
    return hero


def _kpi_row(items):
    width = AVAIL_W / max(len(items), 1)
    inner = width - 3 * mm
    cells = []
    for item in items:
        label, value, meta = item[0], item[1], item[2]
        block = Table(
            [
                [_p(label, 6.5, bold=True, color=MUTED, leading=8)],
                [_p(value, 11, bold=True, color=INK, leading=13)],
                [_p(meta, 6.5, color=MUTED, leading=8)],
            ],
            colWidths=[inner],
            rowHeights=[11, 16, 12],
        )
        block.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), brand.SURFACE_ALT),
            ('BOX', (0, 0), (-1, -1), 0.4, RULE),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (0, 0), 5),
            ('BOTTOMPADDING', (0, -1), (0, -1), 5),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        cells.append(block)
    table = Table([cells], colWidths=[width] * len(items))
    table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 1.5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 1.5),
    ]))
    return table


def _section_head(title, meta=''):
    return Table(
        [[_p(title, 10, bold=True, leading=13), _p(meta, 8, color=MUTED, align=TA_RIGHT, leading=11)]],
        colWidths=[AVAIL_W * 0.62, AVAIL_W * 0.38],
    )


def _data_table(headers, rows, col_widths, *, header=True):
    data = [[_p(h, 7, bold=True, color=MUTED, leading=9) for h in headers]] if header else []
    data.extend(rows)
    table = Table(data, colWidths=col_widths, repeatRows=1 if header else 0)
    cmds = [
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LINEBELOW', (0, 0), (-1, -1), 0.3, RULE),
        ('BACKGROUND', (0, 0), (-1, 0), brand.SURFACE_ALT) if header else ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]
    for i in range(1 if header else 0, len(data)):
        if (i % 2) == (0 if header else 1):
            cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#fafafa')))
    table.setStyle(TableStyle(cmds))
    return table


def _build_pdf(title, subtitle, story) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=L_MARGIN, rightMargin=R_MARGIN,
        topMargin=T_MARGIN, bottomMargin=B_MARGIN,
        title=title, author=brand.PDF_AUTHOR,
    )
    flow = [_hero(title, subtitle), Spacer(1, 7)]
    flow.extend(story)
    doc.build(
        flow,
        canvasmaker=lambda *a, **k: _ReportCanvas(*a, report_title=title, **k),
    )
    return buf.getvalue()


def _xlsx_book(sheets, filename_stem: str):
    wb = Workbook()
    first = True
    for sheet in sheets:
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = sheet['name'][:31]
        ws['A1'] = sheet['title']
        ws['A1'].font = xls.TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(sheet['headers']), 1))
        ws['A2'] = f'{COMPANY_NAME} · {sheet.get("subtitle") or ""} · {_stamp()}'
        ws['A2'].font = xls.MUTED_FONT
        xls.write_header_row(ws, sheet['headers'], row=4)
        for i, row in enumerate(sheet['rows'], start=5):
            xls.write_data_row(ws, i, row)
        xls.apply_column_widths(ws, sheet.get('widths') or [16] * len(sheet['headers']))
        xls.paint_tab(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), f'{filename_stem}_{_utcnow().strftime("%Y%m%d")}.xlsx'


def _prop_badge(prop: dict) -> str:
    name = prop.get('name') or ''
    if prop.get('is_shared') or name == 'Shared':
        return 'Shared store'
    if name == 'Unassigned':
        return 'Unassigned'
    if prop.get('needs_stock'):
        return 'New project'
    linked = bool(prop.get('linked') or prop.get('from_tickets'))
    if linked:
        return 'Linked'
    return 'Standalone'


def _material_rows(q='', department='', property_name=''):
    search = (q or '').strip().lower()
    dept = (department or '').strip()
    site = (property_name or '').strip()
    out = []
    for row in ProcStock.query.order_by(ProcStock.created_at.desc()).all():
        m = row.to_material_dict()
        if dept and (m.get('category') or '') != dept:
            continue
        if site and (m.get('property') or '') != site:
            continue
        if search:
            blob = f"{m.get('material_name') or ''} {m.get('brand') or ''} {m.get('supplier') or ''} {m.get('property') or ''}".lower()
            if search not in blob:
                continue
        out.append(m)
    return out


def _usage_filters(args):
    return dict(
        movement_type=(args.get('type') or '').strip(),
        property_name=args.get('property', ''),
        property_id=(args.get('property_id') or '').strip(),
        department=args.get('department', ''),
        status=args.get('status', ''),
        search=args.get('q', ''),
        limit=1000,
    )


def export_dashboard(fmt, *, range_key='month', grain='month', break_key=None):
    data = svc.dashboard_payload(range_key or 'month', break_key)
    kpis = data['kpis']
    period = data.get('period_label') or data.get('period') or ''
    recent = data.get('recent') or []
    grain = (grain or 'month').strip().lower()
    if grain not in GRAIN_LABEL:
        grain = 'month'
    util = (data.get('charts') or {}).get(grain) or data.get('utilization') or {}
    util_labels = list(util.get('labels') or [])
    util_series = util.get('series') or {}
    daily = data.get('daily') or {}
    daily_labels = list(daily.get('labels') or [])
    daily_issued = [float(v or 0) for v in (daily.get('issued') or [])]
    daily_compare = [float(v or 0) for v in (daily.get('compare') or [])]
    if not any(v > 0 for v in daily_issued):
        daily_labels = util_labels
        n = len(util_labels)
        daily_issued = [0.0] * n
        for vals in util_series.values():
            for i, v in enumerate(vals or []):
                if i < n:
                    daily_issued[i] += float(v or 0)
        daily_compare = [0.0] * n
    if len(daily_compare) < len(daily_issued):
        daily_compare = daily_compare + [0.0] * (len(daily_issued) - len(daily_compare))
    daily_compare = daily_compare[:len(daily_issued)]
    if fmt == 'xlsx':
        util_rows = []
        for trade in TRADE_ORDER:
            vals = list(util_series.get(trade) or [])
            if not any(float(v or 0) > 0 for v in vals):
                continue
            util_rows.append(
                [trade] + [round(float(vals[i] if i < len(vals) else 0), 1) for i in range(len(util_labels))]
            )
        return _xlsx_book([
            {
                'name': 'Overview',
                'title': 'Procurement overview',
                'subtitle': period,
                'headers': ['Metric', 'Value', 'Note'],
                'rows': [
                    ['Items below threshold', kpis['below_threshold']['value'], 'Refill queue'],
                    ['Issued this period', round(kpis['issued']['value'], 1), _pct_vs(kpis['issued']['delta'])],
                    ['Monthly spend (AED)', round(kpis['spend']['value'], 2), _pct_vs(kpis['spend']['delta'])],
                    ['Stock value (AED)', round(kpis['stock_value']['value'], 2), 'On-hand × rate card'],
                ],
                'widths': [28, 18, 36],
            },
            {
                'name': 'Utilization',
                'title': f'Material utilization · {GRAIN_LABEL[grain]}',
                'subtitle': period,
                'headers': ['Trade'] + util_labels,
                'rows': util_rows,
                'widths': [16] + [10] * max(len(util_labels), 1),
            },
            {
                'name': 'Breakdown',
                'title': 'Issued this period',
                'subtitle': period,
                'headers': ['Day', 'This period', 'Last period'],
                'rows': [
                    [lab, round(a, 1), round(b, 1)]
                    for lab, a, b in zip(daily_labels, daily_issued, daily_compare)
                ],
                'widths': [16, 14, 14],
            },
            {
                'name': 'Recent usage',
                'title': 'Recent usage',
                'subtitle': 'Materials issued or returned on service tickets',
                'headers': ['ID', 'Ticket', 'Material', 'Property', 'Qty', 'UOM', 'Status', 'Date'],
                'rows': [[
                    r.get('log_id'), r.get('ticket_id') or '', r.get('material_name'),
                    r.get('property') or '', r.get('qty'), r.get('uom') or '',
                    STOCK_STATUS.get(r.get('status'), r.get('status')),
                    (r.get('created_at') or '')[:16].replace('T', ' '),
                ] for r in recent],
                'widths': [14, 14, 28, 18, 10, 8, 14, 18],
            },
        ], 'procurement_overview')
    chart_h = 44 * mm
    chart_w = AVAIL_W - 12
    active_trades = [
        t for t in TRADE_ORDER
        if any(float(v or 0) > 0 for v in (util_series.get(t) or []))
    ]
    util_legend = [(TRADE_COLORS[t], t) for t in active_trades]
    break_legend = [
        (colors.HexColor('#18181b'), 'This period'),
        (colors.HexColor('#c4c4cc'), 'Last period'),
    ]
    range_chip = _fmt_range_chip(
        data.get('break_period_start') or data.get('period_start'),
        data.get('break_period_end') or data.get('period_end'),
        data.get('break_period_label') or period,
    )
    util_card = _chart_card(
        'Material utilization',
        f"Issued {_fmt_qty(_series_sum(util_series))} · {GRAIN_LABEL[grain]}",
        _CanvasChart(
            chart_w, chart_h,
            lambda c, w, h, lb=util_labels, s=util_series, lg=util_legend: _paint_util(c, w, h, lb, s, lg),
        ),
        AVAIL_W,
    )
    break_card = _chart_card(
        'Breakdown',
        f"{_fmt_qty(sum(daily_issued))} · {range_chip}",
        _CanvasChart(
            chart_w, chart_h,
            lambda c, w, h, lb=daily_labels, a=daily_issued, b=daily_compare, lg=break_legend: _paint_breakdown(c, w, h, lb, a, b, lg),
        ),
        AVAIL_W,
    )
    story = [
        _kpi_row([
            ('Items below threshold', str(kpis['below_threshold']['value']), 'Refill queue'),
            ('Issued this period', f"{kpis['issued']['value']:,.0f}", _pct_vs(kpis['issued']['delta'])),
            ('Monthly spend', _fmt_aed(kpis['spend']['value']), _pct_vs(kpis['spend']['delta'])),
            ('Stock value', _fmt_aed(kpis['stock_value']['value']), 'On-hand × rate card'),
        ]),
        Spacer(1, 8),
        util_card,
        Spacer(1, 6),
        break_card,
        Spacer(1, 8),
        _section_head('Recent usage', f'{len(recent)} movements'),
        Spacer(1, 3),
        _data_table(
            ['ID', 'Ticket', 'Material', 'Property', 'Qty', 'Status', 'Date'],
            [[
                _p(r.get('log_id'), 7),
                _p(r.get('ticket_id') or '—', 7),
                _p(r.get('material_name'), 7.5),
                _p(r.get('property') or '—', 7.5),
                _p(_fmt_qty(r.get('qty'), r.get('uom') or ''), 7.5),
                _status_text(r.get('status')),
                _p((r.get('created_at') or '')[:10], 7.5),
            ] for r in recent] or [[_p('No recent movements'), '', '', '', '', '', '']],
            [18 * mm, 16 * mm, 44 * mm, 26 * mm, 20 * mm, 24 * mm, 20 * mm],
        ),
    ]
    return _build_pdf('Procurement overview', period or 'Stock, usage, and refill', story), 'procurement_overview.pdf'


def export_usage_log(fmt, args):
    rows = svc.usage_log_rows(**_usage_filters(args))
    if fmt == 'xlsx':
        return _xlsx_book([{
            'name': 'Usage log',
            'title': 'Usage log',
            'subtitle': 'Every material issued, returned, or received',
            'headers': ['ID', 'Ticket', 'Material', 'Property', 'Qty', 'UOM', 'Status', 'Issued by', 'Date'],
            'rows': [[
                r.get('log_id'), r.get('ticket_id') or '', r.get('material_name'),
                r.get('property') or '', r.get('qty'), r.get('uom') or '',
                STOCK_STATUS.get(r.get('status'), r.get('status')),
                r.get('submitted_by') or '',
                (r.get('created_at') or '')[:16].replace('T', ' '),
            ] for r in rows],
            'widths': [14, 14, 28, 18, 10, 8, 14, 22, 18],
        }], 'procurement_usage_log')
    story = [
        _section_head('Material movements', f'{len(rows)} movement{"s" if len(rows) != 1 else ""}'),
        Spacer(1, 4),
        _data_table(
            ['ID', 'Material', 'Property', 'Qty', 'Status', 'By', 'Date'],
            [[
                _p(r.get('log_id'), 7),
                _p(r.get('material_name'), 7.5),
                _p(r.get('property') or '—', 7.5),
                _p(_fmt_qty(r.get('qty'), r.get('uom') or ''), 7.5),
                _stock_chip(r.get('status')),
                _p(r.get('submitted_by') or '—', 7),
                _p((r.get('created_at') or '')[:10], 7.5),
            ] for r in rows] or [[_p('No movements match these filters.'), '', '', '', '', '', '']],
            [20 * mm, 42 * mm, 28 * mm, 22 * mm, 30 * mm, 28 * mm, 20 * mm],
        ),
    ]
    return _build_pdf('Usage log', 'Issued, returned, or received into stock', story), 'procurement_usage_log.pdf'


def export_materials(fmt, args):
    rows = _material_rows(
        q=args.get('q', ''),
        department=args.get('department', ''),
        property_name=args.get('property', ''),
    )
    if fmt == 'xlsx':
        return _xlsx_book([{
            'name': 'Property stock',
            'title': 'Property stock',
            'subtitle': 'On-hand quantities across sites',
            'headers': ['#', 'Material', 'Brand', 'Department', 'UOM', 'Unit price (AED)', 'Property', 'Qty'],
            'rows': [[
                i, m.get('material_name'), m.get('brand') or m.get('supplier') or '',
                m.get('category') or 'General', m.get('unit') or 'PCS',
                float(m.get('unit_price') or 0), m.get('property') or '', m.get('quantity') or 0,
            ] for i, m in enumerate(rows, start=1)],
            'widths': [6, 28, 16, 14, 10, 16, 18, 10],
        }], 'procurement_property_stock')
    story = [
        _section_head('Property stock', f'{len(rows)} material{"s" if len(rows) != 1 else ""}'),
        Spacer(1, 4),
        _data_table(
            ['#', 'Material', 'Brand', 'Dept', 'UOM', 'Qty', 'Property', 'Unit price'],
            [[
                _p(str(i), 7.5, color=MUTED),
                _p(m.get('material_name'), 8, bold=True),
                _p(m.get('brand') or m.get('supplier') or '—', 7.5, color=MUTED),
                _p(m.get('category') or 'General', 7.5),
                _p(m.get('unit') or 'PCS', 7.5, color=MUTED),
                _p(str(m.get('quantity') if m.get('quantity') not in (None, '') else '—'), 7.5, align=TA_RIGHT),
                _p(m.get('property') or '—', 7.5),
                _p(_fmt_aed(m.get('unit_price'), 2) if float(m.get('unit_price') or 0) else '—', 7.5, align=TA_RIGHT),
            ] for i, m in enumerate(rows, start=1)] or [[_p('No materials match.'), '', '', '', '', '', '', '']],
            [8 * mm, 38 * mm, 22 * mm, 18 * mm, 12 * mm, 14 * mm, 28 * mm, 28 * mm],
        ),
    ]
    return _build_pdf('Property stock', 'On-hand quantities across sites', story), 'procurement_property_stock.pdf'


def export_properties(fmt):
    rows = svc.properties_with_counts()
    if fmt == 'xlsx':
        return _xlsx_book([{
            'name': 'Properties',
            'title': 'Property management',
            'subtitle': 'Materials by site or location',
            'headers': ['Property', 'Type', 'Materials', 'Items', 'AED value', 'Note'],
            'rows': [[
                p.get('display_name') or p.get('name'),
                _prop_badge(p),
                p.get('materials_count') or 0,
                round(float(p.get('total_quantity') or 0)),
                round(float(p.get('total_value') or 0), 2),
                'Available on every ticket' if p.get('is_shared') or p.get('name') == 'Shared'
                else ('Procurement only' if not (p.get('linked') or p.get('from_tickets')) else 'Linked to tickets'),
            ] for p in rows],
            'widths': [28, 16, 12, 12, 14, 28],
        }], 'procurement_properties')
    cards = []
    width = (AVAIL_W - 8 * mm) / 2
    pair = []
    for p in rows:
        badge = _prop_badge(p)
        val = float(p.get('total_value') or 0)
        val_s = f'{val / 1000:.1f}K' if val >= 1000 else f'{val:,.0f}'
        card = Table(
            [
                [_p(badge.upper(), 6.5, bold=True, color=brand.PRIMARY_DARK, leading=8)],
                [_p(p.get('display_name') or p.get('name'), 10, bold=True, leading=13)],
                [_p(
                    f"{p.get('materials_count') or 0} materials   "
                    f"{round(float(p.get('total_quantity') or 0)):,} items   "
                    f"AED {val_s}",
                    7.5, color=MUTED, leading=10,
                )],
            ],
            colWidths=[width],
        )
        card.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), brand.SURFACE_ALT),
            ('BOX', (0, 0), (-1, -1), 0.4, RULE),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (0, 0), 8),
            ('BOTTOMPADDING', (0, -1), (0, -1), 8),
        ]))
        pair.append(card)
        if len(pair) == 2:
            cards.append(pair)
            pair = []
    if pair:
        pair.append(Spacer(width, 1))
        cards.append(pair)
    grid = []
    for row in cards:
        t = Table([row], colWidths=[width + 4 * mm, width + 4 * mm])
        t.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        grid.append(KeepTogether(t))
    story = [
        _section_head('Sites', f'{len(rows)} propert{"y" if len(rows) == 1 else "ies"}'),
        Spacer(1, 6),
    ]
    story.extend(grid or [_p('No properties yet.')])
    return _build_pdf('Property management', 'Organize and track materials by site', story), 'procurement_properties.pdf'


def export_refill(fmt, args):
    dept = (args.get('department') or '').strip()
    rows = svc.refill_rows()
    if dept:
        rows = [r for r in rows if (r.get('department') or '') == dept]
    out_of_stock = sum(1 for r in rows if float(r.get('qty') or 0) <= 0)
    suggested_value = sum(float(r.get('suggested_qty') or 0) * float(r.get('unit_price') or 0) for r in rows)
    if fmt == 'xlsx':
        return _xlsx_book([{
            'name': 'Refill queue',
            'title': 'Refill queue',
            'subtitle': 'Items at or below minimum',
            'headers': [
                'Material', 'Brand', 'Department', 'Property', 'Status',
                'On hand', 'Minimum', 'UOM', 'Order qty', 'Value (AED)', 'When', 'Last event',
            ],
            'rows': [[
                r.get('name'), r.get('brand') or '', r.get('department') or '',
                r.get('property') or '', STOCK_STATUS.get(r.get('status'), r.get('status')),
                r.get('qty'), r.get('min_qty'), r.get('uom') or 'PCS',
                r.get('suggested_qty'),
                round(float(r.get('suggested_qty') or 0) * float(r.get('unit_price') or 0), 2),
                _fmt_when(r.get('updated_at')),
                r.get('last_event') or '',
            ] for r in rows],
            'widths': [26, 14, 12, 16, 14, 10, 10, 8, 12, 14, 20, 20],
        }], 'procurement_refill_queue')
    story = [
        _kpi_row([
            ('To refill', str(len(rows)), 'Below minimum'),
            ('Out of stock', str(out_of_stock), 'Zero on hand'),
            ('Suggested value', _fmt_aed(suggested_value), 'Order qty × rate card'),
        ]),
        Spacer(1, 12),
        _section_head('Needs refill', f'{len(rows)} item{"s" if len(rows) != 1 else ""}'),
        Spacer(1, 4),
        _data_table(
            ['Material', 'Property', 'Stock', 'Order', 'Value'],
            [[
                Table(
                    [
                        [_p(r.get('name'), 8, bold=True, leading=10)],
                        [_p(
                            f"{r.get('department') or '—'} · {r.get('brand') or '—'} · "
                            f"{STOCK_STATUS.get(r.get('status'), r.get('status') or '')}",
                            7, color=MUTED, leading=9,
                        )],
                    ],
                    colWidths=[52 * mm],
                ),
                _p(r.get('property') or '—', 7.5),
                _p(_fmt_qty(r.get('qty'), r.get('uom') or '') + f" / {_fmt_qty(r.get('min_qty'))}", 7.5),
                _p(_fmt_qty(r.get('suggested_qty'), r.get('uom') or ''), 7.5),
                _p(_fmt_aed(float(r.get('suggested_qty') or 0) * float(r.get('unit_price') or 0)), 7.5, align=TA_RIGHT),
            ] for r in rows] or [[_p('Nothing needs refill.'), '', '', '', '']],
            [54 * mm, 32 * mm, 32 * mm, 28 * mm, 34 * mm],
        ),
    ]
    return _build_pdf('Refill queue', 'Items at or below minimum', story), 'procurement_refill_queue.pdf'


def export_purchase_requests(fmt):
    rows = [
        r.to_dict(with_lines=False)
        for r in ProcPurchaseRequest.query.order_by(ProcPurchaseRequest.created_at.desc()).all()
    ]
    if fmt == 'xlsx':
        return _xlsx_book([{
            'name': 'Purchase requests',
            'title': 'Purchase requests',
            'subtitle': 'Review, approve, order, and receive stock',
            'headers': ['ID', 'Status', 'Property', 'Supplier', 'Total (AED)'],
            'rows': [[
                r.get('id'), PR_STATUS.get(r.get('status'), r.get('status')),
                r.get('property') or '', r.get('supplier') or '',
                round(float(r.get('total_aed') or 0), 2),
            ] for r in rows],
            'widths': [18, 22, 22, 28, 14],
        }], 'procurement_purchase_requests')
    story = [
        _section_head('Requests', f'{len(rows)} request{"s" if len(rows) != 1 else ""}'),
        Spacer(1, 4),
        _data_table(
            ['ID', 'Status', 'Property', 'Supplier', 'Total (AED)'],
            [[
                _p(r.get('id'), 8, bold=True, color=brand.PRIMARY_DARK),
                _pr_chip(r.get('status')),
                _p(r.get('property') or '—', 8),
                _p(r.get('supplier') or '—', 8),
                _p(f"{float(r.get('total_aed') or 0):,.2f}", 8, align=TA_RIGHT),
            ] for r in rows] or [[_p('No requests yet.'), '', '', '', '']],
            [32 * mm, 38 * mm, 32 * mm, 48 * mm, 30 * mm],
        ),
    ]
    return _build_pdf(
        'Purchase requests',
        'Review, approve, order, and receive stock',
        story,
    ), 'procurement_purchase_requests.pdf'


def build_export(kind: str, fmt: str, args) -> tuple[bytes, str, str]:
    if kind not in REPORTS:
        raise ValueError('Unknown export')
    fmt = (fmt or 'xlsx').lower()
    if fmt not in ('pdf', 'xlsx', 'excel'):
        raise ValueError('format must be pdf or xlsx')
    if fmt == 'excel':
        fmt = 'xlsx'
    if kind == 'dashboard':
        payload = export_dashboard(
            fmt,
            range_key=args.get('range', 'month'),
            grain=args.get('grain', 'month'),
            break_key=args.get('break'),
        )
    elif kind == 'usage-log':
        payload = export_usage_log(fmt, args)
    elif kind == 'materials':
        payload = export_materials(fmt, args)
    elif kind == 'properties':
        payload = export_properties(fmt)
    elif kind == 'refill':
        payload = export_refill(fmt, args)
    else:
        payload = export_purchase_requests(fmt)
    data, filename = payload
    if fmt == 'xlsx':
        mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return data, filename, mime
    return data, filename, 'application/pdf'
