"""Local Files storage: folders, items, save-from-module builders."""

from __future__ import annotations

import logging
import os
import re
import uuid
from datetime import datetime
from io import BytesIO
from typing import Optional, Tuple

from flask import current_app
from sqlalchemy.orm import joinedload
from werkzeug.utils import secure_filename

from app.models import (
    FilesFolder,
    FilesItem,
    LeaveEmployee,
    LeaveLog,
    LeavePlan,
    db,
)
from common.datetime_utils import utc_now_naive
from module_files.catalog import DEFAULT_FOLDER_TREE, get_module_catalog

logger = logging.getLogger(__name__)

XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def _files_root() -> str:
    gen = current_app.config.get('GENERATED_DIR') or os.path.join(current_app.root_path, 'generated')
    root = os.path.join(gen, 'files')
    os.makedirs(root, exist_ok=True)
    return root


def ensure_default_folders(created_by: Optional[int] = None) -> dict:
    """Create HR / Leave / Manpower folder tree if missing. Returns path_key → folder."""
    by_key = {f.path_key: f for f in FilesFolder.query.filter(FilesFolder.path_key.isnot(None)).all()}
    now = utc_now_naive()
    for spec in DEFAULT_FOLDER_TREE:
        key = spec['path_key']
        if key in by_key:
            continue
        parent_id = None
        parent_key = spec.get('parent_key')
        if parent_key:
            parent = by_key.get(parent_key)
            if not parent:
                # Parent should already exist from earlier loop order
                parent = FilesFolder.query.filter_by(path_key=parent_key).first()
            parent_id = parent.id if parent else None
        folder = FilesFolder(
            name=spec['name'],
            parent_id=parent_id,
            path_key=key,
            created_by=created_by,
            created_at=now,
            updated_at=now,
        )
        db.session.add(folder)
        db.session.flush()
        by_key[key] = folder
    db.session.commit()
    return by_key


def get_folder_by_path_key(path_key: str, created_by: Optional[int] = None) -> FilesFolder:
    ensure_default_folders(created_by=created_by)
    folder = FilesFolder.query.filter_by(path_key=path_key).first()
    if not folder:
        raise ValueError(f'Folder not found for path_key={path_key}')
    return folder


def build_tree() -> dict:
    ensure_default_folders()
    folders = FilesFolder.query.order_by(FilesFolder.name.asc()).all()
    items = (
        FilesItem.query.options(joinedload(FilesItem.folder))
        .order_by(FilesItem.updated_at.desc())
        .all()
    )
    return {
        'folders': [f.to_dict() for f in folders],
        'items': [i.to_dict() for i in items],
    }


def _safe_store_name(filename: str) -> str:
    base = secure_filename(filename) or 'file.bin'
    return f'{uuid.uuid4().hex[:12]}_{base}'


def save_bytes_to_folder(
    *,
    folder: FilesFolder,
    data: bytes,
    display_name: str,
    filename: str,
    mime_type: str,
    source_module: Optional[str],
    source_kind: Optional[str],
    created_by: Optional[int],
) -> FilesItem:
    folder_dir = os.path.join(_files_root(), str(folder.id))
    os.makedirs(folder_dir, exist_ok=True)
    store_name = _safe_store_name(filename)
    abs_path = os.path.join(folder_dir, store_name)
    with open(abs_path, 'wb') as fh:
        fh.write(data)

    # Store path relative to GENERATED_DIR for portability
    gen = current_app.config.get('GENERATED_DIR') or ''
    stored = abs_path
    if gen and abs_path.startswith(gen):
        stored = os.path.relpath(abs_path, gen)

    now = utc_now_naive()
    item = FilesItem(
        folder_id=folder.id,
        name=display_name,
        filename=filename,
        mime_type=mime_type,
        size_bytes=len(data),
        stored_path=stored,
        source_module=source_module,
        source_kind=source_kind,
        sync_status='local',
        created_by=created_by,
        created_at=now,
        updated_at=now,
    )
    db.session.add(item)
    db.session.commit()
    return item


def resolve_item_abs_path(item: FilesItem) -> str:
    path = item.stored_path or ''
    if os.path.isabs(path) and os.path.isfile(path):
        return path
    gen = current_app.config.get('GENERATED_DIR') or ''
    candidate = os.path.join(gen, path) if gen else path
    if os.path.isfile(candidate):
        return candidate
    # Fallback: under files root
    alt = os.path.join(_files_root(), path)
    if os.path.isfile(alt):
        return alt
    raise FileNotFoundError(f'File missing on disk: {item.filename}')


def create_folder(name: str, parent_id: Optional[int], created_by: Optional[int]) -> FilesFolder:
    name = (name or '').strip()
    if not name:
        raise ValueError('Folder name is required')
    if parent_id is not None:
        parent = db.session.get(FilesFolder, parent_id)
        if not parent:
            raise ValueError('Parent folder not found')
    now = utc_now_naive()
    folder = FilesFolder(
        name=name,
        parent_id=parent_id,
        path_key=None,
        created_by=created_by,
        created_at=now,
        updated_at=now,
    )
    db.session.add(folder)
    db.session.commit()
    return folder


def rename_folder(folder_id: int, name: str) -> FilesFolder:
    name = (name or '').strip()
    if not name:
        raise ValueError('Folder name is required')
    folder = db.session.get(FilesFolder, folder_id)
    if not folder:
        raise ValueError('Folder not found')
    folder.name = name
    folder.updated_at = utc_now_naive()
    db.session.commit()
    return folder


def rename_item(item_id: int, name: str) -> FilesItem:
    name = (name or '').strip()
    if not name:
        raise ValueError('Name is required')
    item = db.session.get(FilesItem, item_id)
    if not item:
        raise ValueError('File not found')
    item.name = name
    # Keep extension on filename if present
    _, ext = os.path.splitext(item.filename or '')
    safe = re.sub(r'[^\w.\- ()]+', '_', name).strip() or item.filename
    if ext and not safe.lower().endswith(ext.lower()):
        safe = f'{safe}{ext}'
    item.filename = safe
    item.updated_at = utc_now_naive()
    if item.sync_status == 'synced':
        item.sync_status = 'local'  # needs re-sync after rename
    db.session.commit()
    return item


def _collect_folder_ids(folder_id: int) -> list[int]:
    """Return folder_id and all descendant folder ids (depth-first)."""
    ids = [folder_id]
    kids = FilesFolder.query.filter_by(parent_id=folder_id).all()
    for kid in kids:
        ids.extend(_collect_folder_ids(kid.id))
    return ids


def delete_item(item_id: int) -> dict:
    """Delete a file from DB + local disk. Returns drive_file_id if present for Drive cleanup."""
    item = db.session.get(FilesItem, item_id)
    if not item:
        raise ValueError('File not found')
    drive_file_id = item.drive_file_id
    disk_path = None
    try:
        disk_path = resolve_item_abs_path(item)
    except FileNotFoundError:
        disk_path = None

    # Remove Drive copy first while we still have a clear intent + id
    drive_removed = None
    if drive_file_id:
        try:
            from module_files import drive_service as _drive
            drive_removed = bool(_drive.delete_on_drive_item(drive_file_id))
        except Exception:
            logger.exception('Drive delete during item delete failed for %s', item_id)
            drive_removed = False

    db.session.delete(item)
    db.session.commit()
    if disk_path and os.path.isfile(disk_path):
        try:
            os.remove(disk_path)
        except OSError:
            logger.exception('Could not remove local file %s', disk_path)
    return {
        'id': item_id,
        'drive_file_id': drive_file_id,
        'drive_removed': drive_removed,
    }


def delete_folder(folder_id: int) -> dict:
    """Delete folder + descendants + items. Blocks seeded catalog folders (path_key set)."""
    folder = db.session.get(FilesFolder, folder_id)
    if not folder:
        raise ValueError('Folder not found')
    if folder.path_key:
        raise ValueError('System folders cannot be deleted. You can rename them instead.')

    folder_ids = _collect_folder_ids(folder_id)
    items = FilesItem.query.filter(FilesItem.folder_id.in_(folder_ids)).all()
    drive_file_ids = [i.drive_file_id for i in items if i.drive_file_id]
    # Parent-first; Drive helper reverses to deepest-first
    drive_folder_ids = []
    for fid in folder_ids:
        f = db.session.get(FilesFolder, fid)
        if f and f.drive_folder_id:
            drive_folder_ids.append(f.drive_folder_id)

    # Remove Drive copies first (before local rows disappear)
    drive_stats = {'drive_deleted': [], 'drive_failed': []}
    if drive_file_ids or drive_folder_ids:
        try:
            from module_files import drive_service as _drive
            drive_stats = _drive.delete_on_drive_ids(
                file_ids=drive_file_ids,
                folder_ids=drive_folder_ids,
            ) or drive_stats
        except Exception:
            logger.exception('Drive delete during folder delete failed for %s', folder_id)

    # Remove local files
    for item in items:
        try:
            path = resolve_item_abs_path(item)
            if os.path.isfile(path):
                os.remove(path)
        except (FileNotFoundError, OSError):
            pass

    for item in items:
        db.session.delete(item)

    for fid in reversed(folder_ids):
        f = db.session.get(FilesFolder, fid)
        if f:
            db.session.delete(f)
    db.session.commit()

    return {
        'id': folder_id,
        'deleted_folder_ids': folder_ids,
        'drive_file_ids': drive_file_ids,
        'drive_folder_ids': drive_folder_ids,
        'drive_stats': drive_stats,
    }


def save_upload(
    *,
    folder_id: int,
    file_storage,
    created_by: Optional[int],
) -> FilesItem:
    folder = db.session.get(FilesFolder, folder_id)
    if not folder:
        raise ValueError('Folder not found')
    raw_name = file_storage.filename or 'upload.bin'
    filename = secure_filename(raw_name) or 'upload.bin'
    data = file_storage.read()
    if not data:
        raise ValueError('Empty file')
    mime = file_storage.mimetype or 'application/octet-stream'
    return save_bytes_to_folder(
        folder=folder,
        data=data,
        display_name=filename,
        filename=filename,
        mime_type=mime,
        source_module='upload',
        source_kind='upload',
        created_by=created_by,
    )


def _build_manpower_bytes(kind: str) -> Tuple[bytes, str, str]:
    from module_hr.manpower_excel import (
        build_manpower_export_bytes,
        build_manpower_template_bytes,
    )

    stamp = datetime.utcnow().strftime('%Y%m%d_%H%M')
    if kind == 'template':
        data = build_manpower_template_bytes()
        filename = f'manpower_template_{stamp}.xlsx'
        display = f'Manpower template ({stamp})'
    elif kind == 'export':
        data = build_manpower_export_bytes()
        filename = f'manpower_export_{stamp}.xlsx'
        display = f'Manpower export ({stamp})'
    else:
        raise ValueError('Invalid kind for manpower')
    return data, filename, display


def _build_leave_bytes(kind: str) -> Tuple[bytes, str, str]:
    from app.models import LEAVE_TRACKER_YEAR
    from module_hr.leave_excel import (
        build_leave_log_template_bytes,
        build_leave_workbook,
    )

    stamp = datetime.utcnow().strftime('%Y%m%d_%H%M')
    if kind == 'template':
        buf = build_leave_log_template_bytes()
        data = buf.getvalue() if isinstance(buf, BytesIO) else bytes(buf)
        filename = f'leave_log_template_{LEAVE_TRACKER_YEAR}_{stamp}.xlsx'
        display = f'Leave log template ({stamp})'
    elif kind == 'export':
        from module_hr.leave_tracker import LEAVE_WINDOW_END, LEAVE_WINDOW_START

        employees = (
            LeaveEmployee.query.filter_by(active=True)
            .options(joinedload(LeaveEmployee.usage))
            .order_by(LeaveEmployee.company.asc(), LeaveEmployee.full_name.asc())
            .all()
        )
        plans = (
            LeavePlan.query.options(joinedload(LeavePlan.employee))
            .filter(
                LeavePlan.start_date <= LEAVE_WINDOW_END,
                LeavePlan.end_date >= LEAVE_WINDOW_START,
            )
            .order_by(LeavePlan.start_date.asc())
            .all()
        )
        logs = (
            LeaveLog.query.options(joinedload(LeaveLog.employee))
            .filter(
                LeaveLog.leave_date >= LEAVE_WINDOW_START,
                LeaveLog.leave_date <= LEAVE_WINDOW_END,
            )
            .order_by(LeaveLog.leave_date.asc(), LeaveLog.id.asc())
            .all()
        )
        buf = build_leave_workbook(employees, plans, logs)
        data = buf.getvalue() if isinstance(buf, BytesIO) else bytes(buf)
        filename = f'leave_tracker_{LEAVE_TRACKER_YEAR}_{stamp}.xlsx'
        display = f'Leave export ({stamp})'
    else:
        raise ValueError('Invalid kind for leave')
    return data, filename, display


def _build_hiring_bytes(kind: str) -> Tuple[bytes, str, str]:
    from app.models import HiringCandidate
    from module_hr.hiring_excel import build_hiring_template_bytes

    stamp = datetime.utcnow().strftime('%Y%m%d_%H%M')
    if kind == 'template':
        data = build_hiring_template_bytes()
        filename = f'hiring_tracker_template_{stamp}.xlsx'
        display = f'Hiring template ({stamp})'
    elif kind == 'export':
        candidates = HiringCandidate.query.order_by(HiringCandidate.updated_at.desc()).all()
        data = build_hiring_template_bytes(candidates=candidates)
        filename = f'hiring_tracker_export_{stamp}.xlsx'
        display = f'Hiring export ({stamp})'
    else:
        raise ValueError('Invalid kind for hiring')
    return data, filename, display


def _build_procurement_bytes(kind: str) -> Tuple[bytes, str, str]:
    import pandas as pd
    from app.models import Submission

    stamp = datetime.utcnow().strftime('%Y%m%d_%H%M')
    if kind == 'template':
        rows = [
            {'Material Name': 'Office Paper A4 Ream', 'Category': 'Stationery', 'Description': '500 sheets per ream', 'Unit': 'ream', 'Quantity': 50, 'Unit Price': 12.50, 'Supplier': 'Gulf Paper Co', 'Notes': 'Monthly supply'},
            {'Material Name': 'Printer Toner Cartridge', 'Category': 'IT Supplies', 'Description': 'Laser printer compatible', 'Unit': 'pcs', 'Quantity': 10, 'Unit Price': 85.00, 'Supplier': 'Tech Supplies LLC', 'Notes': ''},
        ]
        df = pd.DataFrame(rows)
        filename = f'procurement_import_sample_{stamp}.xlsx'
        display = f'Procurement sample ({stamp})'
    elif kind == 'export':
        submissions = (
            Submission.query.filter(Submission.module_type == 'procurement_material')
            .order_by(Submission.created_at.desc())
            .all()
        )
        data_rows = []
        for sub in submissions:
            if sub.form_data:
                data_rows.append({
                    'ID': sub.submission_id,
                    'Material Name': sub.form_data.get('material_name', ''),
                    'Property': sub.form_data.get('property', 'Unassigned'),
                    'Category': sub.form_data.get('category', ''),
                    'Description': sub.form_data.get('description', ''),
                    'Unit': sub.form_data.get('unit', ''),
                    'Quantity': sub.form_data.get('quantity', 0),
                    'Unit Price (AED)': sub.form_data.get('unit_price', 0),
                    'Total Price (AED)': sub.form_data.get('total_price', 0),
                    'Supplier': sub.form_data.get('supplier', ''),
                    'Notes': sub.form_data.get('notes', ''),
                    'Added By': sub.form_data.get('added_by', ''),
                    'Date Added': sub.created_at.strftime('%Y-%m-%d') if sub.created_at else '',
                })
        df = pd.DataFrame(data_rows)
        filename = f'procurement_materials_{stamp}.xlsx'
        display = f'Procurement export ({stamp})'
    else:
        raise ValueError('Invalid kind for procurement')

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Materials')
    return output.getvalue(), filename, display


def _build_qhsi_bytes(kind: str) -> Tuple[bytes, str, str]:
    from module_qhsi.excel_import import build_import_template_bytes

    stamp = datetime.utcnow().strftime('%Y%m%d_%H%M')
    if kind != 'template':
        raise ValueError('Invalid kind for qhsi')
    data = build_import_template_bytes()
    if isinstance(data, BytesIO):
        data = data.getvalue()
    filename = f'qhsi_staff_compliance_template_{stamp}.xlsx'
    display = f'QHSE staff template ({stamp})'
    return data, filename, display


def _build_mmr_bytes(kind: str) -> Tuple[bytes, str, str]:
    import os
    from flask import current_app

    if kind != 'export':
        raise ValueError('Invalid kind for mmr')
    gen = current_app.config.get('GENERATED_DIR') or ''
    path = os.path.join(gen, 'mmr_latest.xlsx')
    if not os.path.isfile(path):
        raise ValueError('No MMR file uploaded yet. Upload a CAFM Excel in Report Generation first.')
    from module_mmr.mmr_service import generate_report_excel, parse_excel

    stamp = datetime.utcnow().strftime('%Y%m%d_%H%M')
    df = parse_excel(path)
    data = generate_report_excel(df)
    filename = f'mmr_report_{stamp}.xlsx'
    display = f'MMR report ({stamp})'
    return data, filename, display


def _build_devices_bytes(kind: str) -> Tuple[bytes, str, str]:
    import pandas as pd

    if kind != 'template':
        raise ValueError('Invalid kind for devices')
    stamp = datetime.utcnow().strftime('%Y%m%d_%H%M')
    rows = [
        {'Device Name': 'LAPTOP-HQ-001', 'Device Type': 'Laptop', 'OS': 'Windows 11 Pro', 'Status': 'online', 'Health': 96, 'Assigned User Email': 'admin@injaaz.ae', 'Serial / Asset Tag': 'AST-10001'},
        {'Device Name': 'DESKTOP-FIN-014', 'Device Type': 'Desktop', 'OS': 'Windows 10', 'Status': 'idle', 'Health': 88, 'Assigned User Email': '', 'Serial / Asset Tag': 'AST-10002'},
    ]
    df = pd.DataFrame(rows)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Devices')
    filename = f'device_import_sample_{stamp}.xlsx'
    display = f'Device sample ({stamp})'
    return output.getvalue(), filename, display


def _build_technicians_bytes(kind: str) -> Tuple[bytes, str, str]:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    if kind != 'template':
        raise ValueError('Invalid kind for technicians')
    stamp = datetime.utcnow().strftime('%Y%m%d_%H%M')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Technicians'
    headers = [
        'Employee ID', 'Full Name', 'Designation', 'Department',
        'Specialization', 'Phone', 'Email', 'Salary', 'Joining Date',
        'Status', 'Supervisor User ID', 'Notes',
    ]
    header_fill = PatternFill('solid', fgColor='125435')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    buf = BytesIO()
    wb.save(buf)
    filename = f'technicians_import_template_{stamp}.xlsx'
    display = f'Technicians template ({stamp})'
    return buf.getvalue(), filename, display


def save_from_module(module: str, kind: str, created_by: Optional[int]) -> Tuple[FilesItem, str]:
    """Build Excel via existing helpers and store in Files. Returns (item, folder_label)."""
    mod = (module or '').strip().lower()
    kind = (kind or '').strip().lower()
    catalog = get_module_catalog(mod)
    if not catalog:
        raise ValueError(f'Unsupported module: {module}')
    kinds = {o['kind'] for o in catalog['options']}
    if kind not in kinds:
        raise ValueError(f'Unsupported kind for {mod}: {kind}')

    builders = {
        'manpower': _build_manpower_bytes,
        'leave': _build_leave_bytes,
        'hiring': _build_hiring_bytes,
        'procurement': _build_procurement_bytes,
        'qhsi': _build_qhsi_bytes,
        'mmr': _build_mmr_bytes,
        'devices': _build_devices_bytes,
        'technicians': _build_technicians_bytes,
    }
    builder = builders.get(mod)
    if not builder:
        raise ValueError(f'Unsupported module: {module}')
    data, filename, display = builder(kind)

    folder = get_folder_by_path_key(catalog['folder_path_key'], created_by=created_by)
    item = save_bytes_to_folder(
        folder=folder,
        data=data,
        display_name=display,
        filename=filename,
        mime_type=XLSX_MIME,
        source_module=mod,
        source_kind=kind,
        created_by=created_by,
    )
    folder_label = catalog.get('folder_label') or folder.name
    return item, folder_label
