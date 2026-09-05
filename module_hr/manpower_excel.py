"""
Manpower Tracker — Excel template, export, and All Trades import.

Export layout: All Trades (every vacancy) plus one sheet per project
(all trades for that site). Import still reads All Trades only.
"""
from __future__ import annotations

import logging
from datetime import date, datetime
from io import BytesIO
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.models import (
    MANPOWER_REQUIREMENT_TYPE_LABELS,
    MANPOWER_REQUIREMENT_TYPES,
    MANPOWER_STATUS_LABELS,
    MANPOWER_STATUSES,
    ManpowerProject,
    ManpowerTrade,
    ManpowerVacancy,
    db,
)
from common.kynvera_excel_brand import (
    HINT_FONT,
    TITLE_FONT,
    InstructionSpec,
    apply_column_widths,
    style_header_row,
    write_data_row,
    write_instructions_sheet,
)

logger = logging.getLogger(__name__)

ALL_TRADES_HEADERS = (
    'Trade',
    'Project',
    'Requirement Type',
    'Replacement Name',
    'Replacement ID',
    'New Candidate Name',
    'Contact Number',
    'Status',
    'Date Joined',
    'Remarks',
    'Hiring Candidate ID',
)

# Column indices in All Trades sheet of the automated workbook (1-based)
# A=Trade, C=Project, F=Req Type, G=Replacement Name, H=Replacement ID,
# I=Candidate, J=Contact, K=Status, L=Date Joined, M=Remarks
_SRC_COL = {
    'trade': 1,
    'project': 3,
    'requirement_type': 6,
    'replacement_name': 7,
    'replacement_employee_id': 8,
    'candidate_name': 9,
    'contact_number': 10,
    'status': 11,
    'date_joined': 12,
    'remarks': 13,
}

HEADER_ALIASES = {
    'trade': ('trade', 'trades', 'designation', 'role'),
    'project': ('project', 'projects', 'site'),
    'requirement_type': (
        'requirement type', 'req type', 'reqt type', 'type', 'requirement',
    ),
    'replacement_name': (
        'replacement name', 'replacement', 'replacing', 'leaver name',
    ),
    'replacement_employee_id': (
        'replacement id', 'replacement employee id', 'replacement emp id',
        'leaver id', 'emp id',
    ),
    'candidate_name': (
        'new candidate name', 'candidate name', 'candidate', 'new candidate',
        'person name',
    ),
    'contact_number': (
        'contact number', 'contact', 'phone', 'mobile', 'telephone',
    ),
    'status': ('status', 'vacancy status'),
    'date_joined': ('date joined', 'joined date', 'joining date', 'join date'),
    'remarks': ('remarks', 'notes', 'comments', 'comment'),
    'hiring_candidate_id': (
        'hiring candidate id', 'candidate id', 'hiring id', 'hiring candidate',
    ),
}


def _norm(value) -> str:
    raw = str(value or '').strip().lower()
    cleaned = ''.join(ch if ch.isalnum() else ' ' for ch in raw)
    return ' '.join(cleaned.split())


def _cell_str(value) -> str:
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _parse_date(raw) -> Optional[date]:
    if raw is None or raw == '':
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    if not s:
        return None
    # Excel serial as float sometimes
    try:
        if isinstance(raw, (int, float)) and not isinstance(raw, bool):
            from openpyxl.utils.datetime import from_excel
            return from_excel(raw).date()
    except Exception:
        pass
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%d.%m.%Y'):
        try:
            return datetime.strptime(s[:10] if fmt == '%Y-%m-%d' else s, fmt).date()
        except ValueError:
            continue
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


def _normalize_status(raw) -> str:
    s = _norm(raw)
    if not s:
        return 'open'
    aliases = {
        'open': 'open',
        'still open': 'open',
        'interviewing': 'interviewing',
        'interview': 'interviewing',
        'in progress': 'interviewing',
        'selected': 'selected',
        'filled': 'filled',
        'fill': 'filled',
        'hired': 'joined',
        'joined': 'joined',
        'on hold': 'on_hold',
        'onhold': 'on_hold',
        'hold': 'on_hold',
    }
    if s in aliases:
        return aliases[s]
    # match labels
    for key, label in MANPOWER_STATUS_LABELS.items():
        if _norm(label) == s or key == s:
            return key
    return 'open'


def _normalize_req_type(raw) -> str:
    s = _norm(raw)
    if not s:
        return 'new'
    if 'replac' in s:
        return 'replacement'
    if s in ('new', 'fresh', 'addition'):
        return 'new'
    for key, label in MANPOWER_REQUIREMENT_TYPE_LABELS.items():
        if _norm(label) == s or key == s:
            return key
    return 'new'


def _style_header_row(ws, row: int, cols: int):
    style_header_row(ws, row, cols)


_INVALID_SHEET_CHARS = set(r':\/?*[]')
_RESERVED_SHEET_TITLES = frozenset({'all trades', 'lists', 'instructions'})
_VACANCY_COL_WIDTHS = [22, 22, 16, 20, 14, 28, 18, 14, 14, 28, 16]


def _safe_sheet_title(name: str, used: set[str]) -> str:
    """Excel sheet names: max 31 chars, no : \\ / ? * [ ]. Unique in the workbook."""
    cleaned = ''.join(ch for ch in (name or 'Project') if ch not in _INVALID_SHEET_CHARS)
    cleaned = ' '.join(cleaned.split()).strip("'")
    if not cleaned:
        cleaned = 'Project'
    if cleaned.lower() in _RESERVED_SHEET_TITLES:
        cleaned = f'{cleaned} project'
    base = cleaned[:31]
    title = base
    n = 2
    used_lower = {u.lower() for u in used}
    while title.lower() in used_lower:
        suffix = f' ({n})'
        title = (base[: max(1, 31 - len(suffix))] + suffix)[:31]
        n += 1
    used.add(title)
    return title


def _vacancy_row_values(v: ManpowerVacancy) -> list:
    status = v.normalized_status()
    req = v.normalized_requirement_type()
    return [
        v.trade.name if v.trade else '',
        v.project.name if v.project else '',
        MANPOWER_REQUIREMENT_TYPE_LABELS.get(req, req),
        v.replacement_name or '',
        v.replacement_employee_id or '',
        v.candidate_name or '',
        v.contact_number or '',
        MANPOWER_STATUS_LABELS.get(status, status),
        v.date_joined.isoformat() if v.date_joined else '',
        v.remarks or '',
        v.hiring_candidate_id or '',
    ]


def _sort_vacancies(vacancies: list[ManpowerVacancy]) -> list[ManpowerVacancy]:
    def key(v: ManpowerVacancy):
        return (
            (v.trade.sort_order if v.trade else 0),
            (v.trade.name if v.trade else ''),
            (v.project.sort_order if v.project else 0),
            (v.project.name if v.project else ''),
            v.sort_order or 0,
            v.id or 0,
        )

    return sorted(vacancies, key=key)


def _projects_for_sheets(
    project_rows: list[ManpowerProject],
    vacancies: Optional[list[ManpowerVacancy]],
) -> list[ManpowerProject]:
    """Active projects, plus any inactive/unlisted project that still has vacancies."""
    vacancy_ids = {
        getattr(v.project, 'id', None)
        for v in (vacancies or [])
        if getattr(v, 'project', None) is not None
    }
    vacancy_ids.discard(None)

    out: list[ManpowerProject] = []
    listed_ids: set[int] = set()
    listed_names: set[str] = set()
    for p in project_rows or []:
        name = (getattr(p, 'name', None) or '').strip()
        if not name:
            continue
        pid = getattr(p, 'id', None)
        active = bool(getattr(p, 'active', True))
        if active or pid in vacancy_ids:
            out.append(p)
            if pid is not None:
                listed_ids.add(pid)
            listed_names.add(name.lower())

    extra: list[ManpowerProject] = []
    for v in vacancies or []:
        proj = getattr(v, 'project', None)
        if not proj or not (getattr(proj, 'name', None) or '').strip():
            continue
        pid = getattr(proj, 'id', None)
        name_key = proj.name.strip().lower()
        if (pid is not None and pid in listed_ids) or name_key in listed_names:
            continue
        extra.append(proj)
        if pid is not None:
            listed_ids.add(pid)
        listed_names.add(name_key)
    return out + extra


def _attach_vacancy_validations(
    ws,
    *,
    n_trades: int,
    n_projects: int,
    n_status: int,
    n_req: int,
) -> None:
    dv_trade = DataValidation(
        type='list',
        formula1=f'Lists!$A$2:$A${n_trades + 1}',
        allow_blank=True,
    )
    dv_project = DataValidation(
        type='list',
        formula1=f'Lists!$B$2:$B${n_projects + 1}',
        allow_blank=True,
    )
    dv_status = DataValidation(
        type='list',
        formula1=f'Lists!$C$2:$C${n_status + 1}',
        allow_blank=True,
    )
    dv_req = DataValidation(
        type='list',
        formula1=f'Lists!$D$2:$D${n_req + 1}',
        allow_blank=True,
    )
    for dv in (dv_trade, dv_project, dv_status, dv_req):
        ws.add_data_validation(dv)
    dv_trade.add('A5:A500')
    dv_project.add('B5:B500')
    dv_req.add('C5:C500')
    dv_status.add('H5:H500')


def _write_vacancy_sheet(
    wb: Workbook,
    *,
    sheet_title: str,
    heading: str,
    subtitle: str,
    index: int,
    vacancies: list[ManpowerVacancy],
    n_trades: int,
    n_projects: int,
    n_status: int,
    n_req: int,
    example: Optional[list] = None,
):
    ws = wb.create_sheet(sheet_title, index)
    ws['A1'] = heading
    ws['A1'].font = TITLE_FONT
    ws['A2'] = subtitle
    ws['A2'].font = HINT_FONT
    ws.merge_cells('A1:K1')
    ws.merge_cells('A2:K2')

    header_row = 4
    for c, h in enumerate(ALL_TRADES_HEADERS, start=1):
        ws.cell(header_row, c, h)
    _style_header_row(ws, header_row, len(ALL_TRADES_HEADERS))
    apply_column_widths(ws, _VACANCY_COL_WIDTHS)
    _attach_vacancy_validations(
        ws,
        n_trades=n_trades,
        n_projects=n_projects,
        n_status=n_status,
        n_req=n_req,
    )

    if example:
        write_data_row(ws, header_row + 1, example, example=True)
        return ws

    for i, v in enumerate(vacancies):
        write_data_row(ws, header_row + 1 + i, _vacancy_row_values(v))
    return ws


def build_manpower_workbook(
    *,
    vacancies: Optional[list[ManpowerVacancy]] = None,
    trades: Optional[list[ManpowerTrade]] = None,
    projects: Optional[list[ManpowerProject]] = None,
    template_only: bool = False,
) -> bytes:
    """All Trades (every vacancy) + one sheet per project + Lists + Instructions."""
    wb = Workbook()

    # Lists sheet first so validations can reference it
    ws_lists = wb.active
    ws_lists.title = 'Lists'
    ws_lists['A1'] = 'Trades'
    ws_lists['B1'] = 'Projects'
    ws_lists['C1'] = 'Status'
    ws_lists['D1'] = 'Req Type'
    _style_header_row(ws_lists, 1, 4)

    trade_rows = trades if trades is not None else (
        ManpowerTrade.query.order_by(ManpowerTrade.sort_order, ManpowerTrade.name).all()
    )
    project_rows = projects if projects is not None else (
        ManpowerProject.query.order_by(ManpowerProject.sort_order, ManpowerProject.name).all()
    )

    for i, t in enumerate(trade_rows, start=2):
        ws_lists.cell(i, 1, t.name if hasattr(t, 'name') else str(t))
    for i, p in enumerate(project_rows, start=2):
        ws_lists.cell(i, 2, p.name if hasattr(p, 'name') else str(p))
    for i, key in enumerate(MANPOWER_STATUSES, start=2):
        ws_lists.cell(i, 3, MANPOWER_STATUS_LABELS[key])
    for i, key in enumerate(MANPOWER_REQUIREMENT_TYPES, start=2):
        ws_lists.cell(i, 4, MANPOWER_REQUIREMENT_TYPE_LABELS[key])

    for col in range(1, 5):
        ws_lists.column_dimensions[get_column_letter(col)].width = 22

    n_trades = max(len(trade_rows), 1)
    n_projects = max(len(project_rows), 1)
    n_status = len(MANPOWER_STATUSES)
    n_req = len(MANPOWER_REQUIREMENT_TYPES)
    sheet_kwargs = dict(
        n_trades=n_trades,
        n_projects=n_projects,
        n_status=n_status,
        n_req=n_req,
    )

    all_rows: list[ManpowerVacancy] = []
    if not template_only and vacancies:
        all_rows = _sort_vacancies(list(vacancies))

    example = None
    if template_only:
        example = [
            'Electrician',
            'Tower A',
            MANPOWER_REQUIREMENT_TYPE_LABELS['new'],
            '',
            '',
            'Sara Ahmed',
            '+971500000000',
            MANPOWER_STATUS_LABELS['open'],
            '',
            'Example row — replace or delete before importing',
            '',
        ]

    _write_vacancy_sheet(
        wb,
        sheet_title='All Trades',
        heading='Kynvera — Manpower Requirement Tracker',
        subtitle=(
            'All projects, all trades. One row per vacancy. Trade and Project are required. '
            'Use Lists sheet values for Status and Requirement Type. '
            'Each project also has its own sheet. Import uses this All Trades sheet.'
        ),
        index=0,
        vacancies=all_rows,
        example=example,
        **sheet_kwargs,
    )

    used_titles = {'All Trades', 'Lists', 'Instructions'}
    project_sheets = _projects_for_sheets(project_rows, None if template_only else vacancies)
    for offset, project in enumerate(project_sheets, start=1):
        project_name = (project.name if hasattr(project, 'name') else str(project)).strip()
        title = _safe_sheet_title(project_name, used_titles)
        pid = getattr(project, 'id', None)
        project_vacancies = [
            v for v in all_rows
            if v.project and (
                (pid is not None and getattr(v.project, 'id', None) == pid)
                or (v.project.name or '').strip().lower() == project_name.lower()
            )
        ]
        project_vacancies = _sort_vacancies(project_vacancies)
        _write_vacancy_sheet(
            wb,
            sheet_title=title,
            heading=f'Kynvera — {project_name}',
            subtitle=(
                f'Project-wise view: all trades for {project_name}. '
                'Same columns as All Trades. To re-import, edit and upload the All Trades sheet.'
            ),
            index=offset,
            vacancies=project_vacancies,
            **sheet_kwargs,
        )

    write_instructions_sheet(wb, _manpower_instruction_spec())

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _manpower_instruction_spec() -> InstructionSpec:
    status_opts = ', '.join(MANPOWER_STATUS_LABELS[k] for k in MANPOWER_STATUSES)
    req_opts = ', '.join(MANPOWER_REQUIREMENT_TYPE_LABELS[k] for k in MANPOWER_REQUIREMENT_TYPES)
    return InstructionSpec(
        title='Manpower vacancies template',
        module_label='HR / Manpower',
        about=(
            'Import vacancies into the Manpower Tracker (one row per open or filled role).',
            'The workbook has All Trades (every vacancy) and a project-wise sheet for each site (all trades on that project).',
            'Trade and Project are required. New trade or project names are created on import if they do not already exist.',
            'Dropdowns read from the Lists sheet (trades, projects, status, requirement type).',
        ),
        how_to=(
            'Open All Trades for the full list, or open a project tab for that site only. Keep the coral header row on row 4.',
            'Pick Trade and Project from the dropdowns (or type a new name to create it on import).',
            'Choose Requirement Type (New or Replacement) and Status from the lists.',
            'For replacements, fill Replacement Name and Replacement ID. For new hires, fill New Candidate Name and Contact Number.',
            'Date Joined is optional (YYYY-MM-DD). Use Hiring Candidate ID to link a hiring-docs candidate.',
            'Save as .xlsx and click Import on the Manpower Tracker page. Import reads All Trades only, not the project sheets.',
        ),
        columns=(
            ('Trade', 'Required. Must match a Lists value or a new trade name will be created.'),
            ('Project', 'Required. Must match a Lists value or a new project name will be created.'),
            ('Requirement Type', f'Required. Allowed: {req_opts}.'),
            ('Replacement Name', 'Optional. Person being replaced (when type is Replacement).'),
            ('Replacement ID', 'Optional. Employee ID of the person being replaced.'),
            ('New Candidate Name', 'Optional. Candidate being considered or hired.'),
            ('Contact Number', 'Optional. Phone for the new candidate.'),
            ('Status', f'Optional (defaults to Open). Allowed: {status_opts}.'),
            ('Date Joined', 'Optional. YYYY-MM-DD, DD/MM/YYYY, or Excel date.'),
            ('Remarks', 'Optional notes.'),
            ('Hiring Candidate ID', 'Optional numeric ID from Hiring Docs. Links the vacancy to that candidate.'),
        ),
        example_headers=ALL_TRADES_HEADERS,
        example_rows=((
            'Electrician',
            'Tower A',
            'New',
            '',
            '',
            'Sara Ahmed',
            '+971500000000',
            'Open',
            '',
            'Example row — replace or delete before importing',
            '',
        ),),
        import_rules=(
            'Rows missing Trade or Project are skipped.',
            'Status and Requirement Type are normalised from common aliases (e.g. Hired → Joined).',
            'Import can append vacancies or replace all existing vacancies, depending on the option you choose in the UI.',
            'The Lists sheet is not imported as data; it only feeds dropdowns.',
            'Project sheets are a site-wise view of the same vacancies. Edits there are not imported — copy them to All Trades first.',
        ),
        extra_sections=(
            (
                'Workbook sheets',
                (
                    'All Trades — every vacancy across every project and trade. This is the sheet used on import.',
                    'One tab per project — the same vacancies filtered to that site, still covering all trades.',
                    'Lists — dropdown values for trades, projects, status, and requirement type.',
                ),
            ),
        ),
    )


def build_manpower_template_bytes() -> bytes:
    return build_manpower_workbook(template_only=True)


def build_manpower_export_bytes() -> bytes:
    vacancies = (
        ManpowerVacancy.query
        .order_by(ManpowerVacancy.id)
        .all()
    )
    return build_manpower_workbook(vacancies=vacancies, template_only=False)


def _map_header_row(ws, row: int = 4) -> dict[str, int]:
    """Map normalized header -> 1-based column index."""
    mapping: dict[str, int] = {}
    for c in range(1, (ws.max_column or 0) + 1):
        key = _norm(ws.cell(row, c).value)
        if not key:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if key in aliases and field not in mapping:
                mapping[field] = c
                break
    return mapping


def _row_has_data(values: dict[str, Any]) -> bool:
    for k in (
        'trade', 'project', 'requirement_type', 'replacement_name',
        'replacement_employee_id', 'candidate_name', 'contact_number',
        'status', 'date_joined', 'remarks',
    ):
        if _cell_str(values.get(k)):
            return True
    return False


def _get_or_create_trade(name: str, cache: dict[str, ManpowerTrade], sort_counter: list) -> ManpowerTrade:
    key = name.strip()
    lookup = key.lower()
    if lookup in cache:
        return cache[lookup]
    existing = ManpowerTrade.query.filter(db.func.lower(ManpowerTrade.name) == lookup).first()
    if existing:
        cache[lookup] = existing
        return existing
    sort_counter[0] += 10
    row = ManpowerTrade(name=key, sort_order=sort_counter[0], active=True)
    db.session.add(row)
    db.session.flush()
    cache[lookup] = row
    return row


def _get_or_create_project(name: str, cache: dict[str, ManpowerProject], sort_counter: list) -> ManpowerProject:
    key = name.strip()
    lookup = key.lower()
    if lookup in cache:
        return cache[lookup]
    existing = ManpowerProject.query.filter(db.func.lower(ManpowerProject.name) == lookup).first()
    if existing:
        cache[lookup] = existing
        return existing
    sort_counter[0] += 10
    row = ManpowerProject(name=key, sort_order=sort_counter[0], active=True)
    db.session.add(row)
    db.session.flush()
    cache[lookup] = row
    return row


def parse_all_trades_rows(file_storage) -> list[dict[str, Any]]:
    """
    Parse vacancy rows from an uploaded workbook.
    Supports the automated Injaaz file (All Trades, forward-filled trade)
    and our flat export/template sheet.
    """
    wb = load_workbook(file_storage, data_only=True)
    sheet_name = None
    for candidate in ('All Trades', 'AllTrades', 'Vacancies', 'Sheet1'):
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # Detect header row (scan first 10 rows)
    header_row = None
    mapping: dict[str, int] = {}
    for r in range(1, min(12, (ws.max_row or 1) + 1)):
        m = _map_header_row(ws, r)
        if 'trade' in m and ('project' in m or 'status' in m):
            header_row = r
            mapping = m
            break

    rows_out: list[dict[str, Any]] = []
    last_trade = ''

    if header_row and mapping.get('trade') and mapping.get('project'):
        # Flat / our template style
        for r in range(header_row + 1, (ws.max_row or 0) + 1):
            values = {
                field: ws.cell(r, col).value
                for field, col in mapping.items()
            }
            trade = _cell_str(values.get('trade'))
            if trade:
                last_trade = trade
            else:
                trade = last_trade
            project = _cell_str(values.get('project'))
            # Skip empty trailing rows
            values['trade'] = trade
            values['project'] = project
            if not _row_has_data(values):
                continue
            if not trade or not project:
                # Incomplete row — skip
                continue
            rows_out.append({
                'trade': trade,
                'project': project,
                'requirement_type': _normalize_req_type(values.get('requirement_type')),
                'replacement_name': _cell_str(values.get('replacement_name')),
                'replacement_employee_id': _cell_str(values.get('replacement_employee_id')),
                'candidate_name': _cell_str(values.get('candidate_name')),
                'contact_number': _cell_str(values.get('contact_number')),
                'status': _normalize_status(values.get('status')),
                'date_joined': _parse_date(values.get('date_joined')),
                'remarks': _cell_str(values.get('remarks')),
                'hiring_candidate_id': _cell_str(values.get('hiring_candidate_id')),
            })
        return rows_out

    # Fallback: automated workbook layout (cols A/C/F–M), data starts row 5
    start = 5
    for r in range(start, (ws.max_row or 0) + 1):
        trade_raw = ws.cell(r, _SRC_COL['trade']).value
        project_raw = ws.cell(r, _SRC_COL['project']).value
        # Detect end-of-data / help text
        a_str = _cell_str(trade_raw).lower()
        if a_str.startswith('how to use') or a_str.startswith('1.') or a_str.startswith('sheet guide'):
            break

        # Gather typed columns — a row counts if any of C,F–M filled
        typed = {
            'project': project_raw,
            'requirement_type': ws.cell(r, _SRC_COL['requirement_type']).value,
            'replacement_name': ws.cell(r, _SRC_COL['replacement_name']).value,
            'replacement_employee_id': ws.cell(r, _SRC_COL['replacement_employee_id']).value,
            'candidate_name': ws.cell(r, _SRC_COL['candidate_name']).value,
            'contact_number': ws.cell(r, _SRC_COL['contact_number']).value,
            'status': ws.cell(r, _SRC_COL['status']).value,
            'date_joined': ws.cell(r, _SRC_COL['date_joined']).value,
            'remarks': ws.cell(r, _SRC_COL['remarks']).value,
        }
        if trade_raw is not None and _cell_str(trade_raw):
            last_trade = _cell_str(trade_raw)

        has_typed = any(_cell_str(v) for v in typed.values())
        if not has_typed:
            continue
        if not last_trade:
            continue
        project = _cell_str(typed['project'])
        if not project:
            # Project may be blank when continuing same project block —
            # look upward is already handled by Excel formulas in O; for import
            # we need project. Skip if no project on this or we can't infer.
            # Try forward-fill project as well.
            continue

        # Forward-fill project within consecutive blank-project rows was not
        # in Excel user cells — project is entered on first row of a block.
        # When subsequent rows leave C blank, Excel fills O via formula.
        # openpyxl data_only may return cached O; try col O (15) as fallback.
        # Actually for blank C rows, project was forward-filled in formula col O.
        rows_out.append({
            'trade': last_trade,
            'project': project,
            'requirement_type': _normalize_req_type(typed['requirement_type']),
            'replacement_name': _cell_str(typed['replacement_name']),
            'replacement_employee_id': _cell_str(typed['replacement_employee_id']),
            'candidate_name': _cell_str(typed['candidate_name']),
            'contact_number': _cell_str(typed['contact_number']),
            'status': _normalize_status(typed['status']),
            'date_joined': _parse_date(typed['date_joined']),
            'remarks': _cell_str(typed['remarks']),
        })

    # Second pass for blank-project continuation rows using forward-fill project
    if not rows_out:
        last_trade = ''
        last_project = ''
        for r in range(start, (ws.max_row or 0) + 1):
            trade_raw = ws.cell(r, _SRC_COL['trade']).value
            project_raw = ws.cell(r, _SRC_COL['project']).value
            a_str = _cell_str(trade_raw).lower()
            if a_str.startswith('how to use'):
                break
            if _cell_str(trade_raw):
                last_trade = _cell_str(trade_raw)
            if _cell_str(project_raw):
                last_project = _cell_str(project_raw)

            typed_vals = [
                ws.cell(r, _SRC_COL[k]).value
                for k in (
                    'requirement_type', 'replacement_name', 'replacement_employee_id',
                    'candidate_name', 'contact_number', 'status', 'date_joined', 'remarks',
                )
            ]
            # Also count project or trade as starting a row
            if not any(_cell_str(v) for v in typed_vals) and not _cell_str(project_raw):
                # Empty row
                continue
            if not last_trade or not last_project:
                continue
            # Only include if requirement type or status present (real vacancy line)
            req = ws.cell(r, _SRC_COL['requirement_type']).value
            status = ws.cell(r, _SRC_COL['status']).value
            if not _cell_str(req) and not _cell_str(status) and not _cell_str(project_raw):
                continue
            rows_out.append({
                'trade': last_trade,
                'project': last_project,
                'requirement_type': _normalize_req_type(req),
                'replacement_name': _cell_str(ws.cell(r, _SRC_COL['replacement_name']).value),
                'replacement_employee_id': _cell_str(
                    ws.cell(r, _SRC_COL['replacement_employee_id']).value
                ),
                'candidate_name': _cell_str(ws.cell(r, _SRC_COL['candidate_name']).value),
                'contact_number': _cell_str(ws.cell(r, _SRC_COL['contact_number']).value),
                'status': _normalize_status(status),
                'date_joined': _parse_date(ws.cell(r, _SRC_COL['date_joined']).value),
                'remarks': _cell_str(ws.cell(r, _SRC_COL['remarks']).value),
            })

    return rows_out


def _parse_automated_with_project_fill(file_storage) -> list[dict[str, Any]]:
    """
    Preferred parser for the automated All Trades sheet:
    forward-fill Trade (A) and Project (C) like Excel merged/continuation rows.
    """
    wb = load_workbook(file_storage, data_only=True)
    if 'All Trades' not in wb.sheetnames:
        return []
    ws = wb['All Trades']
    last_trade = ''
    last_project = ''
    rows_out: list[dict[str, Any]] = []

    for r in range(5, (ws.max_row or 0) + 1):
        trade_raw = _cell_str(ws.cell(r, 1).value)
        project_raw = _cell_str(ws.cell(r, 3).value)
        lower_a = trade_raw.lower()
        if lower_a.startswith('how to use') or lower_a.startswith('sheet guide'):
            break

        if trade_raw:
            last_trade = trade_raw
        if project_raw:
            last_project = project_raw

        req = ws.cell(r, 6).value
        status = ws.cell(r, 11).value
        repl_name = ws.cell(r, 7).value
        repl_id = ws.cell(r, 8).value
        cand = ws.cell(r, 9).value
        contact = ws.cell(r, 10).value
        joined = ws.cell(r, 12).value
        remarks = ws.cell(r, 13).value

        # A vacancy row has at least requirement type or status (Excel always sets both)
        if not _cell_str(req) and not _cell_str(status):
            continue
        if not last_trade or not last_project:
            continue

        rows_out.append({
            'trade': last_trade,
            'project': last_project,
            'requirement_type': _normalize_req_type(req),
            'replacement_name': _cell_str(repl_name),
            'replacement_employee_id': _cell_str(repl_id),
            'candidate_name': _cell_str(cand),
            'contact_number': _cell_str(contact),
            'status': _normalize_status(status),
            'date_joined': _parse_date(joined),
            'remarks': _cell_str(remarks),
        })
    return rows_out


def apply_manpower_import(file_storage, *, replace: bool = False, created_by: Optional[int] = None) -> dict:
    """
    Import vacancies from workbook.
    If replace=True, delete all existing vacancies first (trades/projects kept/updated).
    """
    # Prefer automated layout parser; fall back to generic header mapping
    try:
        file_storage.stream.seek(0)
    except Exception:
        pass
    rows = _parse_automated_with_project_fill(file_storage)
    if not rows:
        try:
            file_storage.stream.seek(0)
        except Exception:
            pass
        rows = parse_all_trades_rows(file_storage)

    if not rows:
        return {
            'created': 0,
            'trades_created': 0,
            'projects_created': 0,
            'deleted': 0,
            'errors': ['No vacancy rows found in workbook'],
        }

    trade_cache: dict[str, ManpowerTrade] = {
        t.name.lower(): t for t in ManpowerTrade.query.all()
    }
    project_cache: dict[str, ManpowerProject] = {
        p.name.lower(): p for p in ManpowerProject.query.all()
    }
    trade_sort = [max((t.sort_order or 0) for t in trade_cache.values()) if trade_cache else 0]
    project_sort = [max((p.sort_order or 0) for p in project_cache.values()) if project_cache else 0]

    deleted = 0
    if replace:
        deleted = ManpowerVacancy.query.delete()
        db.session.flush()

    trades_before = len(trade_cache)
    projects_before = len(project_cache)
    created = 0
    errors: list[str] = []

    # Group sort_order per trade+project
    counters: dict[tuple[int, int], int] = {}

    for i, row in enumerate(rows, start=1):
        try:
            remarks = (row.get('remarks') or '').lower()
            if 'example row' in remarks or '[sample]' in remarks:
                continue
            trade = _get_or_create_trade(row['trade'], trade_cache, trade_sort)
            project = _get_or_create_project(row['project'], project_cache, project_sort)
            key = (trade.id, project.id)
            counters[key] = counters.get(key, 0) + 1
            vac = ManpowerVacancy(
                trade_id=trade.id,
                project_id=project.id,
                requirement_type=row['requirement_type'],
                replacement_name=row['replacement_name'] or None,
                replacement_employee_id=row['replacement_employee_id'] or None,
                candidate_name=row['candidate_name'] or None,
                contact_number=row['contact_number'] or None,
                status=row['status'],
                date_joined=row.get('date_joined'),
                remarks=row['remarks'] or None,
                sort_order=counters[key],
                created_by=created_by,
            )
            db.session.add(vac)
            db.session.flush()
            hid_raw = (row.get('hiring_candidate_id') or '').strip()
            if hid_raw:
                try:
                    hid = int(float(hid_raw)) if '.' in hid_raw else int(hid_raw)
                except (TypeError, ValueError):
                    hid = None
                if hid:
                    from module_hr.staffing_link import assign_candidate_to_vacancy
                    _, _, assign_err = assign_candidate_to_vacancy(hid, vac.id, allow_reassign=True)
                    if assign_err:
                        errors.append(f'Row {i}: link hiring candidate — {assign_err}')
            created += 1
        except Exception as e:
            logger.exception('Manpower import row %s failed', i)
            errors.append(f'Row {i}: {e}')

    db.session.commit()
    return {
        'created': created,
        'deleted': deleted,
        'trades_created': max(0, len(trade_cache) - trades_before),
        'projects_created': max(0, len(project_cache) - projects_before),
        'errors': errors[:20],
    }
