"""
Hiring Document Tracker — Excel template, export, and status import.
"""
from __future__ import annotations

import logging
import re
from io import BytesIO
from typing import Any, Optional

from app.models import (
    HIRING_DOC_LABELS,
    HIRING_DOC_TYPES,
    HIRING_PIPELINE_DEFAULT,
    HIRING_PIPELINE_LABELS,
    HIRING_PIPELINE_STATUSES,
    HIRING_PIPELINE_STEPS,
    HiringCandidate,
    HiringDocument,
    db,
)
from common.datetime_utils import utc_now_naive

logger = logging.getLogger(__name__)

# Short column headers for the wide template (readable + fit dropdowns)
DOC_COLUMN_HEADERS = {
    'passport': 'Passport',
    'emirates_id': 'Emirates ID',
    'photograph': 'Photograph',
    'pcc': 'PCC',
    'education_certificate': 'Education Certificate',
    'offer_letter': 'Offer Letter',
    'insurance': 'Insurance',
    'e_visa': 'E-Visa',
    'contract': 'Contract',
}

PROFILE_HEADERS = (
    'Candidate ID',
    'Full Name',
    'Role',
    'Department',
    'Phone',
    'Email',
    'Replacement Name',
    'Replacement Employee ID',
    'Pipeline Status',
    'Vacancy ID',
)

STATUS_HEADERS = tuple(DOC_COLUMN_HEADERS[dt] for dt in HIRING_DOC_TYPES)
ALL_HEADERS = PROFILE_HEADERS + STATUS_HEADERS + ('Comments',)

DOC_STATUS_TICK = '✓'
DOC_STATUS_CROSS = '✗'
DOC_STATUS_LABELS = (DOC_STATUS_CROSS, DOC_STATUS_TICK)
DOC_STATUS_KEYS = frozenset({'missing', 'uploaded', 'attested', 'verified'})

def _is_template_or_sample_hiring_row(fields: dict) -> bool:
    """Skip blank-template example rows and [SAMPLE] seed people on import."""
    comments = (fields.get('comments') or '').lower()
    ref = str(fields.get('candidate_ref') or '').strip().upper()
    email = (fields.get('email') or '').strip().lower()
    if 'example row' in comments:
        return True
    if '[sample]' in comments:
        return True
    if ref.startswith('HR-CAND-SAMPLE'):
        return True
    if email.endswith('@example.local'):
        return True
    return False


_HIRING_EXAMPLE_ROW = (
    '',
    'Sara Ahmed',
    'Product Designer',
    'Operations',
    '+971500000000',
    'sara.ahmed@example.com',
    'Ali Hassan',
    'EMP-01234',
    HIRING_PIPELINE_LABELS['gathering_documents'],
    '',
    DOC_STATUS_TICK,
    DOC_STATUS_CROSS,
    DOC_STATUS_TICK,
    DOC_STATUS_TICK,
    DOC_STATUS_CROSS,
    DOC_STATUS_CROSS,
    DOC_STATUS_CROSS,
    DOC_STATUS_CROSS,
    DOC_STATUS_CROSS,
    'Example row — replace or delete before importing',
)

# Symbols / glyphs Excel may store for tick / cross
_TICK_GLYPHS = frozenset({
    '✓', '✔', '√', '☑', '✅', '☑️',
})
_CROSS_GLYPHS = frozenset({
    '✗', '✘', '✕', '×', '❌', '☒',
})

FIELD_ALIASES = {
    'candidate_id': ('candidate id', 'id', 'hiring id', 'candidate'),
    'full_name': ('full name', 'name', 'candidate name'),
    'role': ('role', 'position', 'job title', 'role position'),
    'department': ('department', 'dept'),
    'phone': ('phone', 'mobile', 'telephone', 'contact'),
    'email': ('email', 'e mail', 'email address'),
    'replacement_name': ('replacement name', 'replacing', 'replacement'),
    'replacement_employee_id': (
        'replacement employee id', 'replacement emp id', 'replacement id',
        'replacing employee id',
    ),
    'pipeline_status': ('pipeline status', 'pipeline', 'stage', 'hiring stage'),
    'vacancy_id': ('vacancy id', 'manpower vacancy id', 'vacancy', 'project vacancy id'),
    'comments': ('comments', 'comment', 'notes', 'remarks'),
}

DOC_ALIASES = {
    'passport': ('passport', 'passport copy'),
    'emirates_id': ('emirates id', 'emirates id copy', 'eid'),
    'photograph': ('photograph', 'photo', 'photograph white background'),
    'pcc': ('pcc', 'pcc attested', 'police clearance'),
    'education_certificate': ('education certificate', 'education', 'certificate'),
    'offer_letter': ('offer letter', 'offer'),
    'insurance': ('insurance', 'insurance paper'),
    'e_visa': ('e visa', 'evisa', 'visa'),
    'contract': ('contract', 'employment contract'),
}


def normalize_column_name(value) -> str:
    raw = str(value or '').strip().lower()
    cleaned = ''.join(ch if ch.isalnum() else ' ' for ch in raw)
    return ' '.join(cleaned.split())


def _cell_str(value) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and str(value) == 'nan':
        return ''
    s = str(value).strip()
    if s.lower() in ('nan', 'nat', 'none', '-'):
        return ''
    # pandas may give "12.0" for integer IDs
    if re.fullmatch(r'\d+\.0', s):
        return s[:-2]
    return s


def _build_header_map(columns) -> dict:
    """Map normalized header → canonical field or doc_type key."""
    alias_lookup = {}
    for field, aliases in FIELD_ALIASES.items():
        for a in aliases:
            alias_lookup[a] = ('field', field)
    for doc_type, aliases in DOC_ALIASES.items():
        for a in aliases:
            alias_lookup[a] = ('doc', doc_type)
        alias_lookup[normalize_column_name(DOC_COLUMN_HEADERS[doc_type])] = ('doc', doc_type)
        alias_lookup[normalize_column_name(HIRING_DOC_LABELS.get(doc_type, doc_type))] = ('doc', doc_type)

    mapping = {}
    for col in columns:
        n = normalize_column_name(col)
        if not n:
            continue
        if n in alias_lookup:
            mapping[n] = alias_lookup[n]
            continue
        # fuzzy: exact key match
        for field, aliases in FIELD_ALIASES.items():
            if n == field.replace('_', ' ') or n in aliases:
                mapping[n] = ('field', field)
                break
    return mapping


def _normalize_doc_status(raw: str, doc_type: str) -> tuple[Optional[str], Optional[str]]:
    """Return (status_key, error). Blank → (None, None) meaning leave unchanged."""
    raw_s = _cell_str(raw)
    if not raw_s:
        return None, None

    # Tick / cross symbols (normalize_column_name would strip these)
    if raw_s in _TICK_GLYPHS or raw_s.lower() in ('tick', 'check', 'checked', 'yes', 'y'):
        status = 'attested' if doc_type == 'pcc' else 'uploaded'
        return status, None
    if raw_s in _CROSS_GLYPHS or raw_s.lower() in ('cross', 'unchecked', 'no', 'n'):
        return 'missing', None

    s = normalize_column_name(raw_s)
    if not s:
        return None, f'Invalid status "{raw}" for {DOC_COLUMN_HEADERS.get(doc_type, doc_type)}'

    aliases = {
        'missing': 'missing',
        'not started': 'missing',
        'none': 'missing',
        'n a': 'missing',
        'na': 'missing',
        'uploaded': 'uploaded',
        'upload': 'uploaded',
        'received': 'uploaded',
        'done': 'uploaded',
        'complete': 'uploaded',
        'completed': 'uploaded',
        'attested': 'attested',
        'attest': 'attested',
        'verified': 'verified',
        'verify': 'verified',
    }
    status = aliases.get(s)
    if not status:
        return None, f'Invalid status "{raw}" for {DOC_COLUMN_HEADERS.get(doc_type, doc_type)}'

    if status == 'attested' and doc_type != 'pcc':
        # Attested only applies to PCC; treat as uploaded for other docs
        status = 'uploaded'
    return status, None


def _normalize_pipeline(raw: str) -> tuple[Optional[str], Optional[str]]:
    s = _cell_str(raw)
    if not s:
        return None, None
    key = s.strip().lower().replace(' ', '_').replace('-', '_')
    if key in ('onhold',):
        key = 'on_hold'
    if key in (
        'nothired', 'no_hired', 'nohired', 'no_hire', 'nohire',
        'not_hire', 'no_hiried', 'nohiried',
    ):
        key = 'not_hired'
    if key in ('candidateemployee', 'candidate_as_employee', 'file_closed', 'fileclosed', 'closed'):
        key = 'candidate_employee'
    if key in HIRING_PIPELINE_STATUSES:
        return key, None
    label_map = {normalize_column_name(v): k for k, v in HIRING_PIPELINE_LABELS.items()}
    n = normalize_column_name(s)
    if n in label_map:
        return label_map[n], None
    # partial label match
    for k, label in HIRING_PIPELINE_LABELS.items():
        if normalize_column_name(label) == n or n in normalize_column_name(label):
            return k, None
    return None, f'Invalid pipeline status "{raw}"'


def _status_label(status: Optional[str], has_file: bool = False) -> str:
    """Excel display: ✓ = submitted, ✗ = missing."""
    if not status or status == 'missing':
        return DOC_STATUS_CROSS
    if status in ('uploaded', 'attested', 'verified'):
        return DOC_STATUS_TICK
    return DOC_STATUS_CROSS


def _pipeline_label(key: Optional[str]) -> str:
    k = key or HIRING_PIPELINE_DEFAULT
    return HIRING_PIPELINE_LABELS.get(k, k)


def build_hiring_template_bytes(candidates: Optional[list] = None) -> bytes:
    """Blank template (with example row) or prefilled export of candidates."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from openpyxl.worksheet.datavalidation import DataValidation

    from common.kynvera_excel_brand import (
        InstructionSpec,
        apply_column_widths,
        style_data_cell,
        write_header_row,
        write_instructions_sheet,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Candidates'

    write_header_row(ws, ALL_HEADERS, row=1)
    ws.row_dimensions[1].height = 32

    apply_column_widths(ws, {
        'A': 12, 'B': 22, 'C': 18, 'D': 14, 'E': 14, 'F': 24,
        'G': 18, 'H': 18, 'I': 28, 'J': 12, 'K': 12, 'L': 12,
        'M': 12, 'N': 12, 'O': 18, 'P': 12, 'Q': 12, 'R': 12, 'S': 12, 'T': 28,
    })

    if candidates:
        for row_idx, cand in enumerate(candidates, start=2):
            by_type = {d.doc_type: d for d in (cand.documents or [])}
            values = [
                (getattr(cand, 'hr_ref', None) or '').strip() or cand.id,
                cand.full_name or '',
                cand.role or '',
                cand.department or '',
                cand.phone or '',
                cand.email or '',
                cand.replacement_name or '',
                cand.replacement_employee_id or '',
                _pipeline_label(cand.normalized_pipeline_status()),
                getattr(cand, 'assigned_vacancy', None).id if getattr(cand, 'assigned_vacancy', None) else '',
            ]
            for dt in HIRING_DOC_TYPES:
                doc = by_type.get(dt)
                if doc:
                    values.append(_status_label(doc.status, has_file=doc.has_file()))
                else:
                    values.append(_status_label('missing'))
            values.append(cand.comments or '')
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row_idx, col_idx, val)
                style_data_cell(cell)
                if 11 <= col_idx <= 19:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        data_end_row = 1 + len(candidates)
    else:
        for col_idx, val in enumerate(_HIRING_EXAMPLE_ROW, start=1):
            cell = ws.cell(2, col_idx, val)
            style_data_cell(cell, example=True)
            if 11 <= col_idx <= 19:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        data_end_row = 2

    max_dv_row = max(data_end_row + 200, 1000)
    pipe_list = ','.join(HIRING_PIPELINE_LABELS[k] for k in HIRING_PIPELINE_STATUSES)
    status_list = ','.join(DOC_STATUS_LABELS)

    pipe_dv = DataValidation(
        type='list',
        formula1=f'"{pipe_list}"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
    )
    pipe_dv.error = 'Pick a pipeline stage from the list'
    pipe_dv.errorTitle = 'Invalid pipeline'
    pipe_dv.add(f'I2:I{max_dv_row}')
    ws.add_data_validation(pipe_dv)

    status_dv = DataValidation(
        type='list',
        formula1=f'"{status_list}"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
    )
    status_dv.error = f'Use {DOC_STATUS_CROSS} (missing) or {DOC_STATUS_TICK} (submitted)'
    status_dv.errorTitle = 'Invalid status'
    status_dv.add(f'K2:S{max_dv_row}')
    ws.add_data_validation(status_dv)

    pipeline_lines = [HIRING_PIPELINE_LABELS[k] for k in HIRING_PIPELINE_STEPS]
    write_instructions_sheet(wb, InstructionSpec(
        title='Hiring document tracker template',
        module_label='HR / Hiring Docs',
        about=(
            'Import or update hiring candidates and their document checklist (one row per candidate).',
            'Excel does not upload document files — use the candidate detail page for files. This sheet only updates checklist status.',
            'Use the dropdowns for Pipeline Status and each document column.',
        ),
        how_to=(
            'Open the Candidates sheet. Keep the coral header row.',
            'Fill one row per candidate. Full Name is required for new rows.',
            'Use Pipeline Status and document dropdowns (✗ missing, ✓ submitted).',
            'Leave a document cell blank to keep the current status unchanged on update.',
            'Save as .xlsx and click Import Excel on the Hiring dashboard.',
        ),
        columns=(
            ('Candidate ID', 'Optional. System id or any HR reference. Used to match existing rows.'),
            ('Full Name', 'Required for new candidates.'),
            ('Role', 'Job title. Used with Full Name to match when ID/email are missing.'),
            ('Department', 'Optional.'),
            ('Phone', 'Optional.'),
            ('Email', 'Optional. Strong match key when Candidate ID is blank.'),
            ('Replacement Name', 'Optional. Person this hire replaces.'),
            ('Replacement Employee ID', 'Optional.'),
            ('Pipeline Status', 'Pick from the dropdown (stages plus On hold / Not hired).'),
            ('Vacancy ID', 'Optional Manpower vacancy id. Ignored if that vacancy is missing.'),
            ('Passport … Contract', f'{DOC_STATUS_CROSS} = missing (clears file). {DOC_STATUS_TICK} = submitted. Blank = unchanged.'),
            ('Comments', 'Optional. Max 4000 characters.'),
        ),
        example_headers=ALL_HEADERS,
        example_rows=(_HIRING_EXAMPLE_ROW,),
        import_rules=(
            'Match order: DB id → HR reference → Email → Full Name + Role → create new.',
            'If someone was deleted in the UI but still listed in Excel, import recreates them.',
            'Numeric Candidate IDs that are free are restored when possible.',
            'Import shows a confirm dialog: update existing rows and/or add new ones.',
            'Bad rows are reported; other rows still import.',
            f'Dropdowns apply through row {max_dv_row}. Re-download the template if an older file is missing them.',
        ),
        extra_sections=(
            ('Pipeline stages (in order)', (
                *pipeline_lines,
                f'{HIRING_PIPELINE_LABELS["on_hold"]} — pauses the whole process (not a stage).',
                f'{HIRING_PIPELINE_LABELS["not_hired"]} — candidate was not hired (not a stage).',
            )),
        ),
    ))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def read_hiring_dataframe(file_storage):
    """Return pandas DataFrame from uploaded .xlsx / .xls / .csv."""
    import pandas as pd

    filename = (file_storage.filename or '').lower()
    if not filename.endswith(('.xlsx', '.xls', '.csv')):
        raise ValueError('Upload .xlsx, .xls, or .csv')

    if filename.endswith('.csv'):
        return pd.read_csv(file_storage)

    from common.kynvera_excel_brand import read_import_dataframe

    return read_import_dataframe(file_storage, preferred_sheets=('Candidates',))


def parse_hiring_workbook(file_storage) -> list[dict[str, Any]]:
    """Parse spreadsheet into normalized row dicts."""
    import pandas as pd

    df = read_hiring_dataframe(file_storage)
    if df is None or df.empty:
        raise ValueError('Spreadsheet has no data rows')

    # Drop fully empty rows
    df = df.dropna(how='all')
    if df.empty:
        raise ValueError('Spreadsheet has no data rows')

    col_map = {}
    header_map = _build_header_map(df.columns)
    for col in df.columns:
        n = normalize_column_name(col)
        if n in header_map:
            col_map[col] = header_map[n]

    if not any(v == ('field', 'full_name') for v in col_map.values()):
        raise ValueError('Missing required column: Full Name')

    rows = []
    for idx, series in df.iterrows():
        excel_row = int(idx) + 2  # header is row 1; pandas 0-based
        fields: dict[str, Any] = {}
        docs: dict[str, Optional[str]] = {}
        row_errors = []

        for col, kind_key in col_map.items():
            kind, key = kind_key
            raw = _cell_str(series.get(col))
            if kind == 'field':
                if key == 'pipeline_status':
                    pipe, err = _normalize_pipeline(raw)
                    if err:
                        row_errors.append(err)
                    elif pipe is not None:
                        fields['pipeline_status'] = pipe
                elif key == 'candidate_id':
                    if raw:
                        # Accept numeric DB id and/or alphanumeric HR reference.
                        fields['candidate_ref'] = raw.strip()
                        try:
                            fields['candidate_id'] = int(float(raw))
                        except (TypeError, ValueError):
                            # e.g. "SIM-25" — reference only, not a DB id
                            pass
                elif raw:
                    fields[key] = raw
            else:
                status, err = _normalize_doc_status(raw, key)
                if err:
                    row_errors.append(err)
                elif status is not None:
                    docs[key] = status

        # Skip blank example-ish empty name rows
        name = (fields.get('full_name') or '').strip()
        if not name and not fields.get('candidate_id') and not fields.get('candidate_ref') and not fields.get('email'):
            continue

        if _is_template_or_sample_hiring_row(fields):
            continue

        rows.append({
            'excel_row': excel_row,
            'fields': fields,
            'docs': docs,
            'errors': row_errors,
        })

    if not rows:
        raise ValueError('No candidate rows found')
    return rows


def _norm_person_key(name: str, role: str) -> str:
    n = ' '.join((name or '').strip().lower().split())
    r = ' '.join((role or '').strip().lower().split())
    if not n or not r:
        return ''
    return n + '|' + r


def _build_match_indexes():
    """id / hr_ref / email / name|role → candidate."""
    email_index: dict[str, HiringCandidate] = {}
    name_role_index: dict[str, HiringCandidate] = {}
    ref_index: dict[str, HiringCandidate] = {}
    for c in HiringCandidate.query.all():
        em = (c.email or '').strip().lower()
        if em and em not in email_index:
            email_index[em] = c
        key = _norm_person_key(c.full_name or '', c.role or '')
        if key and key not in name_role_index:
            name_role_index[key] = c
        ref = (getattr(c, 'hr_ref', None) or '').strip().lower()
        if ref and ref not in ref_index:
            ref_index[ref] = c
        # Also allow matching exported numeric id stored as text ref
        sid = str(c.id)
        if sid not in ref_index:
            ref_index[sid] = c
    return email_index, name_role_index, ref_index


def resolve_hiring_row_action(
    fields: dict[str, Any],
    email_index: dict[str, HiringCandidate],
    name_role_index: dict[str, HiringCandidate],
    ref_index: Optional[dict[str, HiringCandidate]] = None,
) -> tuple[str, Optional[HiringCandidate], Optional[str], Optional[str]]:
    """
    Decide create vs update for one Excel row.

    Returns (action, candidate, note, match_via)
    match_via: 'id' | 'ref' | 'email' | 'name_role' | None

    Match order:
      1) Candidate ID as DB primary key (if numeric and still present)
      2) Candidate ID / HR reference string (hr_ref)
      3) Email
      4) Full name + role
    """
    cid = fields.get('candidate_id')
    ref = (fields.get('candidate_ref') or '').strip()
    email = (fields.get('email') or '').strip()
    name = (fields.get('full_name') or '').strip()
    role = (fields.get('role') or '').strip()
    ref_index = ref_index or {}

    candidate = None
    note = None
    match_via = None

    if cid is not None:
        candidate = db.session.get(HiringCandidate, cid)
        if candidate:
            match_via = 'id'

    if not candidate and ref:
        candidate = ref_index.get(ref.lower())
        if candidate:
            match_via = 'ref'

    if not candidate and email:
        candidate = email_index.get(email.lower())
        if candidate:
            match_via = 'email'

    if not candidate and name and role:
        candidate = name_role_index.get(_norm_person_key(name, role))
        if candidate:
            match_via = 'name_role'

    if candidate:
        return 'update', candidate, note, match_via

    if cid is not None or ref:
        note = (
            f'Candidate ID "{ref or cid}" not in system — will create as new'
        )

    if not name:
        return 'error', None, 'Full Name is required to create a candidate', None
    if not role:
        return 'error', None, 'Role is required to create a candidate', None
    return 'create', None, note, None


def _is_id_name_conflict(candidate: HiringCandidate, fields: dict[str, Any], match_via: Optional[str]) -> bool:
    """True when Excel reuses an existing ID/ref but with a different person name."""
    if match_via not in ('id', 'ref') or not candidate:
        return False
    old_name = (candidate.full_name or '').strip().lower()
    new_name = (fields.get('full_name') or '').strip().lower()
    return bool(old_name and new_name and old_name != new_name)


def preview_hiring_import(rows: list[dict]) -> dict:
    """Classify rows without writing — used for the confirm dialog."""
    email_index, name_role_index, ref_index = _build_match_indexes()
    will_create = 0
    will_update = 0
    will_skip = 0
    will_rename = 0
    errors: list[dict] = []
    create_names: list[str] = []
    update_names: list[str] = []
    rename_pairs: list[dict] = []
    pending_create_keys: set[str] = set()
    pending_emails: set[str] = set()
    pending_refs: set[str] = set()
    matched_ids: set[int] = set()

    for row in rows:
        excel_row = row.get('excel_row')
        fields = dict(row.get('fields') or {})
        row_errors = list(row.get('errors') or [])
        if row_errors:
            will_skip += 1
            errors.append({'row': excel_row, 'error': '; '.join(row_errors)})
            continue
        comments_val = (fields.get('comments') or '').strip()
        if comments_val and len(comments_val) > 4000:
            will_skip += 1
            errors.append({'row': excel_row, 'error': 'Comments must be 4000 characters or fewer'})
            continue

        action, candidate, note, match_via = resolve_hiring_row_action(
            fields, email_index, name_role_index, ref_index,
        )
        name = (fields.get('full_name') or '').strip() or (candidate.full_name if candidate else 'Candidate')
        email = (fields.get('email') or '').strip().lower()
        key = _norm_person_key(fields.get('full_name') or '', fields.get('role') or '')
        ref = (fields.get('candidate_ref') or '').strip().lower()
        shared_id = fields.get('candidate_id')
        if shared_id is None and ref:
            shared_id = ref

        if action == 'create' and (
            (key and key in pending_create_keys)
            or (email and email in pending_emails)
            or (ref and ref in pending_refs)
        ):
            action = 'update'
            candidate = None
            match_via = None

        if action == 'error':
            will_skip += 1
            errors.append({'row': excel_row, 'error': note or 'Invalid row'})
            continue

        if action == 'update' and candidate and _is_id_name_conflict(candidate, fields, match_via):
            will_rename += 1
            if len(rename_pairs) < 10:
                rename_pairs.append({
                    'id': candidate.id,
                    'shared_id': str(shared_id if shared_id is not None else candidate.id),
                    'from': (candidate.full_name or '').strip(),
                    'to': (fields.get('full_name') or '').strip(),
                    'role': (fields.get('role') or candidate.role or '').strip(),
                })
            # Count as conflict rows; also still "matched" for orphan calc under replace mode
            matched_ids.add(int(candidate.id))
            continue

        if action == 'update':
            will_update += 1
            if candidate and getattr(candidate, 'id', None) is not None:
                matched_ids.add(int(candidate.id))
            if len(update_names) < 5:
                update_names.append(name)
        else:
            will_create += 1
            if len(create_names) < 5:
                create_names.append(name)
            if key:
                pending_create_keys.add(key)
            if email:
                pending_emails.add(email)
            if ref:
                pending_refs.add(ref)

    orphan_candidates = []
    for c in HiringCandidate.query.order_by(HiringCandidate.full_name.asc()).all():
        if c.id in matched_ids:
            continue
        orphan_candidates.append({
            'id': c.id,
            'hr_ref': (getattr(c, 'hr_ref', None) or '').strip() or str(c.id),
            'full_name': c.full_name or '',
            'role': c.role or '',
        })

    return {
        'will_create': will_create,
        'will_update': will_update,
        'will_rename': will_rename,
        'will_skip': will_skip,
        'errors': errors[:20],
        'create_names': create_names,
        'update_names': update_names,
        'rename_pairs': rename_pairs,
        'total_rows': len(rows),
        'orphan_count': len(orphan_candidates),
        'orphan_candidates': orphan_candidates[:40],
        'has_id_conflicts': will_rename > 0,
        'needs_confirm': (
            will_update > 0 or will_create > 0 or will_rename > 0 or len(orphan_candidates) > 0
        ),
    }


def apply_hiring_import(
    rows: list[dict],
    user,
    seed_documents_fn,
    clear_document_file_fn,
    *,
    update_existing: bool = True,
    orphan_action: str = 'keep',
    id_conflict_action: str = 'keep_both',
) -> dict:
    """
    Upsert candidates and apply document statuses.

    Matching: Candidate ID (if present in DB) → hr_ref → email → full name + role.
    Missing/stale IDs create (or match by email/name) so an Excel export that
    still lists a deleted candidate is re-added on import.

    If update_existing is False, matched rows are left unchanged (new rows still created).

    orphan_action:
      - keep: leave app candidates that were not in the Excel alone (merge)
      - delete: remove app candidates that no Excel row matched

    id_conflict_action (same Candidate ID / HR ref, different full name):
      - keep_both: keep the existing person; create Excel person with a new auto ID
      - replace: overwrite the existing row with Excel data
    """
    created = 0
    updated = 0
    skipped = 0
    unchanged = 0
    deleted = 0
    errors: list[dict] = []
    warnings: list[dict] = []
    matched_ids: set[int] = set()

    email_index, name_role_index, ref_index = _build_match_indexes()
    orphan_mode = (orphan_action or 'keep').strip().lower()
    if orphan_mode not in ('keep', 'delete'):
        orphan_mode = 'keep'
    conflict_mode = (id_conflict_action or 'keep_both').strip().lower()
    if conflict_mode not in ('keep_both', 'replace'):
        conflict_mode = 'keep_both'

    for row in rows:
        excel_row = row.get('excel_row')
        fields = dict(row.get('fields') or {})
        docs = dict(row.get('docs') or {})
        row_errors = list(row.get('errors') or [])

        if row_errors:
            errors.append({'row': excel_row, 'error': '; '.join(row_errors)})
            skipped += 1
            continue

        comments_val = (fields.get('comments') or '').strip()
        if comments_val and len(comments_val) > 4000:
            errors.append({'row': excel_row, 'error': 'Comments must be 4000 characters or fewer'})
            skipped += 1
            continue

        try:
            with db.session.begin_nested():
                action, candidate, note, match_via = resolve_hiring_row_action(
                    fields, email_index, name_role_index, ref_index,
                )
                if note and action in ('create', 'update'):
                    warnings.append({'row': excel_row, 'error': note})

                if action == 'error':
                    raise ValueError(note or 'Invalid row')

                is_create = False
                force_new_id = False
                ref_raw = (fields.get('candidate_ref') or '').strip()
                cid = fields.get('candidate_id')

                if (
                    action == 'update'
                    and candidate
                    and _is_id_name_conflict(candidate, fields, match_via)
                ):
                    # Always count the existing ID as matched so orphan-delete
                    # does not remove the person we are keeping (keep_both).
                    matched_ids.add(int(candidate.id))
                    if conflict_mode == 'keep_both':
                        warnings.append({
                            'row': excel_row,
                            'error': (
                                f'Same Candidate ID as '
                                f'"{(candidate.full_name or "").strip()}" — '
                                f'creating Excel person with a new ID'
                            ),
                        })
                        action = 'create'
                        candidate = None
                        force_new_id = True
                    # else: replace — fall through and overwrite existing row

                if action == 'update':
                    if candidate and getattr(candidate, 'id', None) is not None:
                        matched_ids.add(int(candidate.id))
                    if not update_existing:
                        unchanged += 1
                        continue
                    if fields.get('full_name'):
                        candidate.full_name = fields['full_name'].strip()
                    role_val = (fields.get('role') or '').strip()
                    if role_val:
                        candidate.role = role_val
                    for key in (
                        'department', 'phone', 'email', 'replacement_name',
                        'replacement_employee_id', 'comments',
                    ):
                        if key in fields:
                            val = (fields.get(key) or '').strip()
                            setattr(candidate, key, val or None)
                    if fields.get('pipeline_status'):
                        candidate.pipeline_status = fields['pipeline_status']
                    # Keep / set HR reference from Excel when provided
                    if ref_raw and not (getattr(candidate, 'hr_ref', None) or '').strip():
                        candidate.hr_ref = ref_raw
                else:
                    name = (fields.get('full_name') or '').strip()
                    role = (fields.get('role') or '').strip()
                    email = (fields.get('email') or '').strip()
                    # On ID/name conflict keep_both: never reuse colliding PK or hr_ref
                    if force_new_id:
                        hr_ref = None
                        restore_cid = None
                    else:
                        hr_ref = ref_raw or (str(cid) if cid is not None else None)
                        # Avoid unique conflicts if another row already claimed this ref
                        if hr_ref and hr_ref.lower() in ref_index:
                            hr_ref = None
                        restore_cid = cid
                    candidate = HiringCandidate(
                        full_name=name,
                        role=role,
                        department=(fields.get('department') or '').strip() or None,
                        phone=(fields.get('phone') or '').strip() or None,
                        email=email or None,
                        replacement_name=(fields.get('replacement_name') or '').strip() or None,
                        replacement_employee_id=(fields.get('replacement_employee_id') or '').strip() or None,
                        comments=comments_val or None,
                        pipeline_status=fields.get('pipeline_status') or HIRING_PIPELINE_DEFAULT,
                        created_by=user.id if user else None,
                        hr_ref=hr_ref,
                    )
                    # Restore numeric Candidate ID when that PK is free (e.g. deleted then re-imported)
                    if restore_cid is not None and db.session.get(HiringCandidate, restore_cid) is None:
                        candidate.id = restore_cid
                    db.session.add(candidate)
                    db.session.flush()
                    is_create = True
                    matched_ids.add(int(candidate.id))
                    if email:
                        email_index[email.lower()] = candidate
                    key = _norm_person_key(name, role)
                    if key:
                        name_role_index[key] = candidate
                    if hr_ref:
                        ref_index[hr_ref.lower()] = candidate
                    ref_index[str(candidate.id)] = candidate

                # Seed once after create/update. Relationship-backed seed keeps
                # candidate.documents current so we never re-insert the same slot.
                seed_documents_fn(candidate)
                db.session.flush()

                by_type = {d.doc_type: d for d in (candidate.documents or [])}
                for doc_type, status in docs.items():
                    doc = by_type.get(doc_type)
                    if not doc:
                        doc = HiringDocument(
                            doc_type=doc_type,
                            status='missing',
                        )
                        candidate.documents.append(doc)
                        db.session.flush()
                        by_type[doc_type] = doc

                    if status == 'missing':
                        clear_document_file_fn(doc)
                    else:
                        doc.status = status

                candidate.updated_at = utc_now_naive()
                if is_create:
                    created += 1
                else:
                    updated += 1

                # Optional Vacancy ID → assign to manpower vacancy (soft-fail so a
                # stale vacancy from another env does not block candidate import).
                vac_raw = (fields.get('vacancy_id') or '').strip()
                if vac_raw:
                    try:
                        vac_id = int(float(vac_raw)) if '.' in vac_raw else int(vac_raw)
                    except (TypeError, ValueError):
                        vac_id = None
                        warnings.append({
                            'row': excel_row,
                            'error': f'Invalid Vacancy ID "{vac_raw}" — candidate imported without vacancy link',
                        })
                    if vac_id:
                        from module_hr.staffing_link import assign_candidate_to_vacancy
                        _, _, assign_err = assign_candidate_to_vacancy(
                            candidate.id, vac_id, allow_reassign=True,
                        )
                        if assign_err:
                            warnings.append({
                                'row': excel_row,
                                'error': f'{assign_err} — candidate imported without vacancy link',
                            })
        except Exception as e:
            logger.warning('Hiring Excel row %s skipped: %s', excel_row, e)
            errors.append({'row': excel_row, 'error': str(e)})
            skipped += 1

    if orphan_mode == 'delete':
        # When matched_ids is empty and Excel had no valid rows, refuse mass-delete
        if not matched_ids and (created + updated + unchanged) == 0:
            warnings.append({
                'row': None,
                'error': 'No Excel rows matched — skipped deleting app-only candidates for safety',
            })
        else:
            q = HiringCandidate.query
            if matched_ids:
                q = q.filter(~HiringCandidate.id.in_(list(matched_ids)))
            orphans = q.all()
            for orphan in orphans:
                try:
                    vac = getattr(orphan, 'assigned_vacancy', None)
                    if vac is not None:
                        vac.hiring_candidate_id = None
                    db.session.delete(orphan)
                    deleted += 1
                except Exception as e:
                    logger.warning('Could not delete orphan candidate %s: %s', orphan.id, e)
                    warnings.append({
                        'row': None,
                        'error': f'Could not remove {orphan.full_name or orphan.id}: {e}',
                    })

    try:
        db.session.commit()
    except Exception as e:
        logger.exception('Hiring Excel import commit failed')
        db.session.rollback()
        raise ValueError(f'Import failed to save: {e}') from e

    return {
        'created': created,
        'updated': updated,
        'unchanged': unchanged,
        'deleted': deleted,
        'skipped': skipped,
        'errors': errors,
        'warnings': warnings,
        'processed': created + updated,
    }
