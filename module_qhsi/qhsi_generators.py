"""PDF and Excel reports for QHSA unified inspections — Kynvera branded."""
import logging
import os
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.pdfgen.canvas import Canvas

from common.utils import get_image_for_pdf
from common import kynvera_pdf_brand as brand

logger = logging.getLogger(__name__)

_DEPT_LABEL = {'hvac': 'HVAC & MEP', 'civil': 'Civil Works', 'cleaning': 'Cleaning Services'}
PAGE_W, PAGE_H = A4
MARGIN = 1.6 * cm


class _QhsiCanvas(Canvas):
    def __init__(self, *args, **kwargs):
        self._report_title = kwargs.pop('report_title', 'QHSA Site Inspection')
        super().__init__(*args, **kwargs)
        self._saved = []

    def showPage(self):
        self._saved.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        n = len(self._saved)
        for s in self._saved:
            self.__dict__.update(s)
            brand.draw_page_chrome(
                self,
                self._pageNumber,
                n,
                report_title=self._report_title,
                left_margin=MARGIN,
                right_margin=MARGIN,
                footer_left=brand.FOOTER_CONFIDENTIAL,
            )
            Canvas.showPage(self)
        Canvas.save(self)


def _items_from_record(record):
    return record.get('items') or []


def create_excel_report(submission_record, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'QHSA Inspection'
    header_fill = PatternFill('solid', fgColor='FF8E68')
    header_font = Font(color='191B23', bold=True)
    ws.append(['QHSA Site Inspection Report'])
    ws['A1'].font = Font(bold=True, size=14, color='191B23')
    ws.append([])
    meta = [
        ('Project', submission_record.get('project_name', '')),
        ('Visit Date', submission_record.get('visit_date', '')),
        ('Department', _DEPT_LABEL.get(submission_record.get('department'), submission_record.get('department'))),
        ('Inspector', submission_record.get('inspector_name', '')),
        ('Location', submission_record.get('location', '')),
        ('Summary', submission_record.get('summary', '')),
    ]
    for label, val in meta:
        ws.append([label, val])
    ws.append([])
    ws.append(['#', 'Area / System', 'Equipment', 'Severity', 'Description', 'Photos'])
    for c in range(1, 7):
        cell = ws.cell(row=ws.max_row, column=c)
        cell.fill = header_fill
        cell.font = header_font
    for i, item in enumerate(_items_from_record(submission_record), 1):
        area = item.get('area') or item.get('system') or item.get('trade') or ''
        equip = item.get('equipment') or item.get('zone') or ''
        photos = len(item.get('photos') or [])
        ws.append([
            i,
            area,
            equip,
            item.get('severity', ''),
            (item.get('description') or '')[:500],
            photos,
        ])
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb.save(output_path)
    return output_path


def create_pdf_report(submission_record, output_path):
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        topMargin=2.0 * cm,
        bottomMargin=1.8 * cm,
        leftMargin=MARGIN,
        rightMargin=MARGIN,
        title='QHSA Site Inspection',
        author=brand.PDF_AUTHOR,
    )
    styles = getSampleStyleSheet()
    body_style = ParagraphStyle(
        'QhsiBody',
        parent=styles['Normal'],
        fontSize=9,
        textColor=brand.TEXT_DARK,
        leading=12,
    )
    item_style = ParagraphStyle(
        'QhsiItem',
        parent=styles['Heading3'],
        fontSize=10,
        textColor=brand.TEXT_DARK,
        fontName='Helvetica-Bold',
        spaceBefore=8,
        spaceAfter=4,
    )

    body = []
    content_w = PAGE_W - 2 * MARGIN
    body.append(brand.story_header_block('QHSA Site Inspection', brand.COMPANY_NAME, content_w))
    body.append(Spacer(1, 10))

    meta_data = [
        ['Project', submission_record.get('project_name', '-')],
        ['Date', submission_record.get('visit_date', '-')],
        ['Department', _DEPT_LABEL.get(submission_record.get('department'), '-')],
        ['Inspector', submission_record.get('inspector_name', '-')],
        ['Location', submission_record.get('location', '-')],
    ]
    t = Table(meta_data, colWidths=[4 * cm, 12 * cm])
    t.setStyle(TableStyle(brand.meta_table_style()))
    body.append(t)
    body.append(Spacer(1, 14))

    if submission_record.get('summary'):
        body.append(Paragraph(f"<b>Summary</b>", item_style))
        body.append(Paragraph(submission_record.get('summary'), body_style))
        body.append(Spacer(1, 10))

    items = _items_from_record(submission_record)
    if items:
        header_row = [
            Paragraph('<b>#</b>', body_style),
            Paragraph('<b>Area / System</b>', body_style),
            Paragraph('<b>Equipment</b>', body_style),
            Paragraph('<b>Severity</b>', body_style),
        ]
        rows = [header_row]
        for i, item in enumerate(items, 1):
            label = item.get('area') or item.get('system') or item.get('trade') or 'Item'
            equip = item.get('equipment') or item.get('zone') or '—'
            sev = item.get('severity', '—') or '—'
            rows.append([
                Paragraph(str(i), body_style),
                Paragraph(str(label), body_style),
                Paragraph(str(equip), body_style),
                Paragraph(str(sev), body_style),
            ])
        items_tbl = Table(rows, colWidths=[1.2 * cm, 6 * cm, 5 * cm, 3.8 * cm])
        items_tbl.setStyle(TableStyle(brand.data_header_table_style()))
        body.append(items_tbl)
        body.append(Spacer(1, 12))

    for i, item in enumerate(items, 1):
        label = item.get('area') or item.get('system') or item.get('trade') or 'Item'
        equip = item.get('equipment') or item.get('zone') or ''
        body.append(Paragraph(f"{i}. {label} — {equip}", item_style))
        if item.get('description'):
            body.append(Paragraph(item['description'], body_style))
        for ph in (item.get('photos') or [])[:4]:
            url = ph.get('url') if isinstance(ph, dict) else ph
            if not url:
                continue
            try:
                img_data = get_image_for_pdf(url)
                if img_data:
                    body.append(Image(BytesIO(img_data), width=8 * cm, height=6 * cm))
            except Exception as e:
                logger.debug('PDF image skip: %s', e)
        body.append(Spacer(1, 8))

    doc.build(
        body,
        canvasmaker=lambda *a, **kw: _QhsiCanvas(*a, report_title='QHSA Site Inspection', **kw),
    )
    return output_path
