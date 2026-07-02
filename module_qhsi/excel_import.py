"""Parse staff compliance Excel files for QHSE dashboard metrics."""
from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO, StringIO

from module_qhsi.catalog import staff_kit_types

META_ALIASES = {
    'employee_name': (
        'employee name', 'employee', 'name', 'staff name', 'worker name', 'full name',
    ),
    'employee_id': ('employee id', 'emp id', 'staff id', 'id number', 'badge no', 'badge number'),
    'project_name': ('project', 'project name', 'site', 'site name', 'location', 'contract'),
    'record_date': ('record date', 'date', 'inspection date', 'check date', 'visit date'),
    'department': ('department', 'dept', 'division', 'team'),
    'supervisor_name': ('supervisor', 'line manager', 'manager', 'supervisor name'),
    'notes': ('notes', 'remarks', 'overall comments'),
    'line_comments': ('comments', 'item comments', 'line comments'),
    'item': ('item', 'kit item', 'ppe', 'ppe item', 'equipment', 'kit type', 'uniform item'),
    'condition': ('condition', 'status', 'compliance', 'result', 'compliance status', 'ppe status'),
}

CONDITION_OK = frozenset({
    'ok', 'okay', 'compliant', 'good', 'pass', 'passed', 'yes', 'y', '1', 'complete', 'completed',
})
CONDITION_ISSUE = frozenset({
    'issue', 'issues', 'defect', 'damaged', 'repair', 'replacement', 'replace', 'warn', 'warning',
    'non compliant', 'non-compliant', 'fail', 'failed',
})
CONDITION_MISSING = frozenset({
    'missing', 'not issued', 'not available', 'na', 'n/a', 'none', 'absent', 'no kit',
})


def normalize_column_name(value):
    raw = str(value or '').strip().lower()
    cleaned = ''.join(ch if ch.isalnum() else ' ' for ch in raw)
    return ' '.join(cleaned.split())


def _normalize_condition(raw):
    s = normalize_column_name(raw)
    if not s:
        return None
    if s in CONDITION_OK or 'compliant' in s or s.startswith('ok'):
        return 'ok'
    if s in CONDITION_MISSING or 'missing' in s:
        return 'missing'
    if s in CONDITION_ISSUE or 'issue' in s or 'replace' in s:
        return 'issue'
    return None


def _parse_date_cell(value):
    if value is None or (isinstance(value, float) and str(value) == 'nan'):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    if not s or s.lower() in ('nan', 'nat', '-'):
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%d.%m.%Y'):
        try:
            return datetime.strptime(s[:10], fmt).date().isoformat()
        except ValueError:
            continue
    try:
        parsed = datetime.fromisoformat(s.replace('Z', '+00:00')[:10])
        return parsed.date().isoformat()
    except ValueError:
        return None


def _cell_str(value):
    if value is None:
        return ''
    if isinstance(value, float) and str(value) == 'nan':
        return ''
    return str(value).strip()


def _build_kit_label_map():
    out = {}
    for kit in staff_kit_types():
        out[normalize_column_name(kit['label'])] = kit
        out[normalize_column_name(kit['id'].replace('_', ' '))] = kit
    return out


def _match_kit_column(col_name, kit_map):
    n = normalize_column_name(col_name)
    if n in kit_map:
        return kit_map[n]
    for key, kit in kit_map.items():
        if key in n or n in key:
            return kit
    hints = (
        ('uniform shirt', 'uniform_shirt'),
        ('shirt', 'uniform_shirt'),
        ('trouser', 'uniform_trouser'),
        ('pant', 'uniform_trouser'),
        ('coverall', 'uniform_coverall'),
        ('safety shoe', 'safety_shoes'),
        ('boot', 'safety_shoes'),
        ('helmet', 'helmet'),
        ('hard hat', 'helmet'),
        ('hi vis', 'hi_vis_vest'),
        ('vest', 'hi_vis_vest'),
        ('glove', 'gloves'),
        ('goggle', 'goggles'),
        ('ear', 'ear_protection'),
        ('badge', 'id_badge'),
    )
    for hint, kid in hints:
        if hint in n:
            for kit in staff_kit_types():
                if kit['id'] == kid:
                    return kit
    return None


def read_excel_dataframe(file_storage):
    """Return pandas DataFrame from uploaded file."""
    import pandas as pd

    filename = (file_storage.filename or '').lower()
    if not filename.endswith(('.xlsx', '.xls', '.csv')):
        raise ValueError('Upload .xlsx, .xls, or .csv')

    if filename.endswith('.csv'):
        return pd.read_csv(file_storage)

    if filename.endswith('.xlsx'):
        return pd.read_excel(file_storage)

    try:
        file_storage.stream.seek(0)
        return pd.read_excel(file_storage, engine='xlrd')
    except Exception:
        file_storage.stream.seek(0)
        html_text = file_storage.stream.read().decode('utf-8', errors='ignore')
        tables = pd.read_html(StringIO(html_text))
        if not tables:
            raise ValueError('Could not read any table from the file')
        return max(tables, key=lambda t: t.shape[0])


def _map_meta_columns(df):
    normalized = {normalize_column_name(c): c for c in df.columns}
    canonical = {}
    for field, aliases in META_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                canonical[field] = normalized[alias]
                break
    return canonical


def parse_staff_compliance_excel(file_storage):
    """
    Parse Excel into row dicts. Supports:
    - Long format: one row per kit line (employee + item + condition columns)
    - Wide format: one row per employee, kit types as columns with condition in cell
    """
    df = read_excel_dataframe(file_storage)
    if df is None or df.empty:
        raise ValueError('File is empty')

    if hasattr(df.columns, 'levels'):
        flat_cols = []
        for col in df.columns:
            if isinstance(col, tuple):
                flat_cols.append(' '.join(
                    str(p).strip() for p in col
                    if str(p).strip() and str(p).strip().lower() != 'nan'
                ))
            else:
                flat_cols.append(str(col))
        df.columns = flat_cols

    meta = _map_meta_columns(df)
    kit_map = _build_kit_label_map()
    meta_cols = set(meta.values())
    wide_kit_cols = []
    for col in df.columns:
        if col in meta_cols:
            continue
        if _match_kit_column(col, kit_map):
            wide_kit_cols.append(col)

    rows = []
    errors = []
    skipped = 0
    mode = 'wide' if wide_kit_cols and 'item' not in meta else 'long'

    if mode == 'long' and 'employee_name' not in meta:
        raise ValueError(
            'Could not find an Employee Name column. '
            'Use the template or headers: Employee Name, Project, Item, Condition.'
        )

    if mode == 'wide' and 'employee_name' not in meta and not wide_kit_cols:
        raise ValueError('No recognizable employee or kit columns found.')

    for idx, row in df.iterrows():
        line_no = int(idx) + 2

        def meta_val(key, default=''):
            col = meta.get(key)
            if not col:
                return default
            return _cell_str(row.get(col, default))

        employee_name = meta_val('employee_name')
        if not employee_name:
            skipped += 1
            continue

        base = {
            'employee_name': employee_name,
            'employee_id': meta_val('employee_id'),
            'project_name': meta_val('project_name') or '—',
            'record_date': _parse_date_cell(row.get(meta['record_date'])) if meta.get('record_date') else None,
            'department': meta_val('department'),
            'supervisor_name': meta_val('supervisor_name'),
            'notes': meta_val('notes'),
        }

        if mode == 'wide':
            if not wide_kit_cols:
                errors.append(f'Row {line_no}: no kit columns')
                continue
            added = 0
            for col in wide_kit_cols:
                raw = row.get(col)
                if raw is None or (isinstance(raw, float) and str(raw) == 'nan'):
                    continue
                cond = _normalize_condition(raw)
                if not cond:
                    continue
                kit = _match_kit_column(col, kit_map)
                rows.append({
                    **base,
                    'item_type': kit['id'] if kit else normalize_column_name(col).replace(' ', '_'),
                    'item_label': kit['label'] if kit else str(col).strip(),
                    'condition': cond,
                })
                added += 1
            if not added:
                skipped += 1
        else:
            item_col = meta.get('item')
            cond_col = meta.get('condition')
            if not item_col or not cond_col:
                errors.append(f'Row {line_no}: missing Item or Condition column')
                continue
            item_raw = _cell_str(row.get(item_col))
            cond = _normalize_condition(row.get(cond_col))
            if not item_raw or not cond:
                skipped += 1
                continue
            kit = None
            n_item = normalize_column_name(item_raw)
            for k, v in kit_map.items():
                if k == n_item or n_item in k:
                    kit = v
                    break
            line_note = meta_val('line_comments')
            notes = base['notes']
            if line_note:
                notes = (notes + ' — ' + line_note).strip(' —') if notes else line_note
            rows.append({
                **base,
                'notes': notes,
                'item_type': kit['id'] if kit else n_item.replace(' ', '_')[:60],
                'item_label': kit['label'] if kit else item_raw,
                'condition': cond,
            })

    if not rows:
        raise ValueError(
            'No valid compliance rows found. '
            f'Skipped {skipped} empty rows.'
            + (f' Errors: {"; ".join(errors[:5])}' if errors else '')
        )

    return {
        'rows': rows,
        'mode': mode,
        'skipped': skipped,
        'errors': errors[:20],
        'row_count': len(rows),
    }


def stats_from_rows(rows):
    """Aggregate metrics for dashboard from parsed/imported rows."""
    employees = set()
    ok = issue = missing = 0
    projects = set()
    for r in rows:
        employees.add((
            r.get('employee_name', ''),
            r.get('project_name', ''),
            r.get('record_date') or '',
        ))
        projects.add(r.get('project_name') or '')
        c = r.get('condition') or 'ok'
        if c == 'ok':
            ok += 1
        elif c == 'issue':
            issue += 1
        else:
            missing += 1
    total_lines = len(rows)
    non_compliant = issue + missing
    return {
        'employees': len(employees),
        'projects': len([p for p in projects if p and p != '—']),
        'kit_lines': total_lines,
        'compliant': ok,
        'issues': issue,
        'missing': missing,
        'non_compliant': non_compliant,
        'compliance_rate': round(100.0 * ok / total_lines, 1) if total_lines else 0.0,
    }


def build_import_template_bytes():
    """Downloadable .xlsx template (long format)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Staff Compliance'
    headers = [
        'Employee Name', 'Employee ID', 'Project', 'Record Date',
        'Department', 'Supervisor', 'Item', 'Condition', 'Comments',
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='125435')
    ws.append([
        'Ahmed Hassan', 'EMP-1001', 'Site A', date.today().isoformat(),
        'Operations', 'John Smith', 'Uniform — Shirt / Top', 'OK', 'Size L',
    ])
    ws.append([
        'Ahmed Hassan', 'EMP-1001', 'Site A', date.today().isoformat(),
        'Operations', 'John Smith', 'Safety Helmet / Hard Hat', 'OK', '',
    ])
    ws.append([
        'Sara Ali', 'EMP-1002', 'Site B', date.today().isoformat(),
        'Hospitality', 'Jane Doe', 'Hi-Vis Vest / Jacket', 'Missing', 'Re-issue requested',
    ])
    ws.freeze_panes = 'A2'
    for col, width in zip('ABCDEFGHI', [22, 14, 18, 14, 16, 18, 28, 14, 30]):
        ws.column_dimensions[col].width = width

    ws2 = wb.create_sheet('Instructions')
    ws2['A1'] = 'QHSE Excel import — column guide'
    ws2['A3'] = 'Required: Employee Name, Project, Item, Condition'
    ws2['A4'] = 'Condition values: OK, Issue, Missing (or Compliant / Replacement needed / Missing)'
    ws2['A5'] = 'One row per kit item per employee (long format).'
    ws2['A6'] = 'Wide format is also supported: put kit names as column headers with OK/Issue/Missing in cells.'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
