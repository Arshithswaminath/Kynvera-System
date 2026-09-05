"""Integration tests for module_procurement/routes.py (blueprint mounted at /procurement).

Covers: page routes, materials CRUD, recent activity, Excel import/export,
properties, property-materials, material-property assignment, catalog
materials CRUD, and registered properties.

Access rule exercised throughout: every route (except GET
/procurement/api/catalog/materials) requires `user.role == 'admin'` or
`user.access_procurement_module` truthy, else 403. The catalog GET route is
intentionally open to any authenticated user (inspection-form consumers).
"""
from io import BytesIO

import openpyxl
import pytest

from tests.factories import make_user


# ---------------------------------------------------------------------------
# Local fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(autouse=True)
def _clean_procurement_submissions(app):
    """Procurement data lives in the shared, session-scoped DB
    (see conftest.py's session-scoped `app` fixture — the in-memory DB is
    never reset between tests). Isolate each test in this file by clearing
    procurement tables and leftover Submission blobs before it runs."""
    with app.app_context():
        from app.models import db, Submission, TicketMaterial
        from module_procurement.models import (
            ProcGoodsReceiptLine, ProcGoodsReceipt, ProcPurchaseLine,
            ProcPurchaseRequest, ProcMovement, ProcStock, ProcCatalogItem,
            ProcProperty, ProcSupplier, ProcPurchaseDocument, ProcEmailTemplate,
        )
        TicketMaterial.query.filter(
            db.or_(
                TicketMaterial.from_procurement == True,  # noqa: E712
                TicketMaterial.catalog_item_id.isnot(None),
            )
        ).delete(synchronize_session=False)
        ProcPurchaseDocument.query.delete(synchronize_session=False)
        ProcEmailTemplate.query.delete(synchronize_session=False)
        ProcGoodsReceiptLine.query.delete(synchronize_session=False)
        ProcGoodsReceipt.query.delete(synchronize_session=False)
        ProcPurchaseLine.query.delete(synchronize_session=False)
        ProcPurchaseRequest.query.delete(synchronize_session=False)
        ProcMovement.query.delete(synchronize_session=False)
        ProcStock.query.delete(synchronize_session=False)
        ProcCatalogItem.query.delete(synchronize_session=False)
        ProcProperty.query.delete(synchronize_session=False)
        ProcSupplier.query.delete(synchronize_session=False)
        Submission.query.filter(Submission.module_type.in_(
            ['procurement_material', 'procurement_property', 'catalog_material']
        )).delete(synchronize_session=False)
        db.session.commit()
    yield


# A "procurement-enabled" standard user, distinct from the plain
# `standard_user`/`auth_headers` fixtures in conftest.py which do NOT have
# access_procurement_module set.
# ---------------------------------------------------------------------------

@pytest.fixture
def procurement_user(app):
    """A non-admin user with explicit access_procurement_module=True."""
    with app.app_context():
        user, password = make_user(
            username='proc_user',
            access_procurement_module=True,
        )
        user_id = user.id
        yield {'id': user_id, 'username': 'proc_user', 'password': password}
        from app.models import db, User
        u = db.session.get(User, user_id)
        if u:
            db.session.delete(u)
            db.session.commit()


@pytest.fixture
def procurement_auth_headers(client, procurement_user):
    response = client.post('/api/auth/login', json={
        'username': procurement_user['username'],
        'password': procurement_user['password'],
    })
    token = response.get_json().get('access_token')
    return {'Authorization': f'Bearer {token}'}


def make_material(client, headers, **overrides):
    """Create a procurement material via the API and return its submission_id."""
    payload = {'material_name': 'Widget'}
    payload.update(overrides)
    resp = client.post('/procurement/api/materials', headers=headers, json=payload)
    assert resp.status_code == 200, resp.get_json()
    return resp.get_json()['submission_id']


def make_xlsx_bytes(rows, headers=None):
    """Build an in-memory .xlsx file. `rows` is a list of lists (data rows)."""
    if headers is None:
        headers = ['Material Name', 'Category', 'Description', 'Unit',
                    'Quantity', 'Unit Price', 'Supplier', 'Notes']
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Materials'
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Page routes
# ---------------------------------------------------------------------------

class TestProcurementPages:
    def test_dashboard_requires_auth(self, client):
        response = client.get('/procurement/')
        # Page routes are HTML navigations: unauthenticated GETs are silently
        # redirected to /login (302), not a JSON 401/422 like API routes.
        assert response.status_code == 302
        assert '/login' in response.headers.get('Location', '')

    def test_dashboard_admin_ok(self, client, admin_auth_headers):
        response = client.get('/procurement/', headers=admin_auth_headers)
        assert response.status_code == 200

    def test_dashboard_denied_without_access(self, client, auth_headers):
        response = client.get('/procurement/', headers=auth_headers)
        assert response.status_code == 403
        assert response.get_json()['error'] == 'Access denied to Procurement module'

    def test_dashboard_allowed_with_module_access(self, client, procurement_auth_headers):
        response = client.get('/procurement/', headers=procurement_auth_headers)
        assert response.status_code == 200

    def test_materials_page_admin_ok(self, client, admin_auth_headers):
        response = client.get('/procurement/materials', headers=admin_auth_headers)
        assert response.status_code == 200

    def test_materials_page_denied(self, client, auth_headers):
        response = client.get('/procurement/materials', headers=auth_headers)
        assert response.status_code == 403

    def test_add_material_page_admin_ok(self, client, admin_auth_headers):
        response = client.get('/procurement/add-material', headers=admin_auth_headers)
        assert response.status_code == 200

    def test_properties_page_admin_ok(self, client, admin_auth_headers):
        response = client.get('/procurement/properties', headers=admin_auth_headers)
        assert response.status_code == 200

    def test_property_detail_page_admin_ok(self, client, admin_auth_headers):
        response = client.get('/procurement/property/Tower%20A', headers=admin_auth_headers)
        assert response.status_code == 200

    def test_catalog_department_valid_admin_ok(self, client, admin_auth_headers):
        response = client.get('/procurement/catalog/HVAC', headers=admin_auth_headers)
        assert response.status_code == 200

    def test_catalog_department_invalid_redirects(self, client, admin_auth_headers):
        response = client.get('/procurement/catalog/NotADepartment', headers=admin_auth_headers)
        assert response.status_code == 302
        assert '/procurement/' in response.headers.get('Location', '')

    def test_catalog_department_denied_without_access(self, client, auth_headers):
        response = client.get('/procurement/catalog/HVAC', headers=auth_headers)
        assert response.status_code == 403

    def test_suppliers_page_admin_ok(self, client, admin_auth_headers):
        response = client.get('/procurement/suppliers', headers=admin_auth_headers)
        assert response.status_code == 200
        html = response.get_data(as_text=True)
        assert 'Material suppliers' in html
        assert 'Back to Procurement' not in html
        assert html.count('data-module-back') <= 1

    def test_materials_page_shows_qty_and_site(self, client, admin_auth_headers):
        response = client.get('/procurement/materials', headers=admin_auth_headers)
        assert response.status_code == 200
        html = response.get_data(as_text=True)
        assert 'proc-stock-list' in html
        assert 'On-hand quantities across sites' in html
        assert 'All sites' in html

    def test_purchase_requests_page_labels_awaiting_quotation(self, client, admin_auth_headers):
        response = client.get('/procurement/purchase-requests', headers=admin_auth_headers)
        assert response.status_code == 200
        assert 'Awaiting quotation' in response.get_data(as_text=True)

    def test_purchase_requests_page_admin_ok(self, client, admin_auth_headers):
        response = client.get('/procurement/purchase-requests', headers=admin_auth_headers)
        assert response.status_code == 200

    def test_usage_log_page_admin_ok(self, client, admin_auth_headers):
        response = client.get('/procurement/log', headers=admin_auth_headers)
        assert response.status_code == 200

    def test_refill_page_admin_ok(self, client, admin_auth_headers):
        response = client.get('/procurement/refill', headers=admin_auth_headers)
        assert response.status_code == 200

    def test_email_settings_page_admin_ok(self, client, admin_auth_headers):
        response = client.get('/procurement/email-settings', headers=admin_auth_headers)
        assert response.status_code == 200
        html = response.get_data(as_text=True)
        assert 'How to set this up' in html
        assert 'When they send' in html
        assert '{approve_url}' in html
        assert 'proc-mail-scroller' in html

    def test_usage_log_page_denied(self, client, auth_headers):
        response = client.get('/procurement/log', headers=auth_headers)
        assert response.status_code == 403


# ---------------------------------------------------------------------------
# GET/POST /procurement/api/materials
# ---------------------------------------------------------------------------

class TestMaterialsApi:
    def test_requires_auth(self, client):
        response = client.get('/procurement/api/materials')
        assert response.status_code in (401, 422)

    def test_denied_without_module_access(self, client, auth_headers):
        response = client.get('/procurement/api/materials', headers=auth_headers)
        assert response.status_code == 403

    def test_get_empty_list(self, client, admin_auth_headers):
        response = client.get('/procurement/api/materials', headers=admin_auth_headers)
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert data['materials'] == []
        assert data['total'] == 0

    def test_add_material_success(self, client, admin_auth_headers):
        response = client.post('/procurement/api/materials', headers=admin_auth_headers, json={
            'material_name': 'Office Chair',
            'property': 'Tower A',
            'category': 'Furniture',
            'quantity': 3,
            'unit_price': 150.0,
        })
        # Route returns plain jsonify() with no explicit status -> default 200
        # (not 201, despite this being a resource-creation endpoint).
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert data['submission_id'].startswith('PROC-MAT-')
        assert 'message' in data

        listed = client.get('/procurement/api/materials', headers=admin_auth_headers)
        materials = listed.get_json()['materials']
        assert len(materials) == 1
        m = materials[0]
        assert m['material_name'] == 'Office Chair'
        assert m['property'] == 'Tower A'
        assert m['quantity'] == 3
        assert m['unit_price'] == 150.0
        assert m['total_price'] == 450.0
        assert m['id'] == data['submission_id']

    def test_add_material_missing_name(self, client, admin_auth_headers):
        response = client.post('/procurement/api/materials', headers=admin_auth_headers, json={})
        assert response.status_code == 400
        assert response.get_json()['error'] == 'Material name is required'

    def test_add_material_denied_without_access(self, client, auth_headers):
        response = client.post('/procurement/api/materials', headers=auth_headers, json={
            'material_name': 'Widget',
        })
        assert response.status_code == 403

    def test_add_material_allowed_with_module_access(self, client, procurement_auth_headers):
        response = client.post('/procurement/api/materials', headers=procurement_auth_headers, json={
            'material_name': 'Cable Reel',
        })
        assert response.status_code == 200
        assert response.get_json()['success'] is True

    def test_add_material_requires_auth(self, client):
        response = client.post('/procurement/api/materials', json={'material_name': 'X'})
        assert response.status_code in (401, 422)


# ---------------------------------------------------------------------------
# GET /procurement/api/recent-activity
# ---------------------------------------------------------------------------

class TestRecentActivity:
    def test_requires_auth(self, client):
        response = client.get('/procurement/api/recent-activity')
        assert response.status_code in (401, 422)

    def test_denied_without_access(self, client, auth_headers):
        response = client.get('/procurement/api/recent-activity', headers=auth_headers)
        assert response.status_code == 403

    def test_empty_activity(self, client, admin_auth_headers):
        response = client.get('/procurement/api/recent-activity', headers=admin_auth_headers)
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert data['activities'] == []

    def test_lists_recent_materials(self, client, admin_auth_headers):
        make_material(client, admin_auth_headers, material_name='Ladder')
        response = client.get('/procurement/api/recent-activity', headers=admin_auth_headers)
        assert response.status_code == 200
        activities = response.get_json()['activities']
        assert len(activities) == 1
        assert activities[0]['material_name'] == 'Ladder'
        assert 'submitted_by' in activities[0]
        assert 'created_at' in activities[0]

    def test_limit_param_is_respected(self, client, admin_auth_headers):
        for i in range(5):
            make_material(client, admin_auth_headers, material_name=f'Item {i}')
        response = client.get('/procurement/api/recent-activity?limit=2', headers=admin_auth_headers)
        assert response.status_code == 200
        assert len(response.get_json()['activities']) == 2


# ---------------------------------------------------------------------------
# DELETE /procurement/api/materials/<material_id>
# ---------------------------------------------------------------------------

class TestDeleteMaterial:
    def test_requires_auth(self, client):
        response = client.delete('/procurement/api/materials/PROC-MAT-DOESNOTEXIST')
        assert response.status_code in (401, 422)

    def test_denied_without_access(self, client, auth_headers):
        response = client.delete('/procurement/api/materials/PROC-MAT-DOESNOTEXIST', headers=auth_headers)
        assert response.status_code == 403

    def test_delete_unknown_id_404(self, client, admin_auth_headers):
        response = client.delete('/procurement/api/materials/PROC-MAT-DOESNOTEXIST', headers=admin_auth_headers)
        assert response.status_code == 404
        assert response.get_json()['error'] == 'Material not found'

    def test_delete_success(self, client, admin_auth_headers):
        material_id = make_material(client, admin_auth_headers, material_name='Drill')
        response = client.delete(f'/procurement/api/materials/{material_id}', headers=admin_auth_headers)
        assert response.status_code == 200
        assert response.get_json()['success'] is True

        listed = client.get('/procurement/api/materials', headers=admin_auth_headers)
        assert listed.get_json()['total'] == 0


# ---------------------------------------------------------------------------
# POST /procurement/api/import-excel
# ---------------------------------------------------------------------------

class TestImportExcel:
    def test_requires_auth(self, client):
        buf = make_xlsx_bytes([['Widget', 'General', '', 'pcs', 1, 1, '', '']])
        response = client.post(
            '/procurement/api/import-excel',
            data={'file': (buf, 'materials.xlsx')},
            content_type='multipart/form-data',
        )
        assert response.status_code in (401, 422)

    def test_denied_without_access(self, client, auth_headers):
        buf = make_xlsx_bytes([['Widget', 'General', '', 'pcs', 1, 1, '', '']])
        response = client.post(
            '/procurement/api/import-excel',
            headers=auth_headers,
            data={'file': (buf, 'materials.xlsx')},
            content_type='multipart/form-data',
        )
        assert response.status_code == 403

    def test_no_file_provided(self, client, admin_auth_headers):
        response = client.post(
            '/procurement/api/import-excel',
            headers=admin_auth_headers,
            data={},
            content_type='multipart/form-data',
        )
        assert response.status_code == 400
        assert response.get_json()['error'] == 'No file provided'

    def test_non_excel_file_rejected(self, client, admin_auth_headers):
        response = client.post(
            '/procurement/api/import-excel',
            headers=admin_auth_headers,
            data={'file': (BytesIO(b'not an excel file'), 'materials.txt')},
            content_type='multipart/form-data',
        )
        assert response.status_code == 400
        assert 'Invalid file format' in response.get_json()['error']

    def test_import_valid_xlsx(self, client, admin_auth_headers):
        buf = make_xlsx_bytes([
            ['Office Paper A4 Ream', 'Stationery', '500 sheets', 'ream', 50, 12.5, 'Gulf Paper Co', 'Monthly'],
            ['Printer Toner', 'IT Supplies', 'Laser', 'pcs', 10, 85.0, 'Tech Supplies', ''],
        ])
        response = client.post(
            '/procurement/api/import-excel',
            headers=admin_auth_headers,
            data={'file': (buf, 'materials.xlsx')},
            content_type='multipart/form-data',
        )
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert data['imported'] == 2
        assert data['total_rows'] == 2
        assert data['errors'] == []

        listed = client.get('/procurement/api/materials', headers=admin_auth_headers)
        materials = listed.get_json()['materials']
        assert len(materials) == 2
        names = {m['material_name'] for m in materials}
        assert names == {'Office Paper A4 Ream', 'Printer Toner'}
        assert all(m.get('imported_from_excel') is True for m in materials)

    def test_import_missing_material_name_column_400(self, client, admin_auth_headers):
        buf = make_xlsx_bytes(
            [['foo', 1]],
            headers=['Some Other Column', 'Quantity'],
        )
        response = client.post(
            '/procurement/api/import-excel',
            headers=admin_auth_headers,
            data={'file': (buf, 'materials.xlsx')},
            content_type='multipart/form-data',
        )
        assert response.status_code == 400
        assert 'Material Name' in response.get_json()['error']

    def test_import_recognizes_alternate_column_names(self, client, admin_auth_headers):
        """The route lower-cases + maps common header variants (e.g. 'Item' -> material_name)."""
        buf = make_xlsx_bytes(
            [['Cordless Drill', 3, 99.99]],
            headers=['Item', 'Qty', 'Price'],
        )
        response = client.post(
            '/procurement/api/import-excel',
            headers=admin_auth_headers,
            data={'file': (buf, 'materials.xlsx')},
            content_type='multipart/form-data',
        )
        assert response.status_code == 200
        data = response.get_json()
        assert data['imported'] == 1

        listed = client.get('/procurement/api/materials', headers=admin_auth_headers)
        materials = listed.get_json()['materials']
        assert materials[0]['material_name'] == 'Cordless Drill'
        assert materials[0]['quantity'] == 3
        assert materials[0]['unit_price'] == 99.99


# ---------------------------------------------------------------------------
# GET /procurement/api/sample-excel and /procurement/api/export-excel
# ---------------------------------------------------------------------------

class TestSampleAndExportExcel:
    def test_sample_excel_requires_auth(self, client):
        response = client.get('/procurement/api/sample-excel')
        assert response.status_code in (401, 422)

    def test_sample_excel_denied_without_access(self, client, auth_headers):
        response = client.get('/procurement/api/sample-excel', headers=auth_headers)
        assert response.status_code == 403

    def test_sample_excel_download(self, client, admin_auth_headers):
        response = client.get('/procurement/api/sample-excel', headers=admin_auth_headers)
        assert response.status_code == 200
        assert response.headers['Content-Type'] == (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        assert 'attachment' in response.headers.get('Content-Disposition', '')
        assert response.data[:2] == b'PK'
        assert len(response.data) > 0

    def test_export_excel_requires_auth(self, client):
        response = client.get('/procurement/api/export-excel')
        assert response.status_code in (401, 422)

    def test_export_excel_denied_without_access(self, client, auth_headers):
        response = client.get('/procurement/api/export-excel', headers=auth_headers)
        assert response.status_code == 403

    def test_export_excel_empty(self, client, admin_auth_headers):
        response = client.get('/procurement/api/export-excel', headers=admin_auth_headers)
        assert response.status_code == 200
        assert response.headers['Content-Type'] == (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        assert response.data[:2] == b'PK'
        assert len(response.data) > 0

    def test_export_excel_with_property_filter(self, client, admin_auth_headers):
        make_material(client, admin_auth_headers, material_name='Filtered Item', property='Tower A')
        make_material(client, admin_auth_headers, material_name='Other Item', property='Tower B')
        response = client.get(
            '/procurement/api/export-excel?property=Tower%20A',
            headers=admin_auth_headers,
        )
        assert response.status_code == 200
        assert 'Tower_A' in response.headers.get('Content-Disposition', '')


class TestPageExports:
    KINDS = (
        'dashboard', 'usage-log', 'materials', 'properties',
        'refill', 'purchase-requests',
    )

    def test_requires_auth(self, client):
        response = client.get('/procurement/api/export/dashboard?format=pdf')
        assert response.status_code in (401, 422)

    def test_denied_without_access(self, client, auth_headers):
        response = client.get('/procurement/api/export/dashboard?format=pdf', headers=auth_headers)
        assert response.status_code == 403

    def test_unknown_kind(self, client, admin_auth_headers):
        response = client.get('/procurement/api/export/not-a-report?format=pdf', headers=admin_auth_headers)
        assert response.status_code == 400

    def test_bad_format(self, client, admin_auth_headers):
        response = client.get('/procurement/api/export/dashboard?format=csv', headers=admin_auth_headers)
        assert response.status_code == 400

    @pytest.mark.parametrize('kind', KINDS)
    def test_pdf_download(self, client, admin_auth_headers, kind):
        response = client.get(f'/procurement/api/export/{kind}?format=pdf', headers=admin_auth_headers)
        assert response.status_code == 200, response.get_json()
        assert response.data[:4] == b'%PDF'
        assert 'application/pdf' in response.headers.get('Content-Type', '')
        assert 'attachment' in response.headers.get('Content-Disposition', '')

    @pytest.mark.parametrize('kind', KINDS)
    def test_excel_download(self, client, admin_auth_headers, kind):
        response = client.get(f'/procurement/api/export/{kind}?format=xlsx', headers=admin_auth_headers)
        assert response.status_code == 200, response.get_json()
        assert response.data[:2] == b'PK'
        assert 'spreadsheetml' in response.headers.get('Content-Type', '')

    def test_dashboard_respects_range(self, client, admin_auth_headers):
        response = client.get(
            '/procurement/api/export/dashboard?format=xlsx&range=year',
            headers=admin_auth_headers,
        )
        assert response.status_code == 200
        book = openpyxl.load_workbook(BytesIO(response.data))
        assert 'Overview' in book.sheetnames
        assert book['Overview']['A1'].value == 'Procurement overview'

    def test_usage_log_excel_headers(self, client, admin_auth_headers):
        response = client.get(
            '/procurement/api/export/usage-log?format=xlsx',
            headers=admin_auth_headers,
        )
        assert response.status_code == 200
        book = openpyxl.load_workbook(BytesIO(response.data))
        headers = [c.value for c in book.active[4]]
        assert 'Material' in headers
        assert 'Status' in headers

    def test_pdf_keeps_logo_and_details(self, client, admin_auth_headers):
        from pypdf import PdfReader
        response = client.get(
            '/procurement/api/export/refill?format=pdf',
            headers=admin_auth_headers,
        )
        assert response.status_code == 200
        page = PdfReader(BytesIO(response.data)).pages[0]
        text = page.extract_text() or ''
        assert 'Refill queue' in text
        assert 'Generated' in text
        assert 'Procurement' in text
        assert len(page.images) >= 1

    def test_dashboard_pdf_includes_charts(self, client, admin_auth_headers):
        from pypdf import PdfReader
        response = client.get(
            '/procurement/api/export/dashboard?format=pdf&range=month&grain=month',
            headers=admin_auth_headers,
        )
        assert response.status_code == 200, response.get_json()
        text = PdfReader(BytesIO(response.data)).pages[0].extract_text() or ''
        assert 'Material utilization' in text
        assert 'Breakdown' in text
        assert 'Recent usage' in text
        assert 'Monthly' in text


# ---------------------------------------------------------------------------
# GET/POST /procurement/api/properties
# ---------------------------------------------------------------------------

class TestPropertiesApi:
    def test_get_requires_auth(self, client):
        response = client.get('/procurement/api/properties')
        assert response.status_code in (401, 422)

    def test_get_denied_without_access(self, client, auth_headers):
        response = client.get('/procurement/api/properties', headers=auth_headers)
        assert response.status_code == 403

    def test_get_empty(self, client, admin_auth_headers):
        response = client.get('/procurement/api/properties', headers=admin_auth_headers)
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        proc_only = [p for p in data['properties'] if not p.get('from_tickets')]
        assert proc_only == []

    def test_post_requires_name(self, client, admin_auth_headers):
        response = client.post('/procurement/api/properties', headers=admin_auth_headers, json={})
        assert response.status_code == 400
        assert response.get_json()['error'] == 'Property name is required'

    def test_post_blank_name_rejected(self, client, admin_auth_headers):
        response = client.post('/procurement/api/properties', headers=admin_auth_headers, json={'name': '   '})
        assert response.status_code == 400

    def test_post_creates_property(self, client, admin_auth_headers):
        response = client.post('/procurement/api/properties', headers=admin_auth_headers, json={
            'name': 'Marina Heights',
            'address': '123 Marina St',
        })
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert 'Marina Heights' in data['message']

    def test_post_denied_without_access(self, client, auth_headers):
        response = client.post('/procurement/api/properties', headers=auth_headers, json={'name': 'X'})
        assert response.status_code == 403

    def test_properties_aggregate_material_counts(self, client, admin_auth_headers):
        make_material(client, admin_auth_headers, material_name='A', property='Tower C', quantity=2, unit_price=10)
        make_material(client, admin_auth_headers, material_name='B', property='Tower C', quantity=1, unit_price=5)
        response = client.get('/procurement/api/properties', headers=admin_auth_headers)
        assert response.status_code == 200
        props = {p['name']: p for p in response.get_json()['properties']}
        assert props['Tower C']['materials_count'] == 2
        assert props['Tower C']['total_quantity'] == 3
        assert props['Tower C']['total_value'] == 25.0
        assert props['Tower C']['linked'] is False
        assert props['Tower C']['is_shared'] is False


class TestPropertyLinking:
    def test_empty_property_appears_on_list(self, client, admin_auth_headers):
        created = client.post(
            '/procurement/api/properties',
            headers=admin_auth_headers,
            json={'name': 'Empty Site', 'standalone': True},
        )
        assert created.status_code == 200, created.get_json()
        listed = client.get('/procurement/api/properties', headers=admin_auth_headers)
        props = {p['name']: p for p in listed.get_json()['properties']}
        assert 'Empty Site' in props
        assert props['Empty Site']['materials_count'] == 0
        assert props['Empty Site']['total_quantity'] == 0
        assert props['Empty Site']['linked'] is False
        assert created.get_json()['linked'] is False

    def test_create_with_ticket_property_id_links(self, app, client, admin_auth_headers):
        from app.models import db, TicketProperty
        with app.app_context():
            tp = TicketProperty(name='Common Tower', is_active=True)
            db.session.add(tp)
            db.session.commit()
            tp_id = tp.id
        resp = client.post(
            '/procurement/api/properties',
            headers=admin_auth_headers,
            json={'name': 'Common Tower', 'ticket_property_id': tp_id},
        )
        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['linked'] is True
        listed = client.get('/procurement/api/properties', headers=admin_auth_headers)
        row = next(p for p in listed.get_json()['properties'] if p['name'] == 'Common Tower')
        assert row['linked'] is True
        assert row['ticket_property_id'] == tp_id
        picker = client.get('/procurement/api/ticket-properties', headers=admin_auth_headers)
        hit = next(p for p in picker.get_json()['properties'] if p['id'] == tp_id)
        assert hit['linked'] is True
        with app.app_context():
            db.session.delete(db.session.get(TicketProperty, tp_id))
            db.session.commit()

    def test_standalone_does_not_auto_link_matching_name(self, app, client, admin_auth_headers):
        from app.models import db, TicketProperty
        from module_procurement.models import ProcProperty
        with app.app_context():
            tp = TicketProperty(name='Same Name Tower', is_active=True)
            db.session.add(tp)
            db.session.commit()
            tp_id = tp.id
        resp = client.post(
            '/procurement/api/properties',
            headers=admin_auth_headers,
            json={'name': 'Same Name Tower', 'standalone': True},
        )
        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['linked'] is False
        with app.app_context():
            row = ProcProperty.query.filter_by(name='Same Name Tower').first()
            assert row.ticket_property_id is None
            db.session.delete(row)
            db.session.delete(db.session.get(TicketProperty, tp_id))
            db.session.commit()

    def test_link_endpoint_sets_ticket_property(self, app, client, admin_auth_headers):
        from app.models import db, TicketProperty
        created = client.post(
            '/procurement/api/properties',
            headers=admin_auth_headers,
            json={'name': 'Proc Warehouse', 'standalone': True},
        )
        proc_id = created.get_json()['id']
        with app.app_context():
            tp = TicketProperty(name='Ticket Warehouse', is_active=True)
            db.session.add(tp)
            db.session.commit()
            tp_id = tp.id
        linked = client.post(
            f'/procurement/api/properties/{proc_id}/link',
            headers=admin_auth_headers,
            json={'ticket_property_id': tp_id},
        )
        assert linked.status_code == 200, linked.get_json()
        assert linked.get_json()['linked'] is True
        listed = client.get('/procurement/api/properties', headers=admin_auth_headers)
        row = next(p for p in listed.get_json()['properties'] if p['id'] == proc_id)
        assert row['ticket_property_id'] == tp_id
        with app.app_context():
            db.session.delete(db.session.get(TicketProperty, tp_id))
            db.session.commit()

    def test_duplicate_link_rejected(self, app, client, admin_auth_headers):
        from app.models import db, TicketProperty
        first = client.post(
            '/procurement/api/properties',
            headers=admin_auth_headers,
            json={'name': 'First Store', 'standalone': True},
        ).get_json()['id']
        second = client.post(
            '/procurement/api/properties',
            headers=admin_auth_headers,
            json={'name': 'Second Store', 'standalone': True},
        ).get_json()['id']
        with app.app_context():
            tp = TicketProperty(name='Only One Site', is_active=True)
            db.session.add(tp)
            db.session.commit()
            tp_id = tp.id
        ok = client.post(
            f'/procurement/api/properties/{first}/link',
            headers=admin_auth_headers,
            json={'ticket_property_id': tp_id},
        )
        assert ok.status_code == 200, ok.get_json()
        dup = client.post(
            f'/procurement/api/properties/{second}/link',
            headers=admin_auth_headers,
            json={'ticket_property_id': tp_id},
        )
        assert dup.status_code == 400
        with app.app_context():
            db.session.delete(db.session.get(TicketProperty, tp_id))
            db.session.commit()

    def test_ticket_properties_requires_access(self, client, auth_headers):
        resp = client.get('/procurement/api/ticket-properties', headers=auth_headers)
        assert resp.status_code == 403

    def test_ticket_properties_lists_common_sites(self, app, client, admin_auth_headers):
        from app.models import db, TicketProperty
        with app.app_context():
            tp = TicketProperty(name='Picker Site', is_active=True)
            db.session.add(tp)
            db.session.commit()
            tp_id = tp.id
        resp = client.get('/procurement/api/ticket-properties', headers=admin_auth_headers)
        assert resp.status_code == 200, resp.get_json()
        hit = next(p for p in resp.get_json()['properties'] if p['id'] == tp_id)
        assert hit['name'] == 'Picker Site'
        assert hit['linked'] is False
        with app.app_context():
            db.session.delete(db.session.get(TicketProperty, tp_id))
            db.session.commit()

    def test_ticket_sites_appear_on_property_grid(self, app, client, admin_auth_headers):
        from app.models import db, TicketProperty
        with app.app_context():
            tp = TicketProperty(name='CRM Villa 12', is_active=True)
            db.session.add(tp)
            db.session.commit()
            tp_id = tp.id
        listed = client.get('/procurement/api/properties', headers=admin_auth_headers)
        assert listed.status_code == 200
        row = next(
            p for p in listed.get_json()['properties']
            if p.get('ticket_property_id') == tp_id
        )
        assert row['name'] == 'CRM Villa 12'
        assert row['linked'] is True
        assert row['from_tickets'] is True
        assert row['needs_import'] is True
        assert row['id'] is None
        assert row['materials_count'] == 0
        imported = client.post(
            '/procurement/api/properties',
            headers=admin_auth_headers,
            json={'name': 'CRM Villa 12', 'ticket_property_id': tp_id},
        )
        assert imported.status_code == 200, imported.get_json()
        listed2 = client.get('/procurement/api/properties', headers=admin_auth_headers)
        matches = [p for p in listed2.get_json()['properties'] if p['name'] == 'CRM Villa 12']
        assert len(matches) == 1
        assert matches[0]['needs_import'] is False
        assert matches[0]['linked'] is True
        assert matches[0]['id']
        assert matches[0]['theme']['solid'].startswith('#')
        assert 'linear-gradient' in matches[0]['theme']['gradient']
        with app.app_context():
            db.session.delete(db.session.get(TicketProperty, tp_id))
            db.session.commit()

    def test_single_site_project_uses_project_name_on_grid(self, app, client, admin_auth_headers):
        from app.models import db, TicketProject, TicketProperty
        with app.app_context():
            proj = TicketProject(name='Injaaz HQ Building', is_active=True)
            db.session.add(proj)
            db.session.flush()
            tp = TicketProperty(name='HQ Shell Unique', project_id=proj.id, is_active=True)
            db.session.add(tp)
            db.session.commit()
            tp_id = tp.id
            proj_id = proj.id
        listed = client.get('/procurement/api/properties', headers=admin_auth_headers)
        row = next(
            p for p in listed.get_json()['properties']
            if p.get('ticket_property_id') == tp_id
        )
        assert row['name'] == 'HQ Shell Unique'
        assert row['display_name'] == 'Injaaz HQ Building'
        with app.app_context():
            db.session.delete(db.session.get(TicketProperty, tp_id))
            db.session.delete(db.session.get(TicketProject, proj_id))
            db.session.commit()

    def test_new_ticket_project_appears_on_grid_for_stock(self, app, client, admin_auth_headers):
        from app.models import db, TicketProject
        with app.app_context():
            proj = TicketProject(name='Amaan Stock Prompt Unique', is_active=True)
            db.session.add(proj)
            db.session.commit()
            proj_id = proj.id
        listed = client.get('/procurement/api/properties', headers=admin_auth_headers)
        assert listed.status_code == 200
        row = next(
            p for p in listed.get_json()['properties']
            if p.get('ticket_project_id') == proj_id
        )
        assert row['display_name'] == 'Amaan Stock Prompt Unique'
        assert row['needs_stock'] is True
        assert row['needs_import'] is True
        assert row['id'] is None
        imported = client.post(
            '/procurement/api/properties',
            headers=admin_auth_headers,
            json={'name': 'Amaan Stock Prompt Unique'},
        )
        assert imported.status_code == 200, imported.get_json()
        listed2 = client.get('/procurement/api/properties', headers=admin_auth_headers)
        matches = [p for p in listed2.get_json()['properties'] if p['name'] == 'Amaan Stock Prompt Unique']
        assert len(matches) == 1
        assert matches[0]['needs_stock'] is False
        assert matches[0]['id']
        with app.app_context():
            from module_procurement.models import ProcProperty
            row = ProcProperty.query.filter_by(name='Amaan Stock Prompt Unique').first()
            if row:
                db.session.delete(row)
            db.session.delete(db.session.get(TicketProject, proj_id))
            db.session.commit()

    def test_unlinked_proc_card_shows_matching_project_name(self, app, client, admin_auth_headers):
        from app.models import db, TicketProject, TicketProperty
        created = client.post(
            '/procurement/api/properties',
            headers=admin_auth_headers,
            json={'name': 'HQ Match Display', 'standalone': True},
        )
        assert created.status_code == 200, created.get_json()
        with app.app_context():
            proj = TicketProject(name='Injaaz HQ Building', is_active=True)
            db.session.add(proj)
            db.session.flush()
            tp = TicketProperty(name='HQ Match Display', project_id=proj.id, is_active=True)
            db.session.add(tp)
            db.session.commit()
            tp_id = tp.id
            proj_id = proj.id
        listed = client.get('/procurement/api/properties', headers=admin_auth_headers)
        row = next(p for p in listed.get_json()['properties'] if p['name'] == 'HQ Match Display')
        assert row['display_name'] == 'Injaaz HQ Building'
        with app.app_context():
            db.session.delete(db.session.get(TicketProperty, tp_id))
            db.session.delete(db.session.get(TicketProject, proj_id))
            db.session.commit()

    def test_save_property_icon(self, client, admin_auth_headers):
        created = client.post(
            '/procurement/api/properties',
            headers=admin_auth_headers,
            json={'name': 'Icon Site', 'standalone': True},
        )
        proc_id = created.get_json()['id']
        listed = client.get('/procurement/api/properties', headers=admin_auth_headers)
        choices = listed.get_json()['icons']
        assert any(c['icon'] == '🏠' and c['label'] == 'Villa' for c in choices)
        saved = client.post(
            '/procurement/api/properties/icon',
            headers=admin_auth_headers,
            json={'id': proc_id, 'icon': '🏠'},
        )
        assert saved.status_code == 200, saved.get_json()
        assert saved.get_json()['icon'] == '🏠'
        listed2 = client.get('/procurement/api/properties', headers=admin_auth_headers)
        row = next(p for p in listed2.get_json()['properties'] if p['name'] == 'Icon Site')
        assert row['theme']['icon'] == '🏠'

    def test_reject_unknown_property_icon(self, client, admin_auth_headers):
        created = client.post(
            '/procurement/api/properties',
            headers=admin_auth_headers,
            json={'name': 'Icon Site Two', 'standalone': True},
        )
        bad = client.post(
            '/procurement/api/properties/icon',
            headers=admin_auth_headers,
            json={'id': created.get_json()['id'], 'icon': '🚀'},
        )
        assert bad.status_code == 400


# ---------------------------------------------------------------------------
# GET /procurement/api/property-materials/<property_name>
# ---------------------------------------------------------------------------

class TestPropertyMaterials:
    def test_requires_auth(self, client):
        response = client.get('/procurement/api/property-materials/Tower%20A')
        assert response.status_code in (401, 422)

    def test_denied_without_access(self, client, auth_headers):
        response = client.get('/procurement/api/property-materials/Tower%20A', headers=auth_headers)
        assert response.status_code == 403

    def test_empty_for_unknown_property(self, client, admin_auth_headers):
        response = client.get('/procurement/api/property-materials/Nowhere', headers=admin_auth_headers)
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert data['property'] == 'Nowhere'
        assert data['materials'] == []
        assert data['total'] == 0

    def test_filters_by_property(self, client, admin_auth_headers):
        make_material(client, admin_auth_headers, material_name='In Tower D', property='Tower D')
        make_material(client, admin_auth_headers, material_name='In Tower E', property='Tower E')
        response = client.get('/procurement/api/property-materials/Tower%20D', headers=admin_auth_headers)
        assert response.status_code == 200
        data = response.get_json()
        assert data['total'] == 1
        assert data['materials'][0]['material_name'] == 'In Tower D'


# ---------------------------------------------------------------------------
# POST /procurement/api/material-assign-property
# ---------------------------------------------------------------------------

class TestMaterialAssignProperty:
    def test_requires_auth(self, client):
        response = client.post('/procurement/api/material-assign-property', json={})
        assert response.status_code in (401, 422)

    def test_denied_without_access(self, client, auth_headers):
        response = client.post('/procurement/api/material-assign-property', headers=auth_headers, json={
            'material_id': 'X', 'property': 'Y',
        })
        assert response.status_code == 403

    def test_missing_fields_400(self, client, admin_auth_headers):
        response = client.post('/procurement/api/material-assign-property', headers=admin_auth_headers, json={})
        assert response.status_code == 400
        assert 'required' in response.get_json()['error'].lower()

    def test_unknown_material_404(self, client, admin_auth_headers):
        response = client.post('/procurement/api/material-assign-property', headers=admin_auth_headers, json={
            'material_id': 'PROC-MAT-DOESNOTEXIST',
            'property': 'Tower F',
        })
        assert response.status_code == 404
        assert response.get_json()['error'] == 'Material not found'

    def test_assign_success(self, client, admin_auth_headers):
        material_id = make_material(client, admin_auth_headers, material_name='Unassigned Item')
        response = client.post('/procurement/api/material-assign-property', headers=admin_auth_headers, json={
            'material_id': material_id,
            'property': 'Tower F',
        })
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert 'Tower F' in data['message']

        by_property = client.get('/procurement/api/property-materials/Tower%20F', headers=admin_auth_headers)
        materials = by_property.get_json()['materials']
        assert len(materials) == 1
        assert materials[0]['id'] == material_id
        assert materials[0]['property'] == 'Tower F'


# ---------------------------------------------------------------------------
# GET/POST /procurement/api/catalog/materials, PUT/DELETE .../<material_id>
# ---------------------------------------------------------------------------

class TestCatalogMaterials:
    def test_get_requires_auth(self, client):
        response = client.get('/procurement/api/catalog/materials')
        assert response.status_code in (401, 422)

    def test_get_open_to_any_authenticated_user(self, client, auth_headers):
        """Unlike every other route in this module, GET catalog/materials has
        no admin/access_procurement_module check — any authenticated user can
        read it (the route docstring says this is intentional, for inspection
        forms)."""
        response = client.get('/procurement/api/catalog/materials', headers=auth_headers)
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert data['materials'] == {}
        assert data['departments'] == []
        assert data['total'] == 0

    def test_post_requires_auth(self, client):
        response = client.post('/procurement/api/catalog/materials', json={})
        assert response.status_code in (401, 422)

    def test_post_denied_without_access(self, client, auth_headers):
        response = client.post('/procurement/api/catalog/materials', headers=auth_headers, json={
            'department': 'HVAC', 'material_name': 'Compressor',
        })
        assert response.status_code == 403

    def test_post_invalid_department_400(self, client, admin_auth_headers):
        response = client.post('/procurement/api/catalog/materials', headers=admin_auth_headers, json={
            'department': 'NotADept', 'material_name': 'X',
        })
        assert response.status_code == 400
        assert response.get_json()['error'] == 'Invalid department'

    def test_post_missing_material_name_400(self, client, admin_auth_headers):
        response = client.post('/procurement/api/catalog/materials', headers=admin_auth_headers, json={
            'department': 'HVAC',
        })
        assert response.status_code == 400
        assert response.get_json()['error'] == 'Material name is required'

    def test_post_negative_price_400(self, client, admin_auth_headers):
        response = client.post('/procurement/api/catalog/materials', headers=admin_auth_headers, json={
            'department': 'HVAC', 'material_name': 'Compressor', 'unit_price': -1,
        })
        assert response.status_code == 400
        assert response.get_json()['error'] == 'Unit price cannot be negative'

    def test_post_invalid_price_400(self, client, admin_auth_headers):
        response = client.post('/procurement/api/catalog/materials', headers=admin_auth_headers, json={
            'department': 'HVAC', 'material_name': 'Compressor', 'unit_price': 'not-a-number',
        })
        assert response.status_code == 400
        assert response.get_json()['error'] == 'Invalid unit price'

    def test_post_creates_catalog_material(self, client, admin_auth_headers):
        response = client.post('/procurement/api/catalog/materials', headers=admin_auth_headers, json={
            'department': 'HVAC',
            'material_name': 'Compressor 1.5 Ton',
            'brand': 'Daikin',
            'uom': 'PCS',
            'unit_price': 1200,
        })
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert data['id'].startswith('CAT-MAT-')

        listed = client.get('/procurement/api/catalog/materials?department=HVAC', headers=admin_auth_headers)
        materials = listed.get_json()['materials']
        assert 'HVAC' in materials
        assert materials['HVAC'][0]['name'] == 'Compressor 1.5 Ton'
        assert materials['HVAC'][0]['brand'] == 'Daikin'
        assert materials['HVAC'][0]['unit_price'] == 1200

    def test_get_filters_by_query_string(self, client, admin_auth_headers):
        client.post('/procurement/api/catalog/materials', headers=admin_auth_headers, json={
            'department': 'Cleaning', 'material_name': 'Mop Bucket',
        })
        client.post('/procurement/api/catalog/materials', headers=admin_auth_headers, json={
            'department': 'Cleaning', 'material_name': 'Floor Cleaner',
        })
        response = client.get('/procurement/api/catalog/materials?q=mop', headers=admin_auth_headers)
        assert response.status_code == 200
        materials = response.get_json()['materials']
        names = [m['name'] for m in materials.get('Cleaning', [])]
        assert names == ['Mop Bucket']

    def test_put_unknown_id_404(self, client, admin_auth_headers):
        response = client.put('/procurement/api/catalog/materials/CAT-MAT-DOESNOTEXIST', headers=admin_auth_headers, json={
            'material_name': 'New Name',
        })
        assert response.status_code == 404
        assert response.get_json()['error'] == 'Catalog material not found'

    def test_put_requires_auth(self, client):
        response = client.put('/procurement/api/catalog/materials/CAT-MAT-X', json={})
        assert response.status_code in (401, 422)

    def test_put_updates_material(self, client, admin_auth_headers):
        created = client.post('/procurement/api/catalog/materials', headers=admin_auth_headers, json={
            'department': 'Electrical', 'material_name': 'Old Name', 'unit_price': 10,
        })
        material_id = created.get_json()['id']

        response = client.put(f'/procurement/api/catalog/materials/{material_id}', headers=admin_auth_headers, json={
            'material_name': 'New Name',
            'brand': 'ABB',
            'uom': 'BOX',
            'unit_price': 25.5,
        })
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert data['route_version'] == 2
        assert data['material']['id'] == material_id
        assert data['material']['name'] == 'New Name'
        assert data['material']['brand'] == 'ABB'
        assert data['material']['uom'] == 'BOX'
        assert data['material']['unit_price'] == 25.5

    def test_put_blank_name_rejected(self, client, admin_auth_headers):
        created = client.post('/procurement/api/catalog/materials', headers=admin_auth_headers, json={
            'department': 'Plumbing', 'material_name': 'Valve', 'unit_price': 5,
        })
        material_id = created.get_json()['id']
        response = client.put(f'/procurement/api/catalog/materials/{material_id}', headers=admin_auth_headers, json={
            'material_name': '   ',
        })
        assert response.status_code == 400
        assert response.get_json()['error'] == 'Material name is required'

    def test_put_negative_price_rejected(self, client, admin_auth_headers):
        created = client.post('/procurement/api/catalog/materials', headers=admin_auth_headers, json={
            'department': 'Plumbing', 'material_name': 'Valve', 'unit_price': 5,
        })
        material_id = created.get_json()['id']
        response = client.put(f'/procurement/api/catalog/materials/{material_id}', headers=admin_auth_headers, json={
            'unit_price': -10,
        })
        assert response.status_code == 400
        assert response.get_json()['error'] == 'Unit price cannot be negative'

    def test_put_invalid_department_rejected(self, client, admin_auth_headers):
        created = client.post('/procurement/api/catalog/materials', headers=admin_auth_headers, json={
            'department': 'Plumbing', 'material_name': 'Valve', 'unit_price': 5,
        })
        material_id = created.get_json()['id']
        response = client.put(f'/procurement/api/catalog/materials/{material_id}', headers=admin_auth_headers, json={
            'department': 'NotADept',
        })
        assert response.status_code == 400
        assert response.get_json()['error'] == 'Invalid department'

    def test_delete_unknown_id_404(self, client, admin_auth_headers):
        response = client.delete('/procurement/api/catalog/materials/CAT-MAT-DOESNOTEXIST', headers=admin_auth_headers)
        assert response.status_code == 404
        assert response.get_json()['error'] == 'Catalog material not found'

    def test_delete_requires_auth(self, client):
        response = client.delete('/procurement/api/catalog/materials/CAT-MAT-X')
        assert response.status_code in (401, 422)

    def test_delete_denied_without_access(self, client, auth_headers):
        response = client.delete('/procurement/api/catalog/materials/CAT-MAT-X', headers=auth_headers)
        assert response.status_code == 403

    def test_delete_success(self, client, admin_auth_headers):
        created = client.post('/procurement/api/catalog/materials', headers=admin_auth_headers, json={
            'department': 'HVAC', 'material_name': 'Filter', 'unit_price': 3,
        })
        material_id = created.get_json()['id']
        response = client.delete(f'/procurement/api/catalog/materials/{material_id}', headers=admin_auth_headers)
        assert response.status_code == 200
        assert response.get_json()['success'] is True

        listed = client.get('/procurement/api/catalog/materials?department=HVAC', headers=admin_auth_headers)
        materials = listed.get_json()['materials']
        assert materials.get('HVAC', []) == []

    def test_delete_does_not_remove_regular_material_with_same_id_shape(self, client, admin_auth_headers):
        """DELETE on the catalog endpoint filters by module_type='catalog_material',
        so it must not be able to delete a regular procurement material even if
        (hypothetically) IDs collided in shape."""
        material_id = make_material(client, admin_auth_headers, material_name='Regular Material')
        response = client.delete(f'/procurement/api/catalog/materials/{material_id}', headers=admin_auth_headers)
        assert response.status_code == 404
        # Confirm the regular material is untouched.
        listed = client.get('/procurement/api/materials', headers=admin_auth_headers)
        assert listed.get_json()['total'] == 1


# ---------------------------------------------------------------------------
# GET /procurement/api/registered-properties
# ---------------------------------------------------------------------------

class TestRegisteredProperties:
    def test_requires_auth(self, client):
        response = client.get('/procurement/api/registered-properties')
        assert response.status_code in (401, 422)

    def test_denied_without_access(self, client, auth_headers):
        response = client.get('/procurement/api/registered-properties', headers=auth_headers)
        assert response.status_code == 403

    def test_empty(self, client, admin_auth_headers):
        response = client.get('/procurement/api/registered-properties', headers=admin_auth_headers)
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert data['properties'] == []

    def test_lists_created_properties(self, client, admin_auth_headers):
        client.post('/procurement/api/properties', headers=admin_auth_headers, json={
            'name': 'Registered Tower',
            'address': '456 Sample Rd',
            'description': 'A registered property',
        })
        response = client.get('/procurement/api/registered-properties', headers=admin_auth_headers)
        assert response.status_code == 200
        props = response.get_json()['properties']
        assert len(props) == 1
        p = props[0]
        assert p['name'] == 'Registered Tower'
        assert p['address'] == '456 Sample Rd'
        assert p['description'] == 'A registered property'
        assert p['id'].startswith('PROC-PROP-')
        assert p['created_at'] is not None


# ---------------------------------------------------------------------------
# Seed, purchase requests, GRN, issue-to-ticket
# ---------------------------------------------------------------------------

def _catalog_item(client, headers, name='HEPA Filter', department='HVAC', unit_price=85):
    resp = client.post('/procurement/api/catalog/materials', headers=headers, json={
        'department': department, 'material_name': name, 'unit_price': unit_price, 'min_qty': 4,
    })
    assert resp.status_code == 200, resp.get_json()
    return resp.get_json()['id']


def _property_id(client, headers, name='Tower A'):
    resp = client.post('/procurement/api/properties', headers=headers, json={'name': name})
    assert resp.status_code == 200, resp.get_json()
    listed = client.get('/procurement/api/registered-properties', headers=headers)
    props = {p['name']: p['id'] for p in listed.get_json()['properties']}
    return props[name]


def _quote_pdf():
    from reportlab.pdfgen.canvas import Canvas
    buf = BytesIO()
    c = Canvas(buf)
    c.drawString(72, 720, 'Supplier quotation')
    c.save()
    buf.seek(0)
    return buf


def _upload_pr_doc(client, headers, pr_id, kind='quotation', filename='quote.pdf'):
    auth = {'Authorization': headers['Authorization']}
    return client.post(
        f'/procurement/api/purchase-requests/{pr_id}/documents/{kind}',
        headers=auth,
        data={'file': (_quote_pdf(), filename)},
    )


def _send_quotations(client, headers, pr_id):
    return client.post(
        f'/procurement/api/purchase-requests/{pr_id}/send-quotations',
        headers=headers,
        json={},
    )


def _complete_quotation(client, headers, pr_id):
    up = _upload_pr_doc(client, headers, pr_id)
    assert up.status_code == 200, up.get_json()
    sent = _send_quotations(client, headers, pr_id)
    assert sent.status_code == 200, sent.get_json()
    gm = client.post(f'/procurement/api/purchase-requests/{pr_id}/approve', headers=headers, json={})
    assert gm.status_code == 200, gm.get_json()
    return gm


def _assert_both_docs_stamped(client, headers, pr_id, docs, quote_kind='quotation'):
    assert docs['quotation']['has_stamped'] is True
    assert docs['pr_pdf']['has_stamped'] is True
    pr_pdf = client.get(
        f'/procurement/api/purchase-requests/{pr_id}/documents/pr_pdf',
        headers=headers,
    )
    assert pr_pdf.status_code == 200
    assert pr_pdf.data[:4] == b'%PDF'
    quote = client.get(
        f'/procurement/api/purchase-requests/{pr_id}/documents/{quote_kind}?stamped=1',
        headers=headers,
    )
    assert quote.status_code == 200
    assert quote.data[:4] == b'%PDF'


class TestSeedProcurement:
    def test_seed_is_idempotent(self, app):
        from scripts.seed_procurement import seed_procurement
        with app.app_context():
            first = seed_procurement()
            second = seed_procurement()
        assert first['catalog'] == second['catalog']
        assert first['catalog'] >= 80
        assert first['properties'] == second['properties'] == 5
        assert first['suppliers'] == second['suppliers']
        assert first['prs'] == second['prs']


class TestPurchaseRequests:
    def test_create_and_admin_approve_under_threshold(self, client, admin_auth_headers):
        cat_id = _catalog_item(client, admin_auth_headers, unit_price=10)
        prop_id = _property_id(client, admin_auth_headers)
        resp = client.post('/procurement/api/purchase-requests', headers=admin_auth_headers, json={
            'property_id': prop_id,
            'notes': 'Small restock',
            'lines': [{'catalog_id': cat_id, 'qty': 2, 'unit_price': 10}],
        })
        assert resp.status_code == 200, resp.get_json()
        pr = resp.get_json()['request']
        assert pr['status'] == 'procurement_review'
        assert pr['needs_gm'] is False
        approve = client.post(
            f'/procurement/api/purchase-requests/{pr["id"]}/approve',
            headers=admin_auth_headers, json={},
        )
        assert approve.status_code == 200
        assert approve.get_json()['request']['status'] == 'awaiting_quotation'
        quoted = _upload_pr_doc(client, admin_auth_headers, pr['id'])
        assert quoted.status_code == 200, quoted.get_json()
        assert quoted.get_json()['request']['status'] == 'awaiting_quotation'
        assert quoted.get_json()['documents']['quotation']['has_stamped'] is False
        sent = _send_quotations(client, admin_auth_headers, pr['id'])
        assert sent.status_code == 200, sent.get_json()
        assert sent.get_json()['request']['status'] == 'gm_review'
        gm = client.post(
            f'/procurement/api/purchase-requests/{pr["id"]}/approve',
            headers=admin_auth_headers, json={},
        )
        assert gm.status_code == 200
        assert gm.get_json()['request']['status'] == 'approved'
        docs = gm.get_json()['documents']
        assert docs['quotation']['status'] == 'approved'
        assert docs['quotation']['has_stamped'] is True
        assert docs['pr_pdf']['has_original'] is True
        _assert_both_docs_stamped(client, admin_auth_headers, pr['id'], docs)

    def test_create_multiple_lines(self, client, admin_auth_headers):
        cat_a = _catalog_item(client, admin_auth_headers, name='Filter A', unit_price=10)
        cat_b = _catalog_item(client, admin_auth_headers, name='Filter B', unit_price=20)
        prop_id = _property_id(client, admin_auth_headers)
        resp = client.post('/procurement/api/purchase-requests', headers=admin_auth_headers, json={
            'property_id': prop_id,
            'notes': 'Two lines',
            'lines': [
                {'catalog_id': cat_a, 'qty': 2, 'unit_price': 10},
                {'catalog_id': cat_b, 'qty': 3, 'unit_price': 20},
            ],
        })
        assert resp.status_code == 200, resp.get_json()
        pr = resp.get_json()['request']
        assert len(pr['lines']) == 2
        assert pr['total_aed'] == 80
        detail = client.get(f'/procurement/purchase-requests/{pr["id"]}', headers=admin_auth_headers)
        assert detail.status_code == 200
        recv = client.get(f'/procurement/receive/{pr["id"]}', headers=admin_auth_headers)
        assert recv.status_code in (301, 302)
        assert pr['id'] in (recv.headers.get('Location') or '')

    def test_over_threshold_goes_to_gm_review(self, client, admin_auth_headers):
        cat_id = _catalog_item(client, admin_auth_headers, name='Compressor', unit_price=1850)
        prop_id = _property_id(client, admin_auth_headers)
        resp = client.post('/procurement/api/purchase-requests', headers=admin_auth_headers, json={
            'property_id': prop_id,
            'lines': [{'catalog_id': cat_id, 'qty': 1, 'unit_price': 1850}],
        })
        pr = resp.get_json()['request']
        assert pr['needs_gm'] is True
        approve = client.post(
            f'/procurement/api/purchase-requests/{pr["id"]}/approve',
            headers=admin_auth_headers, json={},
        )
        assert approve.status_code == 200
        assert approve.get_json()['request']['status'] == 'awaiting_quotation'
        quoted = _upload_pr_doc(client, admin_auth_headers, pr['id'])
        assert quoted.status_code == 200, quoted.get_json()
        assert quoted.get_json()['request']['status'] == 'awaiting_quotation'
        sent = _send_quotations(client, admin_auth_headers, pr['id'])
        assert sent.status_code == 200, sent.get_json()
        assert sent.get_json()['request']['status'] == 'gm_review'
        assert sent.get_json()['documents']['quotation']['status'] == 'pending_approval'

    def test_module_user_without_procurement_designation_cannot_approve(
        self, client, admin_auth_headers, procurement_auth_headers,
    ):
        cat_id = _catalog_item(client, admin_auth_headers, unit_price=10)
        prop_id = _property_id(client, admin_auth_headers)
        created = client.post('/procurement/api/purchase-requests', headers=admin_auth_headers, json={
            'property_id': prop_id,
            'lines': [{'catalog_id': cat_id, 'qty': 1, 'unit_price': 10}],
        })
        pr_id = created.get_json()['request']['id']
        resp = client.post(
            f'/procurement/api/purchase-requests/{pr_id}/approve',
            headers=procurement_auth_headers, json={},
        )
        assert resp.status_code == 403

    def test_non_gm_cannot_approve_gm_review(self, app, client, admin_auth_headers):
        cat_id = _catalog_item(client, admin_auth_headers, name='Scroll', unit_price=1850)
        prop_id = _property_id(client, admin_auth_headers)
        created = client.post('/procurement/api/purchase-requests', headers=admin_auth_headers, json={
            'property_id': prop_id,
            'lines': [{'catalog_id': cat_id, 'qty': 1, 'unit_price': 1850}],
        })
        pr_id = created.get_json()['request']['id']
        client.post(f'/procurement/api/purchase-requests/{pr_id}/approve', headers=admin_auth_headers, json={})
        _upload_pr_doc(client, admin_auth_headers, pr_id)
        _send_quotations(client, admin_auth_headers, pr_id)
        with app.app_context():
            user, password = make_user(
                username='proc_desig',
                designation='procurement',
                access_procurement_module=True,
            )
            user_id = user.id
        login = client.post('/api/auth/login', json={'username': 'proc_desig', 'password': password})
        headers = {'Authorization': f'Bearer {login.get_json().get("access_token")}'}
        resp = client.post(f'/procurement/api/purchase-requests/{pr_id}/approve', headers=headers, json={})
        assert resp.status_code == 403
        with app.app_context():
            from app.models import db, User
            u = db.session.get(User, user_id)
            if u:
                db.session.delete(u)
                db.session.commit()


class TestPrDocuments:
    def test_pr_pdf_download_is_pdf(self, client, admin_auth_headers):
        cat_id = _catalog_item(client, admin_auth_headers, unit_price=10)
        prop_id = _property_id(client, admin_auth_headers)
        created = client.post('/procurement/api/purchase-requests', headers=admin_auth_headers, json={
            'property_id': prop_id,
            'lines': [{'catalog_id': cat_id, 'qty': 2, 'unit_price': 10}],
            'notes': 'Refill from low-stock queue',
        })
        pr_id = created.get_json()['request']['id']
        client.post(f'/procurement/api/purchase-requests/{pr_id}/approve', headers=admin_auth_headers, json={})
        dl = client.get(
            f'/procurement/api/purchase-requests/{pr_id}/documents/pr_pdf',
            headers=admin_auth_headers,
        )
        assert dl.status_code == 200
        assert dl.data[:4] == b'%PDF'
        assert len(dl.data) > 800

    def test_cannot_order_without_stamped_quotation(self, client, admin_auth_headers):
        cat_id = _catalog_item(client, admin_auth_headers, unit_price=10)
        prop_id = _property_id(client, admin_auth_headers)
        created = client.post('/procurement/api/purchase-requests', headers=admin_auth_headers, json={
            'property_id': prop_id,
            'lines': [{'catalog_id': cat_id, 'qty': 2, 'unit_price': 10}],
        })
        pr_id = created.get_json()['request']['id']
        client.post(f'/procurement/api/purchase-requests/{pr_id}/approve', headers=admin_auth_headers, json={})
        blocked = client.post(
            f'/procurement/api/purchase-requests/{pr_id}/order',
            headers=admin_auth_headers, json={},
        )
        assert blocked.status_code == 400
        quoted = _upload_pr_doc(client, admin_auth_headers, pr_id)
        assert quoted.get_json()['request']['status'] == 'awaiting_quotation'
        still_blocked = client.post(
            f'/procurement/api/purchase-requests/{pr_id}/order',
            headers=admin_auth_headers, json={},
        )
        assert still_blocked.status_code == 400
        sent = _send_quotations(client, admin_auth_headers, pr_id)
        assert sent.status_code == 200
        gm = client.post(
            f'/procurement/api/purchase-requests/{pr_id}/approve',
            headers=admin_auth_headers, json={},
        )
        assert gm.status_code == 200
        ordered = client.post(
            f'/procurement/api/purchase-requests/{pr_id}/order',
            headers=admin_auth_headers, json={},
        )
        assert ordered.status_code == 200
        assert ordered.get_json()['request']['status'] == 'ordered'

    def test_gm_or_token_stamps_over_threshold(self, app, client, admin_auth_headers):
        from unittest.mock import patch
        cat_id = _catalog_item(client, admin_auth_headers, name='Big Unit', unit_price=1850)
        prop_id = _property_id(client, admin_auth_headers)
        created = client.post('/procurement/api/purchase-requests', headers=admin_auth_headers, json={
            'property_id': prop_id,
            'lines': [{'catalog_id': cat_id, 'qty': 1, 'unit_price': 1850}],
        })
        pr_id = created.get_json()['request']['id']
        client.post(f'/procurement/api/purchase-requests/{pr_id}/approve', headers=admin_auth_headers, json={})
        with patch('module_procurement.pr_docs.send_email', return_value=True) as mocked:
            client.put('/procurement/api/email-templates', headers=admin_auth_headers, json={
                'event_key': 'quotation_for_approval',
                'to_emails': 'finance@example.com',
            })
            quoted = _upload_pr_doc(client, admin_auth_headers, pr_id)
            assert quoted.get_json()['request']['status'] == 'awaiting_quotation'
            assert not mocked.called
            sent = _send_quotations(client, admin_auth_headers, pr_id)
            assert sent.get_json()['request']['status'] == 'gm_review'
            assert mocked.called
        gm = client.post(
            f'/procurement/api/purchase-requests/{pr_id}/approve',
            headers=admin_auth_headers, json={},
        )
        assert gm.status_code == 200
        assert gm.get_json()['request']['status'] == 'approved'
        assert gm.get_json()['documents']['quotation']['has_stamped'] is True
        _assert_both_docs_stamped(client, admin_auth_headers, pr_id, gm.get_json()['documents'])

    def test_finance_token_approves_quotation(self, app, client, admin_auth_headers):
        from module_procurement import pr_docs
        from module_procurement.models import ProcPurchaseDocument
        from app.models import db
        cat_id = _catalog_item(client, admin_auth_headers, name='Chiller', unit_price=2000)
        prop_id = _property_id(client, admin_auth_headers)
        created = client.post('/procurement/api/purchase-requests', headers=admin_auth_headers, json={
            'property_id': prop_id,
            'lines': [{'catalog_id': cat_id, 'qty': 1, 'unit_price': 2000}],
        })
        pr_id = created.get_json()['request']['id']
        client.post(f'/procurement/api/purchase-requests/{pr_id}/approve', headers=admin_auth_headers, json={})
        _upload_pr_doc(client, admin_auth_headers, pr_id)
        _send_quotations(client, admin_auth_headers, pr_id)
        with app.app_context():
            doc = ProcPurchaseDocument.query.filter_by(kind='quotation').first()
            raw = pr_docs.issue_approval_token(doc)
            db.session.commit()
        page = client.get(f'/procurement/doc-approve/{raw}')
        assert page.status_code == 200
        detail = client.get(
            f'/procurement/api/purchase-requests/{pr_id}',
            headers=admin_auth_headers,
        ).get_json()
        assert detail['request']['status'] == 'approved'
        assert detail['documents']['quotation']['status'] == 'approved'
        _assert_both_docs_stamped(client, admin_auth_headers, pr_id, detail['documents'])

    def test_invoice_after_receive_waits_for_gm_stamp(self, client, admin_auth_headers):
        cat_id = _catalog_item(client, admin_auth_headers, name='Gasket', unit_price=12)
        prop_id = _property_id(client, admin_auth_headers, name='Plant Room')
        created = client.post('/procurement/api/purchase-requests', headers=admin_auth_headers, json={
            'property_id': prop_id,
            'lines': [{'catalog_id': cat_id, 'qty': 3, 'unit_price': 12}],
        })
        pr_id = created.get_json()['request']['id']
        client.post(f'/procurement/api/purchase-requests/{pr_id}/approve', headers=admin_auth_headers, json={})
        _complete_quotation(client, admin_auth_headers, pr_id)
        recv = client.post(
            f'/procurement/api/purchase-requests/{pr_id}/receive',
            headers=admin_auth_headers, json={},
        )
        assert recv.status_code == 200
        inv = _upload_pr_doc(client, admin_auth_headers, pr_id, kind='invoice', filename='inv.pdf')
        assert inv.status_code == 200, inv.get_json()
        docs = inv.get_json()['documents']
        assert docs['invoice']['status'] == 'pending_approval'
        assert docs['invoice']['has_stamped'] is False
        assert docs['invoice']['has_original'] is True
        approved = client.post(
            f'/procurement/api/purchase-requests/{pr_id}/approve',
            headers=admin_auth_headers, json={},
        )
        assert approved.status_code == 200, approved.get_json()
        stamped = approved.get_json()['documents']['invoice']
        assert stamped['status'] == 'approved'
        assert stamped['has_stamped'] is True
        assert approved.get_json()['request']['status'] == 'closed'
        again = _upload_pr_doc(client, admin_auth_headers, pr_id, kind='invoice', filename='inv-2.pdf')
        assert again.status_code == 400

    def test_email_settings_roundtrip(self, client, admin_auth_headers):
        listed = client.get('/procurement/api/email-templates', headers=admin_auth_headers)
        assert listed.status_code == 200
        keys = [t['event_key'] for t in listed.get_json()['templates']]
        assert 'quotation_for_approval' in keys
        assert all(t['attach_pdf'] is True for t in listed.get_json()['templates'])
        saved = client.put('/procurement/api/email-templates', headers=admin_auth_headers, json={
            'event_key': 'quotation_for_approval',
            'to_emails': 'gm@example.com',
            'cc_emails': 'ops@example.com',
            'subject': 'PR {pr_id}',
            'body': 'Please approve {approve_url}',
            'attach_pdf': False,
        })
        assert saved.status_code == 200, saved.get_json()
        row = next(t for t in saved.get_json()['templates'] if t['event_key'] == 'quotation_for_approval')
        assert row['to_emails'] == 'gm@example.com'
        assert row['cc_emails'] == 'ops@example.com'
        assert '{pr_id}' in row['subject']
        assert row['attach_pdf'] is False
        on_again = client.put('/procurement/api/email-templates', headers=admin_auth_headers, json={
            'event_key': 'quotation_for_approval',
            'to_emails': 'gm@example.com',
            'cc_emails': 'ops@example.com',
            'subject': 'PR {pr_id}',
            'body': 'Please approve {approve_url}',
            'attach_pdf': True,
        })
        assert on_again.status_code == 200, on_again.get_json()
        row = next(t for t in on_again.get_json()['templates'] if t['event_key'] == 'quotation_for_approval')
        assert row['attach_pdf'] is True

    def test_quotation_email_respects_attach_pdf(self, client, admin_auth_headers):
        from unittest.mock import patch
        cat_id = _catalog_item(client, admin_auth_headers, unit_price=10)
        prop_id = _property_id(client, admin_auth_headers)
        created = client.post('/procurement/api/purchase-requests', headers=admin_auth_headers, json={
            'property_id': prop_id,
            'lines': [{'catalog_id': cat_id, 'qty': 1, 'unit_price': 10}],
        })
        pr_id = created.get_json()['request']['id']
        client.post(f'/procurement/api/purchase-requests/{pr_id}/approve', headers=admin_auth_headers, json={})
        client.put('/procurement/api/email-templates', headers=admin_auth_headers, json={
            'event_key': 'quotation_for_approval',
            'to_emails': 'gm@example.com',
            'attach_pdf': False,
        })
        _upload_pr_doc(client, admin_auth_headers, pr_id)
        with patch('module_procurement.pr_docs.send_email', return_value=True) as mocked:
            sent = _send_quotations(client, admin_auth_headers, pr_id)
            assert sent.status_code == 200, sent.get_json()
            assert mocked.called
            assert not mocked.call_args.kwargs.get('attachments')
        client.put('/procurement/api/email-templates', headers=admin_auth_headers, json={
            'event_key': 'quotation_approved',
            'to_emails': 'ops@example.com',
            'attach_pdf': True,
        })
        with patch('module_procurement.pr_docs.send_email', return_value=True) as mocked:
            gm = client.post(
                f'/procurement/api/purchase-requests/{pr_id}/approve',
                headers=admin_auth_headers, json={},
            )
            assert gm.status_code == 200, gm.get_json()
            assert mocked.called
            attached = mocked.call_args.kwargs.get('attachments') or []
            assert len(attached) >= 1
            assert all(p.lower().endswith('.pdf') for p in attached)

    def test_cannot_send_quotations_without_file(self, client, admin_auth_headers):
        cat_id = _catalog_item(client, admin_auth_headers, unit_price=10)
        prop_id = _property_id(client, admin_auth_headers)
        created = client.post('/procurement/api/purchase-requests', headers=admin_auth_headers, json={
            'property_id': prop_id,
            'lines': [{'catalog_id': cat_id, 'qty': 1, 'unit_price': 10}],
        })
        pr_id = created.get_json()['request']['id']
        client.post(f'/procurement/api/purchase-requests/{pr_id}/approve', headers=admin_auth_headers, json={})
        sent = _send_quotations(client, admin_auth_headers, pr_id)
        assert sent.status_code == 400

    def test_up_to_three_quotations_gm_picks_one(self, app, client, admin_auth_headers):
        from app.models import Notification
        cat_id = _catalog_item(client, admin_auth_headers, unit_price=10)
        prop_id = _property_id(client, admin_auth_headers)
        created = client.post('/procurement/api/purchase-requests', headers=admin_auth_headers, json={
            'property_id': prop_id,
            'lines': [{'catalog_id': cat_id, 'qty': 1, 'unit_price': 10}],
        })
        pr_id = created.get_json()['request']['id']
        client.post(f'/procurement/api/purchase-requests/{pr_id}/approve', headers=admin_auth_headers, json={})
        for i in range(3):
            up = _upload_pr_doc(client, admin_auth_headers, pr_id, filename=f'quote-{i}.pdf')
            assert up.status_code == 200, up.get_json()
            assert up.get_json()['documents']['quotation']['has_stamped'] is False
        fourth = _upload_pr_doc(client, admin_auth_headers, pr_id, filename='quote-3.pdf')
        assert fourth.status_code == 400
        sent = _send_quotations(client, admin_auth_headers, pr_id)
        assert sent.status_code == 200
        assert sent.get_json()['request']['status'] == 'gm_review'
        assert sent.get_json()['documents']['quotation']['has_stamped'] is False
        with app.app_context():
            notes = Notification.query.filter_by(title='Quotation ready for approval').all()
            assert any(pr_id in (n.message or '') for n in notes)
            assert any(n.submission_id == pr_id for n in notes)
        gm = client.post(
            f'/procurement/api/purchase-requests/{pr_id}/approve',
            headers=admin_auth_headers,
            json={'quotation_kind': 'quotation_2'},
        )
        assert gm.status_code == 200, gm.get_json()
        docs = gm.get_json()['documents']
        assert docs['quotation']['has_stamped'] is True
        assert docs['quotation']['kind'] == 'quotation_2'
        stamped = [q for q in docs['quotations'] if q['has_stamped']]
        assert len(stamped) == 1
        assert len(docs['quotations']) == 3
        _assert_both_docs_stamped(client, admin_auth_headers, pr_id, docs, quote_kind='quotation_2')

    def test_gm_designation_can_approve_quotation(self, app, client, admin_auth_headers):
        cat_id = _catalog_item(client, admin_auth_headers, unit_price=10)
        prop_id = _property_id(client, admin_auth_headers)
        created = client.post('/procurement/api/purchase-requests', headers=admin_auth_headers, json={
            'property_id': prop_id,
            'lines': [{'catalog_id': cat_id, 'qty': 1, 'unit_price': 10}],
        })
        pr_id = created.get_json()['request']['id']
        client.post(f'/procurement/api/purchase-requests/{pr_id}/approve', headers=admin_auth_headers, json={})
        _upload_pr_doc(client, admin_auth_headers, pr_id)
        sent = _send_quotations(client, admin_auth_headers, pr_id)
        assert sent.get_json()['documents']['quotation']['has_stamped'] is False
        with app.app_context():
            gm, password = make_user(
                username='gm_quote_ok',
                designation='general_manager',
                access_procurement_module=True,
            )
            gm_id = gm.id
        login = client.post('/api/auth/login', json={'username': 'gm_quote_ok', 'password': password})
        headers = {'Authorization': f'Bearer {login.get_json().get("access_token")}'}
        approved = client.post(
            f'/procurement/api/purchase-requests/{pr_id}/approve',
            headers=headers, json={},
        )
        assert approved.status_code == 200, approved.get_json()
        assert approved.get_json()['request']['status'] == 'approved'
        assert approved.get_json()['documents']['quotation']['has_stamped'] is True
        _assert_both_docs_stamped(client, headers, pr_id, approved.get_json()['documents'])
        with app.app_context():
            from app.models import db, User
            u = db.session.get(User, gm_id)
            if u:
                db.session.delete(u)
                db.session.commit()


class TestGoodsReceipt:
    def test_receive_increases_stock(self, client, admin_auth_headers):
        cat_id = _catalog_item(client, admin_auth_headers, name='Pleated Filter', unit_price=28)
        prop_id = _property_id(client, admin_auth_headers, name='Plant Room')
        created = client.post('/procurement/api/purchase-requests', headers=admin_auth_headers, json={
            'property_id': prop_id,
            'lines': [{'catalog_id': cat_id, 'qty': 5, 'unit_price': 28}],
        })
        pr_id = created.get_json()['request']['id']
        client.post(f'/procurement/api/purchase-requests/{pr_id}/approve', headers=admin_auth_headers, json={})
        _complete_quotation(client, admin_auth_headers, pr_id)
        before = client.get('/procurement/api/materials', headers=admin_auth_headers).get_json()['total']
        recv = client.post(
            f'/procurement/api/purchase-requests/{pr_id}/receive',
            headers=admin_auth_headers,
            json={'quantities': {cat_id: 5}},
        )
        assert recv.status_code == 200, recv.get_json()
        assert recv.get_json()['request']['status'] == 'received'
        after = client.get('/procurement/api/materials', headers=admin_auth_headers)
        materials = after.get_json()['materials']
        assert after.get_json()['total'] == before + 1
        match = [m for m in materials if m['material_name'] == 'Pleated Filter']
        assert match
        assert match[0]['quantity'] == 5
        assert match[0]['property'] == 'Plant Room'


class TestIssueToTicket:
    def test_issue_decrements_stock(self, app, client, admin_auth_headers, admin_user):
        from app.models import db, Ticket, TicketMaterial
        reporter_id = admin_user.id
        ticket_public_id = 'TKT-PROC-ISSUE1'
        stock_id = make_material(
            client, admin_auth_headers,
            material_name='MCB 1P 16A', property='Tower A',
            category='Electrical', quantity=10, unit_price=19,
        )
        listed = client.get('/procurement/api/materials', headers=admin_auth_headers)
        row = next(m for m in listed.get_json()['materials'] if m['id'] == stock_id)
        props = client.get('/procurement/api/registered-properties', headers=admin_auth_headers).get_json()['properties']
        prop_id = next(p['id'] for p in props if p['name'] == 'Tower A')
        with app.app_context():
            existing = Ticket.query.filter_by(ticket_id=ticket_public_id).first()
            if existing:
                TicketMaterial.query.filter_by(ticket_id=existing.id).delete()
                db.session.delete(existing)
                db.session.commit()
            t = Ticket(
                ticket_id=ticket_public_id,
                reporter_id=reporter_id,
                title='Board trip',
                project='Marina',
                service_group='Electrical',
                category='Power',
                fault_type='Trip',
                priority='medium',
                work_description='MCB replacement',
                status='open',
                property_name='Tower A',
            )
            db.session.add(t)
            db.session.commit()
            ticket_pk = t.id
        resp = client.post('/procurement/api/issue-to-ticket', headers=admin_auth_headers, json={
            'property_id': prop_id,
            'catalog_id': row['catalog_id'],
            'ticket_id': ticket_public_id,
            'qty': 3,
            'chargeable': True,
        })
        assert resp.status_code == 200, resp.get_json()
        mat = resp.get_json()['material']
        assert mat['from_procurement'] is True
        assert mat['quantity'] == 3
        after = client.get('/procurement/api/materials', headers=admin_auth_headers)
        left = next(m for m in after.get_json()['materials'] if m['id'] == stock_id)
        assert left['quantity'] == 7
        with app.app_context():
            tm = TicketMaterial.query.filter_by(ticket_id=ticket_pk).first()
            assert tm is not None
            assert tm.catalog_item_id is not None
            tm_id = tm.id
        ret = client.post('/procurement/api/return-from-ticket', headers=admin_auth_headers, json={
            'ticket_material_id': tm_id,
            'property_id': prop_id,
        })
        assert ret.status_code == 200, ret.get_json()
        restored = client.get('/procurement/api/materials', headers=admin_auth_headers)
        left = next(m for m in restored.get_json()['materials'] if m['id'] == stock_id)
        assert left['quantity'] == 10
        with app.app_context():
            t = db.session.get(Ticket, ticket_pk)
            if t:
                TicketMaterial.query.filter_by(ticket_id=t.id).delete()
                db.session.delete(t)
                db.session.commit()


def _work_ticket(app, reporter_id, ticket_public_id, property_name):
    from app.models import db, Ticket, TicketMaterial
    with app.app_context():
        existing = Ticket.query.filter_by(ticket_id=ticket_public_id).first()
        if existing:
            TicketMaterial.query.filter_by(ticket_id=existing.id).delete()
            db.session.delete(existing)
            db.session.commit()
        t = Ticket(
            ticket_id=ticket_public_id,
            reporter_id=reporter_id,
            title='Consume materials',
            project='Marina',
            service_group='Electrical',
            category='Power',
            fault_type='Trip',
            priority='medium',
            work_description='Use catalog parts',
            status='work_started',
            property_name=property_name,
        )
        db.session.add(t)
        db.session.commit()
        return t.id


class TestDashboardAndLogApi:
    def test_dashboard_payload_shape(self, client, admin_auth_headers):
        resp = client.get('/procurement/api/dashboard?range=month', headers=admin_auth_headers)
        assert resp.status_code == 200, resp.get_json()
        data = resp.get_json()
        assert data['success'] is True
        assert set(data['kpis']) >= {'below_threshold', 'issued', 'spend', 'stock_value'}
        assert 'labels' in data['utilization']
        assert 'series' in data['utilization']
        assert 'charts' in data
        assert set(data['charts']) >= {'week', 'month', 'year'}
        assert 'labels' in data['daily']
        assert 'issued' in data['daily']
        assert 'compare' in data['daily']
        assert isinstance(data['recent'], list)
        assert isinstance(data['breakdown'], list)
        assert data.get('break_period_start')
        assert data.get('break_period_end')

    def test_period_bounds_three_months(self):
        from calendar import monthrange
        from module_procurement.service import period_bounds
        key, label, start, end, prev_start, prev_end = period_bounds('3m')
        assert key == '3m'
        assert label == 'Last 3 months'
        assert start.day == 1
        assert end.day == monthrange(end.year, end.month)[1]
        assert (end - start).days >= 59
        assert prev_end == start
        assert prev_start.day == 1

    def test_period_bounds_this_month_is_full_calendar_month(self):
        from calendar import monthrange
        from module_procurement.service import period_bounds
        key, label, start, end, prev_start, prev_end = period_bounds('month')
        assert key == 'month'
        assert start.day == 1
        assert start.month == end.month
        assert end.day == monthrange(start.year, start.month)[1]
        assert prev_end == start
        assert prev_start.day == 1

    def test_period_bounds_custom_month(self):
        from module_procurement.service import period_bounds
        key, label, start, end, prev_start, prev_end = period_bounds('2026-03')
        assert key == '2026-03'
        assert start.year == 2026 and start.month == 3 and start.day == 1
        assert end.month == 3 and end.day == 31
        assert prev_start.month == 2
        assert prev_end == start
        assert 'Mar' in label

    def test_dashboard_break_three_months(self, client, admin_auth_headers):
        resp = client.get(
            '/procurement/api/dashboard?range=month&break=3m',
            headers=admin_auth_headers,
        )
        assert resp.status_code == 200, resp.get_json()
        data = resp.get_json()
        assert data['break_period'] == '3m'
        assert data['period'] == 'month'
        assert len(data['daily']['labels']) >= 60

    def test_dashboard_break_custom_month(self, client, admin_auth_headers):
        resp = client.get(
            '/procurement/api/dashboard?range=month&break=2026-03',
            headers=admin_auth_headers,
        )
        assert resp.status_code == 200, resp.get_json()
        data = resp.get_json()
        assert data['break_period'] == '2026-03'
        assert data['break_period_start'] == '2026-03-01'
        assert data['break_period_end'].startswith('2026-03-')

    def test_period_bounds_last_month(self):
        from calendar import monthrange
        from module_procurement.service import period_bounds, _shift_months, _month_start, _utcnow
        key, label, start, end, prev_start, prev_end = period_bounds('last_month')
        assert key == 'last_month'
        assert label == 'Last month'
        assert start.day == 1
        assert end.day == monthrange(start.year, start.month)[1]
        assert start.month == _shift_months(_month_start(_utcnow()), -1).month
        alias = period_bounds('30d')
        assert alias[0] == 'last_month'
        assert alias[2] == start

    def test_dashboard_charts_follow_range(self, client, admin_auth_headers):
        def series_sum(pack):
            return sum(sum(v or []) for v in (pack.get('series') or {}).values())

        month = client.get('/procurement/api/dashboard?range=month', headers=admin_auth_headers).get_json()
        last = client.get('/procurement/api/dashboard?range=last_month', headers=admin_auth_headers).get_json()
        year = client.get('/procurement/api/dashboard?range=year', headers=admin_auth_headers).get_json()
        assert month['period'] == 'month'
        assert last['period'] == 'last_month'
        assert last['period_label'] == 'Last month'
        assert year['period'] == 'year'
        from calendar import monthrange
        y, m = int(month['period_start'][:4]), int(month['period_start'][5:7])
        last_day = monthrange(y, m)[1]
        assert month['period_start'].endswith('-01')
        assert month['period_end'] == f'{y:04d}-{m:02d}-{last_day:02d}'
        ly, lm = int(last['period_start'][:4]), int(last['period_start'][5:7])
        assert last['period_start'].endswith('-01')
        assert last['period_end'] == f'{ly:04d}-{lm:02d}-{monthrange(ly, lm)[1]:02d}'
        assert len(month['daily']['labels']) == last_day
        assert len(month['charts']['month']['labels']) == 1
        assert len(year['charts']['month']['labels']) >= 9
        assert series_sum(month['charts']['week']) == month['kpis']['issued']['value']
        assert series_sum(month['charts']['month']) == month['kpis']['issued']['value']
        assert series_sum(last['charts']['week']) == last['kpis']['issued']['value']
        assert series_sum(year['charts']['month']) == year['kpis']['issued']['value']
        assert month['period_start'] == month['break_period_start']
        assert month['period_end'] == month['break_period_end']
        assert last['period_start'] == last['break_period_start']
        assert last['period_end'] == last['break_period_end']

    def test_dashboard_utilization_clips_outside_period(self, app, client, admin_auth_headers):
        import uuid
        from datetime import timedelta
        from app.models import db
        from module_procurement.models import ProcCatalogItem, ProcMovement
        from module_procurement.service import period_bounds

        def series_sum(pack):
            return sum(sum(v or []) for v in (pack.get('series') or {}).values())

        with app.app_context():
            item = ProcCatalogItem(
                public_id='clip-' + uuid.uuid4().hex[:12],
                name='Clip Valve',
                department='Plumbing',
                uom='PCS',
                unit_price=1,
            )
            db.session.add(item)
            db.session.flush()
            _, _, month_start, _, _, _ = period_bounds('month')
            from module_procurement.service import _utcnow
            now = _utcnow()
            inside = month_start + timedelta(hours=12)
            if inside > now:
                inside = now
            db.session.add(ProcMovement(
                movement_type='issue', catalog_item_id=item.id, qty=-100,
                created_at=month_start - timedelta(days=1),
            ))
            db.session.add(ProcMovement(
                movement_type='issue', catalog_item_id=item.id, qty=-7,
                created_at=inside,
            ))
            db.session.commit()

        month = client.get('/procurement/api/dashboard?range=month', headers=admin_auth_headers).get_json()
        year = client.get('/procurement/api/dashboard?range=year', headers=admin_auth_headers).get_json()
        assert month['kpis']['issued']['value'] == 7
        assert series_sum(month['charts']['week']) == 7
        assert series_sum(month['charts']['month']) == 7
        assert year['kpis']['issued']['value'] == 107
        assert series_sum(year['charts']['month']) == 107

    def test_usage_log_empty(self, client, admin_auth_headers):
        resp = client.get('/procurement/api/usage-log', headers=admin_auth_headers)
        assert resp.status_code == 200
        assert resp.get_json()['success'] is True
        assert resp.get_json()['rows'] == []

    def test_refill_empty(self, client, admin_auth_headers):
        resp = client.get('/procurement/api/refill', headers=admin_auth_headers)
        assert resp.status_code == 200
        assert resp.get_json()['items'] == []

    def test_refill_summary_empty(self, client, admin_auth_headers):
        resp = client.get('/procurement/api/refill/summary', headers=admin_auth_headers)
        assert resp.status_code == 200
        data = resp.get_json()
        assert data['success'] is True
        assert data['total'] == 0
        assert data['out_of_stock'] == 0
        assert data.get('newest_at') in (None, '')

    def test_refill_timestamp_newest_first(self, app, client, admin_auth_headers):
        from datetime import datetime, timedelta, timezone
        from app.models import db
        from module_procurement.models import ProcStock

        older_id = make_material(
            client, admin_auth_headers,
            material_name='Older Low Cable', property='Stamp Site',
            category='Electrical', quantity=0, unit_price=10,
        )
        newer_id = make_material(
            client, admin_auth_headers,
            material_name='Newer Low Cable', property='Stamp Site',
            category='Electrical', quantity=0, unit_price=10,
        )
        with app.app_context():
            older = ProcStock.query.filter_by(public_id=older_id).first()
            newer = ProcStock.query.filter_by(public_id=newer_id).first()
            now = datetime.now(timezone.utc).replace(tzinfo=None)
            older.created_at = now - timedelta(days=3)
            older.updated_at = now - timedelta(days=3)
            newer.created_at = now - timedelta(hours=1)
            newer.updated_at = now - timedelta(hours=1)
            db.session.commit()

        refill = client.get('/procurement/api/refill', headers=admin_auth_headers).get_json()
        named = [i for i in refill['items'] if i['name'] in ('Older Low Cable', 'Newer Low Cable')]
        assert [i['name'] for i in named] == ['Newer Low Cable', 'Older Low Cable']
        assert all(i.get('updated_at') for i in named)
        assert named[0]['updated_at'].endswith('Z')
        assert named[0]['updated_at'] > named[1]['updated_at']
        assert named[0].get('last_event')
        assert named[0].get('edited_at', '').endswith('Z')
        assert named[0].get('open_pr') in (None, {})
        summary = client.get('/procurement/api/refill/summary', headers=admin_auth_headers).get_json()
        assert summary['total'] >= 2
        assert summary['out_of_stock'] >= 2
        assert summary.get('newest_at')


class TestTicketConsume:
    def test_catalog_add_deducts_stock_and_logs(self, app, client, admin_auth_headers, admin_user):
        from app.models import db, Ticket, TicketMaterial
        from module_procurement.models import ProcMovement
        stock_id = make_material(
            client, admin_auth_headers,
            material_name='MCB Consume 16A', property='Consume Site',
            category='Electrical', quantity=10, unit_price=19,
        )
        listed = client.get('/procurement/api/materials', headers=admin_auth_headers)
        row = next(m for m in listed.get_json()['materials'] if m['id'] == stock_id)
        cat_id = row['catalog_id']
        client.put(
            f'/procurement/api/catalog/materials/{cat_id}',
            headers=admin_auth_headers,
            json={'min_qty': 4},
        )
        _work_ticket(app, admin_user.id, 'TKT-PROC-CONSUME1', 'Consume Site')
        resp = client.post(
            '/tickets/api/tickets/TKT-PROC-CONSUME1/materials',
            headers=admin_auth_headers,
            json={
                'material_name': 'MCB Consume 16A',
                'quantity': 3,
                'unit': 'PCS',
                'unit_price': 19,
                'from_procurement': True,
                'procurement_ref': cat_id,
            },
        )
        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['material']['from_procurement'] is True
        after = client.get('/procurement/api/materials', headers=admin_auth_headers)
        left = next(m for m in after.get_json()['materials'] if m['id'] == stock_id)
        assert left['quantity'] == 7
        log = client.get('/procurement/api/usage-log', headers=admin_auth_headers).get_json()
        issues = [r for r in log['rows'] if r['ticket_id'] == 'TKT-PROC-CONSUME1']
        assert len(issues) == 1
        assert issues[0]['qty'] == 3
        assert issues[0]['ticket_title'] == 'Consume materials'
        with app.app_context():
            t = Ticket.query.filter_by(ticket_id='TKT-PROC-CONSUME1').first()
            assert ProcMovement.query.filter_by(ticket_id=t.id, movement_type='issue').count() == 1
            TicketMaterial.query.filter_by(ticket_id=t.id).delete()
            ProcMovement.query.filter_by(ticket_id=t.id).delete()
            db.session.delete(t)
            db.session.commit()

    def test_insufficient_stock_does_not_block(self, app, client, admin_auth_headers, admin_user):
        from app.models import db, Ticket, TicketMaterial
        stock_id = make_material(
            client, admin_auth_headers,
            material_name='Filter Overuse', property='Overuse Site',
            category='HVAC', quantity=2, unit_price=12,
        )
        listed = client.get('/procurement/api/materials', headers=admin_auth_headers)
        row = next(m for m in listed.get_json()['materials'] if m['id'] == stock_id)
        cat_id = row['catalog_id']
        client.put(
            f'/procurement/api/catalog/materials/{cat_id}',
            headers=admin_auth_headers,
            json={'min_qty': 5},
        )
        _work_ticket(app, admin_user.id, 'TKT-PROC-OVERUSE', 'Overuse Site')
        resp = client.post(
            '/tickets/api/tickets/TKT-PROC-OVERUSE/materials',
            headers=admin_auth_headers,
            json={
                'material_name': 'Filter Overuse',
                'quantity': 10,
                'from_procurement': True,
                'procurement_ref': cat_id,
            },
        )
        assert resp.status_code == 200, resp.get_json()
        body = resp.get_json()
        assert body['material']['quantity'] == 2
        assert body['material']['qty_short'] == 8
        after = client.get('/procurement/api/materials', headers=admin_auth_headers)
        left = next(m for m in after.get_json()['materials'] if m['id'] == stock_id)
        assert left['quantity'] == 0
        refill = client.get('/procurement/api/refill', headers=admin_auth_headers).get_json()
        names = [i['name'] for i in refill['items']]
        assert 'Filter Overuse' in names
        with app.app_context():
            t = Ticket.query.filter_by(ticket_id='TKT-PROC-OVERUSE').first()
            tm = TicketMaterial.query.filter_by(ticket_id=t.id).first()
            assert tm is not None
            assert tm.quantity == 2
            assert float(tm.qty_short or 0) == 8
            assert tm.created_at is not None
            from module_procurement.models import ProcMovement
            move = ProcMovement.query.filter_by(ticket_id=t.id, movement_type='issue').first()
            assert move is not None
            assert abs(float(move.qty or 0)) == 2
            TicketMaterial.query.filter_by(ticket_id=t.id).delete()
            db.session.delete(t)
            db.session.commit()

    def test_manual_material_does_not_move_stock(self, app, client, admin_auth_headers, admin_user):
        from app.models import db, Ticket, TicketMaterial
        stock_id = make_material(
            client, admin_auth_headers,
            material_name='Stay Put Filter', property='Manual Site',
            category='HVAC', quantity=8, unit_price=5,
        )
        _work_ticket(app, admin_user.id, 'TKT-PROC-MANUAL', 'Manual Site')
        before_moves = client.get('/procurement/api/usage-log', headers=admin_auth_headers).get_json()['total']
        resp = client.post(
            '/tickets/api/tickets/TKT-PROC-MANUAL/materials',
            headers=admin_auth_headers,
            json={'material_name': 'Shop-bought gasket', 'quantity': 1, 'from_procurement': False},
        )
        assert resp.status_code == 200, resp.get_json()
        after = client.get('/procurement/api/materials', headers=admin_auth_headers)
        left = next(m for m in after.get_json()['materials'] if m['id'] == stock_id)
        assert left['quantity'] == 8
        after_moves = client.get('/procurement/api/usage-log', headers=admin_auth_headers).get_json()['total']
        assert after_moves == before_moves
        with app.app_context():
            t = Ticket.query.filter_by(ticket_id='TKT-PROC-MANUAL').first()
            TicketMaterial.query.filter_by(ticket_id=t.id).delete()
            db.session.delete(t)
            db.session.commit()

    def test_delete_catalog_material_restores_stock(self, app, client, admin_auth_headers, admin_user):
        from app.models import db, Ticket, TicketMaterial
        stock_id = make_material(
            client, admin_auth_headers,
            material_name='Returnable Valve', property='Return Site',
            category='Plumbing', quantity=6, unit_price=40,
        )
        listed = client.get('/procurement/api/materials', headers=admin_auth_headers)
        row = next(m for m in listed.get_json()['materials'] if m['id'] == stock_id)
        cat_id = row['catalog_id']
        _work_ticket(app, admin_user.id, 'TKT-PROC-RETURN', 'Return Site')
        added = client.post(
            '/tickets/api/tickets/TKT-PROC-RETURN/materials',
            headers=admin_auth_headers,
            json={
                'material_name': 'Returnable Valve',
                'quantity': 2,
                'from_procurement': True,
                'procurement_ref': cat_id,
            },
        )
        assert added.status_code == 200, added.get_json()
        mat_id = added.get_json()['material']['id']
        gone = client.delete(
            f'/tickets/api/tickets/TKT-PROC-RETURN/materials/{mat_id}',
            headers=admin_auth_headers,
        )
        assert gone.status_code == 200, gone.get_json()
        restored = client.get('/procurement/api/materials', headers=admin_auth_headers)
        left = next(m for m in restored.get_json()['materials'] if m['id'] == stock_id)
        assert left['quantity'] == 6
        with app.app_context():
            t = Ticket.query.filter_by(ticket_id='TKT-PROC-RETURN').first()
            if t:
                TicketMaterial.query.filter_by(ticket_id=t.id).delete()
                db.session.delete(t)
                db.session.commit()

    def test_refill_creates_purchase_request(self, app, client, admin_auth_headers, admin_user):
        stock_id = make_material(
            client, admin_auth_headers,
            material_name='Low Stock Coil', property='Refill Site',
            category='HVAC', quantity=1, unit_price=80,
        )
        listed = client.get('/procurement/api/materials', headers=admin_auth_headers)
        row = next(m for m in listed.get_json()['materials'] if m['id'] == stock_id)
        cat_id = row['catalog_id']
        client.put(
            f'/procurement/api/catalog/materials/{cat_id}',
            headers=admin_auth_headers,
            json={'min_qty': 10},
        )
        refill = client.get('/procurement/api/refill', headers=admin_auth_headers).get_json()
        match = next(i for i in refill['items'] if i['name'] == 'Low Stock Coil')
        created = client.post('/procurement/api/refill/create-pr', headers=admin_auth_headers, json={
            'items': [{
                'catalog_id': match['catalog_id'],
                'property_id': match['property_id'],
                'qty': match['suggested_qty'],
                'unit_price': match['unit_price'],
            }],
            'property_id': match['property_id'],
        })
        assert created.status_code == 200, created.get_json()
        assert created.get_json()['request']['id']
        assert created.get_json()['request']['status'] in ('submitted', 'procurement_review')
        again = client.get('/procurement/api/refill', headers=admin_auth_headers).get_json()
        still = next(i for i in again['items'] if i['name'] == 'Low Stock Coil')
        assert still.get('open_pr')
        assert still['open_pr']['id'] == created.get_json()['request']['id']
        assert still.get('last_event')
        assert still.get('edited_at')

    def test_bulk_catalog_add_consumes_stock(self, app, client, admin_auth_headers, admin_user):
        from app.models import db, Ticket, TicketMaterial
        stock_id = make_material(
            client, admin_auth_headers,
            material_name='Bulk Cable', property='Bulk Site',
            category='Electrical', quantity=20, unit_price=3,
        )
        listed = client.get('/procurement/api/materials', headers=admin_auth_headers)
        row = next(m for m in listed.get_json()['materials'] if m['id'] == stock_id)
        _work_ticket(app, admin_user.id, 'TKT-PROC-BULK', 'Bulk Site')
        resp = client.post(
            '/tickets/api/tickets/TKT-PROC-BULK/materials/bulk',
            headers=admin_auth_headers,
            json={'items': [{'id': row['catalog_id'], 'name': 'Bulk Cable', 'quantity': 4, 'uom': 'PCS', 'unit_price': 3}]},
        )
        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['added'] == 1
        after = client.get('/procurement/api/materials', headers=admin_auth_headers)
        left = next(m for m in after.get_json()['materials'] if m['id'] == stock_id)
        assert left['quantity'] == 16
        with app.app_context():
            t = Ticket.query.filter_by(ticket_id='TKT-PROC-BULK').first()
            if t:
                TicketMaterial.query.filter_by(ticket_id=t.id).delete()
                db.session.delete(t)
                db.session.commit()


def _stock_qty(client, headers, stock_id):
    listed = client.get('/procurement/api/materials', headers=headers)
    return next(m for m in listed.get_json()['materials'] if m['id'] == stock_id)['quantity']


def _cleanup_ticket(app, ticket_public_id):
    from app.models import db, Ticket, TicketMaterial
    from module_procurement.models import ProcMovement
    with app.app_context():
        t = Ticket.query.filter_by(ticket_id=ticket_public_id).first()
        if not t:
            return
        TicketMaterial.query.filter_by(ticket_id=t.id).delete()
        ProcMovement.query.filter_by(ticket_id=t.id).delete()
        db.session.delete(t)
        db.session.commit()


class TestTicketSiteAndSharedCatalog:
    def test_picker_includes_site_and_shared_not_other_site(
        self, app, client, admin_auth_headers, admin_user,
    ):
        make_material(
            client, admin_auth_headers,
            material_name='Site Only Filter', property='Tower A',
            category='HVAC', quantity=5, unit_price=10,
        )
        make_material(
            client, admin_auth_headers,
            material_name='Other Tower Coil', property='Tower B',
            category='HVAC', quantity=7, unit_price=20,
        )
        make_material(
            client, admin_auth_headers,
            material_name='Shared Tape', property='Shared',
            category='Plumbing', quantity=30, unit_price=3,
        )
        _work_ticket(app, admin_user.id, 'TKT-PROC-PICKER', 'Tower A')
        resp = client.get(
            '/tickets/api/tickets/TKT-PROC-PICKER/catalog-materials',
            headers=admin_auth_headers,
        )
        assert resp.status_code == 200, resp.get_json()
        data = resp.get_json()
        assert data['success'] is True
        assert data['property']['name'] == 'Tower A'
        assert data['shared']['name'] == 'Shared'
        names = {(m['name'], m['pool']) for m in data['materials']}
        assert ('Site Only Filter', 'site') in names
        assert ('Shared Tape', 'shared') in names
        assert ('Other Tower Coil', 'site') not in names
        assert ('Other Tower Coil', 'shared') not in names
        _cleanup_ticket(app, 'TKT-PROC-PICKER')

    def test_same_sku_appears_twice_when_on_site_and_shared(
        self, app, client, admin_auth_headers, admin_user,
    ):
        site_id = make_material(
            client, admin_auth_headers,
            material_name='Dual HEPA', property='Tower A',
            category='HVAC', quantity=4, unit_price=85,
        )
        shared_id = make_material(
            client, admin_auth_headers,
            material_name='Dual HEPA', property='Shared',
            category='HVAC', quantity=8, unit_price=85,
        )
        _work_ticket(app, admin_user.id, 'TKT-PROC-DUAL', 'Tower A')
        resp = client.get(
            '/tickets/api/tickets/TKT-PROC-DUAL/catalog-materials',
            headers=admin_auth_headers,
        )
        rows = [m for m in resp.get_json()['materials'] if m['name'] == 'Dual HEPA']
        pools = {m['pool'] for m in rows}
        assert pools == {'site', 'shared'}
        assert {m['stock_id'] for m in rows} == {site_id, shared_id}
        _cleanup_ticket(app, 'TKT-PROC-DUAL')

    def test_shared_pick_deducts_shared_not_site(
        self, app, client, admin_auth_headers, admin_user,
    ):
        site_id = make_material(
            client, admin_auth_headers,
            material_name='Pool Cable', property='Tower A',
            category='Electrical', quantity=10, unit_price=5,
        )
        shared_id = make_material(
            client, admin_auth_headers,
            material_name='Pool Cable', property='Shared',
            category='Electrical', quantity=10, unit_price=5,
        )
        listed = client.get('/procurement/api/materials', headers=admin_auth_headers)
        cat_id = next(m['catalog_id'] for m in listed.get_json()['materials'] if m['id'] == shared_id)
        _work_ticket(app, admin_user.id, 'TKT-PROC-SHARED', 'Tower A')
        resp = client.post(
            '/tickets/api/tickets/TKT-PROC-SHARED/materials',
            headers=admin_auth_headers,
            json={
                'material_name': 'Pool Cable',
                'quantity': 3,
                'from_procurement': True,
                'procurement_ref': cat_id,
                'stock_id': shared_id,
                'pool': 'shared',
            },
        )
        assert resp.status_code == 200, resp.get_json()
        assert _stock_qty(client, admin_auth_headers, shared_id) == 7
        assert _stock_qty(client, admin_auth_headers, site_id) == 10
        _cleanup_ticket(app, 'TKT-PROC-SHARED')

    def test_site_pick_deducts_site_not_shared(
        self, app, client, admin_auth_headers, admin_user,
    ):
        site_id = make_material(
            client, admin_auth_headers,
            material_name='Site Cable', property='Tower A',
            category='Electrical', quantity=10, unit_price=5,
        )
        shared_id = make_material(
            client, admin_auth_headers,
            material_name='Site Cable', property='Shared',
            category='Electrical', quantity=10, unit_price=5,
        )
        listed = client.get('/procurement/api/materials', headers=admin_auth_headers)
        cat_id = next(m['catalog_id'] for m in listed.get_json()['materials'] if m['id'] == site_id)
        _work_ticket(app, admin_user.id, 'TKT-PROC-SITE', 'Tower A')
        resp = client.post(
            '/tickets/api/tickets/TKT-PROC-SITE/materials',
            headers=admin_auth_headers,
            json={
                'material_name': 'Site Cable',
                'quantity': 2,
                'from_procurement': True,
                'procurement_ref': cat_id,
                'stock_id': site_id,
                'pool': 'site',
            },
        )
        assert resp.status_code == 200, resp.get_json()
        assert _stock_qty(client, admin_auth_headers, site_id) == 8
        assert _stock_qty(client, admin_auth_headers, shared_id) == 10
        _cleanup_ticket(app, 'TKT-PROC-SITE')

    def test_delete_restores_to_issued_property(
        self, app, client, admin_auth_headers, admin_user,
    ):
        site_id = make_material(
            client, admin_auth_headers,
            material_name='Return Cable', property='Tower A',
            category='Electrical', quantity=10, unit_price=5,
        )
        shared_id = make_material(
            client, admin_auth_headers,
            material_name='Return Cable', property='Shared',
            category='Electrical', quantity=10, unit_price=5,
        )
        listed = client.get('/procurement/api/materials', headers=admin_auth_headers)
        cat_id = next(m['catalog_id'] for m in listed.get_json()['materials'] if m['id'] == shared_id)
        _work_ticket(app, admin_user.id, 'TKT-PROC-RESTORE', 'Tower A')
        added = client.post(
            '/tickets/api/tickets/TKT-PROC-RESTORE/materials',
            headers=admin_auth_headers,
            json={
                'material_name': 'Return Cable',
                'quantity': 4,
                'from_procurement': True,
                'procurement_ref': cat_id,
                'stock_id': shared_id,
                'pool': 'shared',
            },
        )
        assert added.status_code == 200, added.get_json()
        mat_id = added.get_json()['material']['id']
        gone = client.delete(
            f'/tickets/api/tickets/TKT-PROC-RESTORE/materials/{mat_id}',
            headers=admin_auth_headers,
        )
        assert gone.status_code == 200, gone.get_json()
        assert _stock_qty(client, admin_auth_headers, shared_id) == 10
        assert _stock_qty(client, admin_auth_headers, site_id) == 10
        _cleanup_ticket(app, 'TKT-PROC-RESTORE')

    def test_unmatched_property_picker_is_shared_only(
        self, app, client, admin_auth_headers, admin_user,
    ):
        make_material(
            client, admin_auth_headers,
            material_name='Tower Only Part', property='Tower A',
            category='HVAC', quantity=6, unit_price=12,
        )
        make_material(
            client, admin_auth_headers,
            material_name='Shared Spare', property='Shared',
            category='Electrical', quantity=15, unit_price=4,
        )
        _work_ticket(app, admin_user.id, 'TKT-PROC-UNKNOWN', 'Nowhere Estate')
        resp = client.get(
            '/tickets/api/tickets/TKT-PROC-UNKNOWN/catalog-materials',
            headers=admin_auth_headers,
        )
        assert resp.status_code == 200, resp.get_json()
        data = resp.get_json()
        assert data['property'] is None
        names = {(m['name'], m['pool']) for m in data['materials']}
        assert ('Shared Spare', 'shared') in names
        assert ('Tower Only Part', 'site') not in names
        listed = client.get('/procurement/api/materials', headers=admin_auth_headers)
        cat_id = next(
            m['catalog_id'] for m in listed.get_json()['materials']
            if m['material_name'] == 'Tower Only Part'
        )
        denied = client.post(
            '/tickets/api/tickets/TKT-PROC-UNKNOWN/materials',
            headers=admin_auth_headers,
            json={
                'material_name': 'Tower Only Part',
                'quantity': 1,
                'from_procurement': True,
                'procurement_ref': cat_id,
            },
        )
        assert denied.status_code == 400
        assert 'not assigned' in (denied.get_json().get('error') or '').lower()
        with app.app_context():
            from app.models import db
            from module_procurement.models import ProcProperty
            assert ProcProperty.query.filter(
                db.func.lower(ProcProperty.name) == 'nowhere estate'
            ).first() is None
        _cleanup_ticket(app, 'TKT-PROC-UNKNOWN')

    def test_add_property_links_ticket_property(self, app, client, admin_auth_headers):
        from app.models import db, TicketProperty
        from module_procurement.models import ProcProperty
        with app.app_context():
            tp = TicketProperty(name='Linked Tower', is_active=True)
            db.session.add(tp)
            db.session.commit()
            tp_id = tp.id
        resp = client.post(
            '/procurement/api/properties',
            headers=admin_auth_headers,
            json={'name': 'Linked Tower'},
        )
        assert resp.status_code == 200, resp.get_json()
        with app.app_context():
            row = ProcProperty.query.filter_by(name='Linked Tower').first()
            assert row is not None
            assert row.ticket_property_id == tp_id
            db.session.delete(row)
            tp = db.session.get(TicketProperty, tp_id)
            if tp:
                db.session.delete(tp)
            db.session.commit()


# ---------------------------------------------------------------------------
# GET /procurement/api/stock-elsewhere  POST /procurement/api/stock-share
# ---------------------------------------------------------------------------

class TestStockShare:
    def test_elsewhere_requires_auth(self, client):
        response = client.get('/procurement/api/stock-elsewhere')
        assert response.status_code in (401, 422)

    def test_share_requires_auth(self, client):
        response = client.post('/procurement/api/stock-share', json={})
        assert response.status_code in (401, 422)

    def test_denied_without_access(self, client, auth_headers):
        response = client.get('/procurement/api/stock-elsewhere', headers=auth_headers)
        assert response.status_code == 403
        response = client.post('/procurement/api/stock-share', headers=auth_headers, json={
            'stock_id': 'X', 'qty': 1,
        })
        assert response.status_code == 403

    def test_missing_fields_400(self, client, admin_auth_headers):
        response = client.post('/procurement/api/stock-share', headers=admin_auth_headers, json={})
        assert response.status_code == 400
        assert 'required' in response.get_json()['error'].lower()

    def test_unknown_material_404(self, client, admin_auth_headers):
        response = client.post('/procurement/api/stock-share', headers=admin_auth_headers, json={
            'stock_id': 'PROC-MAT-DOESNOTEXIST',
            'qty': 1,
        })
        assert response.status_code == 404

    def test_elsewhere_skips_shared_and_dest(self, client, admin_auth_headers):
        shared_id = make_material(
            client, admin_auth_headers,
            material_name='Shared Tape', property='Shared', quantity=8,
        )
        other_id = make_material(
            client, admin_auth_headers,
            material_name='Tower Coil', property='Tower North', quantity=4,
        )
        make_material(
            client, admin_auth_headers,
            material_name='Empty Bin', property='Tower East', quantity=0,
        )
        here_id = make_material(
            client, admin_auth_headers,
            material_name='Already Here', property='Injaz HQ Office', quantity=2,
        )
        listed = client.get(
            '/procurement/api/stock-elsewhere?exclude=Injaz%20HQ%20Office',
            headers=admin_auth_headers,
        )
        assert listed.status_code == 200
        data = listed.get_json()
        ids = {m['id'] for g in data['sources'] for m in g['materials']}
        assert shared_id not in ids
        assert other_id in ids
        assert here_id not in ids
        names = {m['material_name'] for g in data['sources'] for m in g['materials']}
        assert 'Empty Bin' not in names
        assert all(not g['is_shared'] for g in data['sources'])

    def test_share_visible_on_every_property(self, client, admin_auth_headers):
        stock_id = make_material(
            client, admin_auth_headers,
            material_name='Shared Cable', property='Tower North', quantity=10,
        )
        resp = client.post('/procurement/api/stock-share', headers=admin_auth_headers, json={
            'stock_id': stock_id,
            'qty': 4,
        })
        assert resp.status_code == 200, resp.get_json()
        assert _stock_qty(client, admin_auth_headers, stock_id) == 6
        hq = client.get(
            '/procurement/api/property-materials/Injaz%20HQ%20Office',
            headers=admin_auth_headers,
        ).get_json()['materials']
        shared_rows = [m for m in hq if m['material_name'] == 'Shared Cable']
        assert len(shared_rows) == 1
        assert shared_rows[0]['quantity'] == 4
        assert shared_rows[0]['pool'] == 'shared'
        other = client.get(
            '/procurement/api/property-materials/Tower%20West',
            headers=admin_auth_headers,
        ).get_json()['materials']
        assert any(m['material_name'] == 'Shared Cable' and m['pool'] == 'shared' for m in other)

    def test_share_already_shared_400(self, client, admin_auth_headers):
        stock_id = make_material(
            client, admin_auth_headers,
            material_name='Already Shared', property='Shared', quantity=3,
        )
        resp = client.post('/procurement/api/stock-share', headers=admin_auth_headers, json={
            'stock_id': stock_id,
            'qty': 1,
        })
        assert resp.status_code == 400
        assert 'already' in resp.get_json()['error'].lower()

    def test_insufficient_qty_400(self, client, admin_auth_headers):
        stock_id = make_material(
            client, admin_auth_headers,
            material_name='Scarce Filter', property='Tower West', quantity=2,
        )
        resp = client.post('/procurement/api/stock-share', headers=admin_auth_headers, json={
            'stock_id': stock_id,
            'qty': 5,
            'mode': 'site',
            'to_property': 'Injaz HQ Office',
        })
        assert resp.status_code == 200, resp.get_json()
        assert _stock_qty(client, admin_auth_headers, stock_id) == 2
        hq = client.get(
            '/procurement/api/property-materials/Injaz%20HQ%20Office',
            headers=admin_auth_headers,
        ).get_json()['materials']
        rows = [m for m in hq if m['material_name'] == 'Scarce Filter' and m.get('pool') != 'shared']
        assert len(rows) == 1
        assert rows[0]['quantity'] == 5

    def test_qty_above_on_hand_equal_increases_each_site(self, client, admin_auth_headers):
        north_id = make_material(
            client, admin_auth_headers,
            material_name='Overall Extra Pump', property='Tower North', quantity=1,
        )
        resp = client.post('/procurement/api/stock-share', headers=admin_auth_headers, json={
            'stock_id': north_id,
            'qty': 3,
            'mode': 'equal',
            'to_property': 'Injaz HQ Office',
        })
        assert resp.status_code == 200, resp.get_json()
        north = client.get(
            '/procurement/api/property-materials/Tower%20North',
            headers=admin_auth_headers,
        ).get_json()['materials']
        hq = client.get(
            '/procurement/api/property-materials/Injaz%20HQ%20Office',
            headers=admin_auth_headers,
        ).get_json()['materials']

        def site_qty(rows):
            hit = [m for m in rows if m['material_name'] == 'Overall Extra Pump' and m.get('pool') != 'shared']
            assert len(hit) == 1
            return hit[0]['quantity']

        assert site_qty(north) == 3
        assert site_qty(hq) == 3

    def test_merge_when_shared_already_has_sku(self, client, admin_auth_headers):
        source_id = make_material(
            client, admin_auth_headers,
            material_name='PVC Pipe', property='Tower South',
            category='Plumbing', quantity=5,
        )
        shared_id = make_material(
            client, admin_auth_headers,
            material_name='PVC Pipe', property='Shared',
            category='Plumbing', quantity=2,
        )
        resp = client.post('/procurement/api/stock-share', headers=admin_auth_headers, json={
            'stock_id': source_id,
            'qty': 5,
        })
        assert resp.status_code == 200, resp.get_json()
        assert _stock_qty(client, admin_auth_headers, source_id) == 0
        assert _stock_qty(client, admin_auth_headers, shared_id) == 7

    def test_share_copy_within_on_hand_keeps_source(self, client, admin_auth_headers):
        stock_id = make_material(
            client, admin_auth_headers,
            material_name='Copy Keep Pump', property='Tower North', quantity=10,
        )
        resp = client.post('/procurement/api/stock-share', headers=admin_auth_headers, json={
            'stock_id': stock_id,
            'qty': 3,
            'mode': 'copy',
            'to_property': 'Injaz HQ Office',
        })
        assert resp.status_code == 200, resp.get_json()
        assert _stock_qty(client, admin_auth_headers, stock_id) == 10
        hq = client.get(
            '/procurement/api/property-materials/Injaz%20HQ%20Office',
            headers=admin_auth_headers,
        ).get_json()['materials']
        rows = [m for m in hq if m['material_name'] == 'Copy Keep Pump' and m.get('pool') != 'shared']
        assert len(rows) == 1
        assert rows[0]['quantity'] == 3

    def test_share_copy_stacks_on_existing_dest(self, client, admin_auth_headers):
        north_id = make_material(
            client, admin_auth_headers,
            material_name='Copy Stack Valve', property='Tower North', quantity=10,
        )
        make_material(
            client, admin_auth_headers,
            material_name='Copy Stack Valve', property='Injaz HQ Office', quantity=2,
        )
        resp = client.post('/procurement/api/stock-share', headers=admin_auth_headers, json={
            'stock_id': north_id,
            'qty': 3,
            'mode': 'copy',
            'to_property': 'Injaz HQ Office',
        })
        assert resp.status_code == 200, resp.get_json()
        assert _stock_qty(client, admin_auth_headers, north_id) == 10
        hq = client.get(
            '/procurement/api/property-materials/Injaz%20HQ%20Office',
            headers=admin_auth_headers,
        ).get_json()['materials']
        rows = [m for m in hq if m['material_name'] == 'Copy Stack Valve' and m.get('pool') != 'shared']
        assert len(rows) == 1
        assert rows[0]['quantity'] == 5
        stock_id = make_material(
            client, admin_auth_headers,
            material_name='Add Within Coil', property='Tower North', quantity=10,
        )
        resp = client.post('/procurement/api/stock-share', headers=admin_auth_headers, json={
            'stock_id': stock_id,
            'qty': 3,
            'mode': 'add',
            'to_property': 'Injaz HQ Office',
        })
        assert resp.status_code == 200, resp.get_json()
        assert _stock_qty(client, admin_auth_headers, stock_id) == 7
        hq = client.get(
            '/procurement/api/property-materials/Injaz%20HQ%20Office',
            headers=admin_auth_headers,
        ).get_json()['materials']
        rows = [m for m in hq if m['material_name'] == 'Add Within Coil' and m.get('pool') != 'shared']
        assert len(rows) == 1
        assert rows[0]['quantity'] == 3

    def test_share_add_stacks_on_existing_dest(self, client, admin_auth_headers):
        north_id = make_material(
            client, admin_auth_headers,
            material_name='Add Stack Tape', property='Tower North', quantity=8,
        )
        make_material(
            client, admin_auth_headers,
            material_name='Add Stack Tape', property='Injaz HQ Office', quantity=2,
        )
        resp = client.post('/procurement/api/stock-share', headers=admin_auth_headers, json={
            'stock_id': north_id,
            'qty': 3,
            'mode': 'add',
            'to_property': 'Injaz HQ Office',
        })
        assert resp.status_code == 200, resp.get_json()
        assert _stock_qty(client, admin_auth_headers, north_id) == 5
        hq = client.get(
            '/procurement/api/property-materials/Injaz%20HQ%20Office',
            headers=admin_auth_headers,
        ).get_json()['materials']
        rows = [m for m in hq if m['material_name'] == 'Add Stack Tape' and m.get('pool') != 'shared']
        assert len(rows) == 1
        assert rows[0]['quantity'] == 5

    def test_share_add_over_on_hand_400(self, client, admin_auth_headers):
        stock_id = make_material(
            client, admin_auth_headers,
            material_name='Add Over Filter', property='Tower West', quantity=2,
        )
        resp = client.post('/procurement/api/stock-share', headers=admin_auth_headers, json={
            'stock_id': stock_id,
            'qty': 5,
            'mode': 'add',
            'to_property': 'Injaz HQ Office',
        })
        assert resp.status_code == 400
        assert 'insufficient' in resp.get_json()['error'].lower()

    def test_share_this_project_only(self, client, admin_auth_headers):
        stock_id = make_material(
            client, admin_auth_headers,
            material_name='HQ Only Coil', property='Tower North', quantity=10,
        )
        resp = client.post('/procurement/api/stock-share', headers=admin_auth_headers, json={
            'stock_id': stock_id,
            'qty': 3,
            'mode': 'site',
            'to_property': 'Injaz HQ Office',
        })
        assert resp.status_code == 200, resp.get_json()
        assert _stock_qty(client, admin_auth_headers, stock_id) == 10
        hq = client.get(
            '/procurement/api/property-materials/Injaz%20HQ%20Office',
            headers=admin_auth_headers,
        ).get_json()['materials']
        rows = [m for m in hq if m['material_name'] == 'HQ Only Coil']
        assert len(rows) == 1
        assert rows[0]['quantity'] == 3
        assert rows[0].get('pool') != 'shared'

    def test_share_equally_across_sites_that_have_sku(self, client, admin_auth_headers):
        north_id = make_material(
            client, admin_auth_headers,
            material_name='Equal Split Lamp', property='Tower North', quantity=8,
        )
        west_id = make_material(
            client, admin_auth_headers,
            material_name='Equal Split Lamp', property='Tower West', quantity=2,
        )
        resp = client.post('/procurement/api/stock-share', headers=admin_auth_headers, json={
            'stock_id': north_id,
            'qty': 6,
            'mode': 'equal',
            'to_property': 'Injaz HQ Office',
        })
        assert resp.status_code == 200, resp.get_json()
        assert _stock_qty(client, admin_auth_headers, north_id) == 6
        assert _stock_qty(client, admin_auth_headers, west_id) == 6
        hq = client.get(
            '/procurement/api/property-materials/Injaz%20HQ%20Office',
            headers=admin_auth_headers,
        ).get_json()['materials']
        rows = [m for m in hq if m['material_name'] == 'Equal Split Lamp' and m.get('pool') != 'shared']
        assert len(rows) == 1
        assert rows[0]['quantity'] == 6

    def test_add_material_share_equally_existing_sku(self, client, admin_auth_headers):
        make_material(
            client, admin_auth_headers,
            material_name='Equal Add Tape', property='Tower North', quantity=1,
        )
        make_material(
            client, admin_auth_headers,
            material_name='Equal Add Tape', property='Tower West', quantity=1,
        )
        resp = client.post('/procurement/api/materials', headers=admin_auth_headers, json={
            'material_name': 'Equal Add Tape',
            'property': 'Injaz HQ Office',
            'quantity': 6,
            'distribute': 'equal',
        })
        assert resp.status_code == 200, resp.get_json()
        hq = client.get(
            '/procurement/api/property-materials/Injaz%20HQ%20Office',
            headers=admin_auth_headers,
        ).get_json()['materials']
        north = client.get(
            '/procurement/api/property-materials/Tower%20North',
            headers=admin_auth_headers,
        ).get_json()['materials']
        west = client.get(
            '/procurement/api/property-materials/Tower%20West',
            headers=admin_auth_headers,
        ).get_json()['materials']

        def site_qty(rows):
            hit = [m for m in rows if m['material_name'] == 'Equal Add Tape' and m.get('pool') != 'shared']
            assert len(hit) == 1
            return hit[0]['quantity']

        assert site_qty(hq) == 2
        assert site_qty(north) == 3
        assert site_qty(west) == 3


