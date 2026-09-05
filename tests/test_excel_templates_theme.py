"""Kynvera Excel templates: Instructions-first, coral theme, import round-trip."""
from io import BytesIO

from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage

from common.kynvera_excel_brand import HEADER_FILL_HEX
from module_files.catalog import EXCEL_TEMPLATES
from module_files.service import build_excel_template_bytes

TEMPLATE_IDS = [entry['id'] for entry in EXCEL_TEMPLATES]


def _as_upload(data: bytes, name: str = 'template.xlsx') -> FileStorage:
    return FileStorage(
        stream=BytesIO(data),
        filename=name,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


def _header_fill_hex(cell) -> str:
    fg = cell.fill.fgColor
    rgb = getattr(fg, 'rgb', None)
    return str(rgb or '').upper()


def _sheet_has_coral_header(ws) -> bool:
    for row in range(1, 6):
        if HEADER_FILL_HEX in _header_fill_hex(ws.cell(row, 1)):
            return True
    return False


def _instructions_text(ws) -> str:
    parts = []
    for row in ws.iter_rows(max_row=80, max_col=4, values_only=True):
        for val in row:
            if val:
                parts.append(str(val))
    return ' '.join(parts).lower()


class TestExcelTemplateTheme:
    def test_catalog_has_eight_templates(self):
        assert TEMPLATE_IDS == [
            'manpower', 'leave', 'hiring', 'procurement',
            'qhsi', 'devices', 'technicians', 'locations',
        ]

    def test_each_template_starts_with_instructions_and_coral_headers(self, app):
        with app.app_context():
            for tid in TEMPLATE_IDS:
                data, filename, _display = build_excel_template_bytes(tid)
                assert filename.endswith('.xlsx'), tid
                wb = load_workbook(BytesIO(data))
                assert wb.sheetnames[0] == 'Instructions', f'{tid}: {wb.sheetnames}'
                inst_ws = wb['Instructions']
                assert inst_ws['A1'].value == 'Kynvera', tid
                assert str(inst_ws['A1'].font.color.rgb or '').upper() == 'FFFFFFFF', tid
                assert HEADER_FILL_HEX in _header_fill_hex(inst_ws['A1']), tid
                assert not getattr(inst_ws, '_images', None), tid
                inst = _instructions_text(inst_ws)
                assert 'column guide' in inst, tid
                assert 'example' in inst, tid
                data_sheets = [n for n in wb.sheetnames if n != 'Instructions']
                assert data_sheets, tid
                assert any(_sheet_has_coral_header(wb[n]) for n in data_sheets), tid
                wb.close()


class TestExcelTemplateRoundTrip:
    def test_qhsi_parser_reads_data_sheet_not_instructions(self, app):
        from module_qhsi.excel_import import parse_staff_compliance_excel

        with app.app_context():
            data, filename, _ = build_excel_template_bytes('qhsi')
        parsed = parse_staff_compliance_excel(_as_upload(data, filename))
        assert parsed['row_count'] >= 3
        names = {r['employee_name'] for r in parsed['rows']}
        assert 'Ahmed Hassan' in names

    def test_procurement_sample_has_material_name(self, app):
        from common.kynvera_excel_brand import read_import_dataframe

        with app.app_context():
            data, filename, _ = build_excel_template_bytes('procurement')
        df = read_import_dataframe(_as_upload(data, filename), preferred_sheets=('Materials',))
        cols = [str(c).strip().lower() for c in df.columns]
        assert 'material name' in cols
        assert len(df) >= 2

    def test_devices_sample_has_device_name(self, app):
        from common.kynvera_excel_brand import read_import_dataframe

        with app.app_context():
            data, filename, _ = build_excel_template_bytes('devices')
        df = read_import_dataframe(_as_upload(data, filename), preferred_sheets=('Devices',))
        cols = [str(c).strip().lower() for c in df.columns]
        assert 'device name' in cols
        assert len(df) >= 2

    def test_technicians_sheet_is_not_instructions(self, app):
        from common.kynvera_excel_brand import resolve_import_sheet_name

        with app.app_context():
            data, _, _ = build_excel_template_bytes('technicians')
        wb = load_workbook(BytesIO(data), read_only=True)
        name = resolve_import_sheet_name(wb.sheetnames, ('Technicians',))
        assert name == 'Technicians'
        headers = [c.value for c in next(wb[name].iter_rows(min_row=1, max_row=1))]
        wb.close()
        assert 'Employee ID' in headers
        assert 'Full Name' in headers

    def test_hiring_parser_reads_candidates(self, app):
        from common.kynvera_excel_brand import read_import_dataframe

        with app.app_context():
            data, filename, _ = build_excel_template_bytes('hiring')
        df = read_import_dataframe(_as_upload(data, filename), preferred_sheets=('Candidates',))
        names = [str(v).strip() for v in df['Full Name'].tolist() if v is not None]
        assert 'Sara Ahmed' in names

    def test_manpower_parser_reads_all_trades(self, app):
        from module_hr.manpower_excel import parse_all_trades_rows

        with app.app_context():
            data, filename, _ = build_excel_template_bytes('manpower')
        rows = parse_all_trades_rows(_as_upload(data, filename))
        assert rows
        assert rows[0]['trade']
        assert rows[0]['project']

    def test_manpower_export_has_project_sheets(self, app):
        from module_hr.manpower_excel import build_manpower_workbook

        class _Trade:
            def __init__(self, name, sort_order=0):
                self.name = name
                self.sort_order = sort_order

        class _Project:
            def __init__(self, name, pid, sort_order=0, active=True):
                self.id = pid
                self.name = name
                self.sort_order = sort_order
                self.active = active

        class _Vacancy:
            def __init__(self, vid, trade, project):
                self.id = vid
                self.trade = trade
                self.project = project
                self.sort_order = 0
                self.replacement_name = None
                self.replacement_employee_id = None
                self.candidate_name = None
                self.contact_number = None
                self.date_joined = None
                self.remarks = None
                self.hiring_candidate_id = None

            def normalized_status(self):
                return 'open'

            def normalized_requirement_type(self):
                return 'new'

        trade = _Trade('Electrician')
        marina = _Project('Marina Towers', 1, sort_order=10)
        askaan = _Project('Askaan', 2, sort_order=20)
        vacancies = [
            _Vacancy(1, trade, marina),
            _Vacancy(2, trade, askaan),
        ]
        with app.app_context():
            data = build_manpower_workbook(
                vacancies=vacancies,
                trades=[trade],
                projects=[marina, askaan],
                template_only=False,
            )
        wb = load_workbook(BytesIO(data), read_only=True)
        names = wb.sheetnames
        assert names[0] == 'Instructions'
        assert names[1] == 'All Trades'
        assert 'Marina Towers' in names
        assert 'Askaan' in names
        assert 'Lists' in names
        all_projects = [
            row[0].value
            for row in wb['All Trades'].iter_rows(min_row=5, max_row=6, min_col=2, max_col=2)
        ]
        marina_project = next(wb['Marina Towers'].iter_rows(min_row=5, max_row=5, min_col=2, max_col=2))[0].value
        askaan_project = next(wb['Askaan'].iter_rows(min_row=5, max_row=5, min_col=2, max_col=2))[0].value
        wb.close()
        assert set(all_projects) == {'Marina Towers', 'Askaan'}
        assert marina_project == 'Marina Towers'
        assert askaan_project == 'Askaan'

    def test_location_parser_reads_hierarchy_example(self, app):
        from module_ticketing.location_excel import parse_location_xlsx

        with app.app_context():
            data, _, _ = build_excel_template_bytes('locations')
        parsed = parse_location_xlsx(data)
        assert parsed['property']
        assert parsed['zone']
        assert parsed['sub_zone']
        assert parsed['base_unit']
        assert parsed['property'][0]['Property Name'] == 'HQ Building'

    def test_leave_log_sheet_present(self, app):
        with app.app_context():
            data, _, _ = build_excel_template_bytes('leave')
        wb = load_workbook(BytesIO(data), read_only=True)
        assert wb.sheetnames[0] == 'Instructions'
        assert 'Leave Log' in wb.sheetnames
        wb.close()
