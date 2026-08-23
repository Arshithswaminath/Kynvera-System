"""End-to-end ticketing flow: assignment ownership, images, Excel export."""
from io import BytesIO

from openpyxl import load_workbook

from app.models import (
    db, Ticket, TicketNote, TicketImage, TicketProject,
    TicketProjectTeamMember, TicketProjectVendor, TicketVendor, TicketVendorTechnician,
)

PNG_1X1 = (
    b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01'
    b'\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc\x00\x01'
    b'\x00\x00\x00\x05\x00\x01\r\n\xb4\x00\x00\x00\x00IEND\xaeB`\x82'
)


def _make_user(app, *, username, email, full_name, designation, role='user'):
    from app.models import User
    with app.app_context():
        u = User(
            username=username,
            email=email,
            full_name=full_name,
            role=role,
            designation=designation,
            is_active=True,
            password_changed=True,
            access_ticketing=True,
        )
        u.set_password('Pass1234')
        db.session.add(u)
        db.session.commit()
        return u.id


def _make_ticket(app, *, reporter_id, project, ticket_id, status='pending_supervisor',
                 supervisor_id=None, assigned_to_id=None):
    with app.app_context():
        t = Ticket(
            ticket_id=ticket_id,
            reporter_id=reporter_id,
            title='Pump leak',
            project=project,
            service_group='HVAC',
            category='AC',
            fault_type='Leak',
            priority='medium',
            work_description='Water on floor',
            status=status,
            supervisor_id=supervisor_id,
            assigned_to_id=assigned_to_id,
            property_name='Tower A',
            zone='L1',
            sub_zone='Plant',
            base_unit='AHU-1',
        )
        db.session.add(t)
        db.session.commit()
        return t.ticket_id


def _purge_ticket(ticket_public_id: str) -> None:
    t = Ticket.query.filter_by(ticket_id=ticket_public_id).first()
    if not t:
        return
    TicketNote.query.filter_by(ticket_id=t.id).delete()
    TicketImage.query.filter_by(ticket_id=t.id).delete()
    db.session.delete(t)
    db.session.commit()


class TestAssignOwnership:
    def test_project_team_user_becomes_assignee(self, client, admin_auth_headers, app, admin_user):
        tech_id = _make_user(
            app, username='flowtech', email='flowtech@example.com',
            full_name='Flow Tech', designation='technician',
        )
        with app.app_context():
            p = TicketProject(name='Flow Site', is_active=True)
            db.session.add(p)
            db.session.flush()
            db.session.add(TicketProjectTeamMember(project_id=p.id, user_id=tech_id))
            db.session.commit()
        tid = _make_ticket(
            app, reporter_id=admin_user.id, project='Flow Site',
            ticket_id='TKT-FLOW-TEAM', supervisor_id=admin_user.id,
            assigned_to_id=admin_user.id,
        )
        res = client.post(
            f'/tickets/api/tickets/{tid}/assign-technician',
            json={'technician_id': tech_id},
            headers=admin_auth_headers,
        )
        assert res.status_code == 200, res.get_json()
        with app.app_context():
            t = Ticket.query.filter_by(ticket_id=tid).first()
            assert t.assigned_to_id == tech_id
            assert t.technician_id == tech_id
            assert t.status == 'assigned'
            _purge_ticket(tid)

    def test_name_only_vendor_keeps_supervisor(self, client, admin_auth_headers, app, admin_user):
        sid = _make_user(
            app, username='flowsup', email='flowsup@example.com',
            full_name='Flow Sup', designation='supervisor',
        )
        tid = _make_ticket(
            app, reporter_id=admin_user.id, project='Vendor Site',
            ticket_id='TKT-FLOW-VEND', supervisor_id=sid, assigned_to_id=sid,
        )
        res = client.post(
            f'/tickets/api/tickets/{tid}/assign-technician',
            json={
                'technician_name': 'Omar Al Suwaidi',
                'technician_code': 'NAF-101',
                'vendor_company': 'NAFCO',
            },
            headers=admin_auth_headers,
        )
        assert res.status_code == 200, res.get_json()
        with app.app_context():
            t = Ticket.query.filter_by(ticket_id=tid).first()
            assert t.technician_id is None
            assert t.assigned_to_id == sid
            assert t.status == 'assigned'
            _purge_ticket(tid)

    def test_linked_vendor_tech_user_is_assignee(self, client, admin_auth_headers, app, admin_user):
        tech_id = _make_user(
            app, username='vendtech', email='vendtech@example.com',
            full_name='Vendor Tech', designation='technician',
        )
        with app.app_context():
            p = TicketProject(name='Linked Vendor Site', is_active=True)
            db.session.add(p)
            db.session.flush()
            v = TicketVendor(name='Linked Co', is_active=True)
            db.session.add(v)
            db.session.flush()
            db.session.add(TicketVendorTechnician(
                vendor_id=v.id, name='Vendor Tech', speciality='HVAC',
                code='L-1', user_id=tech_id,
            ))
            db.session.add(TicketProjectVendor(project_id=p.id, vendor_id=v.id))
            db.session.commit()
        tid = _make_ticket(
            app, reporter_id=admin_user.id, project='Linked Vendor Site',
            ticket_id='TKT-FLOW-LINK', supervisor_id=admin_user.id,
            assigned_to_id=admin_user.id,
        )
        res = client.post(
            f'/tickets/api/tickets/{tid}/assign-technician',
            json={'technician_id': tech_id, 'vendor_company': 'Linked Co'},
            headers=admin_auth_headers,
        )
        assert res.status_code == 200, res.get_json()
        with app.app_context():
            t = Ticket.query.filter_by(ticket_id=tid).first()
            assert t.assigned_to_id == tech_id
            assert t.technician_id == tech_id
            _purge_ticket(tid)


class TestImageDelete:
    def test_delete_image_ok_and_forbidden(self, client, admin_auth_headers, auth_headers, app, admin_user):
        tid = _make_ticket(
            app, reporter_id=admin_user.id, project='Img Site',
            ticket_id='TKT-FLOW-IMG', supervisor_id=admin_user.id,
            assigned_to_id=admin_user.id,
        )
        up = client.post(
            f'/tickets/api/tickets/{tid}/images',
            headers=admin_auth_headers,
            data={'image': (BytesIO(PNG_1X1), 'shot.png')},
            content_type='multipart/form-data',
        )
        assert up.status_code == 200, up.get_json()
        image_id = up.get_json()['image']['id']

        denied = client.delete(
            f'/tickets/api/tickets/{tid}/images/{image_id}',
            headers=auth_headers,
        )
        assert denied.status_code == 403

        gone = client.delete(
            f'/tickets/api/tickets/{tid}/images/{image_id}',
            headers=admin_auth_headers,
        )
        assert gone.status_code == 200, gone.get_json()
        with app.app_context():
            assert db.session.get(TicketImage, image_id) is None
            _purge_ticket(tid)


class TestTicketRegisterExcel:
    def test_export_filters_and_headers(self, client, admin_auth_headers, app, admin_user):
        open_id = _make_ticket(
            app, reporter_id=admin_user.id, project='Excel Park',
            ticket_id='TKT-XLS-OPEN', status='open',
        )
        closed_id = _make_ticket(
            app, reporter_id=admin_user.id, project='Excel Park',
            ticket_id='TKT-XLS-CLOSED', status='closed',
        )
        res = client.get(
            '/tickets/api/tickets/export',
            query_string={'status': 'open', 'project': 'Excel Park'},
            headers=admin_auth_headers,
        )
        assert res.status_code == 200
        assert 'spreadsheetml' in (res.mimetype or '')
        wb = load_workbook(BytesIO(res.data), data_only=True)
        ws = wb['Tickets']
        headers = [c.value for c in ws[1]]
        ids = [row[0].value for row in ws.iter_rows(min_row=2) if row[0].value]
        wb.close()
        assert headers[0] == 'Ticket ID'
        assert 'Location' in headers
        assert 'Assigned To' in headers
        assert 'TKT-XLS-OPEN' in ids
        assert 'TKT-XLS-CLOSED' not in ids
        with app.app_context():
            _purge_ticket(open_id)
            _purge_ticket(closed_id)

    def test_export_requires_ticketing_access(self, client, auth_headers):
        res = client.get('/tickets/api/tickets/export', headers=auth_headers)
        assert res.status_code == 403


def _login_headers(client, username, password='Pass1234'):
    res = client.post('/api/auth/login', json={'username': username, 'password': password})
    assert res.status_code == 200, res.get_json()
    token = res.get_json().get('access_token')
    return {'Authorization': f'Bearer {token}'}


class TestTechnicianWorkDoneUi:
    def test_work_started_shows_mark_work_done(self, client, app, admin_user):
        tech_id = _make_user(
            app, username='donetech', email='donetech@example.com',
            full_name='Done Tech', designation='technician',
        )
        tid = _make_ticket(
            app, reporter_id=admin_user.id, project='Done Site',
            ticket_id='TKT-DONE-START', status='work_started',
            supervisor_id=admin_user.id, assigned_to_id=tech_id,
        )
        with app.app_context():
            t = Ticket.query.filter_by(ticket_id=tid).first()
            t.technician_id = tech_id
            db.session.commit()

        headers = _login_headers(client, 'donetech')
        page = client.get(f'/tickets/{tid}', headers=headers)
        assert page.status_code == 200, page.data[:500]
        html = page.get_data(as_text=True)
        assert 'id="markWorkCompletedPanel"' in html
        assert 'Mark Work Completed' in html
        assert 'id="workCompletedWaitingPanel"' not in html
        with app.app_context():
            _purge_ticket(tid)

    def test_work_completed_shows_waiting_panel_for_technician(self, client, app, admin_user):
        tech_id = _make_user(
            app, username='waittech', email='waittech@example.com',
            full_name='Wait Tech', designation='technician',
        )
        tid = _make_ticket(
            app, reporter_id=admin_user.id, project='Wait Site',
            ticket_id='TKT-DONE-WAIT', status='work_started',
            supervisor_id=admin_user.id, assigned_to_id=tech_id,
        )
        with app.app_context():
            t = Ticket.query.filter_by(ticket_id=tid).first()
            t.technician_id = tech_id
            db.session.commit()

        headers = _login_headers(client, 'waittech')
        adv = client.post(
            f'/tickets/api/tickets/{tid}/advance',
            json={'resolution_notes': 'Filter replaced'},
            headers=headers,
        )
        assert adv.status_code == 200, adv.get_json()
        assert adv.get_json()['status'] == 'verification'

        page = client.get(f'/tickets/{tid}', headers=headers)
        assert page.status_code == 200, page.data[:500]
        html = page.get_data(as_text=True)
        assert 'id="markWorkCompletedPanel"' not in html
        assert 'id="workCompletedWaitingPanel"' in html
        assert 'Work Completed — sent for verification' in html
        assert 'Filter replaced' in html
        with app.app_context():
            t = Ticket.query.filter_by(ticket_id=tid).first()
            assert t.status == 'verification'
            _purge_ticket(tid)

    def test_supervisor_sees_verify_close_after_work_completed(
        self, client, app, admin_user, admin_auth_headers
    ):
        tech_id = _make_user(
            app, username='vertech', email='vertech@example.com',
            full_name='Ver Tech', designation='technician',
        )
        tid = _make_ticket(
            app, reporter_id=admin_user.id, project='Verify Site',
            ticket_id='TKT-DONE-SUP', status='verification',
            supervisor_id=admin_user.id, assigned_to_id=tech_id,
        )
        with app.app_context():
            t = Ticket.query.filter_by(ticket_id=tid).first()
            t.technician_id = tech_id
            db.session.commit()

        page = client.get(f'/tickets/{tid}', headers=admin_auth_headers)
        assert page.status_code == 200, page.data[:500]
        html = page.get_data(as_text=True)
        assert 'id="workCompletedWaitingPanel"' not in html
        assert 'Under Verification' in html
        assert 'id="beginVerificationBtn"' not in html
        assert 'Provider Verified' not in html
        assert 'id="supCloseBtnActivity"' in html
        with app.app_context():
            _purge_ticket(tid)

    def test_supervisor_verify_closes_ticket(
        self, client, app, admin_user, admin_auth_headers
    ):
        tech_id = _make_user(
            app, username='closetech', email='closetech@example.com',
            full_name='Close Tech', designation='technician',
        )
        tid = _make_ticket(
            app, reporter_id=admin_user.id, project='Close Site',
            ticket_id='TKT-DONE-CLOSE', status='verification',
            supervisor_id=admin_user.id, assigned_to_id=tech_id,
        )
        with app.app_context():
            t = Ticket.query.filter_by(ticket_id=tid).first()
            t.technician_id = tech_id
            db.session.commit()

        sig = (
            'data:image/png;base64,'
            'iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=='
        )
        res = client.post(
            f'/tickets/api/tickets/{tid}/supervisor-close',
            json={
                'signature': sig,
                'signed_by': 'Admin Closer',
                'signed_role': 'Supervisor',
                'markup_pct': 10,
                'verification_notes': 'Looks good',
            },
            headers=admin_auth_headers,
        )
        assert res.status_code == 200, res.get_json()
        payload = res.get_json()
        assert payload['success'] is True
        assert payload['status'] == 'closed'
        with app.app_context():
            t = Ticket.query.filter_by(ticket_id=tid).first()
            assert t.status == 'closed'
            assert t.closed_at is not None
            _purge_ticket(tid)


class TestLocationMapModal:
    def test_detail_renders_location_modal(self, client, admin_auth_headers, app, admin_user):
        import json
        import re

        tid = _make_ticket(
            app, reporter_id=admin_user.id, project='Map Park',
            ticket_id='TKT-LOC-MODAL', status='open',
        )
        page = client.get(f'/tickets/{tid}', headers=admin_auth_headers)
        assert page.status_code == 200, page.data[:500]
        html = page.get_data(as_text=True)
        assert 'id="ticketLocationModal"' in html
        assert 'id="ticketLocationMapHit"' in html
        assert 'id="locModalGoogle"' in html
        assert 'id="locModalApple"' in html
        match = re.search(r'window\.TICKET_LOCATION_MAP = (\{.*?\});', html)
        assert match, 'TICKET_LOCATION_MAP payload missing'
        pin = json.loads(match.group(1))
        assert pin['ticket_id'] == 'TKT-LOC-MODAL'
        assert pin['project'] == 'Map Park'
        assert pin['property_name'] == 'Tower A'
        assert pin['zone'] == 'L1'
        assert pin['sub_zone'] == 'Plant'
        assert pin['base_unit'] == 'AHU-1'
        with app.app_context():
            _purge_ticket(tid)
