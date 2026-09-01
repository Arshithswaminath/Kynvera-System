"""
Manpower Excel import: export round-trip must not use legacy A/C/F–M layout.
"""
from __future__ import annotations

from io import BytesIO

import pytest

pytest.importorskip("openpyxl")
from openpyxl import Workbook  # noqa: E402

from module_hr.manpower_excel import (  # noqa: E402
    ALL_TRADES_HEADERS,
    _parse_automated_with_project_fill,
    load_manpower_import_rows,
    parse_all_trades_rows,
)


def _export_shaped_workbook_bytes(*, with_candidate: bool = True) -> BytesIO:
    """Mirror build_manpower_workbook flat layout (headers row 4, data from 5)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "All Trades"
    ws["A1"] = "Injaaz — Manpower Requirement Tracker"
    for c, h in enumerate(ALL_TRADES_HEADERS, start=1):
        ws.cell(4, c, h)
    values = [
        "Technician",
        "Site A",
        "New",
        "",
        "",
        "John Doe" if with_candidate else "",
        "0501234567" if with_candidate else "",
        "Interviewing",
        "",
        "",
        12 if with_candidate else "",
    ]
    for c, val in enumerate(values, start=1):
        ws.cell(5, c, val)
    # Lists sheet present like real export
    wb.create_sheet("Lists")
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _legacy_automated_workbook_bytes() -> BytesIO:
    """Sparse legacy layout: Trade A, Project C, Req F, Status K (no header map)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "All Trades"
    ws["A1"] = "Automated manpower sheet"
    # No recognizable Trade/Project header row — data starts at row 5
    ws.cell(5, 1, "Electrician")
    ws.cell(5, 3, "Tower B")
    ws.cell(5, 6, "Replacement")
    ws.cell(5, 7, "Alice")
    ws.cell(5, 9, "Bob Candidate")
    ws.cell(5, 10, "055999")
    ws.cell(5, 11, "Selected")
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_legacy_parser_misreads_export_when_candidate_present():
    """Sentinel: legacy A/C/F–M layout corrupts our flat export rows."""
    buf = _export_shaped_workbook_bytes(with_candidate=True)
    bad = _parse_automated_with_project_fill(buf)
    assert bad, "legacy parser should still produce rows (triggering the old bug)"
    row = bad[0]
    assert row["project"] == "New"  # Requirement Type read as Project
    assert row["requirement_type"] != "new" or row["candidate_name"] == ""
    # Candidate name sat in F → treated as requirement type → normalized oddly
    assert row["status"] != "interviewing"  # Hiring Candidate ID 12 → status open


def test_header_parser_reads_flat_export_correctly():
    buf = _export_shaped_workbook_bytes(with_candidate=True)
    rows = parse_all_trades_rows(buf)
    assert len(rows) == 1
    row = rows[0]
    assert row["trade"] == "Technician"
    assert row["project"] == "Site A"
    assert row["requirement_type"] == "new"
    assert row["candidate_name"] == "John Doe"
    assert row["contact_number"] == "0501234567"
    assert row["status"] == "interviewing"
    assert row["hiring_candidate_id"] == "12"


def test_load_prefers_header_parser_for_export_round_trip():
    buf = _export_shaped_workbook_bytes(with_candidate=True)
    rows = load_manpower_import_rows(buf)
    assert len(rows) == 1
    row = rows[0]
    assert row["trade"] == "Technician"
    assert row["project"] == "Site A"
    assert row["requirement_type"] == "new"
    assert row["candidate_name"] == "John Doe"
    assert row["status"] == "interviewing"
    assert row["hiring_candidate_id"] == "12"


def test_load_still_accepts_legacy_automated_layout():
    buf = _legacy_automated_workbook_bytes()
    rows = load_manpower_import_rows(buf)
    assert len(rows) == 1
    row = rows[0]
    assert row["trade"] == "Electrician"
    assert row["project"] == "Tower B"
    assert row["requirement_type"] == "replacement"
    assert row["replacement_name"] == "Alice"
    assert row["candidate_name"] == "Bob Candidate"
    assert row["contact_number"] == "055999"
    assert row["status"] == "selected"
