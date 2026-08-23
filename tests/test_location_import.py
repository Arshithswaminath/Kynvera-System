"""CRM location SpreadsheetML import and location-tree payload."""

from io import BytesIO

from openpyxl import load_workbook

from app.models import Ticket, TicketBaseUnit, TicketProject, TicketProperty, TicketZone, TicketSubZone, db
from module_ticketing.location_catalog import (
    detect_kind,
    import_location_files,
    import_location_rows,
    parse_spreadsheetml,
    place_similar,
)


def _ssml(headers, rows):
    ns = 'xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet" xmlns="urn:schemas-microsoft-com:office:spreadsheet"'
    def row_xml(vals, indexed=True):
        cells = []
        for i, v in enumerate(vals, 1):
            if indexed:
                cells.append(
                    f'<Cell ss:Index="{i}"><Data ss:Type="String">{v}</Data></Cell>'
                )
            else:
                cells.append(f'<Cell><Data ss:Type="String">{v}</Data></Cell>')
        return '<Row>' + ''.join(cells) + '</Row>'
    body = row_xml(headers) + ''.join(row_xml(r) for r in rows)
    xml = (
        f'<?xml version="1.0"?><?mso-application progid="Excel.Sheet"?>'
        f'<Workbook {ns}><Worksheet ss:Name="Worksheet1"><Table>{body}</Table></Worksheet></Workbook>'
    )
    return xml.encode('utf-8')


def _empty_ssml():
    ns = 'xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet" xmlns="urn:schemas-microsoft-com:office:spreadsheet"'
    xml = (
        f'<?xml version="1.0"?><?mso-application progid="Excel.Sheet"?>'
        f'<Workbook {ns}><Worksheet ss:Name="BaseUnit"><Table><Row /></Table></Worksheet></Workbook>'
    )
    return xml.encode('utf-8')


class TestPlaceSimilar:
    def test_exact_and_al_prefix(self):
        assert place_similar('Al Nuaimiya 3', 'Al Nuaimiya 3')
        assert place_similar('Ajman Industrial', 'Ajman Industrial')

    def test_transliteration(self):
        assert place_similar('Nuiymeya 3', 'Al Nuaimiya 3')

    def test_unrelated(self):
        assert not place_similar('Emirates City', 'Al Bustan')
        assert not place_similar('', 'Ajman')


class TestParseSpreadsheet:
    def test_property_headers_and_index_cells(self):
        raw = _ssml(
            ['Property Code', 'Property Name', 'Area', 'City', 'Client Name', 'Status'],
            [['PY0001', 'Orient Tower', 'AL Bustan', 'Ajman', 'Aqaar', 'Active']],
        )
        headers, rows = parse_spreadsheetml(raw)
        assert detect_kind(headers) == 'property'
        assert rows[0]['Property Code'] == 'PY0001'
        assert rows[0]['Area'] == 'AL Bustan'

    def test_empty_base_unit(self):
        headers, rows = parse_spreadsheetml(_empty_ssml())
        assert headers == []
        assert rows == []


class TestLocationImport:
    def test_unique_property_creates_project_and_tree(self, app, admin_auth_headers, client):
        with app.app_context():
            report = import_location_rows(
                property_rows=[{
                    'Property Code': 'PY-TEST-1',
                    'Property Name': 'AQAAR Orient Tower',
                    'Area': 'AL Bustan',
                    'City': 'Ajman',
                    'Client Name': 'Aqaar Community Mangement',
                    'Status': 'Active',
                }],
                zone_rows=[{
                    'Zone Code': 'ZN-TEST-1',
                    'Zone Name': 'Orient Tower A1',
                    'Property': 'AQAAR Orient Tower',
                }],
                sub_zone_rows=[{
                    'Sub Zone Code': 'SZ-TEST-1',
                    'Sub Zone Name': 'Ground Floor',
                    'Property': 'AQAAR Orient Tower',
                    'Zone': 'Orient Tower A1',
                }],
            )
            assert report['counts']['properties_created'] == 1
            assert report['counts']['zones_created'] == 1
            assert report['counts']['sub_zones_created'] == 1
            assert report['counts']['projects_created'] == 1
            assert report['unresolved_count'] == 0
            proj = TicketProject.query.filter_by(name='Aqaar Community Mangement').first()
            assert proj is not None
            prop = TicketProperty.query.filter_by(code='PY-TEST-1').first()
            assert prop.project_id == proj.id
            assert prop.area == 'AL Bustan'
            assert 'AL Bustan' in prop.display_label()

        res = client.get('/tickets/api/settings/location-tree', headers=admin_auth_headers)
        assert res.status_code == 200
        data = res.get_json()
        assert data['success'] is True
        found = None
        for node in data['tree']:
            for p in node.get('properties') or []:
                if p.get('code') == 'PY-TEST-1':
                    found = p
        assert found is not None
        assert found['area'] == 'AL Bustan'
        assert found['label']
        assert found['code'] == 'PY-TEST-1'
        assert found['zones'][0]['code'] == 'ZN-TEST-1'
        assert found['zones'][0]['sub_zones'][0]['name'] == 'Ground Floor'

    def test_duplicate_names_keep_separate_codes(self, app):
        with app.app_context():
            import_location_rows(property_rows=[
                {
                    'Property Code': 'PY-ASK-A',
                    'Property Name': 'ASKAAN',
                    'Area': 'Al Nuaimiya 3',
                    'City': 'Ajman',
                    'Client Name': 'Askaan Properties LLC',
                    'Status': 'Active',
                },
                {
                    'Property Code': 'PY-ASK-B',
                    'Property Name': 'ASKAAN',
                    'Area': 'Emirates City',
                    'City': 'Ajman',
                    'Client Name': 'Askaan Properties LLC',
                    'Status': 'Active',
                },
            ])
            rows = TicketProperty.query.filter_by(name='ASKAAN').all()
            codes = {p.code for p in rows}
            assert 'PY-ASK-A' in codes and 'PY-ASK-B' in codes
            assert len({p.id for p in rows}) >= 2
            labels = {p.display_label() for p in rows if p.code in ('PY-ASK-A', 'PY-ASK-B')}
            assert any('Al Nuaimiya 3' in lb for lb in labels)
            assert any('Emirates City' in lb for lb in labels)

    def test_area_heuristic_attaches_subzone(self, app):
        with app.app_context():
            report = import_location_rows(
                property_rows=[
                    {
                        'Property Code': 'PY-RHA-A',
                        'Property Name': 'RHAs',
                        'Area': 'Ajman Industrial',
                        'Client Name': 'RHAS',
                        'Status': 'Active',
                    },
                    {
                        'Property Code': 'PY-RHA-B',
                        'Property Name': 'RHAs',
                        'Area': 'Emirates City',
                        'Client Name': 'RHAS',
                        'Status': 'Active',
                    },
                ],
                sub_zone_rows=[{
                    'Sub Zone Code': 'SZ-RHA-1',
                    'Sub Zone Name': 'Ajman Industrial',
                    'Property': 'RHAs',
                    'Zone': 'Conqueror Tower',
                }],
            )
            assert report['unresolved_count'] == 0
            prop = TicketProperty.query.filter_by(code='PY-RHA-A').first()
            zone = TicketZone.query.filter_by(property_id=prop.id, name='Conqueror Tower').first()
            assert zone is not None
            sz = TicketSubZone.query.filter_by(code='SZ-RHA-1').first()
            assert sz.zone_id == zone.id

    def test_ambiguous_child_not_copied_to_every_parent(self, app):
        with app.app_context():
            report = import_location_rows(
                property_rows=[
                    {
                        'Property Code': 'PY-DUP-A',
                        'Property Name': 'Twin Site',
                        'Area': 'North',
                        'Client Name': 'Twin Co',
                        'Status': 'Active',
                    },
                    {
                        'Property Code': 'PY-DUP-B',
                        'Property Name': 'Twin Site',
                        'Area': 'South',
                        'Client Name': 'Twin Co',
                        'Status': 'Active',
                    },
                ],
                sub_zone_rows=[{
                    'Sub Zone Code': 'SZ-DUP-X',
                    'Sub Zone Name': 'Lobby',
                    'Property': 'Twin Site',
                    'Zone': 'Unrelated Tower',
                }],
            )
            assert report['unresolved_count'] == 1
            assert report['unresolved'][0]['code'] == 'SZ-DUP-X'
            assert TicketSubZone.query.filter_by(code='SZ-DUP-X').first() is None
            for code in ('PY-DUP-A', 'PY-DUP-B'):
                prop = TicketProperty.query.filter_by(code=code).first()
                zones = list(prop.zones)
                assert zones == [] or all(z.name != 'Unrelated Tower' for z in zones)

    def test_empty_base_unit_file_is_noop(self, app, tmp_path):
        with app.app_context():
            empty = tmp_path / 'BaseUnit.xls'
            empty.write_bytes(_empty_ssml())
            prop = tmp_path / 'Property.xls'
            prop.write_bytes(_ssml(
                ['Property Code', 'Property Name', 'Client Name', 'Status'],
                [['PY-EMPTY-U', 'Solo Site', 'Solo Client', 'Active']],
            ))
            report = import_location_files({'property': prop, 'base_unit': empty})
            assert report['counts']['base_units_created'] == 0
            assert TicketProperty.query.filter_by(code='PY-EMPTY-U').first() is not None

    def test_upsert_by_code_does_not_duplicate(self, app):
        with app.app_context():
            payload = [{
                'Property Code': 'PY-UP-1',
                'Property Name': 'Museum',
                'Area': 'AL Bustan',
                'Client Name': 'Ajman Museum',
                'Status': 'Active',
            }]
            import_location_rows(property_rows=payload)
            import_location_rows(property_rows=[{
                **payload[0],
                'Area': 'Al Bustan 1',
            }])
            rows = TicketProperty.query.filter_by(code='PY-UP-1').all()
            assert len(rows) == 1
            assert rows[0].area == 'Al Bustan 1'


class TestLocationApi:
    def test_import_endpoint_admin_only(self, client, admin_auth_headers, auth_headers, tmp_path):
        raw = _ssml(
            ['Property Code', 'Property Name', 'Client Name', 'Status'],
            [['PY-API-1', 'API Site', 'API Client', 'Active']],
        )
        path = tmp_path / 'Property.xls'
        path.write_bytes(raw)
        denied = client.post(
            '/tickets/api/settings/location-import',
            headers=auth_headers,
            data={'property': (path.open('rb'), 'Property.xls')},
            content_type='multipart/form-data',
        )
        assert denied.status_code == 403

        with path.open('rb') as fh:
            ok = client.post(
                '/tickets/api/settings/location-import',
                headers=admin_auth_headers,
                data={'property': (fh, 'Property.xls')},
                content_type='multipart/form-data',
            )
        assert ok.status_code == 200
        body = ok.get_json()
        assert body['success'] is True
        assert body['counts']['properties_created'] >= 1

    def test_patch_property_relink(self, client, admin_auth_headers, app):
        with app.app_context():
            proj = TicketProject(name='Relink Project', client_name='Relink Co', is_active=True)
            db.session.add(proj)
            db.session.flush()
            prop = TicketProperty(name='Relink Site', code='PY-REL-1', is_active=True)
            db.session.add(prop)
            db.session.commit()
            pid, proj_id = prop.id, proj.id
        res = client.patch(
            f'/tickets/api/settings/properties/{pid}',
            headers=admin_auth_headers,
            json={'project_id': proj_id, 'area': 'Jurf', 'code': 'PY-REL-1'},
        )
        assert res.status_code == 200
        data = res.get_json()['property']
        assert data['project_id'] == proj_id
        assert data['area'] == 'Jurf'
        assert data['label'].startswith('Relink Site')

    def test_create_ticket_stores_location_fks(self, client, admin_auth_headers, app):
        with app.app_context():
            proj = TicketProject(name='WO Project', is_active=True)
            db.session.add(proj)
            db.session.flush()
            prop = TicketProperty(
                name='WO Tower', code='PY-WO-1', area='Marina',
                project_id=proj.id, is_active=True,
            )
            db.session.add(prop)
            db.session.flush()
            zone = TicketZone(name='Podium', property_id=prop.id, is_active=True)
            db.session.add(zone)
            db.session.commit()
            payload = {
                'title': 'AC leak',
                'project': 'WO Project',
                'service_group': 'HVAC',
                'category': 'Cooling',
                'fault_type': 'Leak',
                'priority': 'medium',
                'work_description': 'Water on floor',
                'property_id': prop.id,
                'zone_id': zone.id,
                'property_name': prop.display_label(),
                'zone': zone.display_label(),
            }
        res = client.post('/tickets/api/tickets', headers=admin_auth_headers, json=payload)
        assert res.status_code in (200, 201)
        body = res.get_json()
        assert body.get('success') is True
        with app.app_context():
            t = Ticket.query.filter_by(ticket_id=body['ticket_id']).first()
            assert t.property_id == payload['property_id']
            assert t.zone_id == payload['zone_id']
            assert 'Marina' in (t.property_name or '')
            db.session.delete(t)
            db.session.commit()


class TestProjectLocationWorkspace:
    def test_scoped_tree_is_project_only(self, client, admin_auth_headers, app):
        with app.app_context():
            proj = TicketProject(name='Workspace Proj A', client_name='Client A', is_active=True)
            other = TicketProject(name='Workspace Proj B', is_active=True)
            db.session.add_all([proj, other])
            db.session.flush()
            prop_a = TicketProperty(
                name='Site A', code='PY-WS-A', area='Marina',
                project_id=proj.id, is_active=True,
            )
            prop_b = TicketProperty(
                name='Site B', code='PY-WS-B',
                project_id=other.id, is_active=True,
            )
            unlinked = TicketProperty(name='Loose Site', code='PY-WS-U', is_active=True)
            db.session.add_all([prop_a, prop_b, unlinked])
            db.session.flush()
            zone = TicketZone(name='Podium', code='ZN-WS-A', property_id=prop_a.id, is_active=True)
            db.session.add(zone)
            db.session.commit()
            pid, other_id, unlinked_id = proj.id, other.id, unlinked.id

        res = client.get(
            f'/tickets/api/settings/projects/{pid}/location-tree',
            headers=admin_auth_headers,
        )
        assert res.status_code == 200
        body = res.get_json()
        assert body['success'] is True
        assert body['standalone'] is False
        assert body['project']['id'] == pid
        names = [p['name'] for p in body['properties']]
        assert names == ['Site A']
        assert body['properties'][0]['zones'][0]['name'] == 'Podium'
        assert body['counts']['property'] == 1
        assert body['counts']['zone'] == 1

        other_tree = client.get(
            f'/tickets/api/settings/projects/{other_id}/location-tree',
            headers=admin_auth_headers,
        ).get_json()
        assert [p['name'] for p in other_tree['properties']] == ['Site B']

        stand = client.get(
            '/tickets/api/settings/standalone/location-tree',
            headers=admin_auth_headers,
        )
        assert stand.status_code == 200
        stand_body = stand.get_json()
        assert stand_body['standalone'] is True
        stand_ids = {p['id'] for p in stand_body['properties']}
        assert unlinked_id in stand_ids
        assert pid not in {p.get('project_id') for p in stand_body['properties']}

        missing = client.get(
            '/tickets/api/settings/projects/999999/location-tree',
            headers=admin_auth_headers,
        )
        assert missing.status_code == 404

    def test_workspace_pages_ok(self, client, admin_auth_headers, app):
        with app.app_context():
            proj = TicketProject(name='Workspace Page Proj', client_name='Page Client', is_active=True)
            db.session.add(proj)
            db.session.commit()
            pid = proj.id

        page = client.get(
            f'/tickets/settings/locations/{pid}?from=locations',
            headers=admin_auth_headers,
        )
        assert page.status_code == 200
        html = page.get_data(as_text=True)
        assert 'Workspace Page Proj' in html
        assert 'Elements' in html
        assert 'Property' in html
        assert 'Base Unit' in html
        assert 'tab=locations' in html
        assert 'Excel template' in html
        assert '>Export</button>' in html
        assert '>Import</button>' in html
        assert 'id="locPinModal"' in html
        assert 'id="locPinLat"' in html
        assert 'id="locPinLng"' in html

        from_projects = client.get(
            f'/tickets/settings/locations/{pid}?from=projects',
            headers=admin_auth_headers,
        )
        assert from_projects.status_code == 200
        assert 'tab=projects' in from_projects.get_data(as_text=True)

        stand = client.get(
            '/tickets/settings/locations/standalone?from=locations',
            headers=admin_auth_headers,
        )
        assert stand.status_code == 200
        stand_html = stand.get_data(as_text=True)
        assert 'Unlinked properties' in stand_html
        assert 'Elements' in stand_html
        assert 'Excel template' in stand_html
        assert 'Export' in stand_html
        assert 'Import' in stand_html

        missing = client.get(
            '/tickets/settings/locations/999999',
            headers=admin_auth_headers,
        )
        assert missing.status_code == 404


class TestLocationExcel:
    def test_template_has_four_data_sheets(self, client, admin_auth_headers):
        res = client.get(
            '/tickets/api/settings/locations/excel-template',
            headers=admin_auth_headers,
        )
        assert res.status_code == 200
        assert 'spreadsheetml' in (res.mimetype or '')
        wb = load_workbook(BytesIO(res.data), read_only=True)
        names = wb.sheetnames
        headers = [c.value for c in wb['Property'][1]]
        unit_headers = [c.value for c in wb['Base Unit'][1]]
        wb.close()
        assert names == ['Instructions', 'Property', 'Zone', 'Sub Zone', 'Base Unit']
        assert 'Latitude' in headers
        assert 'Longitude' in headers
        assert 'Latitude' in unit_headers
        assert 'Longitude' in unit_headers

    def test_export_and_import_roundtrip(self, client, admin_auth_headers, app):
        with app.app_context():
            proj = TicketProject(name='Excel HQ', client_name='Kynvera', is_active=True)
            db.session.add(proj)
            db.session.flush()
            prop = TicketProperty(
                name='HQ Building', code='PY-XL-1', area='Ajman',
                city='Ajman', client_name='Kynvera', project_id=proj.id, is_active=True,
            )
            db.session.add(prop)
            db.session.flush()
            zone = TicketZone(name='First Floor', code='ZN-XL-1', property_id=prop.id, is_active=True)
            db.session.add(zone)
            db.session.flush()
            sz = TicketSubZone(name='Meeting Rooms', code='SZ-XL-1', zone_id=zone.id, is_active=True)
            db.session.add(sz)
            db.session.flush()
            unit = TicketBaseUnit(
                name='Boardroom', code='BU-XL-1', sub_zone_id=sz.id, is_active=True,
                latitude=25.4052, longitude=55.5136,
            )
            db.session.add(unit)
            db.session.commit()
            pid = proj.id

        exported = client.get(
            f'/tickets/api/settings/projects/{pid}/locations/export',
            headers=admin_auth_headers,
        )
        assert exported.status_code == 200
        wb = load_workbook(BytesIO(exported.data), data_only=True)
        units = [row[1].value for row in wb['Base Unit'].iter_rows(min_row=2) if row[1].value]
        unit_row = next(row for row in wb['Base Unit'].iter_rows(min_row=2) if row[1].value == 'Boardroom')
        wb.close()
        assert 'Boardroom' in units
        assert '25.4052' in str(unit_row[5].value)
        assert '55.5136' in str(unit_row[6].value)

        from module_ticketing.location_excel import build_location_workbook
        buf = build_location_workbook([{
            'name': 'Annex',
            'code': 'PY-XL-2',
            'area': 'Ajman',
            'city': 'Ajman',
            'client_name': 'Kynvera',
            'status': 'Active',
            'latitude': 25.4111,
            'longitude': 55.5222,
            'zones': [{
                'name': 'Ground',
                'code': 'ZN-XL-2',
                'sub_zones': [{
                    'name': 'Lobby',
                    'code': 'SZ-XL-2',
                    'base_units': [{
                        'name': 'Reception',
                        'code': 'BU-XL-2',
                        'latitude': 25.4112,
                        'longitude': 55.5223,
                    }],
                }],
            }],
        }])
        imported = client.post(
            f'/tickets/api/settings/projects/{pid}/locations/import',
            headers=admin_auth_headers,
            data={'file': (buf, 'locations.xlsx')},
            content_type='multipart/form-data',
        )
        assert imported.status_code == 200, imported.get_json()
        body = imported.get_json()
        assert body['success'] is True
        assert body['counts']['properties_created'] == 1
        assert body['counts']['base_units_created'] == 1

        with app.app_context():
            annex = TicketProperty.query.filter_by(code='PY-XL-2').first()
            assert annex is not None
            assert annex.project_id == pid
            assert annex.latitude == 25.4111
            assert annex.longitude == 55.5222
            reception = TicketBaseUnit.query.filter_by(code='BU-XL-2').first()
            assert reception is not None
            assert reception.latitude == 25.4112
            assert reception.longitude == 55.5223

    def test_xlsx_import_admin_only(self, client, auth_headers):
        from module_ticketing.location_excel import build_location_workbook
        buf = build_location_workbook()
        denied = client.post(
            '/tickets/api/settings/standalone/locations/import',
            headers=auth_headers,
            data={'file': (buf, 'location_template.xlsx')},
            content_type='multipart/form-data',
        )
        assert denied.status_code == 403

