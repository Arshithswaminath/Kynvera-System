import logging
import os
import json
import traceback
from datetime import datetime, timedelta, timezone as dt_timezone
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch, cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO
import base64
from common.utils import get_image_for_pdf
from common.datetime_utils import utc_now_naive

# Try importing PIL for better image handling
try:
    from PIL import Image as PILImage
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

logger = logging.getLogger(__name__)

# Dubai timezone offset (Gulf Standard Time, UTC+4)
DUBAI_OFFSET = timedelta(hours=4)

def get_dubai_time():
    """Get current time in Dubai timezone (GST - Gulf Standard Time, UTC+4)"""
    # Get UTC time and add 4 hours for Dubai time
    utc_now = utc_now_naive()
    dubai_time = utc_now + DUBAI_OFFSET
    return dubai_time

def format_dubai_datetime(dt=None, format_str='%Y-%m-%d %H:%M:%S'):
    """Format datetime in Dubai timezone (GST, UTC+4)"""
    if dt is None:
        dt = get_dubai_time()
    elif isinstance(dt, datetime):
        # If datetime has timezone info, convert to UTC first, then add Dubai offset
        if dt.tzinfo is not None:
            # Convert to UTC
            utc_dt = dt.astimezone(dt_timezone.utc).replace(tzinfo=None)
        else:
            # Assume UTC if naive
            utc_dt = dt
        # Add Dubai offset (UTC+4)
        dt = utc_dt + DUBAI_OFFSET
    else:
        # If not datetime, get current Dubai time
        dt = get_dubai_time()
    return dt.strftime(format_str)

# Try importing professional PDF service, fall back if unavailable
try:
    import sys
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from app.services.professional_pdf_service import (
        create_professional_pdf,
        create_header_with_logo,
        create_info_table,
        create_data_table,
        add_photo_grid,
        add_signatures_section,
        add_section_heading,
        add_item_heading,
        add_paragraph,
        get_professional_styles
    )
    USE_PROFESSIONAL_PDF = True
    logger.info("✅ Professional PDF service loaded successfully")
except Exception as e:
    logger.warning(f"⚠️ Professional PDF service not available: {e}. Using basic PDF generation.")
    USE_PROFESSIONAL_PDF = False

def _write_materials_sheet(ws, materials):
    """Write a 'Materials Used' sheet - matches reference format HVAC_MEP_Injaaz_*.xlsx."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill  = PatternFill('solid', fgColor='125435')
    alt_fill     = PatternFill('solid', fgColor='E3F2FD')  # Light blue for zebra striping
    header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    body_font    = Font(name='Calibri', size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right_align  = Alignment(horizontal='right',  vertical='center', wrap_text=True)
    thin         = Side(style='thin', color='BBDEFB')
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['#', 'Material Name', 'Brand', 'Department', 'UOM', 'Quantity', 'Unit Price (AED)', 'Line Total (AED)']
    # Column widths: A narrow, B widest, G/H wide, C-F medium
    col_widths = [6, 42, 16, 14, 10, 10, 18, 20]

    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "MATERIALS & COST BREAKDOWN"
    title_cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=13)
    title_cell.alignment = center_align
    title_cell.fill = header_fill
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:H2')
    sub_cell = ws['A2']
    sub_cell.value = "Selected inspection materials with pricing and cost totals"
    sub_cell.font = Font(name='Calibri', bold=False, color='125435', size=10)
    sub_cell.alignment = center_align
    sub_cell.fill = PatternFill('solid', fgColor='E8F5E9')
    ws.row_dimensions[2].height = 22

    header_row = 4
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    grand_total = 0.0
    data_start_row = header_row + 1
    for idx, m in enumerate(materials, 1):
        row_idx = data_start_row + idx - 1
        row_fill = PatternFill('solid', fgColor='FFFFFF') if row_idx % 2 == 0 else alt_fill
        qty = float(m.get('quantity', 1) or 0)
        unit_price = float(m.get('unit_price', 0) or 0)
        line_total = qty * unit_price
        grand_total += line_total
        row_data = [
            idx,
            str(m.get('name', '')),
            str(m.get('brand', '') or ''),
            str(m.get('department', '') or ''),
            str(m.get('uom', '') or ''),
            qty,
            unit_price,
            line_total,
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = body_font
            cell.fill      = row_fill
            cell.border    = border
            cell.alignment = center_align if col_idx in (1, 6) else (right_align if col_idx in (7, 8) else left_align)
            if col_idx in (7, 8):
                cell.number_format = '#,##0.00'
        ws.row_dimensions[row_idx].height = 22

    total_row = data_start_row + len(materials)
    for col_idx in range(1, 9):
        cell = ws.cell(row=total_row, column=col_idx, value='' if col_idx < 7 else None)
        cell.font = Font(name='Calibri', bold=True, size=10)
        cell.fill = PatternFill('solid', fgColor='E8F5E9')
        cell.border = border
        cell.alignment = right_align if col_idx in (7, 8) else center_align
    grand_total_cell = ws.cell(row=total_row, column=7, value='Grand Total (AED)')
    grand_total_cell.alignment = right_align
    total_value_cell = ws.cell(row=total_row, column=8, value=grand_total)
    total_value_cell.alignment = right_align
    total_value_cell.number_format = '#,##0.00'
    ws.row_dimensions[total_row].height = 24

    data_end_row = total_row - 1
    if data_end_row >= data_start_row:
        ws.auto_filter.ref = f"A{header_row}:H{data_end_row}"
    ws.freeze_panes = f"A{data_start_row}"


def _add_summary_kpi_cards(ws, start_row, cards):
    """Add KPI card band (4 cards, 2 columns each) to summary sheet."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    thin = Side(style='thin', color='D0D7DE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[start_row].height = 20
    ws.row_dimensions[start_row + 1].height = 28

    groups = [(1, 2), (3, 4), (5, 6), (7, 8)]
    for idx, (c1, c2) in enumerate(groups):
        if idx >= len(cards):
            break
        label, value = cards[idx]
        ws.merge_cells(start_row=start_row, start_column=c1, end_row=start_row, end_column=c2)
        ws.merge_cells(start_row=start_row + 1, start_column=c1, end_row=start_row + 1, end_column=c2)

        lbl_cell = ws.cell(row=start_row, column=c1, value=label)
        lbl_cell.font = Font(name='Calibri', size=9, bold=True, color='125435')
        lbl_cell.alignment = Alignment(horizontal='center', vertical='center')
        lbl_cell.fill = PatternFill('solid', fgColor='E8F5E9')
        lbl_cell.border = border

        val_cell = ws.cell(row=start_row + 1, column=c1, value=value)
        val_cell.font = Font(name='Calibri', size=13, bold=True, color='0F172A')
        val_cell.alignment = Alignment(horizontal='center', vertical='center')
        val_cell.fill = PatternFill('solid', fgColor='FFFFFF')
        val_cell.border = border

        ws.cell(row=start_row, column=c2).fill = PatternFill('solid', fgColor='E8F5E9')
        ws.cell(row=start_row, column=c2).border = border
        ws.cell(row=start_row + 1, column=c2).fill = PatternFill('solid', fgColor='FFFFFF')
        ws.cell(row=start_row + 1, column=c2).border = border

    return start_row + 3


def create_excel_report(data, output_dir):
    """Generate HVAC/MEP Excel report with professional formatting."""
    try:
        # Import professional Excel service
        sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from app.services.professional_excel_service import (
            create_professional_excel_workbook,
            add_logo_and_title,
            add_info_section,
            add_data_table,
            add_section_header,
            finalize_workbook,
            recenter_logo
        )
        
        logger.info(f"Creating professional Excel report in {output_dir}")
        
        # Generate filename
        site_name = data.get('site_name', 'Unknown_Site').replace(' ', '_')
        timestamp = get_dubai_time().strftime('%Y%m%d_%H%M%S')
        excel_filename = f"HVAC_MEP_{site_name}_{timestamp}.xlsx"
        excel_path = os.path.join(output_dir, excel_filename)
        
        # Core data
        items = data.get('items', [])
        materials = data.get('materials_required', []) if isinstance(data.get('materials_required', []), list) else []
        materials_total = 0.0
        for m in materials:
            try:
                materials_total += float(m.get('quantity', 1) or 0) * float(m.get('unit_price', 0) or 0)
            except Exception:
                pass
        priced_materials = sum(
            1 for m in materials
            if (str(m.get('unit_price', '')).strip() not in ('', '0', '0.0', '0.00'))
        )

        # Sheet 1: Executive Summary
        wb, ws_summary = create_professional_excel_workbook(
            title="HVAC & MEP Inspection Report",
            sheet_name="Summary"
        )
        current_row = add_logo_and_title(
            ws_summary,
            title="HVAC & MEP INSPECTION REPORT",
            subtitle=f"Site: {data.get('site_name', 'N/A')}",
            max_columns=8
        )
        summary_info = [
            ('Site Name', data.get('site_name', 'N/A')),
            ('Visit Date', data.get('visit_date', 'N/A')),
            ('Report Generated', format_dubai_datetime() + ' (GST)'),
            ('Inspection Items', str(len(items))),
            ('Selected Materials', str(len(materials))),
            ('Materials Total (AED)', f"{materials_total:,.2f}")
        ]
        current_row = add_info_section(ws_summary, summary_info, current_row, title="Executive Overview", max_columns=8)
        current_row = _add_summary_kpi_cards(ws_summary, current_row, [
            ("Total Items", len(items)),
            ("Total Materials", len(materials)),
            ("Priced Materials", priced_materials),
            ("Materials Value (AED)", f"{materials_total:,.2f}")
        ])
        ws_summary.freeze_panes = "A8"
        finalize_workbook(ws_summary)
        # Lock column widths to match reference format exactly
        for _cl, _w in zip('ABCDEFGH', [23, 32, 17, 10, 18, 10, 23, 10]):
            ws_summary.column_dimensions[_cl].width = _w
        recenter_logo(ws_summary)

        # Sheet 2: Inspection Items
        ws_items = wb.create_sheet(title="Inspection Items")
        items_row = add_logo_and_title(
            ws_items,
            title="HVAC & MEP - INSPECTION ITEMS",
            subtitle=f"Site: {data.get('site_name', 'N/A')}",
            max_columns=8
        )
        items_row = add_info_section(
            ws_items,
            [
                ('Inspector Scope', 'Detailed asset-wise inspection checklist'),
                ('Items Included', str(len(items))),
            ],
            items_row,
            title="Inspection Register",
            max_columns=8
        )
        headers = ['#', 'Asset', 'System', 'Description', 'Quantity', 'Brand', 'Specification', 'Comments']
        table_data = []
        for idx, item in enumerate(items, 1):
            table_data.append([
                idx,
                item.get('asset', 'N/A'),
                item.get('system', 'N/A'),
                item.get('description', 'N/A'),
                float(item.get('quantity', 0) or 0),
                item.get('brand', 'N/A'),
                item.get('specification', 'N/A'),
                item.get('comments', 'N/A')
            ])
        col_widths = {
            'A': 6,   # #
            'B': 18,  # Asset
            'C': 18,  # System
            'D': 30,  # Description
            'E': 10,  # Quantity
            'F': 16,  # Brand
            'G': 22,  # Specification
            'H': 30   # Comments
        }
        table_start_row = add_section_header(ws_items, "Item-wise Findings", items_row, span_columns=8)
        add_data_table(ws_items, headers, table_data, table_start_row, title=None, col_widths=col_widths)
        if table_data:
            header_row = table_start_row
            last_row = header_row + len(table_data)
            ws_items.auto_filter.ref = f"A{header_row}:H{last_row}"
            ws_items.freeze_panes = f"A{header_row + 1}"
        finalize_workbook(ws_items)
        # Lock column widths to match reference format exactly
        for _cl, _w in zip('ABCDEFGH', [16, 47, 13, 47, 15, 14, 20, 31]):
            ws_items.column_dimensions[_cl].width = _w
        recenter_logo(ws_items)

        # Sheet 3: Materials & Cost
        ws_mat = wb.create_sheet(title="Materials & Cost")
        if materials:
            _write_materials_sheet(ws_mat, materials)
        else:
            from openpyxl.styles import Font, Alignment, PatternFill
            ws_mat.merge_cells('A1:H1')
            ws_mat['A1'] = "MATERIALS & COST BREAKDOWN"
            ws_mat['A1'].font = Font(name='Calibri', bold=True, color='125435', size=13)
            ws_mat['A1'].fill = PatternFill('solid', fgColor='E8F5E9')
            ws_mat['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws_mat.merge_cells('A3:H3')
            ws_mat['A3'] = "No materials were selected for this inspection."
        finalize_workbook(ws_mat)
        # Lock column widths to match reference format exactly
        for _cl, _w in zip('ABCDEFGH', [13, 25, 15, 17, 11, 15, 24, 25]):
            ws_mat.column_dimensions[_cl].width = _w

        # Save workbook
        wb.save(excel_path)
        
        if not os.path.exists(excel_path):
            raise Exception(f"Excel file not created at {excel_path}")
        
        logger.info(f"✅ Professional Excel report created: {excel_path}")
        return excel_path
        
    except Exception as e:
        logger.error(f"❌ Excel generation error: {str(e)}")
        raise

def create_pdf_report(data, output_dir):
    """
    Fire Systems inspection PDF — Service Report page 1 + page 2
    (photos + workflow signatures). Shares the ticketing Service Report template.
    """
    try:
        logger.info("Creating Fire Systems inspection Service Report PDF in %s", output_dir)
        os.makedirs(output_dir, exist_ok=True)

        site_name = (data.get('site_name') or 'Unknown_Site').replace(' ', '_')
        timestamp = get_dubai_time().strftime('%Y%m%d_%H%M%S')
        pdf_filename = f"Fire_Systems_{site_name}_{timestamp}.pdf"
        pdf_path = os.path.join(output_dir, pdf_filename)

        from module_hvac_mep.inspection_service_report import build_inspection_service_report_pdf
        build_inspection_service_report_pdf(data, pdf_path)

        logger.info("✅ Fire Systems inspection PDF created: %s", pdf_path)
        return pdf_path
    except Exception as e:
        logger.error("❌ PDF generation error: %s", e)
        raise

